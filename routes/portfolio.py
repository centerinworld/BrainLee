"""
routes/portfolio.py — 보유종목(Portfolio) API

  GET    /api/portfolio
  POST   /api/portfolio/sync-kis
  PATCH  /api/portfolio/{stock_code}/bought-at
  GET    /api/portfolio/transactions
  POST   /api/portfolio/transaction
  POST   /api/portfolio/kakao-parse
  PUT    /api/portfolio/{stock_code}
  DELETE /api/portfolio/{stock_code}
  GET    /api/portfolio/export/excel
  POST   /api/portfolio/import/excel

공개 헬퍼 (scheduler.py 에서 직접 호출):
  sync_kis_executions(db)  → int
  save_portfolio_snapshot(db)  → None
"""

import io
import logging
import sqlite3 as _sl
from datetime import datetime, date as _date_cls

import crud, models
from database import get_db
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)
router = APIRouter()


# ── 지연 임포트 헬퍼 (순환 참조 방지) ──────────────────────────
def _main_collecting() -> dict:
    import main as _m
    return _m._collecting


def _main_bg_ondemand():
    import main as _m
    return _m._bg_ondemand


# ── 수급 단위 환산 (억원) ────────────────────────────────────────
# price_history 컬럼 정책:
#   inst_net_buy_amt / frn_net_buy_amt  → 백만원 단위 (KIS 수집, ÷100 = 억원)
#   inst_net_buy / frn_net_buy          → 수량(주) 단위 (KIS/Naver 수집)
def _to_억(qty, amt, close):
    qty = qty or 0; amt = amt or 0; close = close or 0
    if amt != 0:
        return amt / 100            # KIS amt: 백만원 → 억원
    if qty != 0 and close > 0:
        return qty * close / 1e8   # 수량(주) × 종가(원) → 억원
    return 0.0


def _bavg(rows, s, e, col=1):
    vals = [r[col] or 0 for r in rows[s:e]]
    return sum(vals) / len(vals) if vals else 0


# ═══════════════════════════════════════════════════════
#  스케줄러·API 공용 헬퍼
# ═══════════════════════════════════════════════════════

def sync_kis_executions(db) -> int:
    """
    KIS 당일 체결내역을 조회하여 포트폴리오에 반영.

    평균단가 재계산 원칙 (owner='이효준', broker='KIS' 포지션만 업데이트):
      매도: new_avg = (보유총액 - 매도수량×매도가) / 잔여수량
      매수: new_avg = (보유총액 + 매수수량×매수가) / 총수량
    """
    try:
        from kis_client import kis_client
        executions = kis_client.get_today_executions()
        if not executions:
            return 0
        updated = 0
        today_str = datetime.now().strftime("%Y-%m-%d")
        for ex in executions:
            code  = ex["stock_code"]
            name  = ex["stock_name"]
            qty   = float(ex["quantity"])
            price = float(ex["price"])
            tx    = ex["tx_type"]  # "buy" | "sell"
            if not code or qty == 0 or price == 0:
                continue

            # 중복 체결 방지
            tx_time = ex.get("tx_time", "")
            existing_tx = db.query(models.PortfolioTx).filter(
                models.PortfolioTx.stock_code == code,
                models.PortfolioTx.tx_date   >= datetime.strptime(today_str, "%Y-%m-%d"),
                models.PortfolioTx.quantity  == qty,
                models.PortfolioTx.price     == price,
                models.PortfolioTx.tx_type   == tx,
                models.PortfolioTx.memo      == f"KIS_{tx_time}",
            ).first()
            if existing_tx:
                continue

            holding = db.query(models.Portfolio).filter(
                models.Portfolio.stock_code == code,
                models.Portfolio.broker     == "KIS",
                models.Portfolio.owner      == "이효준",
            ).first()

            if tx == "sell":
                if holding and holding.quantity > 0:
                    remain_qty = holding.quantity - qty
                    if remain_qty > 0:
                        # avg_price는 변경하지 않음 (매도 시 잔여 주식의 평균 매입가는 그대로)
                        holding.quantity = remain_qty
                    else:
                        holding.quantity  = 0.0
                        holding.avg_price = 0.0
                    logger.info(f"[KIS매도] {code} {qty}주@{price:,.0f}원 → 잔여 {holding.quantity}주 평단 {holding.avg_price:,.0f}원")

            elif tx == "buy":
                if holding:
                    new_qty = holding.quantity + qty
                    holding.avg_price = round((holding.avg_price * holding.quantity + price * qty) / new_qty, 2)
                    holding.quantity  = new_qty
                else:
                    holding = models.Portfolio(
                        stock_code=code, stock_name=name,
                        quantity=qty, avg_price=price,
                        broker="KIS", owner="이효준", source="kis",
                    )
                    db.add(holding)
                logger.info(f"[KIS매수] {code} {qty}주@{price:,.0f}원 → 보유 {holding.quantity}주 평단 {holding.avg_price:,.0f}원")

            db.add(models.PortfolioTx(
                stock_code=code, stock_name=name,
                tx_type=tx, quantity=qty, price=price,
                tx_date=datetime.now(), memo=f"KIS_{tx_time}",
            ))
            updated += 1

        db.commit()
        logger.info(f"[KIS체결] {updated}건 포트폴리오 반영")
        return updated
    except Exception as e:
        logger.error(f"[KIS체결] 동기화 오류: {e}")
        db.rollback()
        return 0


def save_portfolio_snapshot(db) -> None:
    """장 마감 후 포트폴리오 평가총액 스냅샷 저장."""
    try:
        today_str = _date_cls.today().isoformat()
        holdings  = db.query(models.Portfolio).filter(models.Portfolio.quantity > 0).all()
        total_eval = 0.0
        for h in holdings:
            price_row = db.query(models.PriceHistory).filter(
                models.PriceHistory.stock_code == h.stock_code
            ).order_by(models.PriceHistory.date.desc()).first()
            cur_price = price_row.close if price_row else h.avg_price
            eval_amt  = round(cur_price   * h.quantity)
            cost_amt  = round(h.avg_price * h.quantity)
            profit    = eval_amt - cost_amt
            pct       = round(profit / cost_amt * 100, 2) if cost_amt else 0.0
            total_eval += eval_amt

            existing = db.query(models.PortfolioSnapshot).filter(
                models.PortfolioSnapshot.snapshot_date == today_str,
                models.PortfolioSnapshot.stock_code    == h.stock_code,
            ).first()
            if existing:
                existing.close_price = cur_price
                existing.quantity    = h.quantity
                existing.avg_price   = h.avg_price
                existing.eval_amount = eval_amt
                existing.profit_amt  = profit
                existing.profit_pct  = pct
            else:
                db.add(models.PortfolioSnapshot(
                    snapshot_date=today_str,
                    stock_code=h.stock_code, stock_name=h.stock_name,
                    close_price=cur_price, quantity=h.quantity, avg_price=h.avg_price,
                    eval_amount=eval_amt, profit_amt=profit, profit_pct=pct,
                ))
        db.commit()
        logger.info(f"[스냅샷] {today_str} 평가총액={total_eval:,.0f}원 저장")
    except Exception as e:
        logger.error(f"[스냅샷] 저장 오류: {e}")
        db.rollback()


# ═══════════════════════════════════════════════════════
#  엔드포인트
# ═══════════════════════════════════════════════════════

@router.get("")
def get_portfolio(db: Session = Depends(get_db)):
    """
    보유종목 목록.
    - 동일 종목은 보유자/증권사 무관하게 합산하여 1행 반환
    - 가중평균단가 = Σ(avg_price × quantity) / Σquantity
    - 평가액 내림차순 정렬
    """
    from sqlalchemy import func as _sqlfunc
    collecting = _main_collecting()
    bg_ondemand = _main_bg_ondemand()

    holdings = db.query(models.Portfolio).filter(models.Portfolio.quantity > 0).all()

    # 직전 거래일 스냅샷 조회
    today_iso = _date_cls.today().isoformat()
    last_snap_date = db.query(_sqlfunc.max(models.PortfolioSnapshot.snapshot_date)).filter(
        models.PortfolioSnapshot.snapshot_date < today_iso
    ).scalar()
    prev_snaps = {
        s.stock_code: s for s in
        db.query(models.PortfolioSnapshot).filter(
            models.PortfolioSnapshot.snapshot_date == last_snap_date
        ).all()
    } if last_snap_date else {}

    # stock_universe에서 종목명 보완용 맵 (NULL stock_name 대비)
    _name_map: dict = {}
    try:
        _conn = _sl.connect("stock.db")
        for _r in _conn.execute("SELECT stock_code, stock_name FROM stock_universe").fetchall():
            _name_map[_r[0]] = _r[1]
        _conn.close()
    except Exception:
        pass

    # 종목코드별 합산
    merged: dict = {}
    for h in holdings:
        code = h.stock_code or h.stock_name
        if code not in merged:
            # stock_name이 None이면 stock_universe에서 보완
            resolved_name = h.stock_name or _name_map.get(h.stock_code) or h.stock_code
            merged[code] = {
                "stock_code": h.stock_code,
                "stock_name": resolved_name,
                "sector":     h.sector or "기타",
                "total_qty":  0.0,
                "total_cost": 0.0,
                "bought_at":  h.bought_at,
            }
        merged[code]["total_qty"]  += h.quantity
        merged[code]["total_cost"] += h.avg_price * h.quantity
        if h.bought_at:
            cur_d = merged[code]["bought_at"]
            if not cur_d or h.bought_at < cur_d:
                merged[code]["bought_at"] = h.bought_at

    result = []
    for code, m in merged.items():
        total_qty  = m["total_qty"]
        total_cost = m["total_cost"]
        avg_price  = round(total_cost / total_qty, 2) if total_qty else 0.0

        price_row = db.query(models.PriceHistory).filter(
            models.PriceHistory.stock_code == m["stock_code"],
            models.PriceHistory.close > 0,
        ).order_by(models.PriceHistory.date.desc()).first()

        # 전 거래일 종가: 오늘 날짜보다 이전 날짜의 최신 레코드 (intraday 혼용 방지)
        prev_row = db.query(models.PriceHistory).filter(
            models.PriceHistory.stock_code == m["stock_code"],
            models.PriceHistory.close > 0,
            models.PriceHistory.date < today_iso,
        ).order_by(models.PriceHistory.date.desc()).first()

        has_price     = price_row is not None
        current_price = price_row.close if price_row else avg_price
        prev_price    = prev_row.close  if prev_row  else current_price

        change_pct  = round((current_price - prev_price) / prev_price * 100, 2) if prev_price else 0.0
        buy_total   = round(total_cost)
        total_value = round(current_price * total_qty)
        profit      = total_value - buy_total
        profit_pct  = round(profit / buy_total * 100, 2) if buy_total else 0.0

        # 전일대비 손익: (현재가 - 전일종가) × 보유수량
        # 스냅샷이 아닌 가격 기반으로 계산 → 당일 신규 편입 종목도 정확히 반영
        daily_profit = round((current_price - prev_price) * total_qty)

        # 주가 없는 종목 자동 수집 트리거
        if not has_price and m["stock_code"] and m["stock_code"] not in collecting:
            import threading as _threading
            _threading.Thread(target=bg_ondemand, args=(m["stock_code"],), daemon=True).start()

        # ── 5일 수급 합계 ────────────────────────────────────────────
        frn_net = inst_net = None
        sc = m["stock_code"]
        try:
            conn = _sl.connect("stock.db")
            sups = conn.execute(
                "SELECT close, frn_net_buy, inst_net_buy, frn_net_buy_amt, inst_net_buy_amt "
                "FROM price_history WHERE stock_code=? AND close>0 "
                "ORDER BY date DESC LIMIT 10", (sc,)
            ).fetchall()
            conn.close()
            # 수급 데이터가 있는 행만 최신 5일 추출
            valid_sup = [r for r in sups if (r[1] or 0) or (r[2] or 0) or (r[3] or 0) or (r[4] or 0)][:5]
            if valid_sup:
                frn_net  = round(sum(_to_억(r[1], r[3], r[0]) for r in valid_sup), 1)
                inst_net = round(sum(_to_억(r[2], r[4], r[0]) for r in valid_sup), 1)
        except Exception:
            pass

        # ── 대차잔고 (buy_candidates/short-sell과 동일 계산) ─────────
        short_data = None
        try:
            conn = _sl.connect("stock.db")
            srows = conn.execute(
                "SELECT bas_dt, borrow_bal_qty FROM short_sell_daily "
                "WHERE stock_code=? AND borrow_bal_qty IS NOT NULL "
                "ORDER BY bas_dt DESC LIMIT 30", (sc,)
            ).fetchall()
            conn.close()
            if len(srows) >= 5:
                def _savg(rs, s, e):
                    vals = [r[1] for r in rs[s:e] if r[1]]
                    return sum(vals) / len(vals) if vals else 0
                today_val  = srows[0][1] or 0
                avg5       = _savg(srows, 0, 5)
                avg5_prev  = _savg(srows, 5, 10)
                avg10      = _savg(srows, 0, 10)
                avg10_prev = _savg(srows, 10, 20)
                short_data = {
                    # 개별 종목 /short-sell API와 동일 필드명
                    "today":       today_val,
                    "avg5":        avg5,
                    "avg5_prev":   avg5_prev,
                    "avg10":       avg10,
                    "avg10_prev":  avg10_prev,
                    "today_signal": "green" if today_val < avg5      else "red",
                    "week_signal":  "green" if avg5      < avg5_prev else "red",
                    "latest_date":  srows[0][0] if srows else None,
                }
        except Exception:
            pass

        # ── 4분면 매매 신호 (추세 × 가치) ───────────────────────────
        # 추세 점수 : +3(완전정배열) ~ -3(완전역배열)
        # 가치 점수 : PBR/PER/ROE 기반  (양수=저평가, 음수=고평가)
        # 판단 매트릭스:
        #   추세+ / 가치+  →  추가매수
        #   추세+ / 가치-  →  익절고려 (오르고 있으나 이미 비쌈)
        #   추세- / 가치+  →  홀딩유지 (가치가 받쳐줌)
        #   추세- / 가치-  →  진매도   (추세도 나쁘고 비쌈)
        #   손절선 도달    →  손절 (최우선)
        trade_signal  = "hold"
        trade_reason  = ""
        trend_score   = 0
        val_score     = 0
        val_detail    = []
        trend_detail  = []
        try:
            conn = _sl.connect("stock.db")
            ph = conn.execute(
                "SELECT close, volume, inst_net_buy, frn_net_buy, "
                "       inst_net_buy_amt, frn_net_buy_amt "
                "FROM price_history WHERE stock_code=? AND close>0 "
                "ORDER BY date DESC LIMIT 270", (sc,)
            ).fetchall()
            # 가치 지표 (stock_universe)
            val_row = conn.execute(
                "SELECT pbr, per, roe, roa FROM stock_universe "
                "WHERE stock_code=? ORDER BY updated_at DESC LIMIT 1", (sc,)
            ).fetchone()
            conn.close()

            if len(ph) >= 20:
                closes = [r[0] for r in ph]
                vols   = [r[1] or 0 for r in ph]
                curr   = closes[0]
                ma5    = sum(closes[:5])   / 5
                ma20   = sum(closes[:20])  / 20
                ma60   = sum(closes[:60])  / 60  if len(closes) >= 60  else ma20
                ma120  = sum(closes[:120]) / 120 if len(closes) >= 120 else ma60
                ma200  = sum(closes[:200]) / 200 if len(closes) >= 200 else ma120
                high52 = max(closes[:252]  if len(closes) >= 252 else closes)
                from_high = (curr - high52) / high52 * 100

                # ATR(14) 기반 손절선
                trs  = [abs(closes[i] - closes[i+1]) for i in range(min(14, len(closes)-1))]
                atr  = sum(trs) / len(trs) if trs else curr * 0.02
                stop_line = curr - 2 * atr

                # RSI(14)
                gains, losses = [], []
                for i in range(1, min(15, len(closes))):
                    d = closes[i-1] - closes[i]
                    (gains if d > 0 else losses).append(abs(d))
                ag  = sum(gains)  / len(gains)  if gains  else 0
                al  = sum(losses) / len(losses) if losses else 1
                rsi = round(100 - 100 / (1 + ag / al)) if al > 0 else 50

                # MACD(12/26)
                def _ema(data, n):
                    k = 2/(n+1); e = data[0]
                    for v in data[1:]: e = v*k + e*(1-k)
                    return e
                c_asc  = list(reversed(closes))
                macd_v = (_ema(c_asc[-12:], 12) - _ema(c_asc[-26:], 26)) if len(c_asc) >= 35 else 0

                avg_vol   = sum(vols[1:21]) / 20 if len(vols) >= 21 else 1
                vol_surge = vols[0] > avg_vol * 1.8

                # 5일 수급 (수량→억원 환산)
                sup5 = ph[:5]
                inst5_억 = sum(_to_억(r[2], r[4], r[0]) for r in sup5)
                frn5_억  = sum(_to_억(r[3], r[5], r[0]) for r in sup5)
                supply_pos = inst5_억 > 0 and frn5_억 > 0
                supply_neg = inst5_억 < 0 and frn5_억 < 0
                lend_surge = short_data and short_data["avg5"] > short_data.get("avg5_prev", 0) * 1.15

                pnl = profit_pct

                # ── 추세 점수 ───────────────────────────────────────
                if curr > ma5 > ma20 > ma60 > ma120:
                    trend_score = 4; trend_detail.append("완전정배열")
                elif curr > ma5 > ma20 > ma60:
                    trend_score = 3; trend_detail.append("정배열")
                elif curr > ma20 > ma60:
                    trend_score = 2; trend_detail.append("중기정배열")
                elif curr > ma20:
                    trend_score = 1; trend_detail.append("MA20위")
                elif curr < ma5 < ma20 < ma60 < ma120:
                    trend_score = -4; trend_detail.append("완전역배열")
                elif curr < ma5 < ma20 < ma60:
                    trend_score = -3; trend_detail.append("역배열")
                elif curr < ma20 < ma60:
                    trend_score = -2; trend_detail.append("중기역배열")
                elif curr < ma20:
                    trend_score = -1; trend_detail.append("MA20이탈")
                else:
                    trend_score = 0;  trend_detail.append("중립")

                if macd_v > 0: trend_detail.append(f"MACD+")
                trend_detail.append(f"RSI{rsi}")
                trend_detail.append(f"고점대비{from_high:.0f}%")

                # ── 가치 점수 ───────────────────────────────────────
                pbr = val_row[0] if val_row and val_row[0] else None
                per = val_row[1] if val_row and val_row[1] else None
                roe = val_row[2] if val_row and val_row[2] else None
                roa = val_row[3] if val_row and val_row[3] else None

                if pbr is not None and 0 < pbr:
                    if   pbr <= 0.5:  val_score += 4; val_detail.append(f"PBR{pbr:.1f}(극저평가)")
                    elif pbr <= 1.0:  val_score += 3; val_detail.append(f"PBR{pbr:.1f}(저평가)")
                    elif pbr <= 2.0:  val_score += 1; val_detail.append(f"PBR{pbr:.1f}(적정)")
                    elif pbr <= 4.0:  val_score -= 1; val_detail.append(f"PBR{pbr:.1f}(다소고평)")
                    else:             val_score -= 2; val_detail.append(f"PBR{pbr:.1f}(고평가)")

                if per is not None and 0 < per < 300:
                    if   per <= 6:    val_score += 4; val_detail.append(f"PER{per:.0f}(극저)")
                    elif per <= 12:   val_score += 3; val_detail.append(f"PER{per:.0f}(저렴)")
                    elif per <= 20:   val_score += 1; val_detail.append(f"PER{per:.0f}(적정)")
                    elif per <= 35:   val_score -= 1; val_detail.append(f"PER{per:.0f}(다소고)")
                    else:             val_score -= 2; val_detail.append(f"PER{per:.0f}(고평가)")
                elif per is not None and per <= 0:
                    val_score -= 2;   val_detail.append("적자기업")

                if roe is not None:
                    if   roe >= 25:   val_score += 3; val_detail.append(f"ROE{roe:.0f}%(탁월)")
                    elif roe >= 15:   val_score += 2; val_detail.append(f"ROE{roe:.0f}%(우량)")
                    elif roe >= 8:    val_score += 1; val_detail.append(f"ROE{roe:.0f}%(양호)")
                    elif roe < 0:     val_score -= 2; val_detail.append(f"ROE{roe:.0f}%(적자)")

                if roa is not None:
                    if   roa >= 10:   val_score += 1; val_detail.append(f"ROA{roa:.0f}%(우량)")
                    elif roa < 0:     val_score -= 1; val_detail.append(f"ROA{roa:.0f}%(마이너스)")

                # 가치 정보 없으면(바이오 등 PER 없는 경우) 점수 0 취급
                has_value_data = (pbr is not None or per is not None or roe is not None)

                # ── 4분면 판단 ──────────────────────────────────────
                reasons = []
                if pnl <= -10 or curr < stop_line:
                    trade_signal = "cut_loss"
                    reasons.append(f"손절선 도달 ({pnl:.1f}%)")
                elif not has_value_data:
                    # 가치 데이터 없음 → 추세만으로 판단
                    if trend_score >= 3:
                        trade_signal = "add_buy"; reasons.append("추세강세(가치데이터없음)")
                    elif trend_score <= -2 and pnl < -3:
                        trade_signal = "real_sell"; reasons.append(f"역배열+손실({pnl:.1f}%)")
                    elif trend_score < 0:
                        trade_signal = "caution"; reasons.append("추세이탈-관망")
                    elif trend_score >= 1:
                        trade_signal = "hold";    reasons.append("추세유지-보유")
                    else:
                        trade_signal = "caution"; reasons.append("중립-관망")
                elif trend_score >= 2 and val_score >= 3:
                    trade_signal = "add_buy"
                    reasons.append(f"✅ 추세정배열+저평가 → 추가매수 유효")
                elif trend_score >= 2 and val_score >= 0:
                    trade_signal = "hold"
                    reasons.append(f"추세양호+적정가치 → 보유유지")
                elif trend_score >= 2 and val_score < 0:
                    trade_signal = "take_profit"
                    reasons.append(f"추세OK BUT 고평가({','.join(val_detail[:2])}) → 익절고려")
                elif trend_score <= -2 and val_score >= 3:
                    trade_signal = "hold_value"
                    reasons.append(f"⚠️ 추세역배열 BUT 저평가({','.join(val_detail[:2])}) → 홀딩유지")
                elif trend_score < 0 and val_score >= 2:
                    trade_signal = "hold_value"
                    reasons.append(f"추세약세 BUT 가치지지({','.join(val_detail[:2])}) → 홀딩")
                elif trend_score <= -2 and val_score <= -1:
                    trade_signal = "real_sell"
                    reasons.append(f"🚨 역배열+고평가({','.join(val_detail[:2])}) → 진매도 검토")
                elif trend_score < 0 and val_score < 0:
                    trade_signal = "real_sell"
                    reasons.append(f"추세이탈+고평가 → 매도 검토")
                elif trend_score == 0 and val_score >= 2:
                    trade_signal = "hold"
                    reasons.append(f"중립추세+가치양호 → 보유")
                elif trend_score == 0 and val_score < 0:
                    trade_signal = "caution"
                    reasons.append(f"중립추세+고평가 → 관망")
                elif pnl > 20 and trend_score < 1:
                    trade_signal = "take_profit"
                    reasons.append(f"고수익({pnl:.0f}%)+추세약화 → 익절고려")
                else:
                    trade_signal = "hold"; reasons.append(f"보유관찰({pnl:.1f}%)")

                # 수급·대차 보조 정보 추가
                if supply_pos:  reasons.append("외인·기관동반매수")
                if supply_neg:  reasons.append("외인·기관동반매도")
                if lend_surge:  reasons.append("대차잔고급증주의")
                if vol_surge:   reasons.append("거래량급증")

                trade_reason = (
                    " | ".join(reasons)
                    + f"  [추세:{trend_score:+d}/가치:{val_score:+d}]"
                    + f"  RSI:{rsi}  손절:{stop_line:,.0f}"
                )
        except Exception as _e:
            logger.debug(f"[포트폴리오신호] {sc}: {_e}")

        result.append({
            "stock_code":    m["stock_code"],
            "stock_name":    m["stock_name"],
            "sector":        m["sector"],
            "quantity":      total_qty,
            "avg_price":     avg_price,
            "current_price": current_price,
            "change_pct":    change_pct,
            "profit":        profit,
            "profit_pct":    profit_pct,
            "total_value":   total_value,
            "buy_total":     buy_total,
            "daily_profit":  daily_profit,
            "has_price":     has_price,
            "collecting":    collecting.get(m["stock_code"]) == "running",
            "bought_at":     m.get("bought_at"),
            "frn_net_buy":   frn_net,
            "inst_net_buy":  inst_net,
            "short_data":    short_data,
            "trade_signal":  trade_signal,
            "trade_reason":  trade_reason,
            "trend_score":   trend_score,
            "val_score":     val_score,
            "val_detail":    val_detail,
            "trend_detail":  trend_detail,
        })

    result.sort(key=lambda x: x["total_value"], reverse=True)
    return result


@router.post("/sync-kis")
def sync_kis_portfolio(db: Session = Depends(get_db)):
    """KIS 당일 체결내역 수동 동기화."""
    cnt = sync_kis_executions(db)
    return {"status": "ok", "synced": cnt}


@router.patch("/{stock_code}/bought-at")
def update_bought_at(stock_code: str, payload: dict, db: Session = Depends(get_db)):
    """종목 매수일 수동 수정."""
    rows = db.query(models.Portfolio).filter(models.Portfolio.stock_code == stock_code).all()
    if not rows:
        raise HTTPException(status_code=404, detail="종목 없음")
    bought_at = payload.get("bought_at", "")
    for row in rows:
        row.bought_at = bought_at
    db.commit()
    return {"status": "ok", "stock_code": stock_code, "bought_at": bought_at}


@router.get("/transactions")
def get_transactions(db: Session = Depends(get_db)):
    """거래 내역 최근 100건"""
    txs = db.query(models.PortfolioTx).order_by(
        models.PortfolioTx.tx_date.desc()
    ).limit(100).all()
    return [
        {"id": t.id, "stock_code": t.stock_code, "stock_name": t.stock_name,
         "tx_type": t.tx_type, "quantity": t.quantity, "price": t.price,
         "tx_date": t.tx_date.strftime("%Y-%m-%d %H:%M") if t.tx_date else "",
         "memo": t.memo}
        for t in txs
    ]


@router.post("/transaction")
def add_transaction(payload: dict, db: Session = Depends(get_db)):
    """매수/매도 거래 추가 + Portfolio 잔고 업데이트"""
    stock_code  = str(payload.get("stock_code", "")).strip()
    stock_name  = str(payload.get("stock_name", stock_code))
    tx_type     = payload.get("tx_type", "buy")
    quantity    = float(payload.get("quantity", 0))
    price       = float(payload.get("price", 0))
    tx_date_str = payload.get("tx_date", "")
    sector      = payload.get("sector", "")
    memo        = payload.get("memo", "")

    if not stock_code or quantity <= 0 or price <= 0:
        raise HTTPException(status_code=400, detail="stock_code, quantity, price 필수")

    try:
        tx_date = datetime.strptime(tx_date_str, "%Y-%m-%d") if tx_date_str else datetime.now()
    except Exception:
        tx_date = datetime.now()

    db.add(models.PortfolioTx(
        stock_code=stock_code, stock_name=stock_name,
        tx_type=tx_type, quantity=quantity, price=price,
        tx_date=tx_date, memo=memo,
    ))

    holding = db.query(models.Portfolio).filter(
        models.Portfolio.stock_code == stock_code
    ).first()

    if tx_type == "buy":
        if holding:
            total_qty = holding.quantity + quantity
            holding.avg_price  = (holding.avg_price * holding.quantity + price * quantity) / total_qty
            holding.quantity   = total_qty
            holding.stock_name = stock_name
            if sector: holding.sector = sector
        else:
            db.add(models.Portfolio(
                stock_code=stock_code, stock_name=stock_name,
                sector=sector, quantity=quantity, avg_price=price,
            ))
    elif tx_type == "sell":
        if holding:
            holding.quantity = max(0.0, holding.quantity - quantity)

    db.commit()
    return {"status": "ok", "stock_code": stock_code, "tx_type": tx_type,
            "quantity": quantity, "price": price}


@router.post("/kakao-parse")
def parse_kakao_message(payload: dict):
    """카카오톡 메시지에서 매수/매도 파싱."""
    import re
    text = payload.get("text", "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="text 필수")

    from ticker_utils import ticker_mapper
    parsed = []
    for line in text.strip().splitlines():
        line = line.strip()
        if not line: continue
        tx_type = None
        if re.search(r'매수', line):   tx_type = "buy"
        elif re.search(r'매도', line): tx_type = "sell"
        if not tx_type: continue

        qty_m   = re.search(r'(\d[\d,]*)\s*주', line)
        price_m = re.search(r'@\s*([\d,]+)|([\d][\d,]+)\s*원|([\d]{4,})', line)
        quantity = float(qty_m.group(1).replace(",", "")) if qty_m   else None
        raw_p    = (price_m.group(1) or price_m.group(2) or price_m.group(3)) if price_m else None
        price    = float(raw_p.replace(",", "")) if raw_p else None

        name_line  = re.sub(r'매수|매도', '', line)
        name_line  = re.sub(r'[\d,]+\s*주|@[\d,]+|[\d,]+\s*원|\d{4,}', '', name_line)
        stock_name = re.sub(r'[^\w가-힣]', ' ', name_line).strip().split()
        stock_name = stock_name[0] if stock_name else ""
        stock_code = ticker_mapper.get_code(stock_name) if stock_name else None

        parsed.append({
            "raw": line, "tx_type": tx_type,
            "stock_name": stock_name, "stock_code": stock_code,
            "quantity": quantity, "price": price,
            "valid": bool(stock_code and quantity and price),
        })
    return {"parsed": parsed, "count": len(parsed)}


@router.put("/{stock_code}")
def update_portfolio(stock_code: str, payload: dict, db: Session = Depends(get_db)):
    """보유종목 편집 저장. 종목명 변경 시 ticker 코드 자동 재조회."""
    import threading
    from ticker_utils import ticker_mapper

    h = db.query(models.Portfolio).filter(models.Portfolio.stock_code == stock_code).first()
    if not h:
        raise HTTPException(status_code=404, detail="종목 없음")

    new_name   = payload.get("stock_name", h.stock_name)
    new_sector = payload.get("sector",     h.sector)
    new_qty    = float(payload.get("quantity",  h.quantity))
    new_price  = float(payload.get("avg_price", h.avg_price))

    collecting  = _main_collecting()
    bg_ondemand = _main_bg_ondemand()
    new_code    = stock_code
    code_changed = False

    if new_name != h.stock_name:
        found = ticker_mapper.get_code(new_name)
        if found and found != stock_code:
            new_code = found
            code_changed = True

    if code_changed:
        db.delete(h)
        db.flush()
        existing = db.query(models.Portfolio).filter(models.Portfolio.stock_code == new_code).first()
        if existing:
            total_qty = existing.quantity + new_qty
            existing.avg_price  = (existing.avg_price * existing.quantity + new_price * new_qty) / total_qty
            existing.quantity   = total_qty
            existing.stock_name = new_name
            if new_sector: existing.sector = new_sector
        else:
            db.add(models.Portfolio(
                stock_code=new_code, stock_name=new_name,
                sector=new_sector, quantity=new_qty, avg_price=new_price,
            ))
        db.query(models.PortfolioTx).filter(
            models.PortfolioTx.stock_code == stock_code
        ).update({"stock_code": new_code, "stock_name": new_name})
        if collecting.get(new_code) != "running":
            threading.Thread(target=bg_ondemand, args=(new_code,), daemon=True).start()
    else:
        # 수량 변경 시 자동 거래내역 생성
        qty_diff = new_qty - h.quantity
        if abs(qty_diff) >= 0.5:
            auto_tx_type = "buy" if qty_diff > 0 else "sell"
            db.add(models.PortfolioTx(
                stock_code=stock_code, stock_name=new_name,
                tx_type=auto_tx_type, quantity=abs(qty_diff),
                price=new_price, tx_date=_date_cls.today(),
                memo="수동편집 자동기록",
            ))
        h.stock_name = new_name
        h.sector     = new_sector
        h.quantity   = new_qty
        h.avg_price  = new_price

    db.commit()
    return {"status": "ok", "code_changed": code_changed,
            "new_code": new_code, "stock_name": new_name}


@router.delete("/{stock_code}")
def delete_portfolio(stock_code: str, db: Session = Depends(get_db)):
    """보유종목 삭제"""
    h = db.query(models.Portfolio).filter(models.Portfolio.stock_code == stock_code).first()
    if h:
        db.delete(h)
        db.commit()
    return {"status": "ok"}


@router.get("/export/excel")
def export_portfolio_excel(db: Session = Depends(get_db)):
    """보유종목을 엑셀 파일로 다운로드"""
    from fastapi.responses import StreamingResponse
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치. pip install openpyxl")

    holdings = db.query(models.Portfolio).filter(models.Portfolio.quantity > 0).all()

    # 현재가 일괄 조회 (N+1 방지)
    codes = [h.stock_code for h in holdings if h.stock_code]
    prices: dict = {}
    if codes:
        from sqlalchemy import func as _sqlfunc
        subq = (
            db.query(
                models.PriceHistory.stock_code,
                _sqlfunc.max(models.PriceHistory.date).label("max_date"),
            )
            .filter(models.PriceHistory.stock_code.in_(codes))
            .group_by(models.PriceHistory.stock_code)
            .subquery()
        )
        rows = db.query(models.PriceHistory).join(
            subq,
            (models.PriceHistory.stock_code == subq.c.stock_code) &
            (models.PriceHistory.date == subq.c.max_date),
        ).all()
        prices = {r.stock_code: r.close for r in rows}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "보유종목"

    header_fill  = PatternFill("solid", fgColor="1E1E2E")
    header_font  = Font(bold=True, color="2DD4BF", size=11)
    data_font    = Font(color="E2E8F0", size=10)
    border_side  = Side(style="thin", color="334155")
    thin_border  = Border(left=border_side, right=border_side,
                          top=border_side, bottom=border_side)
    center_align = Alignment(horizontal="center", vertical="center")
    right_align  = Alignment(horizontal="right",  vertical="center")

    headers = ["종목코드","종목명","섹터","보유수량","매입가(원)",
               "현재가(원)","수익률(%)","손익(원)","평가액(원)","매입총액(원)"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center_align; cell.border = thin_border

    col_widths = [12, 18, 10, 12, 16, 16, 12, 18, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    total_buy = total_val_sum = 0
    for h in holdings:
        current_price = prices.get(h.stock_code, h.avg_price)
        profit_pct  = round((current_price - h.avg_price) / h.avg_price * 100, 2) if h.avg_price else 0
        profit      = round((current_price - h.avg_price) * h.quantity)
        total_value = round(current_price * h.quantity)
        buy_total   = round(h.avg_price   * h.quantity)
        total_buy      += buy_total
        total_val_sum  += total_value

        ws.append([h.stock_code, h.stock_name, h.sector or "",
                   int(h.quantity), int(h.avg_price), int(current_price),
                   profit_pct, profit, total_value, buy_total])

        data_row = ws.max_row
        row_fill = PatternFill("solid", fgColor="1A1A2E" if data_row % 2 == 0 else "0F0F1A")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col)
            cell.font   = data_font; cell.border = thin_border; cell.fill = row_fill
            cell.alignment = right_align if col >= 4 else center_align
            if col == 7 and profit_pct != 0:
                cell.font = Font(color="EF4444" if profit_pct > 0 else "3B82F6", bold=True, size=10)
            if col == 8 and profit != 0:
                cell.font = Font(color="EF4444" if profit > 0 else "3B82F6", size=10)

    total_profit     = total_val_sum - total_buy
    profit_pct_total = round(total_profit / total_buy * 100, 2) if total_buy else 0
    sum_row = len(holdings) + 2
    ws.append(["", "합계", "", "", "", "", profit_pct_total, total_profit, total_val_sum, total_buy])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=sum_row, column=col)
        cell.font  = Font(bold=True, color="2DD4BF", size=10)
        cell.fill  = PatternFill("solid", fgColor="1E1E2E")
        cell.border = thin_border
        cell.alignment = right_align if col >= 4 else center_align

    # 업로드 양식 시트
    ws2 = wb.create_sheet("업로드양식")
    ws2.append(["종목명", "보유수량", "매입가(원)", "섹터(선택)"])
    ws2.append(["삼성전자", 100, 75000, "반도체"])
    ws2.append(["에이엘티", 46698, 12301, "반도체"])
    for col in range(1, 5):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True, color="2DD4BF"); cell.fill = PatternFill("solid", fgColor="1E1E2E")
        cell.alignment = Alignment(horizontal="center")
    for i, w in enumerate([20, 12, 15, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.cell(row=5, column=1).value = "※ 종목명은 정확히 입력하세요 (ticker 자동 조회)"
    ws2.cell(row=6, column=1).value = "※ 현재가/손익/평가액은 시스템이 자동 계산합니다"
    ws2.cell(row=7, column=1).value = "※ 섹터는 비워도 됩니다"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"portfolio_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/import/excel")
async def import_portfolio_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """엑셀 파일로 보유종목 일괄 등록.
    양식: 종목명 | 보유수량 | 매입가(원) | 섹터(선택)
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치")

    from ticker_utils import ticker_mapper
    collecting  = _main_collecting()
    bg_ondemand = _main_bg_ondemand()

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb["업로드양식"] if "업로드양식" in wb.sheetnames else wb.active

    results: dict = {"success": [], "failed": [], "skipped": []}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        stock_name = str(row[0]).strip()
        if stock_name.startswith("※"):
            continue

        try:
            quantity  = float(row[1]) if row[1] else 0
            avg_price = float(str(row[2]).replace(",", "")) if row[2] else 0
            sector    = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        except (ValueError, TypeError):
            results["failed"].append({"name": stock_name, "reason": "수량/매입가 형식 오류"})
            continue

        if quantity <= 0 or avg_price <= 0:
            results["skipped"].append(stock_name)
            continue

        stock_code = ticker_mapper.get_code(stock_name)
        if not stock_code:
            search = ticker_mapper.search(stock_name)
            if search:
                stock_code = search[0].get("code")
                stock_name = search[0].get("name", stock_name)

        if not stock_code:
            results["failed"].append({"name": stock_name, "reason": "종목코드 조회 실패"})
            continue

        holding = db.query(models.Portfolio).filter(
            models.Portfolio.stock_code == stock_code
        ).first()

        if holding:
            holding.quantity   = quantity
            holding.avg_price  = avg_price
            holding.stock_name = stock_name
            if sector: holding.sector = sector
        else:
            db.add(models.Portfolio(
                stock_code=stock_code, stock_name=stock_name,
                sector=sector, quantity=quantity, avg_price=avg_price,
            ))
            db.add(models.PortfolioTx(
                stock_code=stock_code, stock_name=stock_name,
                tx_type="buy", quantity=quantity, price=avg_price,
                tx_date=datetime.now(), memo="엑셀 업로드",
            ))

        crud.add_to_watchlist(db, stock_code)
        results["success"].append({"code": stock_code, "name": stock_name,
                                   "quantity": quantity, "avg_price": avg_price})

    db.commit()

    if results["success"]:
        import threading
        for item in results["success"]:
            code = item["code"]
            if collecting.get(code) != "running":
                threading.Thread(target=bg_ondemand, args=(code,), daemon=True).start()

    return {
        "status": "ok",
        "success_count": len(results["success"]),
        "failed_count":  len(results["failed"]),
        "skipped_count": len(results["skipped"]),
        "success": results["success"],
        "failed":  results["failed"],
    }


@router.post("/recalculate-avg")
def recalculate_avg_price(db: Session = Depends(get_db)):
    """
    portfolio_tx 이력으로 모든 보유종목의 avg_price 재계산.
    - 매수: 가중평균 (avg = (prev_avg * prev_qty + price * qty) / new_qty)
    - 매도: avg_price 변경 없이 수량만 차감
    - 결과: {stock_code: {old_avg, new_avg, quantity, fixed}}
    """
    from collections import defaultdict

    # 전체 거래 이력 (시간순)
    txs = db.query(models.PortfolioTx).order_by(
        models.PortfolioTx.tx_date.asc(),
        models.PortfolioTx.id.asc() if hasattr(models.PortfolioTx, 'id') else models.PortfolioTx.tx_date.asc()
    ).all()

    # 종목별로 재계산
    state: dict = defaultdict(lambda: {"avg": 0.0, "qty": 0.0})
    for tx in txs:
        code = tx.stock_code
        s = state[code]
        qty   = float(tx.quantity or 0)
        price = float(tx.price or 0)

        if tx.tx_type == "buy" and qty > 0 and price > 0:
            new_qty = s["qty"] + qty
            s["avg"] = (s["avg"] * s["qty"] + price * qty) / new_qty if new_qty else price
            s["qty"] = new_qty
        elif tx.tx_type == "sell" and qty > 0:
            # avg_price 변경 없음
            s["qty"] = max(0.0, s["qty"] - qty)
            if s["qty"] == 0:
                s["avg"] = 0.0

    # 현재 보유 종목에 반영
    holdings = db.query(models.Portfolio).filter(models.Portfolio.quantity > 0).all()
    result = []
    for h in holdings:
        s = state.get(h.stock_code)
        if not s:
            continue
        new_avg = round(s["avg"], 2)
        old_avg = h.avg_price
        if abs(new_avg - old_avg) > 1:  # 1원 이상 차이 시 업데이트
            h.avg_price = new_avg
            result.append({
                "stock_code": h.stock_code,
                "stock_name": h.stock_name,
                "old_avg": old_avg,
                "new_avg": new_avg,
                "quantity": h.quantity,
                "fixed": True,
            })
        else:
            result.append({
                "stock_code": h.stock_code,
                "stock_name": h.stock_name,
                "old_avg": old_avg,
                "new_avg": new_avg,
                "quantity": h.quantity,
                "fixed": False,
            })

    db.commit()
    fixed = [r for r in result if r["fixed"]]
    logger.info(f"[재계산] {len(fixed)}/{len(result)}종목 avg_price 수정")
    return {
        "status": "ok",
        "total": len(result),
        "fixed": len(fixed),
        "details": result,
    }
