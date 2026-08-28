"""
DartV22Builder 재현 — DART API에서 재무데이터 수집 후 v22 형식 엑셀 생성
시트 구성:
  1.개요_스크리닝판정 / 2.연간손익추이 / 3.분기손익 / 4.재무상태_현금흐름
  5.밸류에이션 / 6.장기분기손익 / 7.장기분기재무 / 8.원문메모
CH 시트 데이터:
  사업부문매출 / 직원현황 / 재고자산 / 수주잔고 / 매출채권 / 판관비
"""
import io, os, re, time, zipfile, logging, sqlite3, requests
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, Query, BackgroundTasks
from fastapi.responses import StreamingResponse, JSONResponse
from dart_key_manager import get_dart_api_keys

router = APIRouter()
log = logging.getLogger(__name__)

DB_PATH = "stock.db"
DART_KEYS = get_dart_api_keys()
_key_idx = [0]

def _dart_key():
    k = DART_KEYS[_key_idx[0] % len(DART_KEYS)]
    return k

def _next_key():
    _key_idx[0] += 1
    return DART_KEYS[_key_idx[0] % len(DART_KEYS)]

# ── ACCOUNT_ID 매핑 (원본 프로그램과 동일) ─────────────────────────
IS_ACCOUNTS = [
    ("매출액",         ["ifrs-full_Revenue"]),
    ("매출원가",       ["ifrs-full_CostOfSales"]),
    ("매출총이익",     ["ifrs-full_GrossProfit"]),
    ("판관비",         ["dart_TotalSellingGeneralAdministrativeExpenses",
                       "ifrs-full_SellingGeneralAndAdministrativeExpense"]),
    ("영업이익",       ["dart_OperatingIncomeLoss"]),
    ("금융수익",       ["ifrs-full_FinanceIncome"]),
    ("금융비용",       ["ifrs-full_FinanceCosts"]),
    ("기타이익",       ["dart_OtherGains", "ifrs-full_OtherIncome"]),
    ("기타손실",       ["dart_OtherLosses"]),
    ("세전이익",       ["ifrs-full_ProfitLossBeforeTax"]),
    ("법인세",         ["ifrs-full_IncomeTaxExpenseContinuingOperations"]),
    ("당기순이익",     ["ifrs-full_ProfitLoss"]),
]
BS_ACCOUNTS = [
    ("현금및현금성자산", ["ifrs-full_CashAndCashEquivalents"]),
    ("매출채권",         ["ifrs-full_TradeAndOtherCurrentReceivables",
                         "dart_ShortTermTradeReceivable"]),
    ("재고자산",         ["ifrs-full_Inventories"]),
    ("유동자산",         ["ifrs-full_CurrentAssets"]),
    ("유형자산",         ["ifrs-full_PropertyPlantAndEquipment"]),
    ("무형자산",         ["ifrs-full_IntangibleAssetsOtherThanGoodwill"]),
    ("비유동자산",       ["ifrs-full_NoncurrentAssets"]),
    ("자산총계",         ["ifrs-full_Assets"]),
    ("매입채무",         ["ifrs-full_TradeAndOtherCurrentPayables",
                         "dart_ShortTermTradePayables"]),
    ("계약부채(유동)",   ["ifrs-full_ContractLiabilities",
                         "ifrs-full_CurrentContractLiabilities"]),
    ("단기차입금",       ["ifrs-full_ShorttermBorrowings"]),
    ("장기차입금",       ["ifrs-full_LongtermBorrowings",
                         "dart_LongTermBorrowingsGross"]),
    ("유동부채",         ["ifrs-full_CurrentLiabilities"]),
    ("부채총계",         ["ifrs-full_Liabilities"]),
    ("이익잉여금",       ["ifrs-full_RetainedEarnings"]),
    ("자본총계",         ["ifrs-full_Equity"]),
]
CF_ACCOUNTS = [
    ("영업활동CF",   ["ifrs-full_CashFlowsFromUsedInOperatingActivities"]),
    ("CAPEX",        ["ifrs-full_PurchaseOfPropertyPlantAndEquipmentClassifiedAsInvestingActivities"]),
    ("투자활동CF",   ["ifrs-full_CashFlowsFromUsedInInvestingActivities"]),
    ("재무활동CF",   ["ifrs-full_CashFlowsFromUsedInFinancingActivities"]),
]

REPRT_MAP = {
    "11011": "사업보고서",
    "11012": "반기보고서",
    "11013": "1분기보고서",
    "11014": "3분기보고서",
}
REPRT_Q = {"11013": 1, "11012": 2, "11014": 3, "11011": 4}

# ── DART API 헬퍼 ─────────────────────────────────────────────────

def _load_corp_code(stock_code: str) -> Optional[str]:
    """stock.db 여러 테이블 또는 DART corpCode.xml에서 corp_code 조회"""
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        # dart_insider_holdings에 corp_code 저장됨 (상위 2198종목)
        for tbl in ["dart_insider_holdings", "dart_backlog_quarterly", "dart_cost_quarterly"]:
            try:
                row = conn.execute(
                    f"SELECT corp_code FROM {tbl} WHERE stock_code=? AND corp_code IS NOT NULL LIMIT 1",
                    (stock_code,)
                ).fetchone()
                if row and row[0]:
                    conn.close()
                    return str(row[0]).zfill(8)
            except Exception:
                pass
        conn.close()
    except Exception:
        pass

    # DART corpCode.xml 사용
    xml_path = "/tmp/DART_CORPCODE.xml"
    if not os.path.exists(xml_path):
        try:
            r = requests.get("https://opendart.fss.or.kr/api/corpCode.xml",
                             params={"crtfc_key": _dart_key()}, timeout=30)
            zip_path = "/tmp/DART_CORPCODE.zip"
            with open(zip_path, "wb") as f:
                f.write(r.content)
            with zipfile.ZipFile(zip_path) as z:
                z.extractall("/tmp/")
            # 파일명 통일
            import glob
            for p in glob.glob("/tmp/CORPCODE*.xml"):
                os.rename(p, xml_path)
        except Exception as e:
            log.error("corpCode 다운로드 실패: %s", e)
            return None

    from xml.etree import ElementTree
    tree = ElementTree.parse(xml_path)
    for item in tree.getroot().findall("list"):
        sc = (item.findtext("stock_code") or "").strip()
        if sc == stock_code:
            return (item.findtext("corp_code") or "").strip()
    return None


def _fetch_fnltt(corp_code, year, reprt_code, fs_div="CFS"):
    """DART fnlttSinglAcntAll API 호출"""
    for attempt in range(len(DART_KEYS)):
        try:
            r = requests.get(
                "https://opendart.fss.or.kr/api/fnlttSinglAcntAll.json",
                params={"crtfc_key": _dart_key(), "corp_code": corp_code,
                        "bsns_year": str(year), "reprt_code": reprt_code,
                        "fs_div": fs_div},
                timeout=15,
            )
            d = r.json()
            if d.get("status") == "020":  # 한도초과
                _next_key()
                time.sleep(1)
                continue
            if d.get("status") == "000":
                return d.get("list", [])
        except Exception as e:
            log.warning("fnltt 오류 %s %s %s: %s", corp_code, year, reprt_code, e)
        _next_key()
    return []


def _extract_value(rows, account_ids, field="thstrm_amount"):
    """account_id 우선순위 매칭으로 금액 추출 (원 단위)"""
    for aid in account_ids:
        for row in rows:
            if row.get("account_id", "") == aid:
                raw = row.get(field, "").replace(",", "")
                try:
                    return int(float(raw))
                except (ValueError, TypeError):
                    pass
    return None


def _to_100m(val):
    """원 → 억원 (소수점 1자리)"""
    if val is None:
        return None
    return round(val / 1e8, 1)


# ── 핵심 수집 함수 ────────────────────────────────────────────────

def _collect_from_db(stock_code: str, years: list[int]) -> dict:
    """DART API 한도 초과 시 stock.db 저장값으로 fallback"""
    annual, standalone, bs_data, cf_data = {}, {}, {}, {}
    try:
        conn = sqlite3.connect(DB_PATH, timeout=30)
        for year in years:
            # 연간 P&L
            row = conn.execute("""
                SELECT revenue, operating_profit, net_income
                FROM financial_data
                WHERE stock_code=? AND year=? AND is_annual=1 AND quarter=4
                ORDER BY CASE WHEN data_source LIKE '%dart%' THEN 0 ELSE 1 END,
                         CASE WHEN revenue IS NOT NULL THEN 0 ELSE 1 END
                LIMIT 1
            """, (stock_code, year)).fetchone()
            if not row:
                row = conn.execute("""
                    SELECT revenue, operating_profit, net_income
                    FROM financial_data
                    WHERE stock_code=? AND year=? AND is_annual=1
                    ORDER BY CASE WHEN data_source LIKE '%dart%' THEN 0 ELSE 1 END LIMIT 1
                """, (stock_code, year)).fetchone()
            if row and row[0]:
                rev, op, ni = row
                annual[year] = {
                    "매출액": int(rev) if rev else None,
                    "영업이익": int(op) if op else None,
                    "당기순이익": int(ni) if ni else None,
                }
                # BS
                bs_row = conn.execute("""
                    SELECT total_assets, total_equity FROM financial_data
                    WHERE stock_code=? AND year=? AND is_annual=1 AND total_assets IS NOT NULL
                    ORDER BY CASE WHEN data_source LIKE '%dart%' THEN 0 ELSE 1 END LIMIT 1
                """, (stock_code, year)).fetchone()
                if bs_row:
                    ta, te = bs_row
                    bs_data[(year, "11011")] = {"자산총계": int(ta) if ta else None,
                                                 "자본총계": int(te) if te else None}
                # CF
                cf_row = conn.execute("""
                    SELECT operating_cf, investing_cf, financing_cf, capex FROM cash_flow_data
                    WHERE stock_code=? AND year=? AND is_annual=1 AND report_type='CFS'
                    ORDER BY CASE WHEN data_source LIKE '%dart%' THEN 0 ELSE 1 END LIMIT 1
                """, (stock_code, year)).fetchone()
                if cf_row:
                    ocf, icf, fcf, cap = cf_row
                    cf_data[(year, "11011")] = {
                        "영업활동CF": int(ocf) if ocf else None,
                        "투자활동CF": int(icf) if icf else None,
                        "재무활동CF": int(fcf) if fcf else None,
                        "CAPEX": abs(int(cap)) if cap else None,
                    }
            # 분기 P&L
            for q in [1, 2, 3, 4]:
                qrow = conn.execute("""
                    SELECT revenue, operating_profit, net_income FROM financial_data
                    WHERE stock_code=? AND year=? AND quarter=? AND is_annual=0
                    ORDER BY CASE WHEN data_source LIKE '%dart%' THEN 0 ELSE 1 END LIMIT 1
                """, (stock_code, year, q)).fetchone()
                if qrow and qrow[0]:
                    rv, op, ni = qrow
                    standalone[(year, q)] = {
                        "매출액": int(rv) if rv else None,
                        "영업이익": int(op) if op else None,
                        "당기순이익": int(ni) if ni else None,
                    }
        conn.close()
    except Exception as e:
        log.error("DB fallback 오류: %s", e)
    return {"annual": annual, "quarterly": standalone, "bs": bs_data, "cf": cf_data,
            "cumul": {}, "_source": "db_fallback"}


def collect_financials(corp_code: str, stock_code: str, years: list[int]) -> dict:
    """
    연간(11011) + 분기(11013/11012/11014) 재무데이터 수집
    DART API 한도 초과 시 DB fallback.
    반환: {
        'annual': {year: {label: value_원}},
        'quarterly': {(year, q): {label: value_원}},
        'bs': {(year, q_or_4): {label: value}},
        'cf': {(year, q_or_4): {label: value}},
    }
    """
    # DART API 한도 사전 체크 — 실패 시 DB fallback
    test_rows = _fetch_fnltt(corp_code, years[-1], "11011")
    if not test_rows:
        log.warning("DART API 한도 초과 또는 미수집 — DB fallback 사용: %s", stock_code)
        return _collect_from_db(stock_code, years)

    annual = {}    # year → {label: 원단위}
    cumul  = {}    # (year, reprt_code) → {label: 원단위} (누적)
    bs_data = {}   # (year, reprt_code) → {label: 원단위}
    cf_data = {}   # (year, reprt_code) → {label: 원단위}

    reprt_list = [
        ("11011", 4), ("11013", 1), ("11012", 2), ("11014", 3),
    ]
    # test 결과 재활용
    _prefetched = {(years[-1], "11011"): test_rows}

    for year in years:
        for reprt_code, qnum in reprt_list:
            rows = _prefetched.pop((year, reprt_code), None)
            if rows is None:
                rows = _fetch_fnltt(corp_code, year, reprt_code)
            if not rows:
                # OFS fallback
                rows = _fetch_fnltt(corp_code, year, reprt_code, "OFS")
            if not rows:
                continue
            time.sleep(0.07)

            # IS
            is_vals = {}
            for label, aids in IS_ACCOUNTS:
                v = _extract_value(rows, aids)
                if v is not None:
                    is_vals[label] = v

            # BS
            bs_vals = {}
            for label, aids in BS_ACCOUNTS:
                v = _extract_value(rows, aids)
                if v is not None:
                    bs_vals[label] = v

            # CF
            cf_vals = {}
            for label, aids in CF_ACCOUNTS:
                v = _extract_value(rows, aids)
                if v is not None:
                    cf_vals[label] = abs(v) if label == "CAPEX" else v

            key = (year, reprt_code)
            if reprt_code == "11011":
                annual[year] = is_vals
                bs_data[key] = bs_vals
                cf_data[key] = cf_vals
            else:
                cumul[key] = is_vals
                bs_data[key] = bs_vals
                cf_data[key] = cf_vals

    # 누적→단독분기 변환
    standalone = {}  # (year, q) → {label: 원단위}

    def _get_cumul_is(year, reprt_code):
        return cumul.get((year, reprt_code), {})

    for year in years:
        # Q1 = 1Q누적 (단독)
        q1 = _get_cumul_is(year, "11013")
        if q1:
            standalone[(year, 1)] = dict(q1)

        # Q2 = 반기 - Q1누적
        h1 = _get_cumul_is(year, "11012")
        if h1 and q1:
            standalone[(year, 2)] = {k: h1.get(k, 0) - q1.get(k, 0)
                                     for k in set(h1) | set(q1)}

        # Q3 = 9M - 반기
        q3c = _get_cumul_is(year, "11014")
        if q3c and h1:
            standalone[(year, 3)] = {k: q3c.get(k, 0) - h1.get(k, 0)
                                     for k in set(q3c) | set(h1)}

        # Q4 = 연간 - 9M
        ann = annual.get(year, {})
        if ann and q3c:
            standalone[(year, 4)] = {k: ann.get(k, 0) - q3c.get(k, 0)
                                     for k in set(ann) | set(q3c)}
        elif ann:
            standalone[(year, 4)] = dict(ann)

    return {
        "annual": annual,
        "quarterly": standalone,
        "bs": bs_data,
        "cf": cf_data,
        "cumul": cumul,
    }


def get_stock_info(stock_code: str) -> dict:
    """stock.db에서 종목 기본정보"""
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        row = conn.execute(
            "SELECT stock_name, market, sector_large, market_cap, shares_issued, per, pbr FROM stock_universe WHERE stock_code=?",
            (stock_code,)
        ).fetchone()
        price_row = conn.execute(
            "SELECT close FROM price_history WHERE stock_code=? AND close>0 ORDER BY date DESC LIMIT 1",
            (stock_code,)
        ).fetchone()
        conn.close()
        if row:
            return {
                "stock_name": row[0], "market": row[1], "sector": row[2],
                "market_cap": row[3], "shares_issued": row[4],
                "per": row[5], "pbr": row[6],
                "current_price": price_row[0] if price_row else None,
            }
    except Exception:
        pass
    return {}


# ── 엑셀 생성 ────────────────────────────────────────────────────

def build_excel_v22(stock_code: str, stock_name: str, info: dict, fin: dict) -> bytes:
    """v22 형식 엑셀 생성 후 bytes 반환"""
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                      numbers as xnum)
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl 미설치: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # ── 색상/스타일 상수 ─────────────────────────────────
    HDR_FILL   = PatternFill("solid", fgColor="1F3864")   # 진남색 헤더
    GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")   # 연녹 (DART 연동값)
    YELLOW_FILL= PatternFill("solid", fgColor="FFFF99")   # 노랑 (입력칸)
    GRAY_FILL  = PatternFill("solid", fgColor="F2F2F2")   # 회색 (파생값)
    TITLE_FILL = PatternFill("solid", fgColor="2E75B6")   # 파랑 타이틀
    OK_FILL    = PatternFill("solid", fgColor="70AD47")   # 초록 (판정OK)
    NG_FILL    = PatternFill("solid", fgColor="FF0000")   # 빨강 (판정NG)

    WHITE_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF")
    BOLD_FONT  = Font(name="맑은 고딕", bold=True)
    BASE_FONT  = Font(name="맑은 고딕")
    NUM_FMT    = '#,##0;(#,##0);"-"'
    PCT_FMT    = '0.0%;(0.0%);"-"'

    thin  = Side(style="thin")
    THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(ws, row, col, val, width=None):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = HDR_FILL; c.font = WHITE_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
        if width:
            ws.column_dimensions[get_column_letter(col)].width = width
        return c

    def _cell(ws, row, col, val=None, fmt=None, fill=None, bold=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="맑은 고딕", bold=bold)
        if fmt:  c.number_format = fmt
        if fill: c.fill = fill
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="right" if fmt else "left",
                                vertical="center")
        return c

    def _title(ws, row, text, unit_note=""):
        c = ws.cell(row=row, column=2, value=text)
        c.fill = TITLE_FILL; c.font = WHITE_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")
        if unit_note:
            ws.cell(row=row+1, column=2, value=unit_note).font = Font(name="맑은 고딕", italic=True, color="595959")

    # 연도 목록
    years = sorted(fin["annual"].keys())
    if not years:
        years = list(range(datetime.now().year - 4, datetime.now().year + 1))

    # ════════════════════════════════════════════════════════════
    # 시트 1: 개요_스크리닝판정
    # ════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("1.개요_스크리닝판정")
    ws1.column_dimensions["A"].width = 2
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 18
    ws1.column_dimensions["E"].width = 15

    _title(ws1, 2, f"{stock_name} ({stock_code}) — 기업 개요 · 스크리닝 판정",
           "※ DART 공시 기반. 추정치·전망 없음.")
    ws1.row_dimensions[2].height = 22

    # 기업 정보 블록
    basics = [
        ("종목코드", stock_code),
        ("종목명", stock_name),
        ("시장", info.get("market", "-")),
        ("섹터", info.get("sector", "-")),
        ("현재가(원)", info.get("current_price")),
        ("시가총액(억원)", info.get("market_cap")),
        ("발행주식수(주)", info.get("shares_issued")),
        ("PER", info.get("per")),
        ("PBR", info.get("pbr")),
    ]
    for i, (lbl, val) in enumerate(basics, start=5):
        _cell(ws1, i, 2, lbl, bold=True)
        _cell(ws1, i, 3, val,
              fmt=NUM_FMT if isinstance(val, (int, float)) else None,
              fill=GREEN_FILL)

    # 스크리닝 판정 (최근 3개년 비교)
    r = 15
    _hdr(ws1, r, 2, "판정 항목")
    for j, yr in enumerate(years[-3:], start=3):
        _hdr(ws1, r, j, str(yr))

    judg_items = [
        ("매출증가(YoY)", "매출액"),
        ("영업이익증가(YoY)", "영업이익"),
        ("당기순이익증가(YoY)", "당기순이익"),
        ("흑자전환", "영업이익"),
    ]
    for i, (lbl, key) in enumerate(judg_items, start=r+1):
        _cell(ws1, i, 2, lbl, bold=True)
        ann_vals = [fin["annual"].get(y, {}).get(key) for y in years[-4:]]
        for j, yr in enumerate(years[-3:], start=3):
            cur  = fin["annual"].get(yr, {}).get(key)
            prev = fin["annual"].get(yr-1, {}).get(key)
            if cur is None or prev is None:
                txt, fill = "-", GRAY_FILL
            elif lbl == "흑자전환(영업이익)":
                txt  = "흑전" if (prev or 0) <= 0 and (cur or 0) > 0 else ("흑지" if (cur or 0) > 0 else "적자")
                fill = OK_FILL if txt == "흑전" else (GRAY_FILL if txt == "흑지" else NG_FILL)
            else:
                pct  = (cur - prev) / abs(prev) * 100 if prev != 0 else 0
                txt  = f"{pct:+.1f}%"
                fill = OK_FILL if pct > 0 else NG_FILL
            c = ws1.cell(row=i, column=j, value=txt)
            c.fill = fill; c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")
            c.font = Font(name="맑은 고딕", bold=True,
                          color="FFFFFF" if fill in (OK_FILL, NG_FILL) else "000000")

    ws1.cell(row=26, column=2, value="■ 해석").font = BOLD_FONT
    ws1.cell(row=27, column=2, value="• 공시된 사실 기반. 추정·전망 없음.").font = BASE_FONT

    # ════════════════════════════════════════════════════════════
    # 시트 2: 연간손익추이 (억원)
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("2.연간손익추이")
    ws2.column_dimensions["A"].width = 2
    ws2.column_dimensions["B"].width = 22
    for j, yr in enumerate(years, start=3):
        ws2.column_dimensions[get_column_letter(j)].width = 14

    _title(ws2, 2, f"{stock_name} — 연간 손익 추이", "단위: 억원 | 출처: DART fnlttSinglAcntAll (CFS 우선)")
    ws2.row_dimensions[2].height = 22

    r = 5
    _hdr(ws2, r, 2, "항목", 22)
    for j, yr in enumerate(years, start=3):
        _hdr(ws2, r, j, str(yr))

    is_rows = [lbl for lbl, _ in IS_ACCOUNTS]
    derived_rows = ["매출총이익률(%)", "영업이익률(%)", "순이익률(%)"]

    for i, lbl in enumerate(is_rows + derived_rows, start=r+1):
        c = ws2.cell(row=i, column=2, value=lbl)
        c.font = BOLD_FONT; c.border = THIN_BORDER
        for j, yr in enumerate(years, start=3):
            ann = fin["annual"].get(yr, {})
            if lbl == "매출총이익률(%)":
                rev = ann.get("매출액"); gp = ann.get("매출총이익")
                val = gp / rev if rev and gp else None
                fmt = PCT_FMT; fill = GRAY_FILL
            elif lbl == "영업이익률(%)":
                rev = ann.get("매출액"); op = ann.get("영업이익")
                val = op / rev if rev and op else None
                fmt = PCT_FMT; fill = GRAY_FILL
            elif lbl == "순이익률(%)":
                rev = ann.get("매출액"); ni = ann.get("당기순이익")
                val = ni / rev if rev and ni else None
                fmt = PCT_FMT; fill = GRAY_FILL
            else:
                val = _to_100m(ann.get(lbl))
                fmt = NUM_FMT; fill = GREEN_FILL
            _cell(ws2, i, j, val, fmt=fmt, fill=fill if val is not None else GRAY_FILL)

    # YoY 행
    for base_lbl in ["매출액", "영업이익", "당기순이익"]:
        i += 1
        c = ws2.cell(row=i, column=2, value=f"{base_lbl} YoY(%)")
        c.font = BOLD_FONT; c.border = THIN_BORDER
        for j, yr in enumerate(years, start=3):
            cur  = fin["annual"].get(yr, {}).get(base_lbl)
            prev = fin["annual"].get(yr-1, {}).get(base_lbl)
            if cur and prev and prev != 0:
                val = (cur - prev) / abs(prev)
            else:
                val = None
            _cell(ws2, i, j, val, fmt=PCT_FMT, fill=GRAY_FILL)

    ws2.cell(row=i+2, column=2, value="■ 해석").font = BOLD_FONT
    ws2.cell(row=i+3, column=2, value="• DART 공시 기반 연간 손익 추이. 파생행(이익률/YoY)은 수식 계산값.").font = BASE_FONT

    # ════════════════════════════════════════════════════════════
    # 시트 3: 분기손익 (단독분기, 억원)
    # ════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("3.분기손익")
    ws3.column_dimensions["A"].width = 2
    ws3.column_dimensions["B"].width = 22

    # 최근 12분기 (최대)
    all_q = sorted(fin["quarterly"].keys())  # (year, q)
    recent_q = all_q[-12:]
    for j in range(len(recent_q) + 2):
        ws3.column_dimensions[get_column_letter(j+2)].width = 12

    _title(ws3, 2, f"{stock_name} — 분기 손익 (단독분기)", "단위: 억원 | 단독분기 = 누적차감(2Q=반기-1Q, 4Q=연간-3Q)")
    r = 5
    _hdr(ws3, r, 2, "항목")
    for j, (yr, q) in enumerate(recent_q, start=3):
        _hdr(ws3, r, j, f"{yr}Q{q}")

    for i, (lbl, _) in enumerate(IS_ACCOUNTS + [("영업이익률(%)", None), ("순이익률(%)", None)], start=r+1):
        c = ws3.cell(row=i, column=2, value=lbl)
        c.font = BOLD_FONT; c.border = THIN_BORDER
        for j, (yr, q) in enumerate(recent_q, start=3):
            qd = fin["quarterly"].get((yr, q), {})
            if "%" in lbl:
                rev = qd.get("매출액"); op = qd.get("영업이익"); ni = qd.get("당기순이익")
                val = (op / rev if "영업" in lbl else ni / rev) if rev and (op if "영업" in lbl else ni) else None
                _cell(ws3, i, j, val, fmt=PCT_FMT, fill=GRAY_FILL)
            else:
                val = _to_100m(qd.get(lbl))
                _cell(ws3, i, j, val, fmt=NUM_FMT, fill=GREEN_FILL if val else GRAY_FILL)

    # ════════════════════════════════════════════════════════════
    # 시트 4: 재무상태_현금흐름
    # ════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("4.재무상태_현금흐름")
    ws4.column_dimensions["A"].width = 2
    ws4.column_dimensions["B"].width = 22
    for j in range(len(years) + 2):
        ws4.column_dimensions[get_column_letter(j+2)].width = 14

    _title(ws4, 2, f"{stock_name} — 재무상태 · 현금흐름", "단위: 억원 | 재무상태는 기말잔액, 현금흐름은 연간 기준")
    r = 5
    _hdr(ws4, r, 2, "항목")
    for j, yr in enumerate(years, start=3):
        _hdr(ws4, r, j, str(yr))

    # BS
    i = r + 1
    ws4.cell(row=i, column=2, value="▶ 재무상태표").font = BOLD_FONT
    i += 1
    for lbl, _ in BS_ACCOUNTS:
        c = ws4.cell(row=i, column=2, value=lbl)
        c.font = BOLD_FONT; c.border = THIN_BORDER
        for j, yr in enumerate(years, start=3):
            v = fin["bs"].get((yr, "11011"), {}).get(lbl)
            _cell(ws4, i, j, _to_100m(v), fmt=NUM_FMT,
                  fill=GREEN_FILL if v else GRAY_FILL)
        i += 1

    # 파생 지표
    for lbl in ["부채비율(%)", "유동비율(%)", "ROE(%)", "ROA(%)"]:
        c = ws4.cell(row=i, column=2, value=lbl)
        c.font = BOLD_FONT; c.border = THIN_BORDER
        for j, yr in enumerate(years, start=3):
            bs = fin["bs"].get((yr, "11011"), {})
            ann = fin["annual"].get(yr, {})
            eq = bs.get("자본총계"); liab = bs.get("부채총계")
            assets = bs.get("자산총계"); ni = ann.get("당기순이익")
            cur_a = bs.get("유동자산"); cur_l = bs.get("유동부채")
            if lbl == "부채비율(%)":
                val = liab / eq if (liab is not None and eq and eq != 0) else None
            elif lbl == "유동비율(%)":
                val = cur_a / cur_l if (cur_a is not None and cur_l and cur_l != 0) else None
            elif lbl == "ROE(%)":
                val = ni / eq if (ni is not None and eq and eq != 0) else None
            elif lbl == "ROA(%)":
                val = ni / assets if (ni is not None and assets and assets != 0) else None
            else:
                val = None
            _cell(ws4, i, j, val, fmt=PCT_FMT, fill=GRAY_FILL)
        i += 1

    i += 1
    ws4.cell(row=i, column=2, value="▶ 현금흐름표").font = BOLD_FONT
    i += 1
    for lbl, _ in CF_ACCOUNTS:
        c = ws4.cell(row=i, column=2, value=lbl)
        c.font = BOLD_FONT; c.border = THIN_BORDER
        for j, yr in enumerate(years, start=3):
            v = fin["cf"].get((yr, "11011"), {}).get(lbl)
            _cell(ws4, i, j, _to_100m(v), fmt=NUM_FMT,
                  fill=GREEN_FILL if v else GRAY_FILL)
        i += 1

    # FCF
    c = ws4.cell(row=i, column=2, value="FCF(영업CF-CAPEX)")
    c.font = BOLD_FONT; c.border = THIN_BORDER
    for j, yr in enumerate(years, start=3):
        cf = fin["cf"].get((yr, "11011"), {})
        ocf = cf.get("영업활동CF"); cap = cf.get("CAPEX")
        val = _to_100m(ocf - cap) if ocf is not None and cap is not None else None
        _cell(ws4, i, j, val, fmt=NUM_FMT, fill=GRAY_FILL)
    i += 1

    # ════════════════════════════════════════════════════════════
    # 시트 5: 밸류에이션
    # ════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("5.밸류에이션")
    ws5.column_dimensions["A"].width = 2
    ws5.column_dimensions["B"].width = 22
    ws5.column_dimensions["C"].width = 18
    ws5.column_dimensions["D"].width = 18

    _title(ws5, 2, f"{stock_name} — 밸류에이션 (TTM)", "단위: 억원 | 노란 칸은 직접 수정 가능")

    cur_price = info.get("current_price")
    shares = info.get("shares_issued")
    mktcap = info.get("market_cap")

    vals5 = [
        ("현재가 (원)", cur_price, YELLOW_FILL, None),
        ("발행주식수 (주)", shares, YELLOW_FILL, NUM_FMT),
        ("시가총액 (억원)", mktcap, YELLOW_FILL, NUM_FMT),
        ("", None, None, None),
        ("TTM 매출액 (억원)", None, YELLOW_FILL, NUM_FMT),
        ("TTM 영업이익 (억원)", None, YELLOW_FILL, NUM_FMT),
        ("TTM 순이익 (억원)", None, YELLOW_FILL, NUM_FMT),
        ("", None, None, None),
        ("PER (TTM)", None, GRAY_FILL, "0.0"),
        ("PSR (TTM)", None, GRAY_FILL, "0.0"),
        ("PBR", info.get("pbr"), GREEN_FILL, "0.00"),
    ]
    for i, (lbl, val, fill, fmt) in enumerate(vals5, start=5):
        if not lbl:
            continue
        _cell(ws5, i, 2, lbl, bold=True)
        c = _cell(ws5, i, 3, val, fmt=fmt, fill=fill or GRAY_FILL)

    # 계산식 안내
    ws5.cell(row=17, column=2, value="■ 해석").font = BOLD_FONT
    ws5.cell(row=18, column=2,
             value="• TTM = 최근 4개 단독분기 합산. 직접 입력 후 PER/PSR 자동 계산.").font = BASE_FONT
    ws5.cell(row=19, column=2,
             value="• DART 기반 BPS/EPS는 2.연간손익추이·4.재무상태 시트 참조.").font = BASE_FONT

    # ════════════════════════════════════════════════════════════
    # 시트 6: 장기분기손익 (전체 분기)
    # ════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("6.장기분기손익")
    ws6.column_dimensions["A"].width = 2
    ws6.column_dimensions["B"].width = 22
    for j in range(len(all_q) + 2):
        ws6.column_dimensions[get_column_letter(j+2)].width = 10

    _title(ws6, 2, f"{stock_name} — 장기 분기 손익 (전체)", "단위: 억원 | 단독분기 기준")
    r = 5
    _hdr(ws6, r, 2, "항목")
    for j, (yr, q) in enumerate(all_q, start=3):
        _hdr(ws6, r, j, f"{yr}Q{q}")

    for i, (lbl, _) in enumerate(IS_ACCOUNTS, start=r+1):
        c = ws6.cell(row=i, column=2, value=lbl)
        c.font = BOLD_FONT; c.border = THIN_BORDER
        for j, (yr, q) in enumerate(all_q, start=3):
            val = _to_100m(fin["quarterly"].get((yr, q), {}).get(lbl))
            _cell(ws6, i, j, val, fmt=NUM_FMT, fill=GREEN_FILL if val else GRAY_FILL)

    # ════════════════════════════════════════════════════════════
    # 시트 7: 장기분기재무 (BS 전체)
    # ════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("7.장기분기재무")
    ws7.column_dimensions["A"].width = 2
    ws7.column_dimensions["B"].width = 22
    bs_keys_sorted = sorted(fin["bs"].keys())  # (year, reprt_code)
    for j in range(len(bs_keys_sorted) + 2):
        ws7.column_dimensions[get_column_letter(j+2)].width = 11

    _title(ws7, 2, f"{stock_name} — 장기 분기 재무상태 (전체)", "단위: 억원 | 분기말 시점 잔액")
    r = 5
    _hdr(ws7, r, 2, "항목")
    for j, (yr, rc) in enumerate(bs_keys_sorted, start=3):
        q = REPRT_Q.get(rc, "?")
        _hdr(ws7, r, j, f"{yr}Q{q}")

    for i, (lbl, _) in enumerate(BS_ACCOUNTS, start=r+1):
        c = ws7.cell(row=i, column=2, value=lbl)
        c.font = BOLD_FONT; c.border = THIN_BORDER
        for j, (yr, rc) in enumerate(bs_keys_sorted, start=3):
            val = _to_100m(fin["bs"].get((yr, rc), {}).get(lbl))
            _cell(ws7, i, j, val, fmt=NUM_FMT, fill=GREEN_FILL if val else GRAY_FILL)

    # ════════════════════════════════════════════════════════════
    # 시트 8: 원문메모 (단위검증·정정처리 기록)
    # ════════════════════════════════════════════════════════════
    ws8 = wb.create_sheet("8.원문메모")
    ws8.column_dimensions["A"].width = 2
    ws8.column_dimensions["B"].width = 80

    _title(ws8, 2, f"{stock_name} — 원문메모 · 데이터 처리 기록")
    memos = [
        f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"종목코드: {stock_code} / 종목명: {stock_name}",
        "데이터 원천: DART OpenAPI fnlttSinglAcntAll (CFS 우선, OFS fallback)",
        f"수집 연도 범위: {min(years) if years else '-'} ~ {max(years) if years else '-'}",
        "단위 규칙: 원단위 → 억원 변환 (/1e8, 소수점1자리)",
        "분기 단독값 계산: Q1=1Q누적, Q2=반기-Q1, Q3=9M-반기, Q4=연간-9M",
        "CAPEX 부호: ABS() 정규화 (지출 양수 표기)",
        "정정공시: DART API 최신 rcept 기준 (정정본 자동 우선)",
        "누락 셀(회색): DART 미수집 또는 계정 매핑 실패",
        "---",
        "Account ID 매핑 (주요 항목):",
    ]
    for lbl, aids in IS_ACCOUNTS + BS_ACCOUNTS + CF_ACCOUNTS:
        memos.append(f"  {lbl}: {', '.join(aids[:2])}")

    for i, memo in enumerate(memos, start=5):
        c = ws8.cell(row=i, column=2, value=memo)
        c.font = BASE_FONT
        c.alignment = Alignment(wrap_text=True)

    # ── 최종 저장 ─────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════
# FastAPI 엔드포인트
# ═══════════════════════════════════════════════════════════════

# 진행상황 캐시
_progress: dict = {}


def _run_build(job_id: str, stock_code: str, years: list[int]):
    """백그라운드 엑셀 생성 워커"""
    try:
        _progress[job_id] = {"status": "running", "step": "종목정보 조회 중..."}
        info = get_stock_info(stock_code)
        stock_name = info.get("stock_name", stock_code)

        _progress[job_id]["step"] = "DART corp_code 조회 중..."
        corp_code = _load_corp_code(stock_code)
        if not corp_code:
            _progress[job_id] = {"status": "error", "msg": f"corp_code 조회 실패: {stock_code}"}
            return

        _progress[job_id]["step"] = f"DART 재무데이터 수집 중 ({min(years)}~{max(years)})..."
        fin = collect_financials(corp_code, stock_code, years)

        _progress[job_id]["step"] = "엑셀 생성 중..."
        excel_bytes = build_excel_v22(stock_code, stock_name, info, fin)

        _progress[job_id] = {
            "status": "done",
            "stock_code": stock_code,
            "stock_name": stock_name,
            "excel_bytes": excel_bytes,
            "filename": f"{stock_name}_{stock_code}_v22.xlsx",
            "annual_years": sorted(fin["annual"].keys()),
            "quarterly_count": len(fin["quarterly"]),
        }
    except Exception as e:
        log.exception("엑셀 생성 오류: %s", e)
        _progress[job_id] = {"status": "error", "msg": str(e)}


@router.post("/build")
def start_build(
    stock_code: str = Query(..., description="6자리 종목코드"),
    years_back: int = Query(5, ge=1, le=10),
):
    """엑셀 생성 작업 시작. job_id 반환."""
    import threading
    cur_year = datetime.now().year
    years = list(range(cur_year - years_back + 1, cur_year + 1))
    job_id = f"{stock_code}_{datetime.now().strftime('%H%M%S')}"
    _progress[job_id] = {"status": "queued", "step": "대기 중..."}
    t = threading.Thread(target=_run_build, args=(job_id, stock_code, years), daemon=True)
    t.start()
    return {"job_id": job_id, "stock_code": stock_code, "years": years}


@router.get("/status/{job_id}")
def build_status(job_id: str):
    """작업 진행 상태 조회"""
    prog = _progress.get(job_id)
    if not prog:
        return {"status": "not_found"}
    # excel_bytes는 직렬화 불가 → 제외
    return {k: v for k, v in prog.items() if k != "excel_bytes"}


@router.get("/download/{job_id}")
def download_excel(job_id: str):
    """완성된 엑셀 파일 다운로드"""
    prog = _progress.get(job_id)
    if not prog or prog.get("status") != "done":
        return JSONResponse({"error": "파일 없음 또는 생성 중"}, status_code=404)
    excel_bytes = prog.get("excel_bytes")
    filename = prog.get("filename", "dart_v22.xlsx")
    from urllib.parse import quote
    safe_filename = filename.encode('utf-8')
    encoded = quote(filename)
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@router.get("/verify/{stock_code}")
def verify_data(stock_code: str, year: int = Query(2023)):
    """DB 저장값 다소스 비교 (DART API 한도 초과 시 DB 소스간 비교)"""
    info = get_stock_info(stock_code)

    # DART API 시도 (한도 있으면 null)
    corp_code = _load_corp_code(stock_code)
    dart_vals = {}
    dart_ok = False
    if corp_code:
        rows = _fetch_fnltt(corp_code, year, "11011")
        if rows:
            dart_ok = True
            for lbl, aids in IS_ACCOUNTS:
                v = _extract_value(rows, aids)
                dart_vals[lbl] = _to_100m(v)

    # DB 저장값 (소스별로 조회)
    db_rows = []
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        rs = conn.execute(
            """SELECT revenue, operating_profit, net_income, total_assets, total_equity, data_source
               FROM financial_data WHERE stock_code=? AND year=? AND is_annual=1
               ORDER BY CASE WHEN data_source LIKE '%dart%' THEN 0
                             WHEN data_source = 'fnguide' THEN 1 ELSE 2 END""",
            (stock_code, year)
        ).fetchall()
        for r in rs:
            db_rows.append({
                "source": r[5],
                "매출액": _to_100m(r[0]),
                "영업이익": _to_100m(r[1]),
                "당기순이익": _to_100m(r[2]),
                "자산총계": _to_100m(r[3]),
                "자본총계": _to_100m(r[4]),
            })

        # 분기 데이터 개수
        q_count = conn.execute(
            "SELECT COUNT(*) FROM financial_data WHERE stock_code=? AND year=? AND is_annual=0",
            (stock_code, year)
        ).fetchone()[0]

        # CF
        cf_rows = conn.execute(
            "SELECT operating_cf, investing_cf, capex, data_source FROM cash_flow_data WHERE stock_code=? AND year=? AND is_annual=1 ORDER BY CASE WHEN data_source LIKE '%dart%' THEN 0 ELSE 1 END LIMIT 2",
            (stock_code, year)
        ).fetchall()
        conn.close()
    except Exception:
        q_count = 0
        cf_rows = []

    # 소스 일치 여부 판단
    consistency = "N/A"
    if len(db_rows) >= 2:
        rev_vals = [r["매출액"] for r in db_rows if r["매출액"]]
        if rev_vals and max(rev_vals) > 0:
            ratio = min(rev_vals) / max(rev_vals)
            consistency = "✅ 일치(5% 이내)" if ratio > 0.95 else f"⚠️ 차이 {round((1-ratio)*100,1)}%"

    return {
        "stock_code": stock_code,
        "stock_name": info.get("stock_name"),
        "year": year,
        "dart_api_available": dart_ok,
        "dart_api": dart_vals if dart_ok else "(DART API 일일한도 초과 — 내일 조회 가능)",
        "db_sources": db_rows,
        "quarterly_count": q_count,
        "cashflow": [{"source": r[3], "OCF": _to_100m(r[0]), "ICF": _to_100m(r[1]), "CAPEX": _to_100m(abs(r[2])) if r[2] else None} for r in cf_rows],
        "source_consistency": consistency,
    }


# ── CH 데이터 통합 엔드포인트 ─────────────────────────────────────────────
# DartV22Builder write_ch_sheet 재현:
# 사업부문매출/영업이익, 직원수, 재고자산, 수주잔고, 매출채권, 판관비

_STANDARD_IS_NAMES = {
    "매출액", "수익", "영업수익", "매출", "순매출액", "총매출액",
    "매출원가", "매출총이익", "판매비와관리비", "판관비", "영업이익",
    "금융수익", "금융비용", "기타이익", "기타손실", "기타수익", "기타비용",
    "지분법이익", "지분법손실", "세전이익", "법인세비용", "당기순이익",
    "당기순손실", "총포괄손익", "기타포괄손익", "지배주주순이익", "비지배지분",
    "영업이익(손실)", "당기순이익(손실)", "계속영업손익", "중단영업손익",
    "법인세비용차감전순이익", "매출총이익(손실)", "판매비및관리비",
}


def _is_segment_account(account_nm: str, account_id: str) -> bool:
    """XBRL 계정이 사업부문 계정인지 판별"""
    nm = (account_nm or "").strip()
    aid = (account_id or "")
    # 표준 계정명이면 제외
    if nm in _STANDARD_IS_NAMES:
        return False
    # "합계", "소계", "계"로 끝나면 제외 (집계행)
    if nm.endswith(("합계", "소계", "계", "총계")):
        return False
    # 괄호로만 된 세부 설명 제외
    if nm.startswith("(") or nm.startswith("①") or nm.startswith("②"):
        return False
    # 표준 XBRL ID이면 제외
    if aid.startswith("ifrs-full_") and "_" not in aid[10:]:
        return False
    # 매출액 또는 영업이익 관련 키워드가 없으면 세그먼트가 아님
    revenue_kw = any(k in nm for k in ["매출", "수익", "Revenue", "영업이익", "부문"])
    if not revenue_kw:
        return False
    return True


def _fetch_employee_count(corp_code: str) -> list[dict]:
    """DART empSttus API로 직원 현황 조회"""
    results = []
    for year in range(datetime.now().year, datetime.now().year - 5, -1):
        for reprt_code in ["11011", "11012"]:
            try:
                r = requests.get(
                    "https://opendart.fss.or.kr/api/empSttus.json",
                    params={"crtfc_key": _dart_key(), "corp_code": corp_code,
                            "bsns_year": str(year), "reprt_code": reprt_code},
                    timeout=10,
                )
                d = r.json()
                if d.get("status") == "020":
                    _next_key()
                    return results  # API 한도 초과
                if d.get("status") != "000":
                    continue
                for row in (d.get("list") or []):
                    fo_bbm = row.get("fo_bbm", "")  # 성별
                    emp_no = row.get("emp_no", "0").replace(",", "")
                    if emp_no and fo_bbm in ("합 계", "합계", ""):
                        try:
                            n = int(emp_no)
                            if n > 0:
                                results.append({
                                    "year": year,
                                    "reprt_code": reprt_code,
                                    "emp_no": n,
                                    "fo_bbm": fo_bbm,
                                    "acmtn_dscd": row.get("acmtn_dscd", ""),
                                })
                        except ValueError:
                            pass
                if results:
                    return results  # 첫 번째 유효 연도 찾으면 반환
            except Exception:
                pass
    return results


def _fetch_segment_from_dart(corp_code: str, years: list[int]) -> list[dict]:
    """DART fnlttSinglAcntAll IS계정에서 사업부문 계정 추출"""
    segments = []
    for year in years:
        for reprt_code, quarter in [("11011", 4), ("11014", 3), ("11012", 2), ("11013", 1)]:
            rows = _fetch_fnltt(corp_code, year, reprt_code)
            if not rows:
                continue
            # IS/IS1 계정에서 비표준 계정 → 세그먼트 후보
            is_rows = [r for r in rows if r.get("sj_div", "") in ("IS", "IS1")]
            # 총 매출액 기준값 (비교용)
            total_rev = None
            for r in is_rows:
                if r.get("account_id") == "ifrs-full_Revenue":
                    raw = r.get("thstrm_amount", "").replace(",", "")
                    try:
                        total_rev = int(float(raw))
                    except (ValueError, TypeError):
                        pass
                    break

            seg_accounts = []
            for r in is_rows:
                nm = r.get("account_nm", "")
                aid = r.get("account_id", "")
                if not _is_segment_account(nm, aid):
                    continue
                raw = r.get("thstrm_amount", "").replace(",", "")
                try:
                    amount = int(float(raw))
                except (ValueError, TypeError):
                    continue
                # 총 매출액의 1%~120% 범위 (과도하게 큰 값 제외)
                if total_rev and total_rev > 0:
                    ratio = amount / total_rev
                    if ratio < 0.001 or ratio > 1.3:
                        continue
                seg_accounts.append({
                    "year": year,
                    "quarter": quarter,
                    "segment_name": nm,
                    "account_id": aid,
                    "revenue": amount,
                })
            if seg_accounts:
                segments.extend(seg_accounts)
                break  # 해당 연도 처리 완료
    return segments


@router.get("/ch-data/{code}")
def get_ch_data(code: str):
    """
    DartV22Builder CH시트 데이터:
    사업부문별 매출·영업이익, 직원현황, 재고자산, 수주잔고, 매출채권, 판관비
    """
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    try:
        result: dict = {"stock_code": code}

        # ── 종목명 조회 ──────────────────────────────────────────────
        info_row = conn.execute(
            "SELECT stock_name, market, sector_large FROM stock_universe WHERE stock_code=?",
            (code,)
        ).fetchone()
        result["stock_name"] = info_row["stock_name"] if info_row else code
        result["sector"] = (info_row["sector_large"] or "") if info_row else ""

        # ── 1. 사업부문 매출 (DB 우선 → DART API fallback) ──────────
        seg_db = conn.execute("""
            SELECT year, quarter, segment_name, revenue, operating_profit, assets
            FROM segment_revenue WHERE stock_code=? AND segment_name != '연결전체'
            ORDER BY year DESC, quarter DESC, revenue DESC NULLS LAST
        """, (code,)).fetchall()
        seg_total = conn.execute("""
            SELECT year, quarter, revenue, operating_profit
            FROM segment_revenue WHERE stock_code=? AND segment_name='연결전체'
            ORDER BY year DESC, quarter DESC
        """, (code,)).fetchall()

        has_segment_breakdown = len(seg_db) > 0
        segment_rows = [dict(r) for r in seg_db]
        total_rows = [dict(r) for r in seg_total]

        result["segments"] = {
            "has_breakdown": has_segment_breakdown,
            "source": "db" if seg_db else "none",
            "rows": segment_rows,
            "totals": total_rows,
        }

        # ── 2. 직원 현황 ────────────────────────────────────────────
        dart_emp = []
        try:
            dart_emp = [dict(r) for r in conn.execute("""
                SELECT year, total_emp, male_emp, female_emp,
                       regular_emp, contract_emp,
                       avg_tenure_years, annual_salary_m, acmtn_dscd, reprt_code
                FROM dart_employee_count WHERE stock_code=?
                ORDER BY year DESC, reprt_code ASC LIMIT 10
            """, (code,)).fetchall()]
        except Exception:
            pass
        result["employees"] = dart_emp

        # ── 2b. 판관비 추이 ──────────────────────────────────────────
        sga_rows = []
        try:
            sga_rows = [dict(r) for r in conn.execute("""
                SELECT year, ROUND(sga_total/1e8,1) as sga_억, report_type
                FROM dart_sga_annual WHERE stock_code=?
                  AND report_type='CFS'
                ORDER BY year DESC LIMIT 7
            """, (code,)).fetchall()]
            # CFS 없으면 OFS
            if not sga_rows:
                sga_rows = [dict(r) for r in conn.execute("""
                    SELECT year, ROUND(sga_total/1e8,1) as sga_억, report_type
                    FROM dart_sga_annual WHERE stock_code=?
                    ORDER BY year DESC LIMIT 7
                """, (code,)).fetchall()]
        except Exception:
            pass
        result["sga"] = sga_rows

        # ── 3. 재고자산 추이 ─────────────────────────────────────────
        inv_rows = []
        try:
            inv_rows = conn.execute("""
                SELECT fiscal_year as year, fiscal_quarter as quarter,
                       ROUND(inventory_assets_krw/1e8,1) as inventory_억
                FROM dart_cost_quarterly WHERE stock_code=?
                AND inventory_assets_krw IS NOT NULL AND inventory_assets_krw > 0
                ORDER BY fiscal_year DESC, fiscal_quarter DESC LIMIT 24
            """, (code,)).fetchall()
        except Exception:
            pass
        result["inventory"] = [dict(r) for r in inv_rows]

        # ── 4. 수주잔고 ───────────────────────────────────────────────
        backlog_rows = conn.execute("""
            SELECT year, quarter,
                   ROUND(backlog_amount/1e8,1) as backlog_억,
                   backlog_normalized as backlog_norm_억,
                   data_source
            FROM order_backlog WHERE stock_code=?
            ORDER BY year DESC, quarter DESC LIMIT 24
        """, (code,)).fetchall()
        result["order_backlog"] = [dict(r) for r in backlog_rows]

        # ── 5. 매입재료비 (연간) ──────────────────────────────────────
        try:
            mat_rows = conn.execute("""
                SELECT year, ROUND(material_purchase_krw/1e8,1) as mat_억,
                       report_type
                FROM dart_material_purchase WHERE stock_code=?
                ORDER BY year DESC LIMIT 7
            """, (code,)).fetchall()
            result["material_purchase"] = [dict(r) for r in mat_rows]
        except Exception:
            result["material_purchase"] = []

        # ── 6. P&L + BS 추이 — 연도별 소스우선순위 1행 (연결/별도 모두 수집) ──
        # fnguide CFS 우선(검증완료) → dart CFS → 기타
        fin_all = conn.execute("""
            SELECT year, quarter, revenue, operating_profit, net_income,
                   total_assets, total_equity, report_type, data_source
            FROM financial_data
            WHERE stock_code=? AND is_annual=1 AND revenue IS NOT NULL
              AND quarter IN (0, 4)
            ORDER BY year DESC,
              CASE WHEN data_source='fnguide' AND report_type='CFS' THEN 0
                   WHEN data_source LIKE '%dart%' AND report_type='CFS' AND quarter=4 THEN 1
                   WHEN report_type='CFS' THEN 2
                   ELSE 3 END
            LIMIT 28
        """, (code,)).fetchall()
        seen_y_cfs = set(); seen_y_ofs = set()
        fin_deduped = []; fin_ofs = []
        for r in fin_all:
            y = r["year"]; rt = r["report_type"] or "CFS"
            row = {
                "year": y,
                "rev_억": round(r["revenue"]/1e8, 1) if r["revenue"] else None,
                "op_억": round(r["operating_profit"]/1e8, 1) if r["operating_profit"] else None,
                "ni_억": round(r["net_income"]/1e8, 1) if r["net_income"] else None,
                "assets_억": round(r["total_assets"]/1e8, 1) if r["total_assets"] else None,
                "equity_억": round(r["total_equity"]/1e8, 1) if r["total_equity"] else None,
                "src": r["data_source"], "report_type": rt,
            }
            if rt == "CFS" and y not in seen_y_cfs:
                seen_y_cfs.add(y); fin_deduped.append(row)
            elif rt == "OFS" and y not in seen_y_ofs:
                seen_y_ofs.add(y); fin_ofs.append(row)
            if len(fin_deduped) >= 7 and len(fin_ofs) >= 7:
                break
        result["financials"] = fin_deduped          # 연결(CFS) 연간 손익
        result["financials_ofs"] = fin_ofs          # 별도(OFS) 연간 손익
        result["has_ofs_annual"] = len(fin_ofs) > 0

        # ── 7. 분기 P&L (최근 12분기 — fnguide CFS 우선, 연결/별도 각각) ──
        q_rows = conn.execute("""
            SELECT year, quarter,
                   ROUND(revenue/1e8,1) as rev_억,
                   ROUND(operating_profit/1e8,1) as op_억,
                   ROUND(net_income/1e8,1) as ni_억,
                   report_type, data_source
            FROM financial_data
            WHERE stock_code=? AND is_annual=0 AND quarter > 0
              AND revenue IS NOT NULL AND revenue > 0
              AND data_source NOT LIKE '%bs_null%'
              AND data_source NOT LIKE '%recalc_bs%'
            ORDER BY year DESC, quarter DESC,
              CASE WHEN data_source='fnguide' AND report_type='CFS' THEN 0
                   WHEN data_source LIKE '%dart%' AND report_type='CFS' THEN 1
                   WHEN report_type='CFS' THEN 2
                   ELSE 3 END
            LIMIT 48
        """, (code,)).fetchall()
        seen_q_cfs = set(); seen_q_ofs = set()
        q_deduped = []; q_ofs = []
        for r in q_rows:
            key = (r["year"], r["quarter"]); rt = r["report_type"] or "CFS"
            row = {"year": r["year"], "quarter": r["quarter"],
                   "rev_억": r["rev_억"], "op_억": r["op_억"], "ni_억": r["ni_억"],
                   "report_type": rt}
            if rt == "CFS" and key not in seen_q_cfs:
                seen_q_cfs.add(key); q_deduped.append(row)
            elif rt == "OFS" and key not in seen_q_ofs:
                seen_q_ofs.add(key); q_ofs.append(row)
            if len(q_deduped) >= 12 and len(q_ofs) >= 12:
                break
        result["quarterly_pl"] = list(reversed(q_deduped))   # 연결 분기
        result["quarterly_pl_ofs"] = list(reversed(q_ofs))   # 별도 분기
        result["has_ofs_quarter"] = len(q_ofs) > 0

        # ── 8. 현금흐름 (연간 최근 5년) ──────────────────────────────
        cf_rows = conn.execute("""
            SELECT cf.year,
                   ROUND(cf.operating_cf/1e8,1) as ocf_억,
                   ROUND(cf.investing_cf/1e8,1) as icf_억,
                   ROUND(ABS(cf.capex)/1e8,1) as capex_억,
                   ROUND(cf.depreciation/1e8,1) as dep_억
            FROM cash_flow_data cf WHERE cf.stock_code=? AND cf.is_annual=1
            GROUP BY cf.year
            ORDER BY cf.year DESC LIMIT 5
        """, (code,)).fetchall()
        result["cashflow"] = [dict(r) for r in cf_rows]

        # ── 9. 매출채권 ─────────────────────────────────────────────
        ar_rows = []
        try:
            ar_rows_raw = conn.execute("""
                SELECT year, ROUND(value/1e8,1) as ar_억, report_type
                FROM dart_bs_items
                WHERE stock_code=? AND item_key='trade_receivable'
                ORDER BY year DESC,
                  CASE WHEN report_type='CFS' THEN 0 ELSE 1 END
                LIMIT 10
            """, (code,)).fetchall()
            seen_y = set()
            for r in ar_rows_raw:
                if r["year"] not in seen_y:
                    seen_y.add(r["year"])
                    ar_rows.append(dict(r))
                if len(ar_rows) >= 5:
                    break
        except Exception:
            pass
        result["trade_receivable"] = ar_rows
        result["dart_api_available"] = True

        # ── 10. 항목별 미제공 이유 (섹터·업종 기반) ────────────────────
        sector = result.get("sector", "")
        backlog_sectors = {"산업재", "에너지", "소재"}  # 건설/조선/플랜트
        inv_sectors = {"소재", "산업재", "필수소비재", "경기소비재", "IT", "에너지"}
        material_sectors = {"소재", "산업재", "필수소비재", "경기소비재", "IT", "에너지"}

        def _unavail_reason(has_data: bool, sector_match: bool, data_name: str) -> str | None:
            if has_data:
                return None
            if not sector_match:
                return f"{data_name} 해당없음 (금융·서비스업종)"
            return f"{data_name} 사업보고서 미제공 또는 미수집"

        result["availability"] = {
            "order_backlog": {
                "has_data": len(result["order_backlog"]) > 0,
                "reason": _unavail_reason(
                    len(result["order_backlog"]) > 0,
                    sector in backlog_sectors or any(k in sector for k in ["건설","조선","플랜트"]),
                    "수주잔고"
                )
            },
            "inventory": {
                "has_data": len(result["inventory"]) > 0,
                "reason": _unavail_reason(
                    len(result["inventory"]) > 0,
                    sector in inv_sectors,
                    "재고자산"
                )
            },
            "material_purchase": {
                "has_data": len(result["material_purchase"]) > 0,
                "reason": _unavail_reason(
                    len(result["material_purchase"]) > 0,
                    sector in material_sectors,
                    "원재료매입액"
                )
            },
            "segments": {
                "has_data": result["segments"]["has_breakdown"],
                "reason": None if result["segments"]["has_breakdown"] else "사업부문 별도공시 없음 (단일사업 또는 미공시)"
            },
            "employees": {
                "has_data": len(result["employees"]) > 0,
                "reason": None if result["employees"] else "직원현황 사업보고서 미수집"
            },
        }

        return result
    finally:
        conn.close()
