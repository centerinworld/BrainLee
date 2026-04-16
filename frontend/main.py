from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
import models, schemas, crud, processor, screener, ai_analyzer
from database import get_db, engine
import logging
from datetime import datetime
from fastapi.middleware.cors import CORSMiddleware

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 데이터베이스 테이블 생성 (상시 동기화)
models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="주식 분석 백엔드 (프로젝트 안티그래비티)")

# ═══════════════════════════════════════════════════════
#  전역 상태 — startup_event 이전에 반드시 정의
# ═══════════════════════════════════════════════════════
import threading as _th
import time as _tm

_collecting: dict      = {}   # stock_code -> "running"|"done"
_valuation_cache: dict = {}   # stock_code -> {per,pbr,cached_at,...}
_market_info_cache: dict = {} # stock_code -> {market,mktcap,mktcap_rank,cached_at}


def _get_cached_valuation(stock_code: str) -> dict:
    entry = _valuation_cache.get(stock_code)
    if entry and (_tm.time() - entry.get("cached_at", 0)) < 86400:
        return entry
    return {}


def _scrape_naver(stock_code: str) -> dict:
    """네이버 금융 PER/PBR 스크래핑. URL: .naver (구 .nhn 아님), EUC-KR."""
    empty = {"per": None, "pbr": None, "trailing_eps": None,
             "forward_per": None, "source": None}
    if not (stock_code and stock_code.isdigit() and len(stock_code) == 6):
        return empty
    try:
        import requests as _r
        from bs4 import BeautifulSoup
        res = _r.get(
            f"https://finance.naver.com/item/main.naver?code={stock_code}",
            headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                                   "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                     "Referer": "https://finance.naver.com/"},
            timeout=8,
        )
        res.encoding = "euc-kr"
        soup = BeautifulSoup(res.text, "html.parser")
        def _n(sel):
            t = soup.select_one(sel)
            if not t: return None
            s = t.text.strip().replace(",","").replace("N/A","").strip()
            try: return float(s) if s else None
            except: return None
        per, pbr, eps = _n("em#_per"), _n("em#_pbr"), _n("em#_eps")
        if per is None and pbr is None and eps is None:
            logger.warning(f"[Naver] {stock_code}: 파싱 없음")
            return empty
        logger.info(f"[Naver] {stock_code}: PER={per} PBR={pbr}")
        return {"per": per, "pbr": pbr, "trailing_eps": eps, "forward_per": None, "source": "네이버금융"}
    except Exception as e:
        logger.warning(f"[Naver] {stock_code}: {e}")
        return empty


def _get_latest_disclosed_quarter():
    """현재 시점에 DART에 실제 공시된 최신 분기 반환 (year, quarter).
    
    공시 일정 (법정 제출 기한):
      Q1 (1분기보고서): 5월 15일
      Q2 (반기보고서):  8월 14일
      Q3 (3분기보고서): 11월 14일
      Q4 (사업보고서):  3월 31일 (단, 12월 결산법인 기준)
    
    → 3월부터는 전년도 Q4가 공시되기 시작하므로 Q4 포함
    """
    from datetime import date
    m = date.today().month
    d = date.today().day
    y = date.today().year
    if m >= 11:              return (y, 3)
    elif m >= 8:             return (y, 2)
    elif m >= 5:             return (y, 1)
    elif m >= 3:             return (y - 1, 4)   # 3월부터 전년 Q4 공시 시작
    elif m >= 4:             return (y - 1, 4)   # 4월: 전년 Q4 확정
    else:                    return (y - 1, 3)


def _has_any_financial(stock_code: str, db) -> bool:
    """DB에 해당 종목 재무 데이터가 1건이라도 있는지 직접 확인."""
    return db.query(models.FinancialData).filter(
        models.FinancialData.stock_code == stock_code
    ).first() is not None


def _collect_dart_to_db(stock_code: str, db, latest_only: bool = False) -> int:
    """
    DART에서 재무제표를 수집해 DB에 직접 저장 (동기).
    httpx 자기참조 없이 crud.upsert_financial_data 직접 호출.
    반환: 저장 건수
    """
    try:
        import config as _cfg, pandas as pd
        import OpenDartReader
        dart = OpenDartReader(_cfg.DART_API_KEY)
    except Exception as e:
        logger.error(f"[DART] 초기화 실패: {e}")
        return 0

    latest_y, latest_q = _get_latest_disclosed_quarter()
    # ★ 사업보고서(11011)를 먼저 처리해야 Q4 분기 자동계산이 올바르게 됨
    qmap = {"11011": 4, "11013": 1, "11012": 2, "11014": 3}
    # latest_only=True: 최근 2년 / False: 최근 10년
    years = list(range(latest_y, latest_y - (2 if latest_only else 10), -1))

    saved = 0
    for year in years:
        for rcode, qnum in qmap.items():
            # 아직 공시 안 된 분기 스킵
            if year > latest_y or (year == latest_y and qnum > latest_q):
                continue
            # 이미 DB에 있으면 스킵 (연간/분기 구분해서 체크)
            is_annual_flag = (rcode == "11011")
            exists = db.query(models.FinancialData).filter(
                models.FinancialData.stock_code == stock_code,
                models.FinancialData.year       == year,
                models.FinancialData.quarter    == qnum,
                models.FinancialData.is_annual  == is_annual_flag,
            ).first() is not None
            if exists:
                continue

            fn_data = None
            for fs in ["CFS", "OFS"]:
                try:
                    df = dart.finstate_all(stock_code, year, rcode, fs_div=fs)
                    if df is not None and not df.empty:
                        fn_data = df; break
                except Exception:
                    pass
            if fn_data is None or fn_data.empty:
                try:
                    fn_data = dart.finstate(stock_code, year, rcode)
                except Exception:
                    pass
            if fn_data is None or fn_data.empty:
                continue

            m = {k: 0.0 for k in [
                "revenue","operating_profit","net_income","total_assets",
                "total_liabilities","total_equity","capital_stock",
                "eps","bps","dps","cash","total_shares","depreciation_amortization"]}
            for _, row in fn_data.iterrows():
                acc = str(row.get("account_nm","")).replace(" ","")
                vc  = "thstrm_amount"
                if not acc or vc not in row or pd.isna(row[vc]): continue
                try: val = float(str(row[vc]).replace(",",""))
                except: continue
                if   "매출액" in acc or "영업수익" in acc:                             m["revenue"] = val
                elif "영업이익" in acc:                                                m["operating_profit"] = val
                elif "당기순이익" in acc and "주당" not in acc:                        m["net_income"] = val
                elif "자산총계" in acc:                                                m["total_assets"] = val
                elif "부채총계" in acc:                                                m["total_liabilities"] = val
                elif "자본총계" in acc:                                                m["total_equity"] = val
                elif "자본금" in acc:                                                  m["capital_stock"] = val
                elif any(k in acc for k in ["기본주당순이익","주당순이익","기본EPS"]): m["eps"] = val
                elif any(k in acc for k in ["주당순자산","1주당순자산가액"]):           m["bps"] = val
                elif any(k in acc for k in ["주당배당금","주당현금배당금"]):            m["dps"] = val
                elif any(k in acc for k in ["현금및현금성자산","현금성자산"]):          m["cash"] = val
            try:
                fin = schemas.FinancialIngest(
                    stock_code=stock_code, year=year, quarter=qnum,
                    is_annual=(rcode == "11011"), **m)
                crud.upsert_financial_data(db, fin)
                saved += 1
                logger.info(f"[DART] {stock_code} {year}Q{qnum} 저장")

                # ── 사업보고서(Q4 연간) 저장 후 → 4Q 분기값 자동 계산 ──
                # 4Q 분기 = 연간합계 - (Q1 + Q2 + Q3)
                if rcode == "11011":
                    try:
                        q1 = db.query(models.FinancialData).filter(
                            models.FinancialData.stock_code == stock_code,
                            models.FinancialData.year    == year,
                            models.FinancialData.quarter == 1,
                            models.FinancialData.is_annual.is_(False),
                        ).first()
                        q2 = db.query(models.FinancialData).filter(
                            models.FinancialData.stock_code == stock_code,
                            models.FinancialData.year    == year,
                            models.FinancialData.quarter == 2,
                            models.FinancialData.is_annual.is_(False),
                        ).first()
                        q3 = db.query(models.FinancialData).filter(
                            models.FinancialData.stock_code == stock_code,
                            models.FinancialData.year    == year,
                            models.FinancialData.quarter == 3,
                            models.FinancialData.is_annual.is_(False),
                        ).first()
                        # Q4 분기 이미 있으면 스킵
                        q4_exists = db.query(models.FinancialData).filter(
                            models.FinancialData.stock_code == stock_code,
                            models.FinancialData.year    == year,
                            models.FinancialData.quarter == 4,
                            models.FinancialData.is_annual.is_(False),
                        ).first()
                        if q1 and q2 and q3 and not q4_exists:
                            def _sub(a, b, c, annual):
                                """annual - q1 - q2 - q3, None 안전 처리"""
                                if annual is None: return None
                                return (annual or 0) - (a or 0) - (b or 0) - (c or 0)
                            q4_data = schemas.FinancialIngest(
                                stock_code     = stock_code,
                                year           = year,
                                quarter        = 4,
                                is_annual      = False,
                                revenue        = _sub(q1.revenue,           q2.revenue,           q3.revenue,           m["revenue"]),
                                operating_profit= _sub(q1.operating_profit,  q2.operating_profit,  q3.operating_profit,  m["operating_profit"]),
                                net_income     = _sub(q1.net_income,         q2.net_income,         q3.net_income,         m["net_income"]),
                                total_assets   = m["total_assets"],    # 기말 잔액 그대로
                                total_liabilities= m["total_liabilities"],
                                total_equity   = m["total_equity"],
                                capital_stock  = m["capital_stock"],
                                eps            = m["eps"],
                                bps            = m["bps"],
                                dps            = m["dps"],
                                cash           = m["cash"],
                            )
                            crud.upsert_financial_data(db, q4_data)
                            saved += 1
                            logger.info(f"[DART] {stock_code} {year}Q4 분기값 자동 계산 저장")
                        elif q4_exists:
                            logger.debug(f"[DART] {stock_code} {year}Q4 분기 이미 존재")
                        else:
                            logger.info(f"[DART] {stock_code} {year}Q4 계산 불가 (Q1~Q3 미수집)")
                    except Exception as e_q4:
                        logger.warning(f"[DART] {stock_code} {year}Q4 계산 오류: {e_q4}")

            except Exception as e:
                logger.warning(f"[DART] {stock_code} {year}Q{qnum} 저장실패: {e}")

        if latest_only and saved >= 4:
            break  # 최신 모드: 최근 4건이면 충분

    logger.info(f"[DART] {stock_code} 완료: {saved}건")
    return saved


def _get_kis_price(stock_code: str):
    """KIS API 현재가. 실패 시 None."""
    try:
        from kis_client import kis_client
        return kis_client.get_current_price(stock_code)
    except Exception as e:
        logger.warning(f"[KIS] {stock_code}: {e}")
        return None


def _bg_ondemand(stock_code: str):
    """백그라운드 온디맨드: 주가 + 재무 전체."""
    _collecting[stock_code] = "running"
    try:
        from data_collector import DataCollector
        import config as _cfg
        col = DataCollector(dart_api_key=_cfg.DART_API_KEY)
        col.run_ondemand(stock_code)
        logger.info(f"[BG] {stock_code} 완료")
    except Exception as e:
        logger.error(f"[BG] {stock_code}: {e}")
    finally:
        _collecting[stock_code] = "done"


def _bg_collect(stock_code: str):
    _bg_ondemand(stock_code)


def _nightly_update():
    """자정 배치: 주가 전수 + 공시기준 재무."""
    from database import SessionLocal
    db = SessionLocal()
    try:
        codes = set()
        for w in db.query(models.Watchlist).all(): codes.add(w.stock_code)
        for p in db.query(models.Portfolio).filter(models.Portfolio.quantity > 0).all():
            codes.add(p.stock_code)
    finally:
        db.close()
    logger.info(f"[자정배치] {len(codes)}종목")
    try:
        from data_collector import DataCollector
        import config as _cfg
        col = DataCollector(dart_api_key=_cfg.DART_API_KEY)
        col.run_nightly_batch(list(codes))
    except Exception as e:
        logger.error(f"[자정배치] {e}")
    _valuation_cache.clear()
    logger.info("[자정배치] 완료")


def _is_market_hours() -> bool:
    """한국 주식시장 운영 시간 여부 (평일 09:00~15:35)."""
    from datetime import datetime as _dt
    now = _dt.now()
    if now.weekday() >= 5:   # 토(5), 일(6)
        return False
    t = now.hour * 100 + now.minute
    return 900 <= t <= 1535


# ── 매크로 심볼 목록 ──────────────────────────────────────────────
_MACRO_SYMBOLS = {
    "^KS11":    "KOSPI",
    "^KQ11":    "KOSDAQ",
    "^IXIC":    "NASDAQ",
    "^GSPC":    "S&P500",
    "^VIX":     "VIX",
    "GC=F":     "GOLD",
    "CL=F":     "OIL",
    "USDKRW=X": "USD/KRW",
}


def _realtime_fetch_price(stock_code: str, db) -> float | None:
    """KIS API로 현재가를 즉시 조회하고 DB에 upsert 후 종가 반환."""
    try:
        from kis_client import kis_client
        data = kis_client.get_current_price(stock_code)
        if not data:
            return None
        crud.bulk_insert_price_history(db, schemas.PriceIngest(
            stock_code=stock_code,
            prices=[schemas.PriceRecord(
                date=data["date"], open=data["open"], high=data["high"],
                low=data["low"], close=data["close"], volume=data["volume"],
                inst_net_buy=0.0, frn_net_buy=0.0,
            )],
        ))
        return data["close"]
    except Exception as e:
        logger.warning(f"[RT-Price] {stock_code}: {e}")
        return None


def _realtime_fetch_macro(db) -> None:
    from datetime import datetime as _dt2
    _now = _dt2.now()
    if _now.weekday() >= 5:
        return  # 주말 스킵
    # 오늘 매크로 데이터 이미 있으면 스킵 (중복 수집 방지)
    _today = _now.date().isoformat()
    try:
        from sqlalchemy.orm import Session as _S
        _existing = db.query(models.PriceHistory).filter(
            models.PriceHistory.stock_code == "^KS11",
            models.PriceHistory.date >= _today,
        ).first()
        if _existing and _existing.close and _existing.close > 0:
            return  # 오늘 이미 수집됨
    except Exception:
        pass
    """
    Yahoo Finance로 매크로 지수 최신값 조회 후 DB upsert.

    핵심 정책:
    - period="5d"로 넉넉하게 가져와 오늘 데이터가 없으면 최신 데이터를 사용
    - 매크로 심볼은 inst_net_buy/frn_net_buy가 없으므로 기존 값 보존
    - 오늘자 레코드는 upsert(crud 내부), 과거 레코드는 INSERT IGNORE
    """
    import yfinance as yf
    from datetime import datetime as _dt, date as _date
    today = _date.today()

    for symbol, name in _MACRO_SYMBOLS.items():
        try:
            df = yf.download(symbol, period="5d", interval="1d",
                             progress=False, auto_adjust=True)
            if df is None or df.empty:
                logger.warning(f"[RT-Macro] {symbol}: Yahoo 응답 없음")
                continue

            # MultiIndex 열 처리 (yfinance 버전에 따라 다름)
            if hasattr(df.columns, 'get_level_values'):
                try:
                    df.columns = df.columns.get_level_values(0)
                except Exception:
                    pass

            def _gv(col, row):
                v = row.get(col, 0)
                if hasattr(v, 'iloc'):
                    v = v.iloc[0]
                try:
                    return float(v) if v is not None else 0.0
                except (TypeError, ValueError):
                    return 0.0

            prices = []
            for ts, row in df.iterrows():
                try:
                    row_date = ts.date() if hasattr(ts, 'date') else _date.fromisoformat(str(ts)[:10])
                    prices.append(schemas.PriceRecord(
                        date=_dt.combine(row_date, _dt.min.time()),
                        open=_gv("Open", row), high=_gv("High", row),
                        low=_gv("Low", row),   close=_gv("Close", row),
                        volume=_gv("Volume", row),
                        inst_net_buy=0.0, frn_net_buy=0.0,
                    ))
                except Exception as e2:
                    logger.debug(f"[RT-Macro] {symbol} 행 파싱 오류: {e2}")

            if not prices:
                continue

            # 오늘 데이터가 없으면 → 최신 데이터를 오늘 날짜로 추가 upsert
            has_today = any(p.date.date() == today for p in prices)
            if not has_today:
                latest_p = max(prices, key=lambda p: p.date)
                today_record = schemas.PriceRecord(
                    date=_dt.combine(today, _dt.min.time()),
                    open=latest_p.open, high=latest_p.high,
                    low=latest_p.low,   close=latest_p.close,
                    volume=latest_p.volume,
                    inst_net_buy=0.0, frn_net_buy=0.0,
                )
                prices.append(today_record)
                logger.debug(f"[RT-Macro] {symbol}: 오늘 데이터 없음 → 최신값({latest_p.close})을 오늘 날짜로 upsert")

            crud.bulk_insert_price_history(db, schemas.PriceIngest(
                stock_code=symbol, prices=prices))
            latest_close = max(prices, key=lambda p: p.date).close
            logger.info(f"[RT-Macro] {symbol}({name}) {len(prices)}건 저장 / 최신: {latest_close}")

        except Exception as e:
            logger.warning(f"[RT-Macro] {symbol}: {e}")


def _minute_price_loop():
    """
    1분 주기 주가 업데이트 스레드.
    - 장 시간(평일 09:00~15:35)에만 KIS API로 포트폴리오 전 종목 현재가 수집
    - 매크로 지수도 1분마다 Yahoo Finance로 갱신
    """
    from database import SessionLocal
    logger.info("[1분루프] 주가 실시간 업데이트 스레드 시작")
    while True:
        _tm.sleep(60)
        if not _is_market_hours():
            continue
        db = SessionLocal()
        try:
            # 포트폴리오 종목 현재가
            holdings = db.query(models.Portfolio).filter(
                models.Portfolio.quantity > 0
            ).all()
            codes = [h.stock_code for h in holdings if h.stock_code.isdigit() and len(h.stock_code) == 6]
            updated = 0
            for code in codes:
                price = _realtime_fetch_price(code, db)
                if price:
                    updated += 1
                _tm.sleep(1.1)  # KIS 1초 제한

            # 매크로 지수 갱신
            _realtime_fetch_macro(db)
            logger.info(f"[1분루프] 주가갱신 {updated}/{len(codes)}종목 완료")
        except Exception as e:
            logger.error(f"[1분루프] 오류: {e}")
        finally:
            db.close()


def _fetch_index_supply(today_iso: str) -> None:
    """네이버 금융에서 KOSPI/KOSDAQ 수급 스크래핑 후 DB 저장."""
    import requests as _rq, re as _re2, httpx as _hx2
    from bs4 import BeautifulSoup as _BS

    def _pv(s: str) -> float:
        """'+7,270억' or '-19,863억' → float"""
        s   = s.strip()
        neg = "-" in s or "▼" in s
        n   = _re2.sub(r"[^0-9]", "", s)
        return (-float(n) if neg else float(n)) if n else 0.0

    _hd = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                          "AppleWebKit/537.36 Chrome/120.0 Safari/537.36",
           "Referer": "https://finance.naver.com/"}

    for code, symbol in [("KOSPI", "^KS11"), ("KOSDAQ", "^KQ11")]:
        try:
            res = _rq.get(
                f"https://finance.naver.com/sise/sise_index.naver?code={code}",
                headers=_hd, timeout=10)
            res.encoding = res.apparent_encoding
            lines = [l.strip() for l in _BS(res.text, "html.parser").get_text().split("\n") if l.strip()]
            inst = frn = ind = 0.0
            for ln in lines:
                if   ln.startswith("개인")   and "억" in ln: ind  = _pv(ln[2:])
                elif ln.startswith("외국인") and "억" in ln: frn  = _pv(ln[3:])
                elif ln.startswith("기관")   and "억" in ln: inst = _pv(ln[2:])
            if inst != 0 or frn != 0:
                _hx2.post("http://127.0.0.1:8000/api/ingest/investor-trends",
                    json={"stock_code": symbol, "trends": [{
                        "date": today_iso,
                        "inst_net_buy": round(inst, 1),
                        "frn_net_buy":  round(frn,  1),
                    }]}, timeout=10)
                logger.info(f"[수급] {code}: 기관={inst:+,.1f} 외국인={frn:+,.1f}억원")
            else:
                logger.warning(f"[수급] {code}: 파싱값 0 — 페이지 구조 확인 필요")
        except Exception as e_c:
            logger.warning(f"[수급] {code}: {e_c}")


def _five_min_investor_loop():
    """
    수급 업데이트 스레드.
    - 장중 (09:00~15:35)    : 5분(300초)마다 지수수급 + 개별종목수급 + KIS체결내역
    - 장마감 후 (15:35~20:00): 30분(1800초)마다 지수 수급
    - 20:00 이후            : 수집 안 함
    """
    from database import SessionLocal
    from datetime import timedelta as _td2
    import datetime as _dt_mod
    logger.info("[수급루프] 수급 업데이트 스레드 시작")
    _tm.sleep(30)
    _last_supply_time = 0.0

    while True:
        _tm.sleep(60)
        now     = _dt_mod.datetime.now()
        today   = now.date()
        t       = now.hour * 100 + now.minute
        is_wday = now.weekday() < 5
        if not is_wday:
            continue

        # 수집 주기 결정
        if   900 <= t <= 1535:  interval = 300    # 장중: 5분
        elif 1535 < t <= 2000:  interval = 1800   # 장마감~20시: 30분
        else:                   continue           # 20시 이후 없음

        if _tm.time() - _last_supply_time < interval:
            continue
        _last_supply_time = _tm.time()
        today_iso = today.isoformat()

        # 지수 수급 (네이버 스크래핑)
        _fetch_index_supply(today_iso)

        # 장중에만 개별종목 수급 + KIS 체결내역 동기화
        if not _is_market_hours():
            continue
        db = SessionLocal()
        try:
            from kis_client import kis_client

            # ── KIS 체결내역 동기화 (300초마다) ──────────────
            try:
                synced = _sync_kis_executions(db)
                if synced > 0:
                    logger.info(f"[수급루프] KIS 체결 {synced}건 자동 반영")
            except Exception as e_kis:
                logger.warning(f"[수급루프] KIS 체결 오류: {e_kis}")

            # ── 개별종목 수급 (KIS) ───────────────────────────
            holdings = db.query(models.Portfolio).filter(models.Portfolio.quantity > 0).all()
            codes    = [h.stock_code for h in holdings if h.stock_code.isdigit() and len(h.stock_code)==6]
            updated  = 0
            for code in codes:
                try:
                    trends = kis_client.get_investor_trends_bulk(code)
                    if not trends:
                        _tm.sleep(1.1); continue
                    for t2 in trends:
                        try:
                            ds = _dt_mod.datetime.strptime(t2["date"],"%Y-%m-%d")
                            de = ds + _td2(days=1)
                        except ValueError:
                            continue
                        row = db.query(models.PriceHistory).filter(
                            models.PriceHistory.stock_code==code,
                            models.PriceHistory.date>=ds,
                            models.PriceHistory.date<de,
                        ).first()
                        if row:
                            row.inst_net_buy     = t2["inst_net_buy"]
                            row.frn_net_buy      = t2["frn_net_buy"]
                            row.ind_net_buy      = t2.get("ind_net_buy", 0)
                            row.inst_net_buy_amt = t2.get("inst_net_buy_amt", 0)
                            row.frn_net_buy_amt  = t2.get("frn_net_buy_amt", 0)
                            row.ind_net_buy_amt  = t2.get("ind_net_buy_amt", 0)
                    db.commit(); updated += 1
                except Exception as e2:
                    logger.warning(f"[수급루프] {code}: {e2}")
                _tm.sleep(1.1)
            logger.info(f"[수급루프] 개별종목 {updated}/{len(codes)}종목 완료")
        except Exception as e:
            logger.error(f"[수급루프] {e}")
        finally:
            db.close()


def _save_index_history_today():
    """
    장 마감 후 코스피/코스닥 오늘 일봉을 DB에 저장.
    Yahoo Finance에서 최신 데이터를 가져와 upsert.
    """
    from database import SessionLocal
    db = SessionLocal()
    try:
        _realtime_fetch_macro(db)
        logger.info("[장마감] 지수 히스토리 저장 완료")
    except Exception as e:
        logger.error(f"[장마감] {e}")
    finally:
        db.close()


def _sync_kis_executions(db) -> int:
    """
    KIS 당일 체결내역을 조회하여 포트폴리오에 반영.

    평균단가 재계산 원칙 (owner='이효준', broker='KIS' 포지션만 업데이트):
      매도: new_avg = (보유총액 - 매도수량×매도가) / 잔여수량
        예) 1,000원×100주 보유 중 500원×50주 매도
            → (100,000 - 25,000) / 50 = 1,500원
      매수: new_avg = (보유총액 + 매수수량×매수가) / 총수량
    """
    from datetime import datetime as _dt
    try:
        from kis_client import kis_client
        executions = kis_client.get_today_executions()
        if not executions:
            return 0
        updated = 0
        for ex in executions:
            code  = ex["stock_code"]
            name  = ex["stock_name"]
            qty   = float(ex["quantity"])
            price = float(ex["price"])
            tx    = ex["tx_type"]  # "buy" | "sell"
            if not code or qty == 0 or price == 0:
                continue

            # 중복 체결 방지
            tx_time   = ex.get("tx_time", "")
            today_str = _dt.now().strftime("%Y-%m-%d")
            existing_tx = db.query(models.PortfolioTx).filter(
                models.PortfolioTx.stock_code == code,
                models.PortfolioTx.tx_date   >= _dt.strptime(today_str, "%Y-%m-%d"),
                models.PortfolioTx.quantity  == qty,
                models.PortfolioTx.price     == price,
                models.PortfolioTx.tx_type   == tx,
                models.PortfolioTx.memo      == f"KIS_{tx_time}",
            ).first()
            if existing_tx:
                continue

            # 이효준 KIS 포지션 조회
            holding = db.query(models.Portfolio).filter(
                models.Portfolio.stock_code == code,
                models.Portfolio.broker     == "KIS",
                models.Portfolio.owner      == "이효준",
            ).first()

            if tx == "sell":
                if holding and holding.quantity > 0:
                    cur_total  = holding.avg_price * holding.quantity
                    remain_qty = holding.quantity - qty
                    if remain_qty > 0:
                        new_avg = (cur_total - price * qty) / remain_qty
                        holding.avg_price = round(max(new_avg, 0), 2)
                        holding.quantity  = remain_qty
                    else:
                        holding.quantity  = 0.0
                        holding.avg_price = 0.0
                    logger.info(f"[KIS매도] {code} {qty}주@{price:,.0f}원 → 잔여 {holding.quantity}주 평단 {holding.avg_price:,.0f}원")

            elif tx == "buy":
                if holding:
                    cur_total = holding.avg_price * holding.quantity
                    new_qty   = holding.quantity + qty
                    holding.avg_price = round((cur_total + price * qty) / new_qty, 2)
                    holding.quantity  = new_qty
                else:
                    holding = models.Portfolio(
                        stock_code=code, stock_name=name,
                        quantity=qty, avg_price=price,
                        broker="KIS", owner="이효준", source="kis",
                    )
                    db.add(holding)
                logger.info(f"[KIS매수] {code} {qty}주@{price:,.0f}원 → 보유 {holding.quantity}주 평단 {holding.avg_price:,.0f}원")

            # 거래내역 기록
            db.add(models.PortfolioTx(
                stock_code=code, stock_name=name,
                tx_type=tx, quantity=qty, price=price,
                tx_date=_dt.now(), memo=f"KIS_{tx_time}",
            ))
            updated += 1

        db.commit()
        logger.info(f"[KIS체결] {updated}건 포트폴리오 반영")
        return updated
    except Exception as e:
        logger.error(f"[KIS체결] 동기화 오류: {e}")
        db.rollback()
        return 0
    from datetime import datetime as _dt
    try:
        from kis_client import kis_client
        executions = kis_client.get_today_executions()
        if not executions:
            return 0
        updated = 0
        for ex in executions:
            code  = ex["stock_code"]
            name  = ex["stock_name"]
            qty   = float(ex["quantity"])
            price = float(ex["price"])
            tx    = ex["tx_type"]  # "buy" | "sell"
            if not code or qty == 0 or price == 0:
                continue
            # 중복 체결 방지
            tx_time = ex.get("tx_time", "")
            today_str = _dt.now().strftime("%Y-%m-%d")
            existing_tx = db.query(models.PortfolioTx).filter(
                models.PortfolioTx.stock_code == code,
                models.PortfolioTx.tx_date >= _dt.strptime(today_str, "%Y-%m-%d"),
                models.PortfolioTx.quantity == qty,
                models.PortfolioTx.price == price,
                models.PortfolioTx.tx_type == tx,
                models.PortfolioTx.memo == f"KIS_{tx_time}",
            ).first()
            if existing_tx:
                continue

            holding = db.query(models.Portfolio).filter(
                models.Portfolio.stock_code == code
            ).first()

            if tx == "sell":
                if holding and holding.quantity > 0:
                    cur_total  = holding.avg_price * holding.quantity
                    remain_qty = holding.quantity - qty
                    if remain_qty > 0:
                        new_avg = (cur_total - price * qty) / remain_qty
                        holding.avg_price = round(max(new_avg, 0), 2)
                        holding.quantity  = remain_qty
                    else:
                        holding.quantity  = 0.0
                        holding.avg_price = 0.0
                    logger.info(f"[KIS매도] {code} {qty}주@{price:,.0f}원 → 잔여 {holding.quantity}주 평단 {holding.avg_price:,.0f}원")

            elif tx == "buy":
                if holding:
                    cur_total = holding.avg_price * holding.quantity
                    new_qty   = holding.quantity + qty
                    holding.avg_price = round((cur_total + price * qty) / new_qty, 2)
                    holding.quantity  = new_qty
                    holding.source    = "kis"
                else:
                    holding = models.Portfolio(
                        stock_code=code, stock_name=name,
                        quantity=qty, avg_price=price,
                        broker="KIS", owner="이효준", source="kis",
                    )
                    db.add(holding)
                logger.info(f"[KIS매수] {code} {qty}주@{price:,.0f}원 → 보유 {holding.quantity}주 평단 {holding.avg_price:,.0f}원")
            # 거래내역 기록
            db.add(models.PortfolioTx(
                stock_code=code, stock_name=name,
                tx_type=tx, quantity=qty, price=price,
                tx_date=_dt.now(),
                memo=f"KIS_{tx_time}",
            ))
            updated += 1
        db.commit()
        logger.info(f"[KIS체결] {updated}건 포트폴리오 반영")
        return updated
    except Exception as e:
        logger.error(f"[KIS체결] 동기화 오류: {e}")
        db.rollback()
        return 0


def _save_portfolio_snapshot(db) -> None:
    """
    장 마감 후 포트폴리오 평가총액 스냅샷 저장.
    - eval_amount: 평가총액
    - profit_amt: 평가손익 (평가총액 - 매입총액)
    - profit_pct: 수익률
    전일 스냅샷과 비교해 일일 손익 계산 가능.
    """
    from datetime import date as _date_cls
    try:
        today_str = _date_cls.today().isoformat()
        holdings  = db.query(models.Portfolio).filter(models.Portfolio.quantity > 0).all()
        total_eval = 0.0
        for h in holdings:
            price_row = db.query(models.PriceHistory).filter(
                models.PriceHistory.stock_code == h.stock_code
            ).order_by(models.PriceHistory.date.desc()).first()
            cur_price = price_row.close if price_row else h.avg_price
            eval_amt  = round(cur_price  * h.quantity)
            cost_amt  = round(h.avg_price * h.quantity)
            profit    = eval_amt - cost_amt
            pct       = round(profit / cost_amt * 100, 2) if cost_amt else 0.0
            total_eval += eval_amt

            existing = db.query(models.PortfolioSnapshot).filter(
                models.PortfolioSnapshot.snapshot_date == today_str,
                models.PortfolioSnapshot.stock_code    == h.stock_code,
            ).first()
            if existing:
                existing.close_price  = cur_price
                existing.quantity     = h.quantity
                existing.avg_price    = h.avg_price
                existing.eval_amount  = eval_amt
                existing.profit_amt   = profit
                existing.profit_pct   = pct
            else:
                db.add(models.PortfolioSnapshot(
                    snapshot_date = today_str,
                    stock_code    = h.stock_code,
                    stock_name    = h.stock_name,
                    close_price   = cur_price,
                    quantity      = h.quantity,
                    avg_price     = h.avg_price,
                    eval_amount   = eval_amt,
                    profit_amt    = profit,
                    profit_pct    = pct,
                ))
        db.commit()
        logger.info(f"[스냅샷] {today_str} 평가총액={total_eval:,.0f}원 저장")
    except Exception as e:
        logger.error(f"[스냅샷] 저장 오류: {e}")
        db.rollback()


def _closing_scheduler():
    """
    평일 15:40에 장 마감 처리:
    1. 지수 일봉 저장
    2. KIS 체결내역 동기화 → 포트폴리오 업데이트
    3. 포트폴리오 평가총액 스냅샷 저장
    """
    from datetime import datetime as _dt, timedelta as _td
    from database import SessionLocal
    while True:
        now = _dt.now()
        target = now.replace(hour=15, minute=40, second=0, microsecond=0)
        if now >= target:
            target += _td(days=1)
        while target.weekday() >= 5:
            target += _td(days=1)
        _tm.sleep((target - _dt.now()).total_seconds())
        db = SessionLocal()
        try:
            _save_index_history_today()
            cnt = _sync_kis_executions(db)
            logger.info(f"[장마감] KIS 체결 {cnt}건 동기화")
            _save_portfolio_snapshot(db)
        except Exception as e:
            logger.error(f"[장마감 스케줄러] {e}")
        finally:
            db.close()


def _start_scheduler():
    """자정 스케줄러 + 1분 주가 루프 + 5분 수급 루프 + 장마감 저장 모두 기동."""
    from datetime import datetime as _dt, timedelta as _td

    # ── 자정 배치 ──────────────────────────────────────────────────
    def _nightly_loop():
        while True:
            now = _dt.now()
            nxt = (now + _td(days=1)).replace(hour=0, minute=0, second=10, microsecond=0)
            _tm.sleep((nxt - now).total_seconds())
            try: _nightly_update()
            except Exception as e: logger.error(f"[Scheduler] {e}")

    _th.Thread(target=_nightly_loop,           daemon=True, name="NightlyScheduler").start()
    _th.Thread(target=_minute_price_loop,      daemon=True, name="MinutePriceLoop").start()
    _th.Thread(target=_five_min_investor_loop, daemon=True, name="FiveMinInvestorLoop").start()
    _th.Thread(target=_closing_scheduler,      daemon=True, name="ClosingScheduler").start()
    logger.info("[Scheduler] 자정/1분/5분/장마감 스케줄러 모두 시작")

    # ── 서버 시작 시 수급 즉시 1회 수집 ──────────────────────
    def _startup_supply():
        import datetime as _dt
        _tm.sleep(5)  # 서버 완전 기동 대기
        today_iso = _dt.date.today().isoformat()
        logger.info("[시작수급] 서버 기동 후 수급 즉시 수집")
        _fetch_index_supply(today_iso)
    _th.Thread(target=_startup_supply, daemon=True, name="StartupSupply").start()



@app.on_event("startup")
async def startup_event():
    _start_scheduler()
    logger.info("서버 시작 완료")

# [CORS] 로컬 개발 + Cloudflare Tunnel 외부 접속 허용.
# Duck DNS / Cloudflare Tunnel 도메인은 .env의 ALLOWED_ORIGINS에 추가하세요.
# 예: ALLOWED_ORIGINS=https://yourname.duckdns.org,https://xxxx.trycloudflare.com
import os as _os
_extra_origins = [o.strip() for o in _os.getenv("ALLOWED_ORIGINS", "").split(",") if o.strip()]
_allowed_origins = [
    "http://localhost:3000",
    "http://localhost:5173",
    "http://127.0.0.1:3000",
    "http://127.0.0.1:5173",
] + _extra_origins

app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/api/realtime/prices")
def get_realtime_prices(db: Session = Depends(get_db)):
    """
    포트폴리오 전 종목의 최신 현재가 및 등락률을 반환합니다.
    프론트엔드 1분 폴링 전용 경량 API.
    반환: { stock_code: { current_price, change_pct, profit, profit_pct, total_value } }
    """
    holdings = db.query(models.Portfolio).filter(models.Portfolio.quantity > 0).all()
    result = {}
    total_buy_sum   = 0.0
    total_value_sum = 0.0
    total_profit_sum = 0.0

    for h in holdings:
        price_row = db.query(models.PriceHistory).filter(
            models.PriceHistory.stock_code == h.stock_code
        ).order_by(models.PriceHistory.date.desc()).first()

        prev_row = db.query(models.PriceHistory).filter(
            models.PriceHistory.stock_code == h.stock_code
        ).order_by(models.PriceHistory.date.desc()).offset(1).first()

        current_price = price_row.close if price_row else h.avg_price
        prev_price    = prev_row.close  if prev_row  else current_price
        price_date    = price_row.date.strftime("%Y-%m-%d %H:%M") if price_row else ""

        change_pct  = round((current_price - prev_price) / prev_price * 100, 2) if prev_price else 0.0
        profit      = round((current_price - h.avg_price) * h.quantity)
        profit_pct  = round((current_price - h.avg_price) / h.avg_price * 100, 2) if h.avg_price else 0.0
        total_value = round(current_price * h.quantity)
        buy_total   = round(h.avg_price   * h.quantity)

        result[h.stock_code] = {
            "current_price": current_price,
            "change_pct":    change_pct,
            "profit":        profit,
            "profit_pct":    profit_pct,
            "total_value":   total_value,
            "buy_total":     buy_total,
            "price_date":    price_date,
        }
        total_buy_sum    += buy_total
        total_value_sum  += total_value
        total_profit_sum += profit

    total_profit_pct = round(total_profit_sum / total_buy_sum * 100, 2) if total_buy_sum else 0.0
    return {
        "holdings":         result,
        "summary": {
            "total_buy":        round(total_buy_sum),
            "total_value":      round(total_value_sum),
            "total_profit":     round(total_profit_sum),
            "total_profit_pct": total_profit_pct,
        },
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "market_open": _is_market_hours(),
    }


@app.get("/api/realtime/macro")
def get_realtime_macro(db: Session = Depends(get_db)):
    """
    매크로 지표 최신값 반환 (300초 폴링용).
    - 장 시간: Yahoo Finance 즉시 갱신 후 DB 최신값 반환
    - 장 외: DB 저장된 최신값 바로 반환
    반환: { index:{KOSPI,KOSDAQ}, vix:{...}, commodities:{...} }
    """
    if _is_market_hours():
        try:
            _realtime_fetch_macro(db)
        except Exception as e:
            logger.warning(f"[realtime/macro] Yahoo 갱신 실패 (DB값 반환): {e}")
    result = processor.get_macro_status(db)
    result.setdefault("index",       {})
    result.setdefault("vix",         {"value": 0, "change": 0, "date": "-", "history": []})
    result.setdefault("commodities", {})
    return result





@app.get("/")
def read_root():
    return {"message": "주식 분석 백엔드 서버가 작동 중입니다."}


@app.get("/api/dashboard/market-info/{stock_code}")
def get_market_info(stock_code: str, refresh: bool = False, db: Session = Depends(get_db)):
    """
    종목의 시장정보 반환.
    refresh=true 이면 캐시 무시하고 재조회.
    """
    if not (stock_code.isdigit() and len(stock_code) == 6):
        return {"market": None, "mktcap": None, "mktcap_rank": None}

    # ── 캐시 확인 (1시간 유효) ──────────────────────────────────
    cached = _market_info_cache.get(stock_code, {})
    cache_age = _tm.time() - cached.get("cached_at", 0)
    if cached and cache_age < 3600 and not refresh:
        return cached

    result = {"market": None, "mktcap": None, "mktcap_rank": None, "stock_name": None}

    # ── 시장구분: DB 먼저 확인 ──────────────────────────────────
    meta = db.query(models.StockMeta).filter(
        models.StockMeta.stock_code == stock_code
    ).first()
    if meta:
        result["market"]     = meta.market
        result["stock_name"] = meta.stock_name

    # ── 네이버 금융 스크래핑 (시총·순위·시장구분) ───────────────
    try:
        import requests as _req, re as _re_mi
        from bs4 import BeautifulSoup as _BS
        res = _req.get(
            f"https://finance.naver.com/item/main.naver?code={stock_code}",
            headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                                   "AppleWebKit/537.36 Chrome/120.0 Safari/537.36",
                     "Referer": "https://finance.naver.com/"},
            timeout=8,
        )
        res.encoding = res.apparent_encoding  # ★ UTF-8 자동 감지
        soup = _BS(res.text, "html.parser")

        # 시장구분 + 시총순위 — "코스피1위" 형식의 td에서 한번에 파싱
        for tag in soup.find_all("td"):
            txt = tag.get_text(strip=True)
            if "코스피" in txt and len(txt) < 20:
                result["market"] = "KOSPI"
                rank_txt = txt.replace("코스피","").replace("위","").replace(",","").strip()
                try: result["mktcap_rank"] = int(rank_txt)
                except: pass
                break
            elif "코스닥" in txt and len(txt) < 20:
                result["market"] = "KOSDAQ"
                rank_txt = txt.replace("코스닥","").replace("위","").replace(",","").strip()
                try: result["mktcap_rank"] = int(rank_txt)
                except: pass
                break

        # 시장구분 fallback — 타이틀에서 파싱
        if not result["market"]:
            title = soup.find("title")
            title_txt = title.text if title else ""
            if "코스피" in title_txt:   result["market"] = "KOSPI"
            elif "코스닥" in title_txt: result["market"] = "KOSDAQ"

        # 종목명
        for sel in ["div.wrap_company h2 a", "h2.h_company a", "h2.h_company", "title"]:
            t = soup.select_one(sel)
            if t:
                nm = t.get_text(strip=True)
                if ":" in nm: nm = nm.split(":")[0].strip()
                if nm and len(nm) < 30:
                    result["stock_name"] = nm
                    break

        # 시가총액 — table.tb_type1 에서 "시가총액(억)" 행
        for tr in soup.select("table.tb_type1 tr"):
            th = tr.select_one("th"); td = tr.select_one("td")
            if not (th and td): continue
            th_txt = th.get_text(strip=True)
            td_txt = td.get_text(strip=True).replace(",","")
            if "시가총액" in th_txt:
                try:
                    result["mktcap"] = int(td_txt.replace("억","").strip())
                    break
                except: pass

    except Exception as e:
        logger.warning(f"[MarketInfo] {stock_code} 네이버 스크래핑 오류: {e}")

    # ── 시장구분 DB 저장 (최초 1회) ─────────────────────────────
    if result["market"] and not meta:
        try:
            db.add(models.StockMeta(
                stock_code=stock_code,
                stock_name=result.get("stock_name"),
                market=result["market"],
            ))
            db.commit()
            logger.info(f"[MarketInfo] {stock_code} 시장구분 저장: {result['market']}")
        except Exception as e:
            db.rollback()
            logger.warning(f"[MarketInfo] DB 저장 오류: {e}")

    # 캐시 저장
    result["cached_at"] = _tm.time()
    _market_info_cache[stock_code] = result
    return result

@app.post("/api/ingest/fundamentals", response_model=schemas.FinancialIngest)
def ingest_fundamentals(financial: schemas.FinancialIngest, db: Session = Depends(get_db)):
    """
    재무제표 원시 데이터를 수신하여 저장합니다.
    """
    try:
        return crud.upsert_financial_data(db, financial)
    except Exception as e:
        logger.error(f"재무 데이터 수신 중 오류 발생: {e}")
        raise HTTPException(status_code=500, detail="데이터 저장 중 오류가 발생했습니다.")

@app.post("/api/ingest/market-price")
def ingest_market_price(price_ingest: schemas.PriceIngest, db: Session = Depends(get_db)):
    # ★ 휴장일(토/일) 저장 차단
    from datetime import datetime as _dt
    _today = _dt.now()
    if _today.weekday() >= 5:  # 5=토, 6=일
        return {"status": "skip", "reason": "weekend"}
    """
    일일 주가 마감 데이터 등 방대한 시계열 데이터를 수신하여 일괄 저장합니다.
    """
    try:
        crud.bulk_insert_price_history(db, price_ingest)
        return {"status": "success", "count": len(price_ingest.prices)}
    except Exception as e:
        logger.error(f"주가 데이터 수신 중 오류 발생: {e}")
        raise HTTPException(status_code=500, detail="데이터 일괄 저장 중 오류가 발생했습니다.")

@app.post("/api/ingest/sectors")
def ingest_sectors(sector: schemas.SectorMapping, db: Session = Depends(get_db)):
    """
    섹터별 소속 종목 맵핑 데이터를 수신합니다.
    """
    try:
        crud.update_sector_mapping(db, sector)
        return {"status": "success"}
    except Exception as e:
        logger.error(f"섹터 데이터 수신 중 오류 발생: {e}")
        raise HTTPException(status_code=500, detail="데이터 저장 중 오류가 발생했습니다.")

# --- 대시보드 및 결과 조회 API ---

@app.get("/api/dashboard/chart/{stock_code}")
def get_stock_chart(stock_code: str, days: int = 30, db: Session = Depends(get_db)):
    """
    특정 종목의 차트 시계열 데이터를 반환합니다.
    """
    return processor.get_chart_data(db, stock_code, days)

@app.get("/api/dashboard/sectors")
def get_sectors(db: Session = Depends(get_db)):
    """
    섹터별 등락 현황을 반환합니다.
    """
    return processor.get_sector_performance(db)

@app.get("/api/dashboard/screening/triple")
def get_triple_screening(db: Session = Depends(get_db)):
    """
    '3대 스크리닝' 조건에 합부하는 종목 리스트를 반환합니다.
    """
    return screener.apply_triple_screening(db)

# --- AI 분석 리포트 API ---

@app.post("/api/reports/generate/{stock_code}")
def generate_report(stock_code: str, db: Session = Depends(get_db)):
    """
    특정 종목의 AI 분석 리포트를 생성합니다.
    """
    analyzer = ai_analyzer.AIAnalyzer(db)
    report = analyzer.generate_stock_report(stock_code)
    return report

@app.get("/api/reports/latest/{stock_code}")
def get_latest_report(stock_code: str, db: Session = Depends(get_db)):
    """
    최근 생성된 AI 리포트 정보를 조회합니다.
    리포트가 없으면 404 대신 null을 반환하여 프론트 fetch가 실패하지 않도록 합니다.
    """
    analyzer = ai_analyzer.AIAnalyzer(db)
    report = analyzer.get_latest_report(stock_code)
    if not report:
        return None  # 404 대신 200 + null → fetchStockDetail 전체가 중단되지 않음
    return report

@app.get("/api/dashboard/financial-table/{stock_code}")
def get_financial_table(stock_code: str, type: str = "annual", db: Session = Depends(get_db)):
    """재무제표 반환. ?type=annual(기본,연간5년) 또는 ?type=quarter(분기8개)"""
    return processor.get_financial_summary(db, stock_code, data_type=type)

@app.get("/api/dashboard/fundamentals/{stock_code}")
def get_stock_fundamentals(stock_code: str, db: Session = Depends(get_db)):
    """
    종목 핵심 지표.
    1. 재무 DB 없으면 → DART 동기 수집 (crud 직접 저장)
    2. 주가 DB 없으면 → KIS 즉시 저장 + Yahoo 백그라운드
    3. PBR/PER → 캐시 우선, 없으면 네이버금융 동기 스크래핑
    """
    is_kr  = stock_code.isdigit() and len(stock_code) == 6
    is_col = _collecting.get(stock_code) == "running"

    # ── 재무 DB 확인 → 없으면 동기 수집 ─────────────────────────
    data = db.query(models.FinancialData).filter(
        models.FinancialData.stock_code == stock_code
    ).order_by(models.FinancialData.year.desc(),
               models.FinancialData.quarter.desc()).first()

    if not data and is_kr and not is_col:
        logger.info(f"[Fund] {stock_code} 재무없음 → DART 동기수집")
        saved = _collect_dart_to_db(stock_code, db, latest_only=False)
        if saved > 0:
            db.expire_all()
            data = db.query(models.FinancialData).filter(
                models.FinancialData.stock_code == stock_code
            ).order_by(models.FinancialData.year.desc(),
                       models.FinancialData.quarter.desc()).first()

    # ── 주가 DB 확인 → 없으면 KIS 즉시 + Yahoo 백그라운드 ─────────
    if is_kr:
        no_price = db.query(models.PriceHistory).filter(
            models.PriceHistory.stock_code == stock_code
        ).first() is None
        if no_price:
            kis = _get_kis_price(stock_code)
            if kis:
                try:
                    crud.bulk_insert_price_history(db, schemas.PriceIngest(
                        stock_code=stock_code,
                        prices=[schemas.PriceRecord(
                            date=kis["date"], open=kis["open"], high=kis["high"],
                            low=kis["low"], close=kis["close"], volume=kis["volume"],
                            inst_net_buy=0.0, frn_net_buy=0.0)]))
                    logger.info(f"[Fund] KIS 현재가 저장: {stock_code}")
                except Exception as e:
                    logger.warning(f"[Fund] KIS 저장실패: {e}")
            if not is_col:
                _th.Thread(target=_bg_ondemand, args=(stock_code,), daemon=True).start()
                is_col = True

    # ── PBR/PER 캐시 → 없으면 동기 스크래핑 ────────────────────────
    val = _get_cached_valuation(stock_code)
    if not val and is_kr:
        val = _scrape_naver(stock_code)
        if val.get("per") is not None or val.get("pbr") is not None:
            val["cached_at"] = _tm.time()
            _valuation_cache[stock_code] = val

    if not data:
        return {
            "revenue": None, "operating_profit": None, "net_income": None,
            "opm": None, "roe": None,
            "pbr": val.get("pbr"), "per": val.get("per"),
            "forward_per": None, "trailing_eps": val.get("trailing_eps"),
            "source": val.get("source"), "collecting": is_col,
        }

    opm = (
        (data.operating_profit / data.revenue * 100)
        if (data.revenue and data.revenue != 0 and data.operating_profit is not None)
        else 0.0
    )
    roe = getattr(data, "roe", None)
    return {
        "revenue":          data.revenue,
        "operating_profit": data.operating_profit,
        "net_income":       data.net_income,
        "opm":              round(opm, 1),
        "roe":              round(roe, 2) if roe is not None else None,
        "pbr":              val.get("pbr"),
        "per":              val.get("per"),
        "forward_per":      val.get("forward_per"),
        "trailing_eps":     val.get("trailing_eps"),
        "source":           val.get("source"),
        "collecting":       is_col,
    }

@app.get("/api/reports/ready")
def get_ready_reports(db: Session = Depends(get_db)):
    """
    OpenClaw 전송용(텔레그램 등)으로 스크리닝 통과 종목 및 AI 분석 결과를 통합 반환합니다.
    """
    try:
        passed_stocks = screener.apply_triple_screening(db)

        results = []
        analyzer = ai_analyzer.AIAnalyzer(db)

        for stock in passed_stocks:
            report = analyzer.get_latest_report(stock["stock_code"])

            # [버그 1 수정] report는 ORM 객체 또는 None.
            # report.content 속성 접근 전에 None 체크 후 안전하게 추출.
            if report is None:
                ai_report_text = "분석 리포트가 아직 생성되지 않았습니다."
            elif isinstance(report, dict):
                ai_report_text = report.get("content", "분석 리포트가 아직 생성되지 않았습니다.")
            else:
                ai_report_text = getattr(report, "content", "분석 리포트가 아직 생성되지 않았습니다.")

            results.append({
                "stock_code": stock["stock_code"],
                "stock_name": stock["stock_name"],
                "financial_summary": {
                    "revenue": stock["latest_revenue"],
                    "profit": stock["latest_profit"],
                },
                "ai_report": ai_report_text,
            })

        return results
    except Exception as e:
        logger.error(f"최종 리포트 준비 중 오류 발생: {e}")
        raise HTTPException(status_code=500, detail="리포트 생성 중 오류가 발생했습니다.")

# --- 대화형 명령 API (OpenClaw Interaction) ---

@app.get("/api/search")
def search_stocks(q: str):
    """
    프론트엔드 드롭다운용 주식 자동완성(Fuzzy) 검색 API
    """
    from ticker_utils import ticker_mapper
    if not q or len(q.strip()) == 0:
        return []
    return ticker_mapper.search(q.strip())


@app.post("/api/commands/analyze/{stock_name}")
def command_analyze_stock(stock_name: str, db: Session = Depends(get_db)):
    """
    종목 조회 진입점 (검색 즉시 전체 수집).
    1. 종목코드 조회 (ticker_mapper)
    2. watchlist 등록
    3. 주가 히스토리 없으면 KIS 현재가 즉시 저장 → Yahoo 1년치 백그라운드
    4. 재무 없거나 최신분기 없거나 재무관련 공시 있으면 → DART 수집 백그라운드
    5. 즉시 응답 (collecting=True 이면 프론트가 폴링)
    """
    from ticker_utils import ticker_mapper
    # 종목코드로 직접 조회도 허용 (6자리 숫자)
    if stock_name.isdigit() and len(stock_name) == 6:
        stock_code = stock_name
        resolved_name = ticker_mapper.get_name(stock_code) or stock_name
    else:
        stock_code = ticker_mapper.get_code(stock_name)
        resolved_name = stock_name
        if not stock_code:
            # fuzzy search fallback
            results = ticker_mapper.search(stock_name)
            if results:
                stock_code = results[0].get("code")
                resolved_name = results[0].get("name", stock_name)
    if not stock_code:
        raise HTTPException(status_code=404, detail=f"'{stock_name}' 종목을 찾을 수 없습니다.")

    # watchlist 등록
    crud.add_to_watchlist(db, stock_code)
    try: db.commit()
    except: db.rollback()

    is_kr  = stock_code.isdigit() and len(stock_code) == 6
    is_col = _collecting.get(stock_code) == "running"

    if is_kr and not is_col:
        # ── 주가: KIS 현재가 즉시 저장 (DB 유무 무관, 항상 최신값 upsert) ──
        kis = _get_kis_price(stock_code)
        if kis:
            try:
                crud.bulk_insert_price_history(db, schemas.PriceIngest(
                    stock_code=stock_code,
                    prices=[schemas.PriceRecord(
                        date=kis["date"], open=kis["open"], high=kis["high"],
                        low=kis["low"], close=kis["close"], volume=kis["volume"],
                        inst_net_buy=0.0, frn_net_buy=0.0)]))
                logger.info(f"[Analyze] KIS 현재가 즉시 저장: {stock_code} {kis['close']}원")
            except Exception as e:
                logger.warning(f"[Analyze] KIS 저장실패: {e}")

        # ── 히스토리 + 재무 백그라운드 전체 수집 ──────────────────
        # has_history: 30일치 차트 데이터 1건이라도 있는지 확인
        has_history = db.query(models.PriceHistory).filter(
            models.PriceHistory.stock_code == stock_code
        ).count() >= 20  # 20건 미만이면 히스토리 부족으로 간주

        has_financial = db.query(models.FinancialData).filter(
            models.FinancialData.stock_code == stock_code
        ).first() is not None

        # 항상 백그라운드 수집 트리거 (ondemand는 내부적으로 중복 스킵 처리)
        _th.Thread(target=_bg_ondemand, args=(stock_code,), daemon=True).start()
        is_col = True
        logger.info(f"[Analyze] {stock_code} 전체 수집 백그라운드 시작 (히스토리:{has_history}, 재무:{has_financial})")

    is_col = _collecting.get(stock_code) == "running"
    return {
        "status":     "collecting" if is_col else "ready",
        "stock_name": resolved_name,
        "stock_code": stock_code,
        "collecting": is_col,
    }

@app.get("/api/commands/collect-status/{stock_code}")
def get_collect_status(stock_code: str):
    """수집 진행 상태 조회"""
    return {"stock_code": stock_code, "status": _collecting.get(stock_code, "done")}

@app.get("/api/commands/watchlist")
def get_current_watchlist(db: Session = Depends(get_db)):
    """현재 자동 추적 중인 종목 리스트를 반환합니다."""
    return crud.get_watchlist(db)

@app.delete("/api/commands/watchlist/{stock_code}")
def remove_from_watchlist(stock_code: str, db: Session = Depends(get_db)):
    """관심종목에서 특정 종목을 제거합니다."""
    deleted = db.query(models.Watchlist).filter(
        models.Watchlist.stock_code == stock_code
    ).first()
    if not deleted:
        raise HTTPException(status_code=404, detail=f"'{stock_code}' 종목이 관심종목에 없습니다.")
    db.delete(deleted)
    db.commit()
    return {"status": "success", "removed": stock_code}

@app.get("/api/dashboard/macro")
def get_macro_dashboard(db: Session = Depends(get_db)):
    """지수, 환율, 원자재 등 매크로 지표 현황을 반환합니다."""
    result = processor.get_macro_status(db)
    result.setdefault("index",       {})
    result.setdefault("vix",         {"value": 0, "change": 0, "date": "-", "history": []})
    result.setdefault("commodities", {})
    return result

@app.get("/api/dashboard/stats")
def get_db_stats(db: Session = Depends(get_db)):
    """데이터베이스 적재 현황 통계를 반환합니다."""
    return {
        "stock_count": db.query(models.FinancialData.stock_code).distinct().count(),
        "price_records": db.query(models.PriceHistory).count(),
        "db_path": "Applications/stock_dashboard/stock.db",
        "last_update": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# ═══════════════════════════════════════════════════════
#  수급 업데이트 엔드포인트
# ═══════════════════════════════════════════════════════

@app.post("/api/ingest/investor-trends")
def ingest_investor_trends(payload: dict, db: Session = Depends(get_db)):
    """KIS 수급 데이터를 기존 주가 레코드에 업데이트합니다."""
    # 주말 스킵
    if datetime.now().weekday() >= 5:
        return {"status": "skip", "reason": "weekend"}
    try:
        stock_code = payload.get("stock_code")
        trends     = payload.get("trends", [])
        updated    = 0
        from datetime import timedelta as _td2
        for t in trends:
            try:
                ds = datetime.strptime(t["date"], "%Y-%m-%d")
                de = ds + _td2(days=1)
            except ValueError:
                continue
            row = db.query(models.PriceHistory).filter(
                models.PriceHistory.stock_code == stock_code,
                models.PriceHistory.date >= ds,
                models.PriceHistory.date <  de,
            ).first()
            if row:
                row.inst_net_buy     = t.get("inst_net_buy", 0)
                row.frn_net_buy      = t.get("frn_net_buy",  0)
                row.ind_net_buy      = t.get("ind_net_buy",  0)
                row.inst_net_buy_amt = t.get("inst_net_buy_amt", 0)
                row.frn_net_buy_amt  = t.get("frn_net_buy_amt",  0)
                row.ind_net_buy_amt  = t.get("ind_net_buy_amt",  0)
                updated += 1
            else:
                db.add(models.PriceHistory(
                    stock_code       = stock_code,
                    date             = ds,
                    open=0.0, high=0.0, low=0.0, close=0.0, volume=0.0,
                    inst_net_buy     = t.get("inst_net_buy", 0),
                    frn_net_buy      = t.get("frn_net_buy",  0),
                    ind_net_buy      = t.get("ind_net_buy",  0),
                    inst_net_buy_amt = t.get("inst_net_buy_amt", 0),
                    frn_net_buy_amt  = t.get("frn_net_buy_amt",  0),
                    ind_net_buy_amt  = t.get("ind_net_buy_amt",  0),
                ))
                updated += 1
        db.commit()
        return {"status": "success", "updated": updated}
    except Exception as e:
        logger.error(f"수급 업데이트 오류: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ═══════════════════════════════════════════════════════
#  포트폴리오 (보유종목) API
# ═══════════════════════════════════════════════════════

@app.get("/api/portfolio")
def get_portfolio(db: Session = Depends(get_db)):
    """
    보유종목 목록.
    - 동일 종목(stock_code)은 보유자/증권사 무관하게 합산하여 1행 반환
    - 가중평균단가 = Σ(avg_price × quantity) / Σquantity
    - 평가액 내림차순 정렬
    """
    from datetime import date as _date_cls, timedelta as _td_cls
    holdings = db.query(models.Portfolio).filter(models.Portfolio.quantity > 0).all()

    # 전일 스냅샷 조회
    yesterday = (_date_cls.today() - _td_cls(days=1)).isoformat()
    prev_snaps = {
        s.stock_code: s for s in
        db.query(models.PortfolioSnapshot).filter(
            models.PortfolioSnapshot.snapshot_date == yesterday
        ).all()
    }

    # ── 종목코드별 합산 ────────────────────────────────────────
    # 가중평균단가 = Σ(avg_price × quantity) / Σquantity
    # 예) A: 500주@1,000원 + B: 500주@2,000원 → 1,000주@1,500원
    merged: dict = {}
    for h in holdings:
        code = h.stock_code or h.stock_name
        if code not in merged:
            merged[code] = {
                "stock_code": h.stock_code,
                "stock_name": h.stock_name,
                "sector":     h.sector or "기타",
                "total_qty":  0.0,
                "total_cost": 0.0,   # Σ(avg_price × quantity) = 총매입금액
                "bought_at":  h.bought_at,  # 가장 이른 매수일 사용
            }
        merged[code]["total_qty"]  += h.quantity
        merged[code]["total_cost"] += h.avg_price * h.quantity
        # 매수일: 가장 이른 날짜 선택
        if h.bought_at:
            cur_d = merged[code]["bought_at"]
            if not cur_d or h.bought_at < cur_d:
                merged[code]["bought_at"] = h.bought_at

    # ── 현재가 조회 + 계산 ────────────────────────────────────
    result = []
    for code, m in merged.items():
        total_qty  = m["total_qty"]
        total_cost = m["total_cost"]
        avg_price  = round(total_cost / total_qty, 2) if total_qty else 0.0

        price_row = db.query(models.PriceHistory).filter(
            models.PriceHistory.stock_code == m["stock_code"]
        ).order_by(models.PriceHistory.date.desc()).first()

        prev_row = db.query(models.PriceHistory).filter(
            models.PriceHistory.stock_code == m["stock_code"]
        ).order_by(models.PriceHistory.date.desc()).offset(1).first()

        has_price     = price_row is not None
        current_price = price_row.close if price_row else avg_price
        prev_price    = prev_row.close  if prev_row  else current_price

        change_pct  = round((current_price - prev_price) / prev_price * 100, 2) if prev_price else 0.0
        buy_total   = round(total_cost)
        total_value = round(current_price * total_qty)
        profit      = total_value - buy_total
        profit_pct  = round(profit / buy_total * 100, 2) if buy_total else 0.0

        # 전일 대비 손익
        prev_snap    = prev_snaps.get(m["stock_code"])
        prev_value   = getattr(prev_snap, 'eval_amount', None) or buy_total
        daily_profit = round(total_value - prev_value)

        # 주가 없는 종목 자동 수집 트리거
        if not has_price and m["stock_code"] and m["stock_code"] not in _collecting:
            import threading as _threading
            _threading.Thread(target=_bg_collect, args=(m["stock_code"],), daemon=True).start()

        result.append({
            "stock_code":    m["stock_code"],
            "stock_name":    m["stock_name"],
            "sector":        m["sector"],
            "quantity":      total_qty,
            "avg_price":     avg_price,
            "current_price": current_price,
            "change_pct":    change_pct,
            "profit":        profit,
            "profit_pct":    profit_pct,
            "total_value":   total_value,
            "buy_total":     buy_total,
            "daily_profit":  daily_profit,
            "has_price":     has_price,
            "collecting":    _collecting.get(m["stock_code"]) == "running",
            "bought_at":     m.get("bought_at"),
        })

    # 평가액 내림차순 정렬
    result.sort(key=lambda x: x["total_value"], reverse=True)
    return result


@app.post("/api/portfolio/sync-kis")
def sync_kis_portfolio(db: Session = Depends(get_db)):
    """KIS 당일 체결내역 수동 동기화."""
    cnt = _sync_kis_executions(db)
    return {"status": "ok", "synced": cnt}


@app.patch("/api/portfolio/{stock_code}/bought-at")
def update_bought_at(stock_code: str, payload: dict, db: Session = Depends(get_db)):
    """종목 매수일 수동 수정."""
    bought_at = payload.get("bought_at", "")
    rows = db.query(models.Portfolio).filter(
        models.Portfolio.stock_code == stock_code
    ).all()
    if not rows:
        raise HTTPException(status_code=404, detail="종목 없음")
    for row in rows:
        row.bought_at = bought_at
    db.commit()
    return {"status": "ok", "stock_code": stock_code, "bought_at": bought_at}


@app.get("/api/portfolio/transactions")
def get_transactions(db: Session = Depends(get_db)):
    """거래 내역 최근 100건"""
    txs = db.query(models.PortfolioTx).order_by(
        models.PortfolioTx.tx_date.desc()
    ).limit(100).all()
    return [
        {"id": t.id, "stock_code": t.stock_code, "stock_name": t.stock_name,
         "tx_type": t.tx_type, "quantity": t.quantity, "price": t.price,
         "tx_date": t.tx_date.strftime("%Y-%m-%d %H:%M") if t.tx_date else "",
         "memo": t.memo}
        for t in txs
    ]


@app.post("/api/portfolio/transaction")
def add_transaction(payload: dict, db: Session = Depends(get_db)):
    """매수/매도 거래 추가 + Portfolio 잔고 업데이트"""
    from datetime import datetime as dt
    stock_code  = str(payload.get("stock_code", "")).strip()
    stock_name  = str(payload.get("stock_name", stock_code))
    tx_type     = payload.get("tx_type", "buy")
    quantity    = float(payload.get("quantity", 0))
    price       = float(payload.get("price", 0))
    tx_date_str = payload.get("tx_date", "")
    sector      = payload.get("sector", "")
    memo        = payload.get("memo", "")

    if not stock_code or quantity <= 0 or price <= 0:
        raise HTTPException(status_code=400, detail="stock_code, quantity, price 필수")

    try:
        tx_date = dt.strptime(tx_date_str, "%Y-%m-%d") if tx_date_str else dt.now()
    except Exception:
        tx_date = dt.now()

    tx = models.PortfolioTx(
        stock_code=stock_code, stock_name=stock_name,
        tx_type=tx_type, quantity=quantity, price=price,
        tx_date=tx_date, memo=memo
    )
    db.add(tx)

    holding = db.query(models.Portfolio).filter(
        models.Portfolio.stock_code == stock_code
    ).first()

    if tx_type == "buy":
        if holding:
            total_qty = holding.quantity + quantity
            holding.avg_price  = (holding.avg_price * holding.quantity + price * quantity) / total_qty
            holding.quantity   = total_qty
            holding.stock_name = stock_name
            if sector: holding.sector = sector
        else:
            db.add(models.Portfolio(
                stock_code=stock_code, stock_name=stock_name,
                sector=sector, quantity=quantity, avg_price=price
            ))
    elif tx_type == "sell":
        if holding:
            holding.quantity = max(0.0, holding.quantity - quantity)

    db.commit()
    return {"status": "ok", "stock_code": stock_code, "tx_type": tx_type,
            "quantity": quantity, "price": price}


@app.post("/api/portfolio/kakao-parse")
def parse_kakao_message(payload: dict):
    """카카오톡 메시지에서 매수/매도 파싱."""
    import re
    text = payload.get("text", "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="text 필수")

    from ticker_utils import ticker_mapper
    parsed = []
    for line in text.strip().splitlines():
        line = line.strip()
        if not line: continue
        tx_type = None
        if re.search(r'매수', line):   tx_type = "buy"
        elif re.search(r'매도', line): tx_type = "sell"
        if not tx_type: continue

        qty_m   = re.search(r'(\d[\d,]*)\s*주', line)
        price_m = re.search(r'@\s*([\d,]+)|([\d][\d,]+)\s*원|([\d]{4,})', line)
        quantity = float(qty_m.group(1).replace(",",""))  if qty_m   else None
        raw_p    = (price_m.group(1) or price_m.group(2) or price_m.group(3)) if price_m else None
        price    = float(raw_p.replace(",","")) if raw_p else None

        name_line  = re.sub(r'매수|매도', '', line)
        name_line  = re.sub(r'[\d,]+\s*주|@[\d,]+|[\d,]+\s*원|\d{4,}', '', name_line)
        stock_name = re.sub(r'[^\w가-힣]', ' ', name_line).strip().split()
        stock_name = stock_name[0] if stock_name else ""
        stock_code = ticker_mapper.get_code(stock_name) if stock_name else None

        parsed.append({
            "raw": line, "tx_type": tx_type,
            "stock_name": stock_name, "stock_code": stock_code,
            "quantity": quantity, "price": price,
            "valid": bool(stock_code and quantity and price),
        })
    return {"parsed": parsed, "count": len(parsed)}


@app.put("/api/portfolio/{stock_code}")
def update_portfolio(stock_code: str, payload: dict, db: Session = Depends(get_db)):
    """보유종목 편집 저장. 종목명 변경 시 ticker 코드 자동 재조회."""
    import threading
    from ticker_utils import ticker_mapper

    h = db.query(models.Portfolio).filter(models.Portfolio.stock_code == stock_code).first()
    if not h:
        raise HTTPException(status_code=404, detail="종목 없음")

    new_name   = payload.get("stock_name", h.stock_name)
    new_sector = payload.get("sector",     h.sector)
    new_qty    = float(payload.get("quantity",  h.quantity))
    new_price  = float(payload.get("avg_price", h.avg_price))

    code_changed = False
    new_code     = stock_code

    if new_name != h.stock_name:
        found = ticker_mapper.get_code(new_name)
        if found and found != stock_code:
            new_code = found
            code_changed = True

    if code_changed:
        db.delete(h)
        db.flush()
        existing = db.query(models.Portfolio).filter(
            models.Portfolio.stock_code == new_code).first()
        if existing:
            total_qty = existing.quantity + new_qty
            existing.avg_price  = (existing.avg_price * existing.quantity + new_price * new_qty) / total_qty
            existing.quantity   = total_qty
            existing.stock_name = new_name
            if new_sector: existing.sector = new_sector
        else:
            db.add(models.Portfolio(
                stock_code=new_code, stock_name=new_name,
                sector=new_sector, quantity=new_qty, avg_price=new_price))
        db.query(models.PortfolioTx).filter(
            models.PortfolioTx.stock_code == stock_code
        ).update({"stock_code": new_code, "stock_name": new_name})
        # 새 코드 온디맨드 수집
        if _collecting.get(new_code) != "running":
            threading.Thread(target=_bg_ondemand, args=(new_code,), daemon=True).start()
    else:
        h.stock_name = new_name
        h.sector     = new_sector
        h.quantity   = new_qty
        h.avg_price  = new_price

    db.commit()
    return {"status": "ok", "code_changed": code_changed,
            "new_code": new_code, "stock_name": new_name}


@app.delete("/api/portfolio/{stock_code}")
def delete_portfolio(stock_code: str, db: Session = Depends(get_db)):
    """보유종목 삭제"""
    h = db.query(models.Portfolio).filter(models.Portfolio.stock_code == stock_code).first()
    if h:
        db.delete(h)
        db.commit()
    return {"status": "ok"}


# ═══════════════════════════════════════════════════════
#  포트폴리오 엑셀 다운로드 / 업로드
# ═══════════════════════════════════════════════════════

@app.get("/api/portfolio/export/excel")
def export_portfolio_excel(db: Session = Depends(get_db)):
    """보유종목을 엑셀 파일로 다운로드"""
    import io
    from fastapi.responses import StreamingResponse
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치. pip install openpyxl")

    holdings = db.query(models.Portfolio).filter(models.Portfolio.quantity > 0).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "보유종목"

    # 스타일 정의
    header_fill  = PatternFill("solid", fgColor="1E1E2E")
    header_font  = Font(bold=True, color="2DD4BF", size=11)
    data_font    = Font(color="E2E8F0", size=10)
    border_side  = Side(style="thin", color="334155")
    thin_border  = Border(left=border_side, right=border_side,
                          top=border_side, bottom=border_side)
    center_align = Alignment(horizontal="center", vertical="center")
    right_align  = Alignment(horizontal="right",  vertical="center")

    # 헤더
    headers = ["종목코드","종목명","섹터","보유수량","매입가(원)",
               "현재가(원)","수익률(%)","손익(원)","평가액(원)","매입총액(원)"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border

    # 열 너비
    col_widths = [12,18,10,12,16,16,12,18,18,18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 데이터
    for h in holdings:
        price_row = db.query(models.PriceHistory).filter(
            models.PriceHistory.stock_code == h.stock_code
        ).order_by(models.PriceHistory.date.desc()).first()
        current_price = price_row.close if price_row else h.avg_price
        profit_pct  = round((current_price - h.avg_price) / h.avg_price * 100, 2) if h.avg_price else 0
        profit      = round((current_price - h.avg_price) * h.quantity)
        total_value = round(current_price * h.quantity)
        buy_total   = round(h.avg_price * h.quantity)

        row = [h.stock_code, h.stock_name, h.sector or "",
               int(h.quantity), int(h.avg_price), int(current_price),
               profit_pct, profit, total_value, buy_total]
        ws.append(row)

        # 셀 스타일
        data_row = ws.max_row
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=data_row, column=col)
            cell.font   = data_font
            cell.border = thin_border
            cell.alignment = right_align if col >= 4 else center_align
            # 배경 교대
            if data_row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="1A1A2E")
            else:
                cell.fill = PatternFill("solid", fgColor="0F0F1A")
            # 수익률/손익 색상
            if col == 7 and profit_pct != 0:
                cell.font = Font(color="EF4444" if profit_pct > 0 else "3B82F6", bold=True, size=10)
            if col == 8 and profit != 0:
                cell.font = Font(color="EF4444" if profit > 0 else "3B82F6", size=10)

    # 합계 행
    total_buy    = sum(int(h.avg_price * h.quantity) for h in holdings)
    total_val    = sum(
        round((db.query(models.PriceHistory).filter(models.PriceHistory.stock_code==h.stock_code)
               .order_by(models.PriceHistory.date.desc()).first() or type('x',(),{'close':h.avg_price})()).close * h.quantity)
        for h in holdings
    )
    total_profit = total_val - total_buy
    profit_pct_total = round(total_profit / total_buy * 100, 2) if total_buy else 0

    sum_row = len(holdings) + 2
    ws.append(["","합계","","","","",profit_pct_total, total_profit, total_val, total_buy])
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=sum_row, column=col)
        cell.font   = Font(bold=True, color="2DD4BF", size=10)
        cell.fill   = PatternFill("solid", fgColor="1E1E2E")
        cell.border = thin_border
        cell.alignment = right_align if col >= 4 else center_align

    # 두 번째 시트: 업로드 양식
    ws2 = wb.create_sheet("업로드양식")
    ws2.append(["종목명","보유수량","매입가(원)","섹터(선택)"])
    ws2.append(["삼성전자", 100, 75000, "반도체"])
    ws2.append(["에이엘티", 46698, 12301, "반도체"])
    for col in range(1, 5):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True, color="2DD4BF")
        cell.fill = PatternFill("solid", fgColor="1E1E2E")
        cell.alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["D"].width = 12
    # 안내문
    ws2.cell(row=5, column=1).value = "※ 종목명은 정확히 입력하세요 (ticker 자동 조회)"
    ws2.cell(row=6, column=1).value = "※ 현재가/손익/평가액은 시스템이 자동 계산합니다"
    ws2.cell(row=7, column=1).value = "※ 섹터는 비워도 됩니다"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import datetime as dt
    filename = f"portfolio_{dt.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.post("/api/portfolio/import/excel")
async def import_portfolio_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """엑셀 파일로 보유종목 일괄 등록.
    양식: 종목명 | 보유수량 | 매입가(원) | 섹터(선택)
    """
    import io
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치")

    from ticker_utils import ticker_mapper

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))

    # "업로드양식" 또는 첫 번째 시트 사용
    ws = wb["업로드양식"] if "업로드양식" in wb.sheetnames else wb.active

    results = {"success": [], "failed": [], "skipped": []}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        stock_name = str(row[0]).strip()
        # 안내문 행 스킵
        if stock_name.startswith("※"):
            continue

        try:
            quantity  = float(row[1]) if row[1] else 0
            avg_price = float(str(row[2]).replace(",","")) if row[2] else 0
            sector    = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        except (ValueError, TypeError):
            results["failed"].append({"name": stock_name, "reason": "수량/매입가 형식 오류"})
            continue

        if quantity <= 0 or avg_price <= 0:
            results["skipped"].append(stock_name)
            continue

        # 종목코드 조회
        stock_code = ticker_mapper.get_code(stock_name)
        if not stock_code:
            search = ticker_mapper.search(stock_name)
            if search:
                stock_code = search[0].get("code")
                stock_name = search[0].get("name", stock_name)

        if not stock_code:
            results["failed"].append({"name": stock_name, "reason": "종목코드 조회 실패"})
            continue

        # 기존 보유 확인 → 업데이트 or 신규 등록
        holding = db.query(models.Portfolio).filter(
            models.Portfolio.stock_code == stock_code
        ).first()

        if holding:
            # 이미 있으면 수량/매입가 업데이트
            holding.quantity  = quantity
            holding.avg_price = avg_price
            if sector: holding.sector = sector
            holding.stock_name = stock_name
        else:
            db.add(models.Portfolio(
                stock_code=stock_code, stock_name=stock_name,
                sector=sector, quantity=quantity, avg_price=avg_price
            ))
            # 거래 내역에도 추가
            from datetime import datetime as dt
            db.add(models.PortfolioTx(
                stock_code=stock_code, stock_name=stock_name,
                tx_type="buy", quantity=quantity, price=avg_price,
                tx_date=dt.now(), memo="엑셀 업로드"
            ))

        # watchlist 등록
        crud.add_to_watchlist(db, stock_code)
        results["success"].append({"code": stock_code, "name": stock_name,
                                    "quantity": quantity, "avg_price": avg_price})

    db.commit()

    # 성공 종목 주가 백그라운드 수집
    if results["success"]:
        import threading
        for item in results["success"]:
            code = item["code"]
            if _collecting.get(code) != "running":
                threading.Thread(target=_bg_collect, args=(code,), daemon=True).start()

    return {
        "status": "ok",
        "success_count": len(results["success"]),
        "failed_count":  len(results["failed"]),
        "skipped_count": len(results["skipped"]),
        "success": results["success"],
        "failed":  results["failed"],
    }




# ═══════════════════════════════════════════════════════════════════
#  보고서 API
# ═══════════════════════════════════════════════════════════════════

@app.get("/api/reports/stock/{stock_code}")
def get_stock_reports(stock_code: str):
    """종목코드별 보고서 목록 — 코드 없으면 종목명으로 fallback."""
    import sqlite3 as _sl
    conn = _sl.connect("stock.db")
    # 먼저 종목코드로 조회
    rows = conn.execute("""
        SELECT id, channel_id, stock_name, report_date,
               file_name, saved_name, file_size, caption
        FROM report_files
        WHERE stock_code=?
        ORDER BY report_date DESC LIMIT 50
    """, (stock_code,)).fetchall()
    # 없으면 stock.db의 종목명으로 fallback 조회
    if not rows:
        # watchlist 또는 listed_company_info에서 종목명 조회
        name_row = conn.execute(
            "SELECT stock_name FROM watchlist WHERE stock_code=? LIMIT 1",
            (stock_code,)
        ).fetchone()
        if not name_row:
            name_row = conn.execute(
                "SELECT corp_name FROM listed_company_info WHERE stock_code=? LIMIT 1",
                (stock_code,)
            ).fetchone()
        if name_row:
            sname = name_row[0][:4]  # 앞 4글자로 매칭
            rows = conn.execute("""
                SELECT id, channel_id, stock_name, report_date,
                       file_name, saved_name, file_size, caption
                FROM report_files
                WHERE stock_name LIKE ? OR file_name LIKE ?
                ORDER BY report_date DESC LIMIT 50
            """, (f"%{sname}%", f"%{sname}%")).fetchall()
    conn.close()
    return [{"id":r[0],"channel_id":r[1],"stock_name":r[2],"report_date":r[3],
             "file_name":r[4],"saved_name":r[5],"file_size":r[6],"caption":r[7]} for r in rows]


@app.get("/api/reports/download/{report_id}")
def download_report(report_id: int):
    """보고서 파일 다운로드."""
    from fastapi.responses import FileResponse
    import sqlite3 as _sl
    conn = _sl.connect("stock.db")
    row = conn.execute(
        "SELECT file_path, saved_name, mime_type FROM report_files WHERE id=?",
        (report_id,)
    ).fetchone()
    conn.close()
    if not row: raise HTTPException(status_code=404, detail="파일 없음")
    from pathlib import Path as _P
    fp = _P(row[0])
    if not fp.exists(): raise HTTPException(status_code=404, detail="파일이 로컬에 없습니다")
    return FileResponse(str(fp), filename=row[1], media_type=row[2] or "application/octet-stream")


@app.get("/api/reports/sectors")
def get_report_sectors():
    """섹터별 보고서 통계 (종목코드 없는 것만)."""
    import sqlite3 as _sl
    conn = _sl.connect("stock.db")
    rows = conn.execute("""
        SELECT sector, COUNT(*) as cnt, MAX(report_date) as latest
        FROM report_files
        WHERE (stock_code IS NULL OR stock_code='')
          AND sector != '' AND sector IS NOT NULL
        GROUP BY sector ORDER BY cnt DESC
    """).fetchall()
    conn.close()
    return [{"sector":r[0],"count":r[1],"latest":r[2]} for r in rows]


@app.get("/api/reports/sector/{sector:path}")
def get_sector_reports(sector: str, limit: int = 50):
    """특정 섹터 보고서 목록 (종목코드 없는 것만)."""
    import sqlite3 as _sl
    conn = _sl.connect("stock.db")
    rows = conn.execute("""
        SELECT id, channel_id, stock_name, report_date,
               file_name, saved_name, file_size, caption
        FROM report_files
        WHERE sector=? AND (stock_code IS NULL OR stock_code='')
        ORDER BY report_date DESC, id DESC LIMIT ?
    """, (sector, limit)).fetchall()
    conn.close()
    return [{"id":r[0],"channel_id":r[1],"stock_name":r[2],"report_date":r[3],
             "file_name":r[4],"saved_name":r[5],"file_size":r[6],"caption":r[7]} for r in rows]


@app.get("/api/telegram/channels")
def get_telegram_channels():
    import sqlite3 as _sl
    conn = _sl.connect("stock.db")
    rows = conn.execute(
        "SELECT id,channel_id,channel_name,is_active,last_sync FROM telegram_channels ORDER BY id"
    ).fetchall()
    conn.close()
    return [{"id":r[0],"channel_id":r[1],"channel_name":r[2],"is_active":bool(r[3]),"last_sync":r[4]} for r in rows]


@app.post("/api/telegram/channels")
def add_telegram_channel(payload: dict):
    import sqlite3 as _sl
    ch_id = payload.get("channel_id","").strip()
    if not ch_id: raise HTTPException(status_code=400, detail="channel_id 필수")
    conn = _sl.connect("stock.db")
    conn.execute("INSERT OR IGNORE INTO telegram_channels (channel_id,channel_name) VALUES (?,?)",
                 (ch_id, payload.get("channel_name", ch_id)))
    conn.commit(); conn.close()
    return {"status":"ok","channel_id":ch_id}


@app.delete("/api/telegram/channels/{channel_id}")
def delete_telegram_channel(channel_id: str):
    import sqlite3 as _sl
    conn = _sl.connect("stock.db")
    conn.execute("UPDATE telegram_channels SET is_active=0 WHERE channel_id=?", (channel_id,))
    conn.commit(); conn.close()
    return {"status":"ok"}


@app.post("/api/telegram/collect")
def trigger_collect(payload: dict = {}):
    import threading as _thr, subprocess as _sp
    channel = payload.get("channel_id","")
    def _run():
        cmd = ["/Applications/stock_dashboard/venv/bin/python3",
               "/Applications/stock_dashboard/telegram_collector.py"]
        if channel: cmd += ["--channel", channel]
        _sp.run(cmd, capture_output=True)
    _thr.Thread(target=_run, daemon=True).start()
    return {"status":"collecting","channel":channel or "전체"}

# ═══════════════════════════════════════════════════════════════════
#  추세추종 API  (/api/trend/*)
# ═══════════════════════════════════════════════════════════════════

@app.get("/api/trend/holdings")
def get_trend_holdings(db: Session = Depends(get_db)):
    import sqlite3 as _sl
    conn = _sl.connect("stock.db")
    rows = conn.execute("""
        SELECT id, stock_name, buy_price, current_price,
               quantity, profit_pct, sell_price, sold_at, is_active,
               hold_days, sector, updated_at, entry_date, sold_price, strategy
        FROM peak_holding ORDER BY entry_date DESC
    """).fetchall()
    conn.close()
    return [{
        "id": r[0], "stock_code": "", "stock_name": r[1],
        "buy_price": r[2], "current_price": r[3], "quantity": r[4],
        "profit_pct": r[5],
        "profit": round((r[3] - r[2]) * r[4]) if r[3] and r[2] and r[4] else 0,
        "total_value": round(r[3] * r[4]) if r[3] and r[4] else 0,
        "sell_price": r[6], "sold_at": r[7],
        "is_active": bool(r[8]), "hold_days": r[9],
        "sector": r[10], "updated_at": r[11],
        "entry_date": r[12], "strategy": r[14],
    } for r in rows]

@app.get("/api/trend/trades")
def get_trend_trades(db: Session = Depends(get_db)):
    import sqlite3 as _sl
    conn = _sl.connect("stock.db")
    rows = conn.execute("""
        SELECT id, stock_code, stock_name, tx_type, price,
               quantity, total_amount, profit, profit_pct, tx_at
        FROM peak_trade ORDER BY tx_at DESC LIMIT 100
    """).fetchall()
    conn.close()
    return [{
        "id": r[0], "stock_code": r[1], "stock_name": r[2],
        "tx_type": r[3], "price": r[4], "quantity": r[5],
        "total_amount": r[6], "profit": r[7], "profit_pct": r[8], "tx_at": r[9],
    } for r in rows]

@app.get("/api/trend/summary")
def get_trend_summary(db: Session = Depends(get_db)):
    import sqlite3 as _sl
    conn = _sl.connect("stock.db")
    row = conn.execute("""
        SELECT COUNT(*) as total,
               SUM(CASE WHEN profit > 0 THEN 1 ELSE 0 END) as wins,
               SUM(profit) as total_profit
        FROM peak_holding WHERE is_active=0
    """).fetchone()
    conn.close()
    total = row[0] or 0
    wins  = row[1] or 0
    return {
        "total_trades": total,
        "win_count":    wins,
        "win_rate":     round(wins/total*100, 1) if total > 0 else None,
        "total_profit": row[2] or 0,
    }


# ═══════════════════════════════════════════════════════════════════
#  시그널 보드 API
# ═══════════════════════════════════════════════════════════════════

@app.on_event("startup")
def init_signal():
    try:
        from signal_engine import init_signal_db
        init_signal_db()
        logger.info("Smart Score 100 시그널 DB 초기화 완료")
    except Exception as e:
        logger.warning(f"시그널 DB 초기화 실패: {e}")


@app.get("/api/signals/market")
def get_market_signals():
    """종합현황 시그널 반환."""
    try:
        import sqlite3 as _sl
        from signal_engine import calc_market_signals
        conn = _sl.connect("stock.db")
        results = calc_market_signals(conn)
        conn.commit(); conn.close()
        return results
    except Exception as e:
        logger.error(f"[시그널/시장] {e}")
        return []


@app.get("/api/signals/stock/{stock_code}")
def get_stock_signals(stock_code: str):
    """개별종목 Smart Score 시그널 반환."""
    try:
        import sqlite3 as _sl
        from signal_engine import calc_stock_signals
        conn = _sl.connect("stock.db")
        result = calc_stock_signals(stock_code, conn)
        conn.commit(); conn.close()
        return result  # {smart_score, verdict, one_liner, signals, flags}
    except Exception as e:
        logger.error(f"[시그널/종목] {stock_code}: {e}")
        return {"smart_score":0,"verdict":"⚪ 계산불가","verdict_color":"gray",
                "one_liner":"데이터 부족","signals":[],"flags":{}}


@app.get("/api/signals/config")
def get_signal_configs():
    """시그널 설정 목록 반환."""
    import sqlite3 as _sl
    conn = _sl.connect("stock.db")
    rows = conn.execute(
        "SELECT id,scope,name,label,description,logic_type,params,weight,is_active,sort_order "
        "FROM signal_config ORDER BY scope,sort_order"
    ).fetchall()
    conn.close()
    return [{"id":r[0],"scope":r[1],"name":r[2],"label":r[3],"description":r[4],
             "logic_type":r[5],"params":r[6],"weight":r[7],"is_active":bool(r[8]),"sort_order":r[9]}
            for r in rows]


@app.put("/api/signals/config/{config_id}")
def update_signal_config(config_id: int, payload: dict):
    """시그널 설정 수정."""
    import sqlite3 as _sl
    conn = _sl.connect("stock.db")
    allowed = ['label','description','params','weight','is_active','sort_order']
    sets = []
    vals = []
    for k in allowed:
        if k in payload:
            sets.append(f"{k}=?")
            vals.append(payload[k])
    if not sets:
        conn.close()
        raise HTTPException(status_code=400, detail="수정할 필드 없음")
    vals.append(config_id)
    conn.execute(f"UPDATE signal_config SET {','.join(sets)} WHERE id=?", vals)
    conn.commit(); conn.close()
    return {"status": "ok"}


@app.post("/api/signals/config")
def add_signal_config(payload: dict):
    """시그널 설정 추가."""
    import sqlite3 as _sl
    conn = _sl.connect("stock.db")
    conn.execute("""
        INSERT INTO signal_config (scope,name,label,description,logic_type,params,weight,sort_order)
        VALUES (?,?,?,?,?,?,?,?)
    """, (
        payload.get("scope","stock"),
        payload.get("name","custom"),
        payload.get("label","새 시그널"),
        payload.get("description",""),
        payload.get("logic_type","manual"),
        payload.get("params","{}"),
        payload.get("weight",1),
        payload.get("sort_order",99),
    ))
    conn.commit(); conn.close()
    return {"status": "ok"}


@app.delete("/api/signals/config/{config_id}")
def delete_signal_config(config_id: int):
    """시그널 삭제."""
    import sqlite3 as _sl
    conn = _sl.connect("stock.db")
    conn.execute("UPDATE signal_config SET is_active=0 WHERE id=?", (config_id,))
    conn.commit(); conn.close()
    return {"status": "ok"}


@app.post("/api/signals/manual/{config_id}")
def set_manual_signal(config_id: int, payload: dict):
    """수동 입력 시그널 값 설정 (Fear&Greed 등)."""
    import sqlite3 as _sl
    from datetime import date as _d
    conn = _sl.connect("stock.db")
    conn.execute("""
        INSERT OR REPLACE INTO signal_result
        (config_id, stock_code, signal, value, description, calc_date)
        VALUES (?,?,?,?,?,?)
    """, (
        config_id, '',
        payload.get("signal","yellow"),
        payload.get("value"),
        payload.get("description",""),
        _d.today().isoformat(),
    ))
    conn.commit(); conn.close()
    return {"status": "ok"}
