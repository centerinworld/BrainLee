#!/usr/bin/env python3
from __future__ import annotations

import json
import os
import subprocess
import sqlite3
import tempfile
from io import StringIO
from io import BytesIO
from datetime import date, datetime, timedelta
from pathlib import Path
import re
import xml.etree.ElementTree as ET

import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup

ROOT = Path(__file__).resolve().parents[2]
DB_PATH = ROOT / "stock.db"
HS_TRADE_DB_PATH = ROOT / "hs_trade_lab" / "data" / "hs_trade_lab.db"


CUSTOMS_SECTOR_QUANT_SPECS = [
    ("public:23:1", 1, "자동차 완성차 수출입", "자동차_완성차", ["8703"], "관세청 HS 8703 월별 합산. 완성차 수출액·수입액·무역수지·단가를 섹터 총량으로 제공."),
    ("public:23:2", 2, "자동차 부품 수출입", "자동차_부품", ["8708"], "관세청 HS 8708 월별 합산. 자동차부품 수출입 업황 보조 지표."),
    ("public:23:3", 3, "이차전지 리튬이온 수출입", "이차전지_리튬이온", ["850760"], "관세청 HS 850760 월별 합산. 배터리 셀/모듈 수출입 사이클 보조 지표."),
    ("public:23:4", 4, "메모리 반도체 수출입", "반도체_메모리", ["854232"], "관세청 HS 854232 월별 합산. DRAM/NAND/HBM 포함 메모리 반도체 수출입 지표."),
    ("public:23:5", 5, "시스템 반도체 수출입", "반도체_시스템", ["854231"], "관세청 HS 854231 월별 합산. 비메모리/시스템 반도체 수출입 지표."),
    ("public:23:6", 6, "반도체 제조장비 수출입", "반도체_제조장비", ["8486"], "관세청 HS 8486 월별 합산. 반도체 장비 사이클 보조 지표."),
    ("public:23:7", 7, "조선 상선 수출입", "조선_상선", ["8901"], "관세청 HS 8901 월별 합산. 선박 인도/수출 사이클 보조 지표."),
    ("public:23:8", 8, "철강 72/73류 수출입", "철강_72_73류", ["72", "73"], "관세청 HS 72+73류 월별 합산. 철강 제품 수출입과 단가 보조 지표."),
    ("public:23:9", 9, "화장품 수출입", "화장품", ["3304"], "관세청 HS 3304 월별 합산. 화장품 수출 업황 보조 지표."),
    ("public:23:10", 10, "의약품 수출입", "의약품", ["3004"], "관세청 HS 3004 월별 합산. 완제의약품 수출입 업황 보조 지표."),
    ("public:23:11", 11, "OLED/평판디스플레이 모듈 수출입", "디스플레이_OLED", ["8524", "901380"], "관세청 HS 8524+901380 월별 합산. OLED·평판디스플레이 모듈 수출입 사이클 보조 지표."),
    ("public:23:12", 12, "PCB 인쇄회로 수출입", "전자부품_PCB", ["8534"], "관세청 HS 8534 월별 합산. PCB/기판 업황 보조 지표."),
    ("public:23:13", 13, "MLCC/다층세라믹콘덴서 수출입", "전자부품_MLCC", ["853224"], "관세청 HS 853224 월별 합산. MLCC 수출입과 단가 보조 지표."),
    ("public:23:14", 14, "정유 석유제품 수출입", "정유_석유제품", ["2710"], "관세청 HS 2710 월별 합산. 석유제품 수출입·수입단가·무역수지 보조 지표."),
    ("public:23:15", 15, "원유 수입", "에너지_원유", ["2709"], "관세청 HS 2709 월별 합산. 원유 수입액·수입단가 보조 지표."),
    ("public:23:16", 16, "LNG 수입", "에너지_LNG", ["271111"], "관세청 HS 271111 월별 합산. LNG 수입액·수입단가 보조 지표."),
    ("public:23:17", 17, "석유화학 합성수지 수출입", "석유화학_합성수지", ["3901", "3902", "3903", "3904", "3905", "3906", "3907"], "관세청 HS 3901~3907 월별 합산. 합성수지/화학제품 사이클 보조 지표."),
    ("public:23:18", 18, "구리 원재료 수출입", "비철_구리", ["7403", "7408"], "관세청 HS 7403+7408 월별 합산. 구리 원재료·동선 수출입과 단가 보조 지표."),
    ("public:23:19", 19, "알루미늄 원재료 수출입", "비철_알루미늄", ["7601", "7606"], "관세청 HS 7601+7606 월별 합산. 알루미늄 원재료·판재 수출입과 단가 보조 지표."),
    ("public:23:20", 20, "후판/열연강판 수출입", "철강_후판열연", ["7208", "7225"], "관세청 HS 7208+7225 월별 합산. 조선·철강 후판/열연강판 단가 보조 지표."),
    ("public:23:21", 21, "선박용 디젤엔진 수출입", "조선_선박엔진", ["840810"], "관세청 HS 840810 월별 합산. 선박용 엔진 수출입 보조 지표."),
    ("public:23:22", 22, "공작기계 수출입", "기계_공작기계", ["8456", "8457", "8458", "8459", "8460", "8461", "8462", "8463", "8464", "8465", "8466"], "관세청 HS 8456~8466 월별 합산. 설비투자·기계 업황 보조 지표."),
    ("public:23:23", 23, "산업용 로봇 수출입", "기계_산업용로봇", ["847950"], "관세청 HS 847950 월별 합산. 자동화/로봇 업황 보조 지표."),
    ("public:23:24", 24, "의료기기 수출입", "헬스케어_의료기기", ["9018"], "관세청 HS 9018 월별 합산. 의료기기 수출입 업황 보조 지표."),
    ("public:23:25", 25, "진단시약 수출입", "헬스케어_진단시약", ["3822"], "관세청 HS 3822 월별 합산. 진단시약 수출입 업황 보조 지표."),
    ("public:23:26", 26, "백신/바이오의약품 수출입", "헬스케어_바이오의약품", ["3002"], "관세청 HS 3002 월별 합산. 백신·혈청·바이오의약품 수출입 보조 지표."),
    ("public:23:27", 27, "타이어 수출입", "자동차_타이어", ["4011"], "관세청 HS 4011 월별 합산. 타이어 수출입 업황 보조 지표."),
    ("public:23:28", 28, "비료 수출입", "화학_비료", ["3102", "3103", "3104", "3105"], "관세청 HS 3102~3105 월별 합산. 비료 수출입과 원가 사이클 보조 지표."),
    ("public:23:29", 29, "의류 수출입", "소비재_의류", ["61", "62"], "관세청 HS 61+62류 월별 합산. 의류 수출입 업황 보조 지표."),
    ("public:23:30", 30, "식품 가공품 수출입", "소비재_식품가공", ["1905", "2106"], "관세청 HS 1905+2106 월별 합산. 가공식품 수출입 보조 지표."),
]


def now_kst() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def safe_float(value) -> float | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        return None


def get_ecos_key() -> str:
    key = (os.getenv("ECOS_API_KEY", "") or os.getenv("BOK_ECOS_API_KEY", "")).strip()
    if key:
        return key
    for p in [
        "/Users/brainlee/Downloads/한국은행ECOS.txt",
        "/Applications/stock_dashboard/한국은행ECOS.txt",
    ]:
        try:
            if os.path.exists(p):
                value = Path(p).read_text(encoding="utf-8").strip()
                if value:
                    return value
        except Exception:
            pass
    return ""


def ecos_fetch_series(key: str, stat_code: str, cycle: str, start_t: str, end_t: str, item_code: str = "") -> list[dict]:
    suffix = f"/{item_code}" if item_code else ""
    url = f"https://ecos.bok.or.kr/api/StatisticSearch/{key}/json/kr/1/10000/{stat_code}/{cycle}/{start_t}/{end_t}{suffix}"
    r = requests.get(url, timeout=20)
    r.raise_for_status()
    js = r.json()
    return (js.get("StatisticSearch") or {}).get("row") or []


def init_tables(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS quant_major_indicator_catalog (
            indicator_key TEXT PRIMARY KEY,
            epic_category_code INTEGER,
            epic_sub_code INTEGER,
            epic_indicator_name TEXT NOT NULL,
            frequency TEXT,
            base_unit TEXT,
            status TEXT NOT NULL,
            replacement_family TEXT NOT NULL,
            source_system TEXT,
            collector_path TEXT,
            exactness TEXT,
            priority TEXT,
            notes TEXT,
            enabled INTEGER DEFAULT 1,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS quant_major_indicator_series (
            id INTEGER PRIMARY KEY,
            indicator_key TEXT NOT NULL,
            period TEXT NOT NULL,
            series_name TEXT NOT NULL,
            value REAL,
            unit TEXT,
            source_name TEXT,
            source_detail TEXT,
            quality TEXT DEFAULT 'raw',
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(indicator_key, period, series_name, source_name)
        );

        CREATE INDEX IF NOT EXISTS idx_qmi_series_key_period
            ON quant_major_indicator_series(indicator_key, period);
        CREATE INDEX IF NOT EXISTS idx_qmi_catalog_priority
            ON quant_major_indicator_catalog(priority, status);
        """
    )


def seed_catalog(conn: sqlite3.Connection) -> list[dict]:
    rows = conn.execute(
        """
        select category_code, category_name, sub_code, indicator_name, frequency, unit,
               status, replacement_family, source_system, collector_path, exactness, priority, notes
        from epic_indicator_replacement_plan
        order by priority, category_code, sub_code, indicator_name
        """
    ).fetchall()

    seeded = []
    for row in rows:
        key = f"epic:{row[0]}:{row[2]}"
        conn.execute(
            """
            INSERT INTO quant_major_indicator_catalog (
                indicator_key, epic_category_code, epic_sub_code, epic_indicator_name,
                frequency, base_unit, status, replacement_family, source_system,
                collector_path, exactness, priority, notes, enabled, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
            ON CONFLICT(indicator_key) DO UPDATE SET
                epic_indicator_name=excluded.epic_indicator_name,
                frequency=excluded.frequency,
                base_unit=excluded.base_unit,
                status=CASE
                    WHEN excluded.status='new_collector_needed'
                     AND quant_major_indicator_catalog.status IN ('ready_existing', 'ready_existing_partial')
                    THEN quant_major_indicator_catalog.status
                    ELSE excluded.status
                END,
                replacement_family=excluded.replacement_family,
                source_system=CASE
                    WHEN excluded.status='new_collector_needed'
                     AND quant_major_indicator_catalog.status IN ('ready_existing', 'ready_existing_partial')
                    THEN quant_major_indicator_catalog.source_system
                    ELSE excluded.source_system
                END,
                collector_path=CASE
                    WHEN excluded.status='new_collector_needed'
                     AND quant_major_indicator_catalog.status IN ('ready_existing', 'ready_existing_partial')
                    THEN quant_major_indicator_catalog.collector_path
                    ELSE excluded.collector_path
                END,
                exactness=CASE
                    WHEN excluded.status='new_collector_needed'
                     AND quant_major_indicator_catalog.status IN ('ready_existing', 'ready_existing_partial')
                    THEN quant_major_indicator_catalog.exactness
                    ELSE excluded.exactness
                END,
                priority=excluded.priority,
                notes=CASE
                    WHEN excluded.status='new_collector_needed'
                     AND quant_major_indicator_catalog.status IN ('ready_existing', 'ready_existing_partial')
                    THEN quant_major_indicator_catalog.notes
                    ELSE excluded.notes
                END,
                enabled=1,
                updated_at=excluded.updated_at
            """,
            (
                key, row[0], row[2], row[3], row[4], row[5], row[6], row[7], row[8],
                row[9], row[10], row[11], row[12], now_kst(),
            ),
        )
        seeded.append({"indicator_key": key, "name": row[3], "status": row[6], "priority": row[11]})
    conn.commit()
    return seeded


def upsert_series(conn: sqlite3.Connection, indicator_key: str, rows: list[dict]) -> int:
    count = 0
    for row in rows:
        series_name = row.get("series_name") or row.get("label") or "value"
        source_name = row.get("source_name") or row.get("source") or "unknown"
        conn.execute(
            """
            INSERT INTO quant_major_indicator_series (
                indicator_key, period, series_name, value, unit, source_name, source_detail, quality, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(indicator_key, period, series_name, source_name) DO UPDATE SET
                value=excluded.value,
                unit=excluded.unit,
                source_detail=excluded.source_detail,
                quality=excluded.quality,
                updated_at=excluded.updated_at
            """,
            (
                indicator_key,
                row["period"],
                series_name,
                row.get("value"),
                row.get("unit"),
                source_name,
                row.get("source_detail", ""),
                row.get("quality", "raw"),
                now_kst(),
            ),
        )
        count += 1
    conn.commit()
    return count


def update_catalog_status(
    conn: sqlite3.Connection,
    indicator_key: str,
    *,
    status: str,
    source_system: str,
    collector_path: str,
    exactness: str,
    notes: str,
) -> None:
    conn.execute(
        """
        UPDATE quant_major_indicator_catalog
           SET status = ?,
               source_system = ?,
               collector_path = ?,
               exactness = ?,
               notes = ?,
               updated_at = ?
         WHERE indicator_key = ?
        """,
        (status, source_system, collector_path, exactness, notes, now_kst(), indicator_key),
    )
    conn.commit()


def upsert_custom_catalog(
    conn: sqlite3.Connection,
    *,
    indicator_key: str,
    epic_category_code: int,
    epic_sub_code: int,
    epic_indicator_name: str,
    frequency: str,
    base_unit: str,
    status: str,
    replacement_family: str,
    source_system: str,
    collector_path: str,
    exactness: str,
    priority: str,
    notes: str,
) -> None:
    conn.execute(
        """
        INSERT INTO quant_major_indicator_catalog (
            indicator_key, epic_category_code, epic_sub_code, epic_indicator_name,
            frequency, base_unit, status, replacement_family, source_system,
            collector_path, exactness, priority, notes, enabled, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
        ON CONFLICT(indicator_key) DO UPDATE SET
            epic_category_code=excluded.epic_category_code,
            epic_sub_code=excluded.epic_sub_code,
            epic_indicator_name=excluded.epic_indicator_name,
            frequency=excluded.frequency,
            base_unit=excluded.base_unit,
            status=excluded.status,
            replacement_family=excluded.replacement_family,
            source_system=excluded.source_system,
            collector_path=excluded.collector_path,
            exactness=excluded.exactness,
            priority=excluded.priority,
            notes=excluded.notes,
            enabled=1,
            updated_at=excluded.updated_at
        """,
        (
            indicator_key,
            epic_category_code,
            epic_sub_code,
            epic_indicator_name,
            frequency,
            base_unit,
            status,
            replacement_family,
            source_system,
            collector_path,
            exactness,
            priority,
            notes,
            now_kst(),
        ),
    )
    conn.commit()


def derive_market_breadth_from_price_history(conn: sqlite3.Connection, start_date: str = "2021-01-01") -> dict[str, list[dict]]:
    """Derive market breadth and volume expansion indicators from local daily prices.

    We keep this as a derived indicator instead of an official KRX feed so that
    downstream users can distinguish price-history-based breadth from exchange
    published advance/decline counts.
    """
    query = """
        WITH latest_universe AS (
            SELECT stock_code, MAX(base_date) AS base_date
              FROM stock_universe
             GROUP BY stock_code
        ),
        universe AS (
            SELECT u.stock_code, u.market, COALESCE(u.stock_type, '') AS stock_type
              FROM stock_universe u
              JOIN latest_universe lu
                ON u.stock_code = lu.stock_code
               AND u.base_date = lu.base_date
             WHERE u.market IN ('KOSPI', 'KOSDAQ')
               AND COALESCE(u.stock_type, '보통주') IN ('', '보통주')
        )
        SELECT p.stock_code,
               DATE(p.date) AS period,
               u.market,
               p.close,
               p.volume,
               COALESCE(p.trade_amount, 0) AS trade_amount
          FROM price_history p
          JOIN universe u ON p.stock_code = u.stock_code
         WHERE DATE(p.date) >= ?
           AND p.close > 0
         ORDER BY p.stock_code, DATE(p.date)
    """
    df = pd.read_sql_query(query, conn, params=(start_date,))
    if df.empty:
        return {}

    df["close"] = pd.to_numeric(df["close"], errors="coerce")
    df["volume"] = pd.to_numeric(df["volume"], errors="coerce")
    df["trade_amount"] = pd.to_numeric(df["trade_amount"], errors="coerce").fillna(0)
    df = df.dropna(subset=["close"])
    df = df.sort_values(["stock_code", "period"])

    grouped = df.groupby("stock_code", sort=False)
    df["prev_close"] = grouped["close"].shift(1)
    df["return_pct"] = (df["close"] / df["prev_close"] - 1.0) * 100.0
    df["prev_20_high"] = grouped["close"].transform(lambda s: s.shift(1).rolling(20, min_periods=15).max())
    df["prev_20_low"] = grouped["close"].transform(lambda s: s.shift(1).rolling(20, min_periods=15).min())
    df["prev_20_volume_avg"] = grouped["volume"].transform(lambda s: s.shift(1).rolling(20, min_periods=10).mean())
    df["volume_ratio_20d"] = df["volume"] / df["prev_20_volume_avg"]
    df["is_new_high_20d"] = df["close"] >= df["prev_20_high"]
    df["is_new_low_20d"] = df["close"] <= df["prev_20_low"]
    df["is_volume_3x"] = df["volume_ratio_20d"] >= 3

    result: dict[str, list[dict]] = {
        "public:21:1": [],
        "public:21:2": [],
        "public:21:3": [],
        "public:21:4": [],
    }

    for (market, period), g in df.groupby(["market", "period"], sort=True):
        valid = g.dropna(subset=["return_pct"])
        total = int(len(valid))
        if total < 500:
            # Incomplete local collection days are kept out of the breadth series.
            continue
        advance = int((valid["return_pct"] > 0).sum())
        decline = int((valid["return_pct"] < 0).sum())
        unchanged = int((valid["return_pct"] == 0).sum())
        advance_ratio = advance / total * 100.0 if total else None
        decline_ratio = decline / total * 100.0 if total else None
        median_return = float(valid["return_pct"].median()) if total else None
        new_high_count = int(valid["is_new_high_20d"].fillna(False).sum())
        new_low_count = int(valid["is_new_low_20d"].fillna(False).sum())
        volume_3x_count = int(valid["is_volume_3x"].fillna(False).sum())
        avg_volume_ratio = float(valid["volume_ratio_20d"].replace([float("inf"), -float("inf")], pd.NA).dropna().median()) if valid["volume_ratio_20d"].notna().any() else None
        total_trade_amount_100m = float(valid["trade_amount"].sum() / 100_000_000.0) if valid["trade_amount"].sum() else None

        breadth_key = "public:21:1" if market == "KOSPI" else "public:21:2"
        volume_key = "public:21:3" if market == "KOSPI" else "public:21:4"
        common = {
            "period": period,
            "source_name": "local_price_history",
            "source_detail": "price_history + latest stock_universe common-stock universe",
            "quality": "derived_market_breadth",
        }
        result[breadth_key].extend([
            {**common, "series_name": "상승종목수", "value": advance, "unit": "종목"},
            {**common, "series_name": "하락종목수", "value": decline, "unit": "종목"},
            {**common, "series_name": "보합종목수", "value": unchanged, "unit": "종목"},
            {**common, "series_name": "상승종목비율", "value": round(advance_ratio, 2) if advance_ratio is not None else None, "unit": "%"},
            {**common, "series_name": "하락종목비율", "value": round(decline_ratio, 2) if decline_ratio is not None else None, "unit": "%"},
            {**common, "series_name": "중앙수익률", "value": round(median_return, 2) if median_return is not None else None, "unit": "%"},
            {**common, "series_name": "커버종목수", "value": total, "unit": "종목"},
        ])
        result[volume_key].extend([
            {**common, "series_name": "20일신고가수", "value": new_high_count, "unit": "종목"},
            {**common, "series_name": "20일신저가수", "value": new_low_count, "unit": "종목"},
            {**common, "series_name": "거래량3배종목수", "value": volume_3x_count, "unit": "종목"},
            {**common, "series_name": "거래량20일중앙배율", "value": round(avg_volume_ratio, 2) if avg_volume_ratio is not None else None, "unit": "배"},
            {**common, "series_name": "총거래대금", "value": round(total_trade_amount_100m, 0) if total_trade_amount_100m is not None else None, "unit": "억원"},
            {**common, "series_name": "커버종목수", "value": total, "unit": "종목"},
        ])

    for key, rows in result.items():
        print(f"[MarketBreadth] {key}: {len(rows)}행")
    return result


def collect_ecos_macro_quant_extensions() -> dict[str, list[dict]]:
    key = get_ecos_key()
    if not key:
        return {}
    end_t = date.today().strftime("%Y%m")
    start_t = "201001"
    specs = [
        ("public:20:101", "511Y002", "M", "FME", "소비자심리지수", "지수", "ECOS 소비자동향조사"),
        ("public:20:102", "513Y001", "M", "E2000", "경제심리지수_순환변동치", "지수", "ECOS 경제심리지수"),
        ("public:20:103", "512Y013", "M", "C0000/AA", "제조업_업황실적BSI", "지수", "ECOS 기업경기조사"),
        ("public:20:103", "512Y013", "M", "C0000/AD", "제조업_신규수주실적BSI", "지수", "ECOS 기업경기조사"),
        ("public:20:103", "512Y014", "M", "C0000/BA", "제조업_업황전망BSI", "지수", "ECOS 기업경기조사"),
        ("public:20:104", "901Y026", "M", "I33A", "제조업_재고율", "%", "ECOS 제조업 재고율"),
        ("public:20:105", "901Y033", "M", "A00/2", "전산업생산지수_SA", "지수", "ECOS 전산업생산지수"),
    ]
    result: dict[str, list[dict]] = {}
    for indicator_key, stat_code, cycle, item_code, series_name, unit, source_name in specs:
        try:
            rows = ecos_fetch_series(key, stat_code, cycle, start_t, end_t, item_code)
        except Exception as exc:
            print(f"[ECOS-MacroExtension] skip {indicator_key}/{item_code}: {exc}")
            continue
        for row in rows:
            period = str(row.get("TIME") or "")
            value = safe_float(row.get("DATA_VALUE"))
            if not period or value is None:
                continue
            result.setdefault(indicator_key, []).append({
                "period": f"{period[:4]}-{period[4:6]}",
                "series_name": series_name,
                "value": value,
                "unit": unit,
                "source_name": source_name,
                "source_detail": f"{stat_code}/{item_code}",
                "quality": "official_ecos",
            })
    for key_, rows in result.items():
        print(f"[ECOS-MacroExtension] {key_}: {len(rows)}행")
    return result


def collect_bok_base_rate() -> list[dict]:
    key = get_ecos_key()
    if not key:
        return []
    rows = ecos_fetch_series(key, "722Y001", "M", "199905", date.today().strftime("%Y%m"), "0101000")
    out = []
    for row in rows:
        ym = str(row.get("TIME") or "")
        if len(ym) != 6:
            continue
        out.append(
            {
                "period": f"{ym[:4]}-{ym[4:6]}",
                "series_name": "base_rate_pct",
                "value": float(str(row.get("DATA_VALUE") or "0").replace(",", "")),
                "unit": "%",
                "source_name": "ECOS_722Y001_0101000",
                "source_detail": "한국은행 기준금리 (월)",
                "quality": "official",
            }
        )
    return out


def collect_bok_market_liquidity() -> list[dict]:
    key = get_ecos_key()
    if not key:
        return []
    start_ym = "200810"
    end_ym = date.today().strftime("%Y%m")
    dep = ecos_fetch_series(key, "901Y056", "M", start_ym, end_ym, "S23A")
    crd = ecos_fetch_series(key, "901Y056", "M", start_ym, end_ym, "S23E")
    dep_map = {str(r.get("TIME")): float(str(r.get("DATA_VALUE") or "0").replace(",", "")) / 100_000_000.0 for r in dep}
    crd_map = {str(r.get("TIME")): float(str(r.get("DATA_VALUE") or "0").replace(",", "")) / 100_000_000.0 for r in crd}
    out = []
    for ym in sorted(set(dep_map) | set(crd_map)):
        if len(ym) != 6:
            continue
        period = f"{ym[:4]}-{ym[4:6]}"
        if ym in dep_map:
            out.append(
                {
                    "period": period,
                    "series_name": "customer_deposit_100m",
                    "value": round(dep_map[ym], 2),
                    "unit": "억원",
                    "source_name": "ECOS_901Y056_S23A",
                    "source_detail": "투자자 예탁금",
                    "quality": "official",
                }
            )
        if ym in crd_map:
            out.append(
                {
                    "period": period,
                    "series_name": "credit_balance_100m",
                    "value": round(crd_map[ym], 2),
                    "unit": "억원",
                    "source_name": "ECOS_901Y056_S23E",
                    "source_detail": "신용공여/신용융자 잔고",
                    "quality": "official",
                }
            )
    return out


def collect_monthly_short_balance(conn: sqlite3.Connection) -> list[dict]:
    sql = """
    WITH month_last AS (
      SELECT substr(bas_dt,1,6) AS ym, max(bas_dt) AS bas_dt
      FROM short_sell_daily
      GROUP BY substr(bas_dt,1,6)
    ),
    month_dates AS (
      SELECT
        bas_dt,
        substr(bas_dt,1,4)||'-'||substr(bas_dt,5,2)||'-'||substr(bas_dt,7,2) AS dt
      FROM month_last
    ),
    latest_close AS (
      SELECT
        stock_code,
        substr(date,1,10) AS dt,
        close,
        ROW_NUMBER() OVER (
            PARTITION BY stock_code, substr(date,1,10)
            ORDER BY date DESC
        ) AS rn
      FROM price_history
      WHERE close > 0
        AND substr(date,1,10) IN (SELECT dt FROM month_dates)
    )
    SELECT
      ml.ym,
      ROUND(SUM(ss.borrow_bal_qty), 0) AS borrow_bal_qty_sum,
      ROUND(SUM(ss.borrow_bal_qty * lc.close) / 1000000.0, 2) AS borrow_bal_million_krw
    FROM month_last ml
    JOIN short_sell_daily ss
      ON ss.bas_dt = ml.bas_dt
     AND ss.borrow_bal_qty IS NOT NULL
     AND ss.borrow_bal_qty > 0
    JOIN latest_close lc
      ON lc.stock_code = ss.stock_code
     AND lc.dt = substr(ss.bas_dt,1,4)||'-'||substr(ss.bas_dt,5,2)||'-'||substr(ss.bas_dt,7,2)
     AND lc.rn = 1
    GROUP BY ml.ym
    ORDER BY ml.ym
    """
    out = []
    for ym, qty_sum, bal_million in conn.execute(sql).fetchall():
        period = f"{ym[:4]}-{ym[4:6]}"
        out.append(
            {
                "period": period,
                "series_name": "borrow_balance_million_krw",
                "value": float(bal_million or 0),
                "unit": "백만원",
                "source_name": "short_sell_daily_x_price_history",
                "source_detail": "대차잔고주수 × 월말종가 합산",
                "quality": "derived_from_market_data",
            }
        )
        out.append(
            {
                "period": period,
                "series_name": "borrow_balance_qty",
                "value": float(qty_sum or 0),
                "unit": "주",
                "source_name": "short_sell_daily",
                "source_detail": "월말 대차잔고주수 합산",
                "quality": "official_or_exchange_derived",
            }
        )
    return out


def collect_kpx_monthly_smp() -> list[dict]:
    url = "https://new.kpx.or.kr/menu.es?mid=a10404080300"
    r = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    table = soup.find("table")
    if not table:
        return []

    month_headers: list[int] = []
    rows = table.find_all("tr")
    if not rows:
        return []

    header_cells = [c.get_text(" ", strip=True) for c in rows[0].find_all(["th", "td"])]
    for cell in header_cells[1:]:
        m = re.search(r"(\d{1,2})월", cell)
        if m:
            month_headers.append(int(m.group(1)))
    if not month_headers:
        return []

    out: list[dict] = []
    current_region: str | None = None
    region_map = {
        "육지 SMP": "land_smp_krw_per_kwh",
        "제주 SMP": "jeju_smp_krw_per_kwh",
        "통합 SMP": "integrated_smp_krw_per_kwh",
    }

    for tr in rows[1:]:
        cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
        if not cells:
            continue
        if cells[0] in region_map:
            current_region = cells[0]
            year_token = cells[1] if len(cells) > 1 else ""
            values = cells[2:]
        else:
            year_token = cells[0]
            values = cells[1:]

        if current_region not in region_map:
            continue
        year_match = re.search(r"(\d{2})년", year_token)
        if not year_match:
            continue
        year = 2000 + int(year_match.group(1))

        for month, value_text in zip(month_headers, values):
            if value_text in {"", "-", "…"}:
                continue
            value = safe_float(value_text)
            if value is None:
                continue
            out.append(
                {
                    "period": f"{year:04d}-{month:02d}",
                    "series_name": region_map[current_region],
                    "value": round(value, 2),
                    "unit": "원/kWh",
                    "source_name": "KPX_MONTHLY_SMP",
                    "source_detail": f"전력거래소 월별 계통한계가격(SMP) - {current_region}",
                    "quality": "official",
                }
            )
    return out


def collect_kosis_online_shopping_total() -> list[dict]:
    base = "https://kosis.kr/visual/economyBoard"
    payload = {
        "unitySrvcIdArr": "599",
        "stdIdctIdArr": "511",
        "clsfGroupCdArr": "",
        "clsfCdArr": "",
        "cyclSe": "M",
        "regionArr": "00",
        "spclBefore": "",
        "spclIncrease": "",
    }
    r = requests.post(
        f"{base}/selectDetailDataList.do?lang=ko",
        data=payload,
        timeout=30,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Referer": f"{base}/economyJipyo.do?lang=ko",
        },
    )
    r.raise_for_status()
    js = r.json()
    out: list[dict] = []
    for row in js.get("data", []):
        ym = str(row.get("wrtPnttm") or "")
        if len(ym) != 6:
            continue
        value = safe_float(row.get("vl"))
        if value is None:
            continue
        out.append(
            {
                "period": f"{ym[:4]}-{ym[4:6]}",
                "series_name": "online_shopping_total_million_krw",
                "value": value,
                "unit": row.get("unit") or "백만원",
                "source_name": "KOSIS_ECONOMYBOARD_599_511",
                "source_detail": "KOSIS 경제상황판 온라인쇼핑 거래액",
                "quality": "official",
            }
        )
    return out


def month_range_desc(start_ym: str, end_ym: str) -> list[str]:
    sy, sm = int(start_ym[:4]), int(start_ym[4:6])
    ey, em = int(end_ym[:4]), int(end_ym[4:6])
    out: list[str] = []
    y, m = ey, em
    while (y > sy) or (y == sy and m >= sm):
        out.append(f"{y}{m:02d}")
        m -= 1
        if m == 0:
            y -= 1
            m = 12
    return out


def parse_kosis_stat_html_table(table_html: str) -> list[dict]:
    soup = BeautifulSoup(table_html, "html.parser")
    periods: list[str] = []
    for cell in soup.select("thead th"):
        text = cell.get_text(" ", strip=True)
        m = re.search(r"(\d{4})\.(\d{2})", text)
        if m:
            periods.append(f"{m.group(1)}-{m.group(2)}")
    rows: list[dict] = []
    current_group1 = ""
    current_group2 = ""
    for tr in soup.select("tbody tr"):
        cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
        if len(cells) < 3:
            continue
        axis_count = len(cells) - len(periods)
        if axis_count >= 3:
            if cells[0]:
                current_group1 = cells[0]
                current_group2 = ""
            if cells[1]:
                current_group2 = cells[1]
            group1, group2, channel = current_group1, current_group2, cells[2]
            data_cells = cells[3:]
        else:
            if cells[0]:
                current_group1 = cells[0]
            group1, group2, channel = current_group1, "", cells[1]
            data_cells = cells[2:]
        for period, raw in zip(periods, data_cells):
            value = safe_float(raw)
            if value is None:
                continue
            rows.append(
                {
                    "period": period,
                    "group1": group1,
                    "group2": group2,
                    "channel": channel,
                    "value": value,
                }
            )
    return rows


def parse_kosis_single_axis_table(table_html: str, axis_name: str) -> list[dict]:
    """Parse a KOSIS statHtml table with one row dimension and time columns."""
    soup = BeautifulSoup(table_html, "html.parser")
    periods: list[str] = []
    for cell in soup.select("thead th"):
        text = cell.get_text(" ", strip=True)
        m = re.search(r"(\d{4})\.(\d{2})", text)
        if m:
            periods.append(f"{m.group(1)}-{m.group(2)}")
    rows: list[dict] = []
    if not periods:
        return rows
    for tr in soup.select("tbody tr"):
        cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
        if len(cells) < len(periods) + 1:
            continue
        item = cells[0].strip()
        if not item or item == axis_name:
            continue
        for period, raw in zip(periods, cells[-len(periods):]):
            value = safe_float(raw)
            if value is None:
                continue
            rows.append({"period": period, "item": item, "value": value})
    return rows




def collect_kosis_pharmacy_goods_sales_proxy() -> dict[str, list[dict]]:
    """Collect official KOSIS medicine retail sales as a pharmacy proxy.

    EPIC asks for pharmacy card spending. Exact card merchant-category data is
    not available in the current public/connected sources. KOSIS DT_1K41002
    provides monthly retail sales by goods group, including "의약품"; this is a
    useful official proxy but not pharmacy card exact.
    """
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return {}

    source_name = "KOSIS_DT_1K41002"
    source_detail = "KOSIS 서비스업동향조사 재별 및 상품군별 판매액"
    periods = month_range_desc("202001", "202604")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        try:
            page = browser.new_page(viewport={"width": 1400, "height": 900})
            page.goto(
                "https://kosis.kr/statHtml/statHtml.do?orgId=101&tblId=DT_1K41002",
                wait_until="domcontentloaded",
                timeout=30000,
            )
            page.wait_for_timeout(5000)
            centers = [frame for frame in page.frames if frame.name == "iframe_centerMenu"]
            if not centers:
                return {}
            response_text = centers[0].evaluate(
                """async (periods) => {
                  const form = document.querySelector('#ParamInfo');
                  if (!form) return JSON.stringify({errCode: 1, errMsg: 'ParamInfo missing'});
                  const fd = new FormData(form);
                  let fieldList = JSON.parse(fd.get('fieldList'));
                  fieldList = fieldList.filter((x) => x.targetId !== 'PRD');
                  fieldList.unshift({targetId:'PRD', targetValue:'', prdValue:'M,' + periods.join(',') + ',@'});
                  fd.set('fieldList', JSON.stringify(fieldList));
                  fd.set('jsonStr', '');
                  fd.set('isFirst', 'Y');
                  fd.set('colAxis', 'TIME');
                  fd.set('rowAxis', 'A');
                  fd.set('viewKind', '1');
                  fd.set('view', 'table');
                  fd.set('diviSearchYn', 'N');
                  fd.set('orderStr', 'OV_L1_ID,TIME');
                  fd.set('startNum', '1');
                  fd.set('endNum', String(periods.length * 40));
                  fd.set('lastChk', 'N');
                  fd.set('classAllArr', JSON.stringify([{objVarId:'A', ovlSn:'1'}]));
                  fd.set('classSet', JSON.stringify([{objVarId:'A', ovlSn:'1', visible:'true'}]));
                  const body = new URLSearchParams();
                  for (const [k, v] of fd.entries()) body.append(k, v);
                  const resp = await fetch('/statHtml/html.do', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'},
                    body: body.toString()
                  });
                  return await resp.text();
                }""",
                periods,
            )
        finally:
            browser.close()

    try:
        payload = json.loads(response_text)
    except Exception:
        return {}
    html = (payload.get("result") or [""])[0]
    if not html:
        return {}
    parsed = parse_kosis_single_axis_table(html, "상품군별")
    rows: list[dict] = []
    for row in parsed:
        if row.get("item") != "의약품":
            continue
        rows.append(
            {
                "period": row["period"],
                "series_name": "medicine_retail_sales_million_krw",
                "value": row["value"],
                "unit": "백만원",
                "source_name": source_name,
                "source_detail": f"{source_detail} - 의약품",
                "quality": "official_medicine_retail_sales_proxy_not_pharmacy_card_exact",
            }
        )
    return {"epic:16:113": rows}
def collect_kosis_retail_store_sales_proxy() -> dict[str, list[dict]]:
    """Collect official KOSIS retail sales by store type as card-spending proxies.

    EPIC labels these as card-spending estimates, but the public exact card data
    is not available. This function intentionally marks the output as a proxy:
    official monthly retail sales by store type from KOSIS DT_1K41003.
    """
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return {}

    source_name = "KOSIS_DT_1K41003"
    source_detail = "KOSIS 서비스업동향조사 소매업태별 판매액"
    store_codes = ["A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8"]
    periods = month_range_desc("202001", "202604")
    field_list = [
        {"targetId": "PRD", "targetValue": "", "prdValue": "M," + ",".join(periods) + ",@"},
        {"targetId": "ITM_ID", "targetValue": "T1", "prdValue": ""},
    ]
    field_list.extend({"targetId": "OV_L1_ID", "targetValue": code, "prdValue": ""} for code in store_codes)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        try:
            page = browser.new_page(viewport={"width": 1400, "height": 900})
            page.goto(
                "https://kosis.kr/statHtml/statHtml.do?orgId=101&tblId=DT_1K41003",
                wait_until="domcontentloaded",
                timeout=30000,
            )
            page.wait_for_timeout(5000)
            centers = [frame for frame in page.frames if frame.name == "iframe_centerMenu"]
            if not centers:
                return {}
            response_text = centers[0].evaluate(
                """async (arg) => {
                  const {fieldList, endNum} = arg;
                  const form = document.querySelector('#ParamInfo');
                  if (!form) return JSON.stringify({errCode: 1, errMsg: 'ParamInfo missing'});
                  const fd = new FormData(form);
                  fd.set('fieldList', JSON.stringify(fieldList));
                  fd.set('jsonStr', '');
                  fd.set('isFirst', 'Y');
                  fd.set('colAxis', 'TIME');
                  fd.set('rowAxis', 'A');
                  fd.set('viewKind', '1');
                  fd.set('view', 'table');
                  fd.set('diviSearchYn', 'N');
                  fd.set('orderStr', 'OV_L1_ID,TIME');
                  fd.set('startNum', '1');
                  fd.set('endNum', String(endNum));
                  fd.set('lastChk', 'N');
                  fd.set('classAllArr', JSON.stringify([{objVarId:'A', ovlSn:'1'}]));
                  fd.set('classSet', JSON.stringify([{objVarId:'A', ovlSn:'1', visible:'true'}]));
                  const body = new URLSearchParams();
                  for (const [k, v] of fd.entries()) body.append(k, v);
                  const resp = await fetch('/statHtml/html.do', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'},
                    body: body.toString()
                  });
                  return await resp.text();
                }""",
                {"fieldList": field_list, "endNum": len(periods) * len(store_codes)},
            )
        finally:
            browser.close()

    try:
        payload = json.loads(response_text)
    except Exception:
        return {}
    result = payload.get("result") or []
    if not result:
        return {}
    parsed = parse_kosis_single_axis_table(result[0], "업태별")
    if not parsed:
        return {}

    out: dict[str, list[dict]] = {
        "epic:2:93": [],
        "epic:2:94": [],
        "epic:2:95": [],
        "epic:2:96": [],
        "epic:2:97": [],
    }
    discount_super_acc: dict[str, float] = {}
    item_map = {
        "백화점": ("epic:2:93", "department_store_sales_million_krw"),
        "편의점": ("epic:2:95", "convenience_store_sales_million_krw"),
        "면세점": ("epic:2:96", "duty_free_store_sales_million_krw"),
        "무점포 소매": ("epic:2:97", "nonstore_retail_sales_million_krw"),
    }
    for row in parsed:
        item = row["item"]
        if item in {"대형마트", "슈퍼마켓 및 잡화점"}:
            discount_super_acc[row["period"]] = discount_super_acc.get(row["period"], 0.0) + row["value"]
            continue
        spec = item_map.get(item)
        if not spec:
            continue
        key, series_name = spec
        out[key].append(
            {
                "period": row["period"],
                "series_name": series_name,
                "value": row["value"],
                "unit": "백만원",
                "source_name": source_name,
                "source_detail": f"{source_detail} - {item}",
                "quality": "official_retail_sales_proxy",
            }
        )
    for period, value in sorted(discount_super_acc.items()):
        out["epic:2:94"].append(
            {
                "period": period,
                "series_name": "discount_store_supermarket_sales_million_krw",
                "value": value,
                "unit": "백만원",
                "source_name": source_name,
                "source_detail": f"{source_detail} - 대형마트+슈퍼마켓 및 잡화점 합산",
                "quality": "official_retail_sales_proxy_derived_sum",
            }
        )
    return out


def collect_kosis_service_industry_index_proxy() -> dict[str, list[dict]]:
    """Collect quarterly service production indices as broad card-spending proxies.

    These are not card-spending exact values. We only map industries where the
    KOSIS industry label is close enough to be useful as a directional proxy.
    """
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return {}

    periods: list[str] = []
    for year in range(2021, 2027):
        for quarter in range(1, 5):
            if year == 2026 and quarter > 1:
                continue
            periods.append(f"{year}{quarter:02d}")
    periods = list(reversed(periods))
    source_name = "KOSIS_DT_1KC2023"
    source_detail = "KOSIS 서비스업동향조사 시도별 서비스업생산지수"

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        try:
            page = browser.new_page(viewport={"width": 1400, "height": 900})
            page.goto(
                "https://kosis.kr/statHtml/statHtml.do?orgId=101&tblId=DT_1KC2023",
                wait_until="domcontentloaded",
                timeout=30000,
            )
            page.wait_for_timeout(5000)
            centers = [frame for frame in page.frames if frame.name == "iframe_centerMenu"]
            if not centers:
                return {}
            response_text = centers[0].evaluate(
                """async (periods) => {
                  const form = document.querySelector('#ParamInfo');
                  if (!form) return JSON.stringify({errCode: 1, errMsg: 'ParamInfo missing'});
                  const fd = new FormData(form);
                  let fieldList = JSON.parse(fd.get('fieldList'));
                  fieldList = fieldList.filter((x) => x.targetId !== 'PRD');
                  fieldList.unshift({targetId:'PRD', targetValue:'', prdValue:'Q,' + periods.join(',') + ',@'});
                  fd.set('fieldList', JSON.stringify(fieldList));
                  fd.set('jsonStr', '');
                  fd.set('isFirst', 'Y');
                  fd.set('colAxis', 'TIME,ITEM');
                  fd.set('rowAxis', 'SGG,A');
                  fd.set('viewKind', '1');
                  fd.set('view', 'table');
                  fd.set('diviSearchYn', 'N');
                  fd.set('orderStr', 'OV_L1_ID,OV_L2_ID,TIME,CHAR_ITM_ID');
                  fd.set('startNum', '1');
                  fd.set('endNum', String(periods.length * 2 * 17 * 14));
                  fd.set('lastChk', 'N');
                  const body = new URLSearchParams();
                  for (const [k, v] of fd.entries()) body.append(k, v);
                  const resp = await fetch('/statHtml/html.do', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'},
                    body: body.toString()
                  });
                  return await resp.text();
                }""",
                periods,
            )
        finally:
            browser.close()

    try:
        payload = json.loads(response_text)
    except Exception:
        return {}
    html = (payload.get("result") or [""])[0]
    if not html:
        return {}

    soup = BeautifulSoup(html, "html.parser")
    parsed_periods: list[str] = []
    first_header = soup.select_one("thead tr")
    if not first_header:
        return {}
    for th in first_header.find_all("th"):
        title = th.get("title") or th.get_text(" ", strip=True)
        match = re.search(r"(\d{4})\.(\d)/4", title)
        if match:
            parsed_periods.append(f"{match.group(1)}-Q{match.group(2)}")
    if not parsed_periods:
        return {}

    out: dict[str, list[dict]] = {"epic:11:156": [], "epic:13:20": []}
    industry_map = {
        "숙박 및 음식점업": ("epic:11:156", "food_service_production_index"),
        "교육 서비스업": ("epic:13:20", "education_service_production_index"),
    }
    current_region = ""
    for tr in soup.select("tbody tr"):
        cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
        if len(cells) < 2 + len(parsed_periods) * 2:
            continue
        if cells[0]:
            current_region = cells[0]
        industry = cells[1]
        spec = industry_map.get(industry)
        if not spec or not current_region:
            continue
        indicator_key, prefix = spec
        values = cells[-len(parsed_periods) * 2:]
        for idx, period in enumerate(parsed_periods):
            nominal = safe_float(values[idx * 2])
            real = safe_float(values[idx * 2 + 1])
            if nominal is not None:
                out[indicator_key].append(
                    {
                        "period": period,
                        "series_name": f"{prefix}_nominal_{current_region}",
                        "value": nominal,
                        "unit": "2020=100",
                        "source_name": source_name,
                        "source_detail": f"{source_detail} - {current_region} {industry} 경상지수",
                        "quality": "official_quarterly_service_index_proxy",
                    }
                )
            if real is not None:
                out[indicator_key].append(
                    {
                        "period": period,
                        "series_name": f"{prefix}_real_{current_region}",
                        "value": real,
                        "unit": "2020=100",
                        "source_name": source_name,
                        "source_detail": f"{source_detail} - {current_region} {industry} 불변지수",
                        "quality": "official_quarterly_service_index_proxy",
                    }
                )
    return out


def collect_kosis_online_shopping_breakdown() -> dict[str, list[dict]]:
    """Collect DT_1KE10071 through KOSIS' official statHtml renderer.

    KOSIS OpenAPI requires an API key, but the public statHtml page renders this
    table through a same-origin html.do JSON response. We initialize that page
    with Playwright and request the official table response from inside the page
    session to avoid brittle hand-built cookies.
    """
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return {}

    source_name = "KOSIS_DT_1KE10071"
    source_detail = "KOSIS 온라인쇼핑동향조사 온라인쇼핑몰 판매매체별/상품군별거래액"
    item_codes = [
        "000", "001", "002", "003", "004", "005", "006", "007", "008", "009",
        "010", "011", "012", "013", "014", "015", "016", "017", "018", "019",
        "020", "021", "022", "023", "0021", "0022",
    ]
    channel_codes = ["00", "20", "10"]
    periods = month_range_desc("201701", "202604")
    field_list = [
        {"targetId": "PRD", "targetValue": "", "prdValue": "M," + ",".join(periods) + ",@"},
        {"targetId": "ITM_ID", "targetValue": "T20", "prdValue": ""},
    ]
    field_list.extend({"targetId": "OV_L1_ID", "targetValue": code, "prdValue": ""} for code in item_codes)
    field_list.extend({"targetId": "OV_L2_ID", "targetValue": code, "prdValue": ""} for code in channel_codes)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        try:
            page = browser.new_page(viewport={"width": 1400, "height": 900})
            page.goto(
                "https://kosis.kr/statHtml/statHtml.do?orgId=101&tblId=DT_1KE10071",
                wait_until="domcontentloaded",
                timeout=30000,
            )
            page.wait_for_timeout(6000)
            centers = [frame for frame in page.frames if frame.name == "iframe_centerMenu"]
            if not centers:
                return {}
            response_text = centers[0].evaluate(
                """async (arg) => {
                  const {fieldList, endNum} = arg;
                  const form = document.querySelector('#ParamInfo');
                  if (!form) return JSON.stringify({errCode: 1, errMsg: 'ParamInfo missing'});
                  const fd = new FormData(form);
                  fd.set('fieldList', JSON.stringify(fieldList));
                  fd.set('jsonStr', '');
                  fd.set('isFirst', 'Y');
                  fd.set('colAxis', 'TIME');
                  fd.set('rowAxis', 'A,B');
                  fd.set('viewKind', '1');
                  fd.set('view', 'table');
                  fd.set('diviSearchYn', 'N');
                  fd.set('orderStr', 'OV_L1_ID,OV_L2_ID,TIME');
                  fd.set('startNum', '1');
                  fd.set('endNum', String(endNum));
                  fd.set('lastChk', 'N');
                  fd.set('classAllArr', JSON.stringify([{objVarId:'A', ovlSn:'1'}, {objVarId:'B', ovlSn:'2'}]));
                  fd.set('classSet', JSON.stringify([{objVarId:'A', ovlSn:'1', visible:'true'}, {objVarId:'B', ovlSn:'2', visible:'true'}]));
                  const body = new URLSearchParams();
                  for (const [k, v] of fd.entries()) body.append(k, v);
                  const resp = await fetch('/statHtml/html.do', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'},
                    body: body.toString()
                  });
                  return await resp.text();
                }""",
                {"fieldList": field_list, "endNum": len(periods) * len(item_codes) * len(channel_codes)},
            )
        finally:
            browser.close()

    payload = json.loads(response_text)
    result = payload.get("result") or []
    if not result:
        return {}
    parsed = parse_kosis_stat_html_table(result[0])
    if not parsed:
        return {}

    out: dict[str, list[dict]] = {
        "epic:2:22": [],
        "epic:2:23": [],
        "epic:8:14": [],
        "epic:8:15": [],
        "epic:12:5": [],
        "epic:12:6": [],
    }
    apparel_groups = {"의복", "신발", "가방", "패션용품 및 액세서리"}
    apparel_acc: dict[tuple[str, str], float] = {}

    for row in parsed:
        group = row["group2"] if row["group2"] and row["group2"] != "소계" else row["group1"]
        channel = row["channel"]
        if not group or channel not in {"인터넷쇼핑", "모바일쇼핑"}:
            continue
        channel_slug = "internet" if channel == "인터넷쇼핑" else "mobile"
        channel_key = "epic:2:22" if channel_slug == "internet" else "epic:2:23"
        out[channel_key].append(
            {
                "period": row["period"],
                "series_name": f"{channel_slug}_{group}_million_krw",
                "value": row["value"],
                "unit": "백만원",
                "source_name": source_name,
                "source_detail": f"{source_detail} - {channel}/{group}",
                "quality": "official",
            }
        )
        if group == "화장품":
            target_key = "epic:8:14" if channel_slug == "internet" else "epic:8:15"
            out[target_key].append(
                {
                    "period": row["period"],
                    "series_name": f"{channel_slug}_cosmetics_million_krw",
                    "value": row["value"],
                    "unit": "백만원",
                    "source_name": source_name,
                    "source_detail": f"{source_detail} - {channel}/화장품",
                    "quality": "official",
                }
            )
        if group in apparel_groups:
            apparel_acc[(row["period"], channel_slug)] = apparel_acc.get((row["period"], channel_slug), 0.0) + row["value"]

    for (period, channel_slug), value in sorted(apparel_acc.items()):
        target_key = "epic:12:5" if channel_slug == "internet" else "epic:12:6"
        out[target_key].append(
            {
                "period": period,
                "series_name": f"{channel_slug}_apparel_fashion_million_krw",
                "value": value,
                "unit": "백만원",
                "source_name": source_name,
                "source_detail": f"{source_detail} - 의복+신발+가방+패션용품 및 액세서리 합산",
                "quality": "official_derived_sum",
            }
        )
    return out


def collect_macao_ggr(start_year: int = 2010) -> list[dict]:
    """Collect Macao monthly gross gaming revenue from DICJ official XML files."""
    out: list[dict] = []
    month_map = {
        "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04", "May": "05", "Jun": "06",
        "Jul": "07", "Aug": "08", "Sept": "09", "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
    }
    current_year = date.today().year
    for year in range(start_year, current_year + 1):
        url = f"https://www.dicj.gov.mo/web/en/information/DadosEstat_mensal/{year}/report_en.xml"
        try:
            r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=20)
            if r.status_code != 200 or not r.text.lstrip().startswith("<?xml"):
                continue
            root = ET.fromstring(r.content)
        except Exception:
            continue
        for rec in root.findall(".//RECORD"):
            cells = [(d.text or "").strip() for d in rec.findall("DATA")]
            if len(cells) < 2:
                continue
            month = month_map.get(cells[0])
            value = safe_float(cells[1])
            if not month or value is None:
                continue
            out.append(
                {
                    "period": f"{year}-{month}",
                    "series_name": "macao_ggr_mop_million",
                    "value": value,
                    "unit": "MOP million",
                    "source_name": "DICJ_MACAO_GGR",
                    "source_detail": f"DICJ Monthly Gross Revenue from Games of Fortune ({year})",
                    "quality": "official",
                }
            )
    return out


def collect_kto_foreign_visitors_purpose_file() -> list[dict]:
    """Collect KTO inbound visitor purpose monthly file from data.go.kr.

    The public file currently covers only a short 12-month window, so this is
    intentionally treated as a partial connection in the catalog.
    """
    url = (
        "https://www.data.go.kr/cmm/cmm/fileDownload.do?"
        "atchFileId=FILE_000000003027530&fileDetailSn=1&insertDataPrcus=N"
    )
    try:
        r = requests.get(
            url,
            headers={
                "User-Agent": "Mozilla/5.0",
                "Referer": "https://www.data.go.kr/data/15136332/fileData.do?recommendDataYn=Y",
            },
            timeout=30,
        )
        r.raise_for_status()
        df = pd.read_csv(BytesIO(r.content), encoding="cp949")
    except Exception:
        return []

    date_col = "기준연월"
    if date_col not in df.columns:
        return []

    purpose_cols = [c for c in ["공용", "관광", "기타", "상용", "유학연수"] if c in df.columns]
    out: list[dict] = []
    for _, row in df.iterrows():
        try:
            period = pd.to_datetime(row[date_col]).strftime("%Y-%m")
        except Exception:
            continue
        total = 0.0
        seen = False
        for col in purpose_cols:
            value = safe_float(row.get(col))
            if value is None:
                continue
            seen = True
            total += value
            out.append(
                {
                    "period": period,
                    "series_name": f"inbound_visitors_{col}",
                    "value": value,
                    "unit": "명",
                    "source_name": "DATA_GO_KTO_15136332",
                    "source_detail": "한국관광공사_방한 외래관광객 목적별 월별 집계",
                    "quality": "official_partial_file",
                }
            )
        if seen:
            out.append(
                {
                    "period": period,
                    "series_name": "inbound_visitors_total",
                    "value": total,
                    "unit": "명",
                    "source_name": "DATA_GO_KTO_15136332",
                    "source_detail": "목적별 방한 외래관광객 합계(공용+관광+기타+상용+유학연수)",
                    "quality": "official_partial_file_derived_sum",
                }
            )
    return out




def _parse_korean_count_phrase(text: str) -> float | None:
    """Parse Korean count phrases like '8만4616명' or '15만 5000여 명'."""
    text = text.replace(",", "").replace("여", "")
    m = re.search(r"(\d+)\s*만\s*(\d+)?\s*명", text)
    if m:
        return float(int(m.group(1)) * 10000 + int(m.group(2) or 0))
    m = re.search(r"(\d+)\s*명", text)
    if m:
        return float(m.group(1))
    return None


def collect_modetour_press_release_package_outbound() -> list[dict]:
    """Collect Modetour package traveler counts from company press releases.

    This is intentionally partial: only months with explicit numeric package
    traveler counts in Modetour's Newswire press releases are persisted. We do
    not infer missing months and we do not treat broad KTO outbound data as
    company package exact values.
    """
    company_url = "https://www.newswire.co.kr/?md=A10&act=article&no=25926"
    headers = {"User-Agent": "Mozilla/5.0"}
    links: list[str] = []
    try:
        r = requests.get(company_url, headers=headers, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        for a in soup.select("a[href]"):
            href = a.get("href") or ""
            title = a.get_text(" ", strip=True)
            if "newsRead.php" not in href:
                continue
            if "모두투어" not in title and "송출객" not in title and "패키지" not in title:
                continue
            if href.startswith("/"):
                href = "https://www.newswire.co.kr" + href
            links.append(href)
    except Exception:
        links = []

    # Keep known historical disclosure link as a fallback because the company
    # page only exposes recent releases.
    links.append("https://www.newswire.co.kr/newsRead.php?no=990946")
    seen_links: set[str] = set()
    rows: list[dict] = []
    for url in links:
        if url in seen_links:
            continue
        seen_links.add(url)
        try:
            r = requests.get(url, headers=headers, timeout=20)
            if r.status_code != 200:
                continue
            soup = BeautifulSoup(r.text, "html.parser")
            text = soup.get_text(" ", strip=True)
        except Exception:
            continue
        if "모두투어" not in text or "송출객" not in text:
            continue

        # Prefer publication year plus article-stated month, e.g. 2024-05.
        title = soup.find("title")
        pub_year = None
        m_pub = re.search(r"(20\d{2})[.-](\d{1,2})[.-](\d{1,2})", text)
        if m_pub:
            pub_year = int(m_pub.group(1))
        if pub_year is None:
            m_meta = re.search(r"(20\d{2})", text)
            pub_year = int(m_meta.group(1)) if m_meta else datetime.now().year

        # Sentences with explicit package traveler count.
        for m in re.finditer(r"(\d{1,2})월[^.。]{0,80}?해외\s*패키지\s*송출객수는\s*([^.,。]+?명)", text):
            month = int(m.group(1))
            count = _parse_korean_count_phrase(m.group(2))
            if count is None:
                continue
            # If article was published in Jan but discusses Dec, use previous year.
            year = pub_year - 1 if pub_year and month == 12 and datetime.now().month == 1 else pub_year
            period = f"{year}-{month:02d}"
            rows.append(
                {
                    "period": period,
                    "series_name": "modetour_package_outbound_press_release",
                    "value": count,
                    "unit": "명",
                    "source_name": "Newswire_Modetour_press_release",
                    "source_detail": f"모두투어 보도자료 숫자 파싱: {url}",
                    "quality": "company_press_release_partial_exact",
                }
            )
    # Deduplicate by period, keep first source.
    dedup: dict[tuple[str, str], dict] = {}
    for row in rows:
        dedup.setdefault((row["period"], row["series_name"]), row)
    return [dedup[k] for k in sorted(dedup)]
def collect_kto_korean_outbound_transport_monthly() -> list[dict]:
    """Collect KTO monthly Korean outbound travelers by departure transport.

    This is a broad official travel-demand proxy. It is not Modetour company
    outbound traffic and should not be used as a package-tour exact value.
    """
    url = (
        "https://www.data.go.kr/cmm/cmm/fileDownload.do"
        "?atchFileId=FILE_000000003027495&fileDetailSn=1&insertDataPrcus=N"
    )
    try:
        response = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
        response.raise_for_status()
        df = pd.read_csv(BytesIO(response.content), encoding="cp949")
    except Exception as exc:
        print(f"[KTO-Outbound] 수집 실패: {exc}")
        return []

    if "기준연월" not in df.columns:
        print("[KTO-Outbound] 예상 컬럼 불일치")
        return []
    value_cols = [c for c in df.columns if c != "기준연월"]
    for col in value_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    out: list[dict] = []
    for _, row in df.iterrows():
        period = str(row["기준연월"])[:7]
        if not re.match(r"\d{4}-\d{2}", period):
            continue
        total = float(sum(row[col] for col in value_cols))
        out.append(
            {
                "period": period,
                "series_name": "korean_outbound_total",
                "value": total,
                "unit": "명",
                "source_name": "DATA_GO_KTO_15136315",
                "source_detail": "한국관광공사_국민 해외관광객 교통수단별 월별 집계 - 전체 합산",
                "quality": "official_kto_outbound_travel_demand_proxy",
            }
        )
        air_total = float(sum(row[col] for col in value_cols if "공항" in col))
        sea_total = float(sum(row[col] for col in value_cols if "항구" in col))
        out.append(
            {
                "period": period,
                "series_name": "korean_outbound_air",
                "value": air_total,
                "unit": "명",
                "source_name": "DATA_GO_KTO_15136315",
                "source_detail": "한국관광공사_국민 해외관광객 교통수단별 월별 집계 - 공항 합산",
                "quality": "official_kto_outbound_travel_demand_proxy",
            }
        )
        out.append(
            {
                "period": period,
                "series_name": "korean_outbound_sea",
                "value": sea_total,
                "unit": "명",
                "source_name": "DATA_GO_KTO_15136315",
                "source_detail": "한국관광공사_국민 해외관광객 교통수단별 월별 집계 - 항구 합산",
                "quality": "official_kto_outbound_travel_demand_proxy",
            }
        )
        for col in value_cols:
            out.append(
                {
                    "period": period,
                    "series_name": f"korean_outbound_{col}",
                    "value": float(row[col]),
                    "unit": "명",
                    "source_name": "DATA_GO_KTO_15136315",
                    "source_detail": f"한국관광공사_국민 해외관광객 교통수단별 월별 집계 - {col}",
                    "quality": "official_kto_outbound_travel_demand_proxy",
                }
            )
    if out:
        print(f"[KTO-Outbound] {len(out)}행 수집")
    return out


def collect_kto_regional_visitors_monthly(start_ym: str = "202001") -> list[dict]:
    """Collect KTO DataLab monthly visitor trend by Korean province/city.

    The DataLab page explicitly recommends using these big-data visitor
    counts as a trend indicator rather than an absolute total. We therefore
    keep the series quality as official_trend and document the caveat in the
    catalog instead of treating it as a census-grade population total.
    """
    session = requests.Session()
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://datalab.visitkorea.or.kr/datalab/portal/loc/getAreaDataForm.do",
        "X-Requested-With": "XMLHttpRequest",
    }
    init_url = "https://datalab.visitkorea.or.kr/portal/getSrchDteDivInitVal.do"
    data_url = "https://datalab.visitkorea.or.kr/visualize/getTempleteData.do"

    try:
        init = session.post(
            init_url,
            data={"menuGrpCd": "1", "menuCd": "10102030000002020091512"},
            headers=headers,
            timeout=20,
        )
        init.raise_for_status()
        init_rows = init.json().get("list") or []
    except Exception as exc:
        print(f"[KTO-Regional] init failed: {exc}")
        return []

    end_ym = ""
    for row in init_rows:
        if str(row.get("cyclDiv") or "") == "1":
            end_ym = str(row.get("endDteVal") or "")
            break
    if not re.fullmatch(r"\d{6}", end_ym or ""):
        end_ym = date.today().strftime("%Y%m")

    def month_iter(start: str, end: str):
        y, m = int(start[:4]), int(start[4:6])
        ey, em = int(end[:4]), int(end[4:6])
        while (y, m) <= (ey, em):
            yield f"{y:04d}{m:02d}"
            m += 1
            if m == 13:
                y += 1
                m = 1

    out: list[dict] = []
    for ym in month_iter(start_ym, end_ym):
        payload = {
            "qid": "LN_01_01_016",
            "SGG_CD": "47830",
            "SGG_NM": "경상북도 고령군",
            "tabDiv": "1",
            "BASE_YM1": ym,
            "BASE_YM2": ym,
            "srchAreaDate": "1",
            "sggIntgYnFlag": "N",
            "dispYn": "Y",
        }
        try:
            response = session.post(data_url, data=payload, headers=headers, timeout=20)
            response.raise_for_status()
            rows = response.json().get("list") or []
        except Exception as exc:
            print(f"[KTO-Regional] {ym} failed: {exc}")
            continue

        period = f"{ym[:4]}-{ym[4:6]}"
        for row in rows:
            name = str(row.get("SIDO_NM") or "").strip()
            value = safe_float(row.get("TOU_NUM"))
            if not name or value is None:
                continue
            out.append(
                {
                    "period": period,
                    "series_name": f"regional_visitors_{name}",
                    "value": round(value / 1000.0, 3),
                    "unit": "천명",
                    "source_name": "KTO_DATALAB_LN_01_01_016",
                    "source_detail": f"한국관광공사 DataLab 지역별 관광현황 월간 시도 방문자수 추세 - {name}",
                    "quality": "official_trend",
                }
            )

    return out


def collect_itstat_iptv_subscribers_annual(start_year: int = 2009) -> list[dict]:
    """Collect official annual IPTV subscriber counts from ITSTAT.

    EPIC's target is monthly IPTV subscribers. The public ICT statistics table
    available without credentials is annual terminal-count data, so keep this
    as a partial official series rather than pretending it is monthly exact.
    """
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return []

    end_year = datetime.now().year - 1
    years = list(range(end_year, start_year - 1, -1))
    if not years:
        return []

    url = "https://www.itstat.go.kr/statHtml/statHtml.do?orgId=006&tblId=DT_164_27"
    response_text = ""
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        try:
            page = browser.new_page(viewport={"width": 1400, "height": 900})
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(7000)
            centers = [frame for frame in page.frames if frame.name == "iframe_centerMenu"]
            if not centers:
                return []
            response_text = centers[0].evaluate(
                """async (years) => {
                  const form = document.querySelector('#ParamInfo');
                  if (!form) return JSON.stringify({errCode: 1, errMsg: 'ParamInfo missing'});
                  const fd = new FormData(form);
                  let fieldList = JSON.parse(fd.get('fieldList'));
                  fieldList = fieldList.filter((x) => x.targetId !== 'PRD');
                  fieldList.unshift({
                    targetId: 'PRD',
                    targetValue: '',
                    prdValue: 'Y,' + years.join(',') + ',@'
                  });
                  fd.set('fieldList', JSON.stringify(fieldList));
                  fd.set('jsonStr', '');
                  fd.set('isFirst', 'Y');
                  fd.set('colAxis', 'TIME');
                  fd.set('rowAxis', 'O_1');
                  fd.set('viewKind', '1');
                  fd.set('view', 'table');
                  fd.set('diviSearchYn', 'N');
                  fd.set('orderStr', 'OV_L1_ID,TIME');
                  fd.set('startNum', '1');
                  fd.set('endNum', String(years.length * 20));
                  fd.set('lastChk', 'N');
                  const body = new URLSearchParams();
                  for (const [k, v] of fd.entries()) body.append(k, v);
                  const resp = await fetch('/statHtml/html.do', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'},
                    body: body.toString()
                  });
                  return await resp.text();
                }""",
                [str(year) for year in years],
            )
        finally:
            browser.close()

    try:
        payload = json.loads(response_text)
    except Exception:
        return []
    html = (payload.get("result") or [""])[0]
    if not html:
        return []

    soup = BeautifulSoup(html, "html.parser")
    header_years: list[int] = []
    for th in soup.select("thead th"):
        text = th.get_text(" ", strip=True)
        if re.fullmatch(r"\d{4}", text):
            header_years.append(int(text))
    if not header_years:
        return []

    out: list[dict] = []
    for tr in soup.select("tbody tr"):
        cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
        if len(cells) < len(header_years) + 2:
            continue
        category = cells[0].strip()
        subcategory = cells[1].strip()
        if "IPTV" not in category or subcategory != "소계":
            continue
        values = cells[-len(header_years):]
        for year, raw in zip(header_years, values):
            value = safe_float(raw)
            if value is None:
                continue
            out.append(
                {
                    "period": str(year),
                    "series_name": "iptv_subscribers_terminal_count",
                    "value": round(value, 0),
                    "unit": "명",
                    "source_name": "ITSTAT_DT_164_27",
                    "source_detail": "ICT통계포털 유료방송 가입자(단자기준) - IPTV 소계, 연간",
                    "quality": "official_annual_partial",
                }
            )
        break
    return out


def collect_eia_us_rig_count_monthly() -> list[dict]:
    """Collect official monthly U.S. rotary rig count from EIA.

    EPIC's target is weekly Baker Hughes rig count. The Baker Hughes page is
    the exact source, but it is not reliably reachable from this environment.
    EIA republishes U.S. drilling activity monthly, so this is stored as an
    official monthly proxy and not used as a weekly exact replacement.
    """
    url = "https://www.eia.gov/dnav/ng/hist/e_ertrr0_xr0_nus_cm.htm"
    try:
        response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
        response.raise_for_status()
        tables = pd.read_html(StringIO(response.text))
        df = next((table for table in tables if "Year" in table.columns and "Jan" in table.columns), None)
        if df is None:
            return []
    except Exception as exc:
        print(f"[EIA-RigCount] failed: {exc}")
        return []

    month_map = {
        "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
        "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
        "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
    }
    out: list[dict] = []
    for _, row in df.iterrows():
        year = safe_float(row.get("Year"))
        if year is None:
            continue
        for month_name, month in month_map.items():
            value = safe_float(row.get(month_name))
            if value is None:
                continue
            out.append(
                {
                    "period": f"{int(year):04d}-{month}",
                    "series_name": "us_crude_oil_and_natural_gas_rotary_rigs",
                    "value": value,
                    "unit": "개",
                    "source_name": "EIA_E_ERTRR0_XR0_NUS_CM",
                    "source_detail": "EIA U.S. Crude Oil and Natural Gas Rotary Rigs in Operation, monthly history",
                    "quality": "official_monthly_proxy",
                }
            )
    return out


def collect_bdry_dry_bulk_shipping_proxy_monthly() -> list[dict]:
    """Collect BDRY ETF monthly prices as a clearly-labelled BDI proxy.

    The EPIC target is the daily Baltic Dry Index. Public free CSV endpoints for
    BDI/BCI/BPI/BSI are currently unreliable from this environment, while BDRY
    (Breakwave Dry Bulk Shipping ETF) is a transparent dry-bulk freight proxy.
    Store it only as a partial monthly proxy, never as the exact BDI index.
    """
    try:
        import yfinance as yf
    except Exception:
        return []

    try:
        df = yf.download("BDRY", period="10y", interval="1mo", progress=False, auto_adjust=True)
    except Exception:
        return []
    if df is None or df.empty:
        return []

    def col(name: str):
        try:
            s = df[(name, "BDRY")]
            return s.iloc[:, 0] if hasattr(s, "columns") else s
        except Exception:
            pass
        try:
            s = df[name]
            return s.iloc[:, 0] if hasattr(s, "columns") else s
        except Exception:
            return None

    close = col("Close")
    volume = col("Volume")
    if close is None:
        return []

    out: list[dict] = []
    for idx, value in close.dropna().items():
        try:
            period = pd.to_datetime(idx).strftime("%Y-%m")
            close_value = float(value)
        except Exception:
            continue
        out.append(
            {
                "period": period,
                "series_name": "bdry_etf_monthly_close_usd",
                "value": round(close_value, 4),
                "unit": "USD",
                "source_name": "Yahoo Finance BDRY",
                "source_detail": "BDRY ETF monthly adjusted close; dry bulk freight proxy for BDI",
                "quality": "market_proxy_monthly",
            }
        )
        if volume is not None:
            try:
                vol_value = float(volume.loc[idx])
            except Exception:
                vol_value = None
            if vol_value is not None:
                out.append(
                    {
                        "period": period,
                        "series_name": "bdry_etf_monthly_volume",
                        "value": vol_value,
                        "unit": "주",
                        "source_name": "Yahoo Finance BDRY",
                        "source_detail": "BDRY ETF monthly trading volume; liquidity reference only",
                        "quality": "market_proxy_monthly",
                    }
                )
    return out


def collect_kline_dry_bulk_indices_weekly() -> dict[str, list[dict]]:
    """Collect weekly dry bulk indices embedded in K-Line's public IR chart.

    K-Line republishes BDI plus BCI/BPI/BSI/BHSI series in a Highcharts page.
    This is closer to the EPIC Baltic index targets than ETF proxies, but it is
    still a third-party IR republication rather than a direct Baltic Exchange
    licensed feed, so keep the catalog status partial.
    """
    url = "https://www.kline.co.jp/en/ir/finance/shipping.html"
    try:
        response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
        response.raise_for_status()
        text = response.text
    except Exception as exc:
        print(f"[KLineDryBulk] failed: {exc}")
        return {}

    def block_for(component_id: str) -> str:
        start = text.find(component_id)
        if start < 0:
            return ""
        end = text.find("</script>", start)
        return text[start:end if end > start else len(text)]

    def decode_arg_after(block: str, comment: str):
        pos = block.find(f"// {comment}")
        if pos < 0:
            return None
        quote = block.find('"', pos)
        if quote < 0:
            return None
        try:
            raw, _ = json.JSONDecoder().raw_decode(block[quote:])
            return json.loads(raw)
        except Exception:
            return None

    result: dict[str, list[dict]] = {}
    specs = {
        "graphComponent-1": {"BDI": ("epic:7:14", "baltic_dry_index_bdi")},
        "graphComponent-2": {
            "BCI": ("epic:7:15", "baltic_capesize_index_bci"),
            "BPI": ("epic:7:16", "baltic_panamax_index_bpi"),
            "BSI": ("epic:7:17", "baltic_supramax_index_bsi"),
        },
    }
    for component_id, index_map in specs.items():
        block = block_for(component_id)
        labels = decode_arg_after(block, "label")
        affiliates = decode_arg_after(block, "affiliate")
        tabledata = decode_arg_after(block, "tabledata")
        if not labels or not affiliates or not tabledata:
            continue
        for series_idx, series_label in enumerate(affiliates):
            if series_label not in index_map or series_idx >= len(tabledata):
                continue
            indicator_key, series_name = index_map[series_label]
            rows: list[dict] = []
            for label, value in zip(labels, tabledata[series_idx]):
                if value is None:
                    continue
                try:
                    period = pd.to_datetime(label).strftime("%Y-%m-%d")
                    number = float(value)
                except Exception:
                    continue
                rows.append(
                    {
                        "period": period,
                        "series_name": series_name,
                        "value": number,
                        "unit": "Index",
                        "source_name": "K-Line Shipping Market Information",
                        "source_detail": f"K-Line public IR chart republication of {series_label}; weekly Friday label",
                        "quality": "third_party_weekly_index_republication",
                    }
                )
            if rows:
                result[indicator_key] = rows
    return result


def derive_video_subscription_proxy_from_online_breakdown(
    online_rows: dict[str, list[dict]]
) -> dict[str, list[dict]]:
    """Derive a public proxy for video subscription spending from KOSIS.

    EPIC asks for video subscription spend/count. We do not have exact card
    transaction data. KOSIS DT_1KE10071 exposes monthly online-shopping
    transaction amount for "문화 및 레저서비스"; summing internet+mobile gives
    a broad official proxy for online content/leisure service spending.
    """
    acc: dict[str, float] = {}
    source_details: dict[str, list[str]] = {}
    for source_key in ("epic:2:22", "epic:2:23"):
        for row in online_rows.get(source_key, []):
            if "문화 및 레저서비스" not in str(row.get("series_name", "")):
                continue
            period = row["period"]
            acc[period] = acc.get(period, 0.0) + float(row.get("value") or 0)
            source_details.setdefault(period, []).append(row.get("source_detail", ""))

    out = {"epic:3:97": []}
    for period, value in sorted(acc.items()):
        out["epic:3:97"].append(
            {
                "period": period,
                "series_name": "online_culture_leisure_services_million_krw",
                "value": round(value, 2),
                "unit": "백만원",
                "source_name": "KOSIS_DT_1KE10071",
                "source_detail": "KOSIS 온라인쇼핑 문화 및 레저서비스 거래액(인터넷+모바일) 합산; 영상구독 지출 exact 아님",
                "quality": "official_online_content_spending_proxy",
            }
        )
    return out



def derive_music_subscription_proxy_from_online_breakdown(
    online_rows: dict[str, list[dict]]
) -> dict[str, list[dict]]:
    """Derive a conservative public proxy for music subscription spending.

    EPIC asks for music subscription spend/count. Public exact card transaction
    data is not wired. KOSIS DT_1KE10071 does not split music streaming alone,
    so we reuse the official monthly online-shopping "문화 및 레저서비스"
    amount and label it as a broad content/leisure proxy.
    """
    acc: dict[str, float] = {}
    for source_key in ("epic:2:22", "epic:2:23"):
        for row in online_rows.get(source_key, []):
            if "문화 및 레저서비스" not in str(row.get("series_name", "")):
                continue
            period = row.get("period")
            value = row.get("value")
            if not period or value is None:
                continue
            acc[period] = acc.get(period, 0.0) + float(value)

    out = {"epic:3:98": []}
    for period, value in sorted(acc.items()):
        out["epic:3:98"].append(
            {
                "period": period,
                "series_name": "online_culture_leisure_services_million_krw",
                "value": round(value, 2),
                "unit": "백만원",
                "source_name": "KOSIS_DT_1KE10071",
                "source_detail": "KOSIS 온라인쇼핑 문화 및 레저서비스 거래액(인터넷+모바일) 합산; 음원 구독 지출 exact 아님",
                "quality": "official_online_content_spending_proxy_not_music_exact",
            }
        )
    return out



def derive_child_education_proxy_from_service_index(
    service_rows: dict[str, list[dict]]
) -> dict[str, list[dict]]:
    """Reuse KOSIS education service index as a very broad child-education proxy.

    This is not early-childhood card spending. It is only a public quarterly
    education-service demand proxy until a card/merchant-category source is
    connected.
    """
    out = {"epic:13:21": []}
    for row in service_rows.get("epic:13:20", []):
        src_name = str(row.get("series_name", ""))
        # Keep the national series first; regional rows remain useful but are
        # clearly labeled in series_name/source_detail.
        out["epic:13:21"].append(
            {
                "period": row.get("period"),
                "series_name": src_name.replace("education_service_production_index", "child_education_broad_proxy_index"),
                "value": row.get("value"),
                "unit": row.get("unit") or "2020=100",
                "source_name": row.get("source_name") or "KOSIS_DT_1KC2023",
                "source_detail": (row.get("source_detail") or "") + " / 유아교육 카드결제 exact 아님",
                "quality": "official_quarterly_education_service_proxy_not_child_card_exact",
            }
        )
    out["epic:13:21"] = [r for r in out["epic:13:21"] if r.get("period") and r.get("value") is not None]
    return out
def derive_online_consumption_proxies_from_online_breakdown(
    online_rows: dict[str, list[dict]]
) -> dict[str, list[dict]]:
    """Derive conservative monthly consumption proxies from KOSIS online shopping.

    These EPIC targets are card-spending style indicators, but exact public card
    merchant-category feeds are not wired yet. We only map KOSIS product groups
    that are close enough for directional monitoring and label every row as a
    proxy so downstream users do not mistake it for card exact data.
    """
    specs = {
        "epic:11:155": {
            "series_name": "online_food_service_million_krw",
            "keywords": ("음식서비스",),
            "detail": "KOSIS 온라인쇼핑 음식서비스 거래액(인터넷+모바일) 합산; 제과/커피/패스트푸드 카드 결제 exact 아님",
            "quality": "official_online_food_service_proxy",
        },
        "epic:13:22": {
            "series_name": "online_books_stationery_million_krw",
            "keywords": ("서적", "사무·문구"),
            "detail": "KOSIS 온라인쇼핑 서적+사무·문구 거래액(인터넷+모바일) 합산; 교육용품 카드 결제 exact 아님",
            "quality": "official_online_education_goods_proxy",
        },
    }
    acc: dict[str, dict[str, float]] = {key: {} for key in specs}
    for source_key in ("epic:2:22", "epic:2:23"):
        for row in online_rows.get(source_key, []):
            name = str(row.get("series_name", ""))
            period = row.get("period")
            value = row.get("value")
            if not period or value is None:
                continue
            for indicator_key, spec in specs.items():
                if any(keyword in name for keyword in spec["keywords"]):
                    acc[indicator_key][period] = acc[indicator_key].get(period, 0.0) + float(value)

    out: dict[str, list[dict]] = {key: [] for key in specs}
    for indicator_key, period_values in acc.items():
        spec = specs[indicator_key]
        for period, value in sorted(period_values.items()):
            out[indicator_key].append(
                {
                    "period": period,
                    "series_name": spec["series_name"],
                    "value": round(value, 2),
                    "unit": "백만원",
                    "source_name": "KOSIS_DT_1KE10071",
                    "source_detail": spec["detail"],
                    "quality": spec["quality"],
                }
            )
    return out


def collect_worldbank_iron_ore_monthly() -> list[dict]:
    """Collect World Bank Pink Sheet monthly iron ore CFR spot price.

    This is a high-quality public proxy for China iron ore import conditions,
    but the EPIC field requests a weekly China import price, so catalog status
    should remain partial/proxy rather than exact.
    """
    url = (
        "https://thedocs.worldbank.org/en/doc/"
        "74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/related/"
        "CMO-Historical-Data-Monthly.xlsx"
    )
    try:
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=40)
        r.raise_for_status()
        df = pd.read_excel(BytesIO(r.content), sheet_name="Monthly Prices", header=None)
    except Exception:
        return []

    header_row = None
    iron_col = None
    for idx in range(min(len(df), 20)):
        for col in range(df.shape[1]):
            value = str(df.iat[idx, col]).strip()
            if value == "Iron ore, cfr spot":
                header_row = idx
                iron_col = col
                break
        if iron_col is not None:
            break
    if header_row is None or iron_col is None:
        return []

    out: list[dict] = []
    for idx in range(header_row + 2, len(df)):
        raw_period = str(df.iat[idx, 0]).strip()
        if not re.fullmatch(r"\d{4}M\d{2}", raw_period):
            continue
        value = safe_float(df.iat[idx, iron_col])
        if value is None:
            continue
        period = f"{raw_period[:4]}-{raw_period[-2:]}"
        out.append(
            {
                "period": period,
                "series_name": "iron_ore_cfr_spot_usd_dmtu",
                "value": value,
                "unit": "$/dmtu",
                "source_name": "WORLD_BANK_PINK_SHEET",
                "source_detail": "World Bank Commodity Price Data monthly Iron ore, cfr spot",
                "quality": "official_proxy_monthly",
            }
        )
    return out


def collect_steelbenchmarker_china_latest() -> dict[str, list[dict]]:
    """Collect latest China steel benchmark prices from SteelBenchmarker PDF.

    The public PDF reliably exposes the latest Mainland China HRB/CRC/Rebar/Plate
    prices, but historical table parsing needs a separate validator. Keep these
    indicators partial until we can backfill a stable time series.
    """
    url = "https://www.steelbenchmarker.com/history.pdf"
    try:
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
        r.raise_for_status()
        with tempfile.NamedTemporaryFile(suffix=".pdf") as f:
            f.write(r.content)
            f.flush()
            text = subprocess.check_output(["pdftotext", f.name, "-"], text=True, timeout=20)
    except Exception:
        return {}

    date_match = re.search(r"([A-Z][a-z]+ \d{1,2}, \d{4})\ndollars per metric tonne", text)
    try:
        period = pd.to_datetime(date_match.group(1)).strftime("%Y-%m-%d") if date_match else date.today().isoformat()
    except Exception:
        period = date.today().isoformat()

    block = re.search(
        r"Region: Mainland China\*\*\*\s*"
        r"Hot-rolled band:\s*Cold-rolled coil:\s*Rebar:\s*Standard plate:\s*"
        r"\n\s*([\d,]+)\s*\n\s*([\d,]+)\s*\n\s*([\d,]+)\s*\n\s*([\d,]+)",
        text,
    )
    if not block:
        return {}

    hrb, crc, rebar, plate = [safe_float(v) for v in block.groups()]
    specs = {
        "epic:1:25": ("china_hrb_usd_mt", hrb, "Hot-rolled band"),
        "epic:1:26": ("china_crc_usd_mt", crc, "Cold-rolled coil"),
        "epic:1:27": ("china_plate_usd_mt", plate, "Standard plate"),
        "epic:1:29": ("china_rebar_usd_mt", rebar, "Rebar"),
    }
    out: dict[str, list[dict]] = {}
    for key, (series_name, value, label) in specs.items():
        if value is None:
            continue
        out[key] = [
            {
                "period": period,
                "series_name": series_name,
                "value": value,
                "unit": "$/mt",
                "source_name": "STEELBENCHMARKER_PDF",
                "source_detail": f"SteelBenchmarker latest Mainland China {label} price",
                "quality": "public_report_latest_only",
            }
        ]
    return out


def _fetch_sunsirs_html(url: str) -> str:
    """Fetch SunSirs pages that require a lightweight HW_CHECK cookie challenge."""
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 AppleWebKit/537.36 Chrome/125 Safari/537.36"})
    try:
        response = session.get(url, timeout=20)
        response.raise_for_status()
    except Exception:
        return ""

    cookie_match = re.search(r'var _0x2 = "([0-9a-f]+)"', response.text)
    if cookie_match:
        session.cookies.set("HW_CHECK", cookie_match.group(1), domain="www.sunsirs.com", path="/")
        try:
            response = session.get(url, timeout=20)
            response.raise_for_status()
        except Exception:
            return ""
    return response.text


def collect_sunsirs_china_steel_daily() -> dict[str, list[dict]]:
    """Collect recent public China G.I and Wire Rod spot prices from SunSirs.

    SunSirs exposes only the recent public table without login. Keep these
    indicators partial until a licensed or official long-history feed is found.
    """
    specs = {
        "epic:1:28": {
            "commodity_id": 301,
            "commodity": "Galvanized sheet",
            "series_name": "china_galvanized_sheet_spot_price",
            "detail": "Variety: HDG; Grade: DX51D+Z; Size: 1.0*1250*C",
        },
        "epic:1:30": {
            "commodity_id": 740,
            "commodity": "Wire Rod",
            "series_name": "china_wire_rod_spot_price",
            "detail": "Grade: HPB235; Diameter: Φ8",
        },
    }
    out: dict[str, list[dict]] = {key: [] for key in specs}
    for indicator_key, spec in specs.items():
        url = f"https://www.sunsirs.com/uk/prodetail-{spec['commodity_id']}.html"
        html = _fetch_sunsirs_html(url)
        if not html:
            continue
        text = BeautifulSoup(html, "html.parser").get_text("\n", strip=True)
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        for idx in range(len(lines) - 3):
            if lines[idx] != spec["commodity"]:
                continue
            if lines[idx + 1] != "Steel":
                continue
            value = safe_float(lines[idx + 2])
            if value is None:
                continue
            raw_date = lines[idx + 3]
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw_date):
                continue
            out[indicator_key].append(
                {
                    "period": raw_date,
                    "series_name": spec["series_name"],
                    "value": value,
                    "unit": "RMB/ton",
                    "source_name": "SunSirs",
                    "source_detail": f"SunSirs China {spec['commodity']} daily public table; {spec['detail']}; {url}",
                    "quality": "third_party_recent_public_price_partial",
                }
            )
        # The first rows are latest-first; de-duplicate while preserving dates.
        seen: set[str] = set()
        deduped: list[dict] = []
        for row in out[indicator_key]:
            if row["period"] in seen:
                continue
            seen.add(row["period"])
            deduped.append(row)
        out[indicator_key] = sorted(deduped, key=lambda r: r["period"])
    return out


def _fetch_dart_document_text(rcept_no: str) -> str:
    """Fetch a DART disclosure document as flattened text using existing key rotation."""
    try:
        from collectors.dart_backlog_collector import _fetch_document_with_key_rotation

        return _fetch_document_with_key_rotation(rcept_no) or ""
    except Exception:
        try:
            from collectors.dart_contract_collector import _fetch_dart_document

            return _fetch_dart_document(rcept_no) or ""
        except Exception:
            return ""


def _first_number_after(text: str, anchor_pattern: str, *, require_thousands: bool = False) -> float | None:
    m = re.search(anchor_pattern, text, flags=re.I | re.S)
    if not m:
        return None
    tail = text[m.end(): m.end() + 900]
    pattern = r"(?<![\d.-])(-?\d{1,3}(?:,\d{3})+)(?![\d.-])" if require_thousands else r"(?<![\d.-])(-?\d{1,3}(?:,\d{3})+|-?\d+(?:\.\d+)?)(?![\d.-])"
    n = re.search(pattern, tail)
    return safe_float(n.group(1)) if n else None


def collect_dart_casino_monthly(conn: sqlite3.Connection) -> dict[str, list[dict]]:
    """Collect casino monthly sales/drop/hold from DART fair-disclosure texts.

    Visitor counts are intentionally not filled here because the sampled DART
    disclosure body does not provide them. We only persist fields that are
    directly visible in the disclosure text: casino sales and table drop.
    """
    companies = {
        "034230": {
            "name": "파라다이스",
            "sales_key": "epic:9:18",
            "drop_key": "epic:9:23",
            "hold_key": "epic:9:25",
        },
        "114090": {
            "name": "GKL",
            "sales_key": "epic:9:20",
            "drop_key": "epic:9:21",
            "hold_key": "epic:9:22",
        },
        "032350": {
            "name": "롯데관광개발",
            "sales_key": "epic:9:35",
            "drop_key": "epic:9:36",
            "hold_key": "epic:9:38",
        },
    }
    out: dict[str, list[dict]] = {cfg[k]: [] for cfg in companies.values() for k in ("sales_key", "drop_key", "hold_key")}
    seen: set[tuple[str, str]] = set()
    rows = conn.execute(
        """
        SELECT stock_code, corp_name, rcept_no, rcept_dt, report_nm
          FROM dart_disclosures
         WHERE stock_code IN ('034230','114090','032350')
           AND report_nm LIKE '%영업%실적%공정공시%'
         ORDER BY rcept_dt ASC, rcept_no ASC
        """
    ).fetchall()
    for row in rows:
        code = str(row["stock_code"])
        cfg = companies.get(code)
        if not cfg:
            continue
        text = _fetch_dart_document_text(str(row["rcept_no"]))
        if not text:
            continue
        text = re.sub(r"\s+", " ", text)
        period_m = re.search(r"당기실적\s+(\d{4})-(\d{2})-\d{2}\s*~\s*(\d{4})-(\d{2})-\d{2}", text)
        if not period_m:
            continue
        period = f"{period_m.group(1)}-{period_m.group(2)}"
        if (code, period) in seen:
            # Keep the first regular monthly disclosure and avoid duplicate same-month revisions.
            continue

        sales = _first_number_after(text, r"카지노\s*매출액", require_thousands=True)
        drop = _first_number_after(text, r"(테이블\s*)?드[랍롭]액", require_thousands=True)
        if sales is None and drop is None:
            continue

        source_detail = f"DART {cfg['name']} 영업잠정실적 {row['rcept_dt']} {row['rcept_no']}"
        if sales is not None:
            out[cfg["sales_key"]].append(
                {
                    "period": period,
                    "series_name": f"{cfg['name']}_casino_sales",
                    "value": sales,
                    "unit": "백만원",
                    "source_name": "DART_casino_monthly_disclosure",
                    "source_detail": source_detail,
                    "quality": "official_disclosure",
                }
            )
        if drop is not None:
            out[cfg["drop_key"]].append(
                {
                    "period": period,
                    "series_name": f"{cfg['name']}_table_drop",
                    "value": drop,
                    "unit": "백만원",
                    "source_name": "DART_casino_monthly_disclosure",
                    "source_detail": source_detail,
                    "quality": "official_disclosure",
                }
            )
        if sales is not None and drop and drop > 0:
            out[cfg["hold_key"]].append(
                {
                    "period": period,
                    "series_name": f"{cfg['name']}_hold_rate",
                    "value": round(sales / drop * 100.0, 4),
                    "unit": "%",
                    "source_name": "DART_casino_monthly_disclosure",
                    "source_detail": source_detail,
                    "quality": "derived_from_official_disclosure",
                }
            )
        seen.add((code, period))
    return out


def collect_seoul_subway_monthly() -> dict[str, list[dict]]:
    """Collect Seoul subway monthly passengers by line from Seoul Open Data CSV files."""
    page_url = "https://data.seoul.go.kr/dataList/OA-12914/S/1/datasetView.do"
    download_url = "https://datafile.seoul.go.kr/bigfile/iot/inf/nio_download.do?&useCache=false"
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0", "Referer": page_url})
    try:
        html = session.get(page_url, timeout=30).text
    except Exception:
        return {}

    files: list[tuple[int, str]] = []
    for seq, filename in re.findall(r"downloadFile\('(\d+)'\).*?>(CARD_SUBWAY_MONTH_\d{4,6}\.csv)<", html):
        try:
            files.append((int(seq), filename))
        except Exception:
            continue
    if not files:
        return {}

    by_line_frames: list[pd.DataFrame] = []
    total_frames: list[pd.DataFrame] = []
    form_base = {"infId": "OA-12914", "seqNo": "", "infSeq": "3"}
    for seq, filename in sorted(set(files), key=lambda x: x[1]):
        try:
            r = session.post(download_url, data={**form_base, "seq": str(seq)}, timeout=60)
            r.raise_for_status()
            df = pd.read_csv(BytesIO(r.content), encoding="utf-8-sig", usecols=range(6), dtype={"사용일자": str})
        except Exception:
            continue
        if not {"사용일자", "노선명", "승차총승객수", "하차총승객수"}.issubset(set(df.columns)):
            continue
        df["period"] = df["사용일자"].astype(str).str.slice(0, 6).str.replace(r"^(\d{4})(\d{2})$", r"\1-\2", regex=True)
        df["승차총승객수"] = pd.to_numeric(df["승차총승객수"], errors="coerce").fillna(0)
        df["하차총승객수"] = pd.to_numeric(df["하차총승객수"], errors="coerce").fillna(0)
        grouped = df.groupby(["period", "노선명"], as_index=False)[["승차총승객수", "하차총승객수"]].sum()
        by_line_frames.append(grouped)
        total = df.groupby(["period"], as_index=False)[["승차총승객수", "하차총승객수"]].sum()
        total_frames.append(total)

    if not by_line_frames:
        return {}
    by_line = pd.concat(by_line_frames, ignore_index=True).groupby(["period", "노선명"], as_index=False)[["승차총승객수", "하차총승객수"]].sum()
    total_df = pd.concat(total_frames, ignore_index=True).groupby(["period"], as_index=False)[["승차총승객수", "하차총승객수"]].sum()
    out: dict[str, list[dict]] = {"epic:22:10": [], "epic:22:9": []}
    for _, row in by_line.iterrows():
        line = str(row["노선명"]).strip()
        period = str(row["period"])
        board = float(row["승차총승객수"])
        alight = float(row["하차총승객수"])
        for metric, value, label in [
            ("boarding", board, "승차"),
            ("alighting", alight, "하차"),
            ("total", board + alight, "승하차"),
        ]:
            out["epic:22:10"].append(
                {
                    "period": period,
                    "series_name": f"{line}_{metric}",
                    "value": value,
                    "unit": "명",
                    "source_name": "SEOUL_OPEN_DATA_OA_12914",
                    "source_detail": f"서울 열린데이터광장 지하철호선별 역별 승하차 인원 정보 {line} {label}",
                    "quality": "official_city_open_data",
                }
            )
    for _, row in total_df.iterrows():
        period = str(row["period"])
        board = float(row["승차총승객수"])
        alight = float(row["하차총승객수"])
        for metric, value, label in [
            ("subway_boarding", board, "지하철 승차"),
            ("subway_alighting", alight, "지하철 하차"),
            ("subway_total", board + alight, "지하철 승하차"),
        ]:
            out["epic:22:9"].append(
                {
                    "period": period,
                    "series_name": metric,
                    "value": value,
                    "unit": "명",
                    "source_name": "SEOUL_OPEN_DATA_OA_12914",
                    "source_detail": f"서울 열린데이터광장 지하철호선별 역별 승하차 인원 정보 - {label}",
                    "quality": "official_partial_subway_only",
                }
            )
    return out


def collect_kric_rail_line_passenger_monthly(start_year: int = 2022) -> list[dict]:
    """Collect monthly passenger counts by KORAIL route from KRIC.

    The EPIC target is "rail passengers by route". KRIC's page is an official
    Korea Railroad Corporation general-rail statistic, but it does not cover
    every urban/private rail service. Keep catalog status partial and label the
    source precisely so this is not mistaken for all-rail national coverage.
    """
    url = "https://www.kric.go.kr/jsp/industry/rss/raillinepassmonList.jsp"
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0", "Referer": url})
    out: list[dict] = []
    current = date.today()
    train_cols = [
        "KTX",
        "새마을",
        "무궁화",
        "통근열차",
        "누리로",
        "KTX-산천",
        "KTX-호남",
        "KTX-이음",
        "KTX-청룡",
        "ITX-청춘열차",
        "ITX-새마을",
        "ITX-마음",
    ]

    for year in range(start_year, current.year + 1):
        max_month = 12 if year < current.year else current.month
        for month in range(1, max_month + 1):
            period = f"{year}-{month:02d}"
            for page_no in (1, 2, 3):
                try:
                    r = session.post(
                        url,
                        data={
                            "q_fdate": str(year),
                            "q_month": str(month),
                            "q_excel_flag": "Y",
                            "pageNo": str(page_no),
                        },
                        timeout=30,
                    )
                    r.raise_for_status()
                except Exception:
                    break

                soup = BeautifulSoup(r.text, "html.parser")
                text = soup.get_text(" ", strip=True)
                if "검색된 자료가 없습니다" in text:
                    break
                rows = soup.select("table.listtbl_c100 tbody tr")
                if not rows:
                    break

                parsed_any = False
                for tr in rows:
                    cells = [c.get_text(" ", strip=True) for c in tr.find_all(["td", "th"])]
                    if len(cells) < 2 or "검색된 자료" in " ".join(cells):
                        continue
                    line_name = cells[0].strip()
                    if not line_name:
                        continue
                    total = safe_float(cells[1])
                    if total is None:
                        continue
                    parsed_any = True
                    out.append(
                        {
                            "period": period,
                            "series_name": f"{line_name}_total",
                            "value": total,
                            "unit": "명",
                            "source_name": "KRIC_KORAIL_RAIL_LINE_PASSENGER_MONTHLY",
                            "source_detail": "철도산업정보센터 노선별 여객수송(월) - 한국철도공사 일반철도",
                            "quality": "official_korail_general_rail_partial",
                        }
                    )
                    for idx, label in enumerate(train_cols, start=2):
                        if idx >= len(cells):
                            continue
                        value = safe_float(cells[idx])
                        if value is None:
                            continue
                        out.append(
                            {
                                "period": period,
                                "series_name": f"{line_name}_{label}",
                                "value": value,
                                "unit": "명",
                                "source_name": "KRIC_KORAIL_RAIL_LINE_PASSENGER_MONTHLY",
                                "source_detail": f"철도산업정보센터 노선별 여객수송(월) - {label}",
                                "quality": "official_korail_general_rail_partial",
                            }
                        )

                if not parsed_any or page_no == 3:
                    break
    return out


def collect_mtrace_pork_auction_price_monthly(start_ym: str = "202001") -> list[dict]:
    """Collect monthly Korean pork carcass auction price from MTRACE.

    EPIC names the field as a daily wholesale price, but the public official
    table exposed by data.go.kr/MTRACE is monthly. Store the all-grade/all-market
    auction price as a partial official monthly proxy and keep the label clear.
    """
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return []

    end_ym = date.today().strftime("%Y%m")
    periods = month_range_desc(start_ym, end_ym)
    if not periods:
        return []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        try:
            page = browser.new_page(viewport={"width": 1400, "height": 900})
            page.goto(
                "https://mtrace.go.kr/statHtml/statHtml.do?orgId=323&tblId=DT_APGS_016&conn_path=I2",
                wait_until="domcontentloaded",
                timeout=30000,
            )
            page.wait_for_timeout(8000)
            response_text = page.evaluate(
                """async (periods) => {
                  const form = document.querySelector('#ParamInfo');
                  if (!form) return JSON.stringify({errCode: 1, errMsg: 'ParamInfo missing'});
                  const fd = new FormData(form);
                  let fieldList = JSON.parse(fd.get('fieldList'));
                  fieldList = fieldList.filter((x) => x.targetId !== 'PRD');
                  fieldList.unshift({targetId:'PRD', targetValue:'', prdValue:'M,' + periods.join(',') + ',@'});
                  fd.set('fieldList', JSON.stringify(fieldList));
                  fd.set('jsonStr', '');
                  fd.set('isFirst', 'Y');
                  fd.set('colAxis', 'TIME');
                  fd.set('rowAxis', 'ITEM,A,B');
                  fd.set('viewKind', '1');
                  fd.set('view', 'table');
                  fd.set('diviSearchYn', 'N');
                  fd.set('orderStr', 'CHAR_ITM_ID,OV_L1_ID,OV_L2_ID,TIME');
                  fd.set('startNum', '1');
                  fd.set('endNum', String(periods.length * 2 * 24 * 20));
                  fd.set('lastChk', 'N');
                  const body = new URLSearchParams();
                  for (const [k, v] of fd.entries()) body.append(k, v);
                  const resp = await fetch('/statHtml/html.do', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'},
                    body: body.toString()
                  });
                  return await resp.text();
                }""",
                periods,
            )
        finally:
            browser.close()

    try:
        payload = json.loads(response_text)
    except Exception:
        return []
    html = (payload.get("result") or [""])[0]
    if not html:
        return []
    soup = BeautifulSoup(html, "html.parser")
    parsed_periods: list[str] = []
    for th in soup.select("thead th"):
        label = th.get("title") or th.get_text(" ", strip=True)
        match = re.search(r"(\d{4})\.\s*(\d{1,2})", label)
        if match:
            parsed_periods.append(f"{match.group(1)}-{int(match.group(2)):02d}")
    if not parsed_periods:
        return []

    out: list[dict] = []
    current_item = ""
    current_grade = ""
    for tr in soup.select("tbody tr"):
        cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
        if len(cells) < 3 + len(parsed_periods):
            continue
        if cells[0]:
            current_item = cells[0]
        if cells[1]:
            current_grade = cells[1]
        market = cells[2]
        if "경락가격" not in current_item or current_grade != "전체" or market != "전체":
            continue
        values = cells[-len(parsed_periods):]
        for period, raw_value in zip(parsed_periods, values):
            value = safe_float(raw_value)
            if value is None:
                continue
            # The official table currently returns an apparent bad point
            # (e.g. 2025-11 = 547 won/kg). Keep the series clean instead of
            # persisting an obviously unusable wholesale price.
            if value < 2000 or value > 15000:
                continue
            out.append(
                {
                    "period": period,
                    "series_name": "pork_carcass_auction_price_all_grade_all_market",
                    "value": value,
                    "unit": "원/kg",
                    "source_name": "MTRACE_DT_APGS_016",
                    "source_detail": "축산물품질평가원/MTRACE 돼지도체 도매시장별 등급별 경락가격 - 전체 등급/전체 도매시장",
                    "quality": "official_monthly_proxy",
                }
            )
        break
    return out


LOCAL_MARKET_STRUCTURE_INDICATORS = [
    {
        "indicator_key": "public:21:5",
        "epic_category_code": 21,
        "epic_sub_code": 5,
        "epic_indicator_name": "시장 프로그램 매매 수급",
        "frequency": "Daily",
        "base_unit": "억원",
        "replacement_family": "program_trading_flow",
        "source_system": "KIS/Kiwoom/KRX program trading tables",
        "exactness": "broker_or_exchange_daily_aggregate",
        "priority": "p1",
        "notes": "broker_program_market_daily에서 KOSPI/KOSDAQ별 프로그램·차익·비차익 순매수 금액을 일별 집계. 시장 수급/패시브 흐름 판단용.",
    },
    {
        "indicator_key": "public:21:6",
        "epic_category_code": 21,
        "epic_sub_code": 6,
        "epic_indicator_name": "종목 프로그램 매매 집중도",
        "frequency": "Daily",
        "base_unit": "억원/%/종목",
        "replacement_family": "program_trading_concentration",
        "source_system": "broker_program_stock_daily",
        "exactness": "broker_or_exchange_stock_level_derived",
        "priority": "p1",
        "notes": "broker_program_stock_daily에서 종목별 프로그램 순매수 총액, 순매수/순매도 종목 수, 상위 10개 집중도를 계산.",
    },
    {
        "indicator_key": "public:20:106",
        "epic_category_code": 20,
        "epic_sub_code": 106,
        "epic_indicator_name": "신용잔고·외국인 보유 월간 총량",
        "frequency": "Monthly",
        "base_unit": "억원/%/종목",
        "replacement_family": "credit_foreign_positioning",
        "source_system": "kiwoom_credit_balance + margin_balance_daily + foreign_holding_daily",
        "exactness": "monthly_snapshot_derived_from_collected_daily",
        "priority": "p1",
        "notes": "월말 기준 신용잔고 금액/수량, 키움 신용/대주 잔고, 외국인 보유율 평균·커버 종목 수를 집계. 과열/수급 리스크 판단용.",
    },
    {
        "indicator_key": "public:20:107",
        "epic_category_code": 20,
        "epic_sub_code": 107,
        "epic_indicator_name": "투자자별 순매수 총량",
        "frequency": "Daily",
        "base_unit": "주",
        "replacement_family": "investor_flow_total",
        "source_system": "investor_trading_daily / kiwoom_investor_daily",
        "exactness": "daily_aggregate_from_collected_investor_flow",
        "priority": "p1",
        "notes": "일별 개인·기관·외국인 순매수 수량을 시장 전체로 합산. 금액 기준이 아닌 수량 기준이므로 해석 시 대형주 편향을 별도 점검.",
    },
    {
        "indicator_key": "public:20:108",
        "epic_category_code": 20,
        "epic_sub_code": 108,
        "epic_indicator_name": "공매도·대차잔고 시장 총량",
        "frequency": "Daily/Monthly",
        "base_unit": "억원/주/종목",
        "replacement_family": "short_lending_market_total",
        "source_system": "short_sell_daily + short_rank_daily",
        "exactness": "official_or_exchange_derived",
        "priority": "p1",
        "notes": "공매도 금액/수량과 대차잔고 금액/수량을 시장 전체로 집계. short_rank_daily 월말 스냅샷도 함께 저장.",
    },
]


def _dt_to_period(value: str, monthly: bool = False) -> str:
    text = str(value or "").strip()
    digits = re.sub(r"\D", "", text)
    if len(digits) >= 8:
        return f"{digits[:4]}-{digits[4:6]}" if monthly else f"{digits[:4]}-{digits[4:6]}-{digits[6:8]}"
    if re.fullmatch(r"\d{4}-\d{2}(-\d{2})?", text):
        return text[:7] if monthly else text[:10]
    return text[:7] if monthly else text


def collect_local_market_structure_indicators(conn: sqlite3.Connection) -> dict[str, list[dict]]:
    """Expose already-collected market structure tables on the quant page.

    These are not new upstream downloads. They turn high-value local tables
    collected by KIS/Kiwoom/KRX/public-data jobs into auditable quant series.
    """
    result: dict[str, list[dict]] = {spec["indicator_key"]: [] for spec in LOCAL_MARKET_STRUCTURE_INDICATORS}

    def has_table(table_name: str) -> bool:
        row = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
            (table_name,),
        ).fetchone()
        return bool(row)

    if has_table("broker_program_market_daily"):
        rows = conn.execute(
            """
            SELECT source, dt, market,
                   SUM(COALESCE(prog_net_buy_amt, 0)) AS prog,
                   SUM(COALESCE(arb_net_buy_amt, 0)) AS arb,
                   SUM(COALESCE(non_arb_net_buy_amt, 0)) AS non_arb
              FROM broker_program_market_daily
             WHERE dt IS NOT NULL
             GROUP BY source, dt, market
             ORDER BY dt, source, market
            """
        ).fetchall()
        for row in rows:
            period = _dt_to_period(row["dt"])
            market = str(row["market"] or "ALL")
            source = str(row["source"] or "unknown")
            common = {
                "period": period,
                "unit": "억원",
                "source_name": "broker_program_market_daily",
                "source_detail": f"{source} {market} 프로그램 매매 일별 집계",
                "quality": "broker_or_exchange_daily_aggregate",
            }
            result["public:21:5"].extend([
                {**common, "series_name": f"{source}_{market}_program_net_buy_100m", "value": round(float(row["prog"] or 0), 4)},
                {**common, "series_name": f"{source}_{market}_arbitrage_net_buy_100m", "value": round(float(row["arb"] or 0), 4)},
                {**common, "series_name": f"{source}_{market}_non_arbitrage_net_buy_100m", "value": round(float(row["non_arb"] or 0), 4)},
            ])

    if has_table("broker_program_stock_daily"):
        rows = conn.execute(
            """
            WITH base AS (
                SELECT source, dt, COALESCE(market_channel, 'ALL') AS market_channel,
                       stock_code,
                       COALESCE(net_buy_amt_krw, 0) AS net_buy_amt_krw
                  FROM broker_program_stock_daily
                 WHERE dt IS NOT NULL
                   AND stock_code IS NOT NULL
            ),
            agg AS (
                SELECT source, dt, market_channel,
                       COUNT(*) AS cover_count,
                       SUM(CASE WHEN net_buy_amt_krw > 0 THEN 1 ELSE 0 END) AS net_buy_stock_count,
                       SUM(CASE WHEN net_buy_amt_krw < 0 THEN 1 ELSE 0 END) AS net_sell_stock_count,
                       SUM(net_buy_amt_krw) / 100000000.0 AS net_buy_100m,
                       SUM(ABS(net_buy_amt_krw)) / 100000000.0 AS abs_net_buy_100m
                  FROM base
                 GROUP BY source, dt, market_channel
            ),
            ranked AS (
                SELECT source, dt, market_channel, ABS(net_buy_amt_krw) AS abs_amt,
                       ROW_NUMBER() OVER (
                           PARTITION BY source, dt, market_channel
                           ORDER BY ABS(net_buy_amt_krw) DESC
                       ) AS rn
                  FROM base
            ),
            top10 AS (
                SELECT source, dt, market_channel,
                       SUM(abs_amt) / 100000000.0 AS top10_abs_100m
                  FROM ranked
                 WHERE rn <= 10
                 GROUP BY source, dt, market_channel
            )
            SELECT a.*, COALESCE(t.top10_abs_100m, 0) AS top10_abs_100m
              FROM agg a
              LEFT JOIN top10 t
                ON t.source = a.source
               AND t.dt = a.dt
               AND t.market_channel = a.market_channel
             ORDER BY a.dt, a.source, a.market_channel
            """
        ).fetchall()
        for row in rows:
            period = _dt_to_period(row["dt"])
            source = str(row["source"] or "unknown")
            channel = str(row["market_channel"] or "ALL")
            abs_total = float(row["abs_net_buy_100m"] or 0)
            top10 = float(row["top10_abs_100m"] or 0)
            concentration = (top10 / abs_total * 100.0) if abs_total else None
            common = {
                "period": period,
                "source_name": "broker_program_stock_daily",
                "source_detail": f"{source} {channel} 종목별 프로그램 매매 일별 파생지표",
                "quality": "broker_or_exchange_stock_level_derived",
            }
            result["public:21:6"].extend([
                {**common, "series_name": f"{source}_{channel}_program_stock_net_buy_100m", "value": round(float(row["net_buy_100m"] or 0), 4), "unit": "억원"},
                {**common, "series_name": f"{source}_{channel}_program_stock_abs_net_buy_100m", "value": round(abs_total, 4), "unit": "억원"},
                {**common, "series_name": f"{source}_{channel}_program_net_buy_stock_count", "value": int(row["net_buy_stock_count"] or 0), "unit": "종목"},
                {**common, "series_name": f"{source}_{channel}_program_net_sell_stock_count", "value": int(row["net_sell_stock_count"] or 0), "unit": "종목"},
                {**common, "series_name": f"{source}_{channel}_program_cover_count", "value": int(row["cover_count"] or 0), "unit": "종목"},
            ])
            if concentration is not None:
                result["public:21:6"].append({
                    **common,
                    "series_name": f"{source}_{channel}_program_top10_abs_concentration_pct",
                    "value": round(concentration, 4),
                    "unit": "%",
                })

    if has_table("kiwoom_credit_balance"):
        rows = conn.execute(
            """
            WITH month_last AS (
                SELECT substr(dt,1,6) AS ym, stock_code, MAX(dt) AS dt
                  FROM kiwoom_credit_balance
                 WHERE dt IS NOT NULL
                 GROUP BY substr(dt,1,6), stock_code
            ),
            snap AS (
                SELECT ml.ym, k.stock_code,
                       COALESCE(k.credit_balance_qty, 0) AS qty,
                       COALESCE(k.credit_balance_amt, 0) AS amt,
                       k.credit_ratio
                  FROM month_last ml
                  JOIN kiwoom_credit_balance k
                    ON k.stock_code = ml.stock_code
                   AND k.dt = ml.dt
            )
            SELECT ym,
                   COUNT(*) AS cover_count,
                   SUM(qty) AS credit_qty_sum,
                   SUM(amt) AS credit_amt_sum,
                   AVG(CASE WHEN credit_ratio IS NOT NULL THEN credit_ratio END) AS avg_credit_ratio
              FROM snap
             GROUP BY ym
             ORDER BY ym
            """
        ).fetchall()
        for row in rows:
            common = {
                "period": _dt_to_period(row["ym"] + "01", monthly=True),
                "source_name": "kiwoom_credit_balance",
                "source_detail": "월말 종목별 신용잔고 스냅샷 합산",
                "quality": "monthly_snapshot_derived_from_kiwoom_daily",
            }
            result["public:20:106"].extend([
                {**common, "series_name": "kiwoom_credit_balance_qty_sum", "value": float(row["credit_qty_sum"] or 0), "unit": "주"},
                {**common, "series_name": "kiwoom_credit_balance_amt_sum_raw", "value": float(row["credit_amt_sum"] or 0), "unit": "원천단위"},
                {**common, "series_name": "kiwoom_credit_balance_cover_count", "value": int(row["cover_count"] or 0), "unit": "종목"},
            ])
            if row["avg_credit_ratio"] is not None:
                result["public:20:106"].append({
                    **common,
                    "series_name": "kiwoom_avg_credit_ratio",
                    "value": round(float(row["avg_credit_ratio"]), 4),
                    "unit": "%",
                })

    if has_table("margin_balance_daily"):
        rows = conn.execute(
            """
            WITH month_last AS (
                SELECT substr(dt,1,6) AS ym, stock_code, MAX(dt) AS dt
                  FROM margin_balance_daily
                 WHERE dt IS NOT NULL
                 GROUP BY substr(dt,1,6), stock_code
            ),
            snap AS (
                SELECT ml.ym, m.stock_code,
                       COALESCE(m.credit_balance, 0) AS credit_balance,
                       COALESCE(m.credit_amount, 0) AS credit_amount,
                       m.credit_ratio,
                       COALESCE(m.short_balance, 0) AS short_balance
                  FROM month_last ml
                  JOIN margin_balance_daily m
                    ON m.stock_code = ml.stock_code
                   AND m.dt = ml.dt
            )
            SELECT ym,
                   COUNT(*) AS cover_count,
                   SUM(credit_balance) AS credit_balance_sum,
                   SUM(credit_amount) AS credit_amount_sum,
                   AVG(CASE WHEN credit_ratio IS NOT NULL THEN credit_ratio END) AS avg_credit_ratio,
                   SUM(short_balance) AS short_balance_sum
              FROM snap
             GROUP BY ym
             ORDER BY ym
            """
        ).fetchall()
        for row in rows:
            common = {
                "period": _dt_to_period(row["ym"] + "01", monthly=True),
                "source_name": "margin_balance_daily",
                "source_detail": "키움 ka10013 월말 종목별 신용/대주 잔고 스냅샷 합산",
                "quality": "monthly_snapshot_derived_from_kiwoom_margin_daily",
            }
            result["public:20:106"].extend([
                {**common, "series_name": "kiwoom_margin_credit_balance_qty_sum", "value": float(row["credit_balance_sum"] or 0), "unit": "주"},
                {**common, "series_name": "kiwoom_margin_credit_amount_sum_raw", "value": float(row["credit_amount_sum"] or 0), "unit": "원천단위"},
                {**common, "series_name": "kiwoom_margin_short_balance_qty_sum", "value": float(row["short_balance_sum"] or 0), "unit": "주"},
                {**common, "series_name": "kiwoom_margin_cover_count", "value": int(row["cover_count"] or 0), "unit": "종목"},
            ])
            if row["avg_credit_ratio"] is not None:
                result["public:20:106"].append({
                    **common,
                    "series_name": "kiwoom_margin_avg_credit_ratio",
                    "value": round(float(row["avg_credit_ratio"]), 4),
                    "unit": "%",
                })

    if has_table("foreign_holding_daily"):
        rows = conn.execute(
            """
            WITH month_last AS (
                SELECT substr(bas_dt,1,6) AS ym, stock_code, MAX(bas_dt) AS bas_dt
                  FROM foreign_holding_daily
                 WHERE bas_dt IS NOT NULL
                 GROUP BY substr(bas_dt,1,6), stock_code
            ),
            snap AS (
                SELECT ml.ym, f.stock_code,
                       COALESCE(f.frgn_hold_qty, 0) AS hold_qty,
                       f.frgn_hold_pct
                  FROM month_last ml
                  JOIN foreign_holding_daily f
                    ON f.stock_code = ml.stock_code
                   AND f.bas_dt = ml.bas_dt
            )
            SELECT ym,
                   COUNT(*) AS cover_count,
                   SUM(hold_qty) AS foreign_hold_qty_sum,
                   AVG(CASE WHEN frgn_hold_pct IS NOT NULL THEN frgn_hold_pct END) AS avg_foreign_hold_pct
              FROM snap
             GROUP BY ym
             ORDER BY ym
            """
        ).fetchall()
        for row in rows:
            common = {
                "period": _dt_to_period(row["ym"] + "01", monthly=True),
                "source_name": "foreign_holding_daily",
                "source_detail": "월말 종목별 외국인 보유 현황 스냅샷 합산",
                "quality": "monthly_snapshot_derived_from_public_daily",
            }
            result["public:20:106"].extend([
                {**common, "series_name": "foreign_hold_qty_sum", "value": float(row["foreign_hold_qty_sum"] or 0), "unit": "주"},
                {**common, "series_name": "foreign_holding_cover_count", "value": int(row["cover_count"] or 0), "unit": "종목"},
            ])
            if row["avg_foreign_hold_pct"] is not None:
                result["public:20:106"].append({
                    **common,
                    "series_name": "avg_foreign_hold_pct",
                    "value": round(float(row["avg_foreign_hold_pct"]), 4),
                    "unit": "%",
                })

    if has_table("investor_trading_daily"):
        rows = conn.execute(
            """
            SELECT bas_dt,
                   COUNT(*) AS cover_count,
                   SUM(COALESCE(indv_net, 0)) AS indv_net,
                   SUM(COALESCE(inst_net, 0)) AS inst_net,
                   SUM(COALESCE(frgn_net, 0)) AS frgn_net
              FROM investor_trading_daily
             WHERE bas_dt IS NOT NULL
             GROUP BY bas_dt
             ORDER BY bas_dt
            """
        ).fetchall()
        for row in rows:
            period = _dt_to_period(row["bas_dt"])
            common = {
                "period": period,
                "source_name": "investor_trading_daily",
                "source_detail": "종목별 투자자 순매수 수량 시장 전체 합산",
                "quality": "daily_aggregate_from_collected_investor_flow",
            }
            result["public:20:107"].extend([
                {**common, "series_name": "individual_net_buy_qty_sum", "value": float(row["indv_net"] or 0), "unit": "주"},
                {**common, "series_name": "institution_net_buy_qty_sum", "value": float(row["inst_net"] or 0), "unit": "주"},
                {**common, "series_name": "foreign_net_buy_qty_sum", "value": float(row["frgn_net"] or 0), "unit": "주"},
                {**common, "series_name": "investor_flow_cover_count", "value": int(row["cover_count"] or 0), "unit": "종목"},
            ])

    if has_table("short_sell_daily"):
        rows = conn.execute(
            """
            SELECT bas_dt,
                   COUNT(*) AS cover_count,
                   SUM(COALESCE(short_qty, 0)) AS short_qty_sum,
                   SUM(COALESCE(short_amt, 0)) / 100000000.0 AS short_amt_100m,
                   SUM(COALESCE(borrow_bal_qty, 0)) AS borrow_bal_qty_sum,
                   SUM(COALESCE(borrow_bal_amt, 0)) / 100000000.0 AS borrow_bal_amt_100m
              FROM short_sell_daily
             WHERE bas_dt IS NOT NULL
             GROUP BY bas_dt
             ORDER BY bas_dt
            """
        ).fetchall()
        for row in rows:
            common = {
                "period": _dt_to_period(row["bas_dt"]),
                "source_name": "short_sell_daily",
                "source_detail": "일별 공매도/대차잔고 시장 전체 합산",
                "quality": "official_or_exchange_derived",
            }
            result["public:20:108"].extend([
                {**common, "series_name": "short_sell_qty_sum", "value": float(row["short_qty_sum"] or 0), "unit": "주"},
                {**common, "series_name": "short_sell_amt_100m", "value": round(float(row["short_amt_100m"] or 0), 4), "unit": "억원"},
                {**common, "series_name": "borrow_balance_qty_sum", "value": float(row["borrow_bal_qty_sum"] or 0), "unit": "주"},
                {**common, "series_name": "borrow_balance_amt_100m", "value": round(float(row["borrow_bal_amt_100m"] or 0), 4), "unit": "억원"},
                {**common, "series_name": "short_sell_cover_count", "value": int(row["cover_count"] or 0), "unit": "종목"},
            ])

    if has_table("short_rank_daily"):
        rows = conn.execute(
            """
            WITH month_last AS (
                SELECT substr(bas_dt,1,6) AS ym, MAX(bas_dt) AS bas_dt
                  FROM short_rank_daily
                 WHERE bas_dt IS NOT NULL
                 GROUP BY substr(bas_dt,1,6)
            )
            SELECT ml.ym,
                   COUNT(*) AS cover_count,
                   SUM(COALESCE(s.lnb_rman_stck_cnt, 0)) AS lending_remain_qty,
                   SUM(COALESCE(s.lnb_bal, 0)) / 100000000.0 AS lending_balance_100m
              FROM month_last ml
              JOIN short_rank_daily s ON s.bas_dt = ml.bas_dt
             GROUP BY ml.ym
             ORDER BY ml.ym
            """
        ).fetchall()
        for row in rows:
            common = {
                "period": _dt_to_period(row["ym"] + "01", monthly=True),
                "source_name": "short_rank_daily",
                "source_detail": "월말 대차거래 잔고 랭킹 테이블 시장 전체 합산",
                "quality": "monthly_snapshot_derived_from_exchange_lending_rank",
            }
            result["public:20:108"].extend([
                {**common, "series_name": "monthly_lending_remain_qty_sum", "value": float(row["lending_remain_qty"] or 0), "unit": "주"},
                {**common, "series_name": "monthly_lending_balance_100m", "value": round(float(row["lending_balance_100m"] or 0), 4), "unit": "억원"},
                {**common, "series_name": "monthly_lending_cover_count", "value": int(row["cover_count"] or 0), "unit": "종목"},
            ])

    for key, rows in result.items():
        print(f"[LocalMarketStructure] {key}: {len(rows)}행")
    return result


def collect_skipjack_import_unit_price_from_hs() -> list[dict]:
    """Collect monthly skipjack tuna import unit-price proxies from HS trade DB.

    EPIC's target is "가다랑어 어가추이". We do not yet have a verified fishery
    auction-price source, so this stores official customs value/weight as a
    clearly-labelled partial proxy instead of pretending it is the exact price.
    """
    if not HS_TRADE_DB_PATH.exists():
        return []

    hs_codes = {
        "0303430000": ("frozen_skipjack_import_usd_per_kg", "냉동 가다랑어 수입단가"),
        "0302330000": ("fresh_chilled_skipjack_import_usd_per_kg", "신선/냉장 가다랑어 수입단가"),
    }
    out: list[dict] = []

    with sqlite3.connect(HS_TRADE_DB_PATH) as hs_conn:
        hs_conn.row_factory = sqlite3.Row
        by_hs_rows = hs_conn.execute(
            """
            SELECT
                period_ym,
                hs_code,
                SUM(import_value) AS import_value_usd,
                SUM(import_weight) AS import_weight_kg
            FROM customs_monthly_record
            WHERE hs_code IN ('0303430000', '0302330000')
              AND import_value > 0
              AND import_weight > 0
            GROUP BY period_ym, hs_code
            ORDER BY period_ym, hs_code
            """
        ).fetchall()
        combined_rows = hs_conn.execute(
            """
            SELECT
                period_ym,
                SUM(import_value) AS import_value_usd,
                SUM(import_weight) AS import_weight_kg
            FROM customs_monthly_record
            WHERE hs_code IN ('0303430000', '0302330000')
              AND import_value > 0
              AND import_weight > 0
            GROUP BY period_ym
            ORDER BY period_ym
            """
        ).fetchall()

    def append_unit_price(
        period_ym: str,
        series_name: str,
        import_value_usd: float,
        import_weight_kg: float,
        source_detail: str,
        quality: str,
    ) -> None:
        if not period_ym or import_weight_kg <= 0:
            return
        unit_price = import_value_usd / import_weight_kg
        # Tiny lots can create meaningless unit-price spikes. Keep broad enough
        # bounds for tuna trade while blocking obvious parsing/unit failures.
        if unit_price < 0.3 or unit_price > 40:
            return
        out.append(
            {
                "period": period_ym,
                "series_name": series_name,
                "value": round(unit_price, 4),
                "unit": "USD/kg",
                "source_name": "HS_TRADE_LAB_CUSTOMS_MONTHLY",
                "source_detail": source_detail,
                "quality": quality,
            }
        )

    for row in by_hs_rows:
        spec = hs_codes.get(str(row["hs_code"]))
        if not spec:
            continue
        series_name, label = spec
        append_unit_price(
            str(row["period_ym"]),
            series_name,
            float(row["import_value_usd"] or 0),
            float(row["import_weight_kg"] or 0),
            f"관세청 월별 수출입 HS {row['hs_code']} {label}: 수입금액/수입중량",
            "official_customs_unit_price_proxy",
        )

    for row in combined_rows:
        append_unit_price(
            str(row["period_ym"]),
            "skipjack_import_usd_per_kg_all_hs",
            float(row["import_value_usd"] or 0),
            float(row["import_weight_kg"] or 0),
            "관세청 월별 수출입 HS 0303430000+0302330000 가다랑어 합산: 수입금액/수입중량",
            "official_customs_unit_price_proxy_derived_sum",
        )

    return out


def collect_customs_sector_quant_extensions() -> dict[str, list[dict]]:
    """Build sector-level export/import indicators from local customs HS records."""
    if not HS_TRADE_DB_PATH.exists():
        return {}

    result: dict[str, list[dict]] = {}
    hs_conn = sqlite3.connect(HS_TRADE_DB_PATH)
    hs_conn.row_factory = sqlite3.Row
    try:
        for indicator_key, _sub_code, _name, label, prefixes, _notes in CUSTOMS_SECTOR_QUANT_SPECS:
            where = " OR ".join(["hs_code LIKE ?" for _ in prefixes])
            params = [f"{prefix}%" for prefix in prefixes]
            rows = hs_conn.execute(
                f"""
                SELECT period_ym,
                       SUM(export_value) AS export_value_usd,
                       SUM(import_value) AS import_value_usd,
                       SUM(trade_balance) AS trade_balance_usd,
                       SUM(export_weight) AS export_weight_kg,
                       SUM(import_weight) AS import_weight_kg
                  FROM customs_monthly_record
                 WHERE period_ym LIKE '____-__'
                   AND ({where})
                 GROUP BY period_ym
                 ORDER BY period_ym
                """,
                params,
            ).fetchall()

            series: list[dict] = []
            for row in rows:
                period = str(row["period_ym"])
                export_value = float(row["export_value_usd"] or 0)
                import_value = float(row["import_value_usd"] or 0)
                trade_balance = float(row["trade_balance_usd"] or (export_value - import_value))
                export_weight = float(row["export_weight_kg"] or 0)
                import_weight = float(row["import_weight_kg"] or 0)
                source_detail = f"관세청 수출입 HS prefix {','.join(prefixes)} 월별 합산"
                common = {
                    "period": period,
                    "source_name": "hs_trade_lab.customs_monthly_record",
                    "source_detail": source_detail,
                    "quality": "official_customs_hs_prefix_aggregate",
                }
                series.extend([
                    {**common, "series_name": f"{label}_수출액", "value": round(export_value / 1_000_000.0, 3), "unit": "백만달러"},
                    {**common, "series_name": f"{label}_수입액", "value": round(import_value / 1_000_000.0, 3), "unit": "백만달러"},
                    {**common, "series_name": f"{label}_무역수지", "value": round(trade_balance / 1_000_000.0, 3), "unit": "백만달러"},
                ])
                if export_weight > 0:
                    series.append({**common, "series_name": f"{label}_수출단가", "value": round(export_value / export_weight, 4), "unit": "USD/kg"})
                if import_weight > 0:
                    series.append({**common, "series_name": f"{label}_수입단가", "value": round(import_value / import_weight, 4), "unit": "USD/kg"})
            result[indicator_key] = series
            print(f"[CustomsSector] {indicator_key} {label}: {len(series)}행")
    finally:
        hs_conn.close()
    return result


def get_customs_service_key() -> str:
    """Return the Korea Customs public-data service key from env files or env."""
    for key in ("CUSTOMS_ITEMTRADE_SERVICE_KEY", "PUBLIC_DATA_API_KEY", "CUSTOMS_API_KEY"):
        value = os.environ.get(key, "").strip().strip('"').strip("'")
        if value:
            return value
    for path in (ROOT / "hs_trade_lab" / ".env", ROOT / ".env"):
        try:
            if not path.exists():
                continue
            for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
                if "=" not in line or line.strip().startswith("#"):
                    continue
                key, _, value = line.partition("=")
                if key.strip() in ("CUSTOMS_ITEMTRADE_SERVICE_KEY", "PUBLIC_DATA_API_KEY", "CUSTOMS_API_KEY"):
                    value = value.strip().strip('"').strip("'")
                    if value:
                        return value
        except Exception:
            continue
    return ""


def _parse_customs_xml_items(xml_text: str) -> tuple[str, str, list[dict[str, str]]]:
    root = ET.fromstring(xml_text)
    result_code = root.findtext(".//resultCode", default="")
    result_msg = root.findtext(".//resultMsg", default="")
    items: list[dict[str, str]] = []
    for item in root.findall(".//item"):
        rows: dict[str, str] = {}
        for child in item:
            rows[child.tag] = child.text or ""
        items.append(rows)
    return result_code, result_msg, items


def collect_vietnam_country_product_exports() -> dict[str, list[dict]]:
    """Collect Vietnam country x product exports from Korea Customs nnewtempertrade.

    These are not Vietnam's global exports. They are Korea Customs' country x
    unified-product export statistics for exports to Vietnam. The catalog keeps
    this distinction explicit so the dashboard does not treat it as an exact
    EPIC global Vietnam export feed.
    """
    service_key = get_customs_service_key()
    if not service_key:
        return {}

    specs = {
        "epic:12:10": {
            "label": "베트남_의류신발_수출",
            "codes": {
                "13020101": "가죽제 의류",
                "13020201": "모피 의류",
                "13020301": "편물제 의류",
                "13020401": "직물제 의류",
                "13020590": "기타 의류부속품",
                "13030301": "신발류",
            },
        },
        "epic:15:11": {
            "label": "베트남_IT제품_수출",
            "codes": {
                "32050301": "전자계산기",
                "32050302": "전자계산기 부품",
                "33010101": "유선통신기기",
                "33010102": "유선통신기기부품",
                "33029001": "기타 무선통신기기",
                "33029002": "기타 무선통신기기 부품",
                "33030101": "컴퓨터",
                "33030290": "기타 컴퓨터 주변기기",
                "33030301": "정보기기 부품",
                "34010101": "메모리반도체",
                "34010102": "시스템반도체",
                "34010103": "집적회로반도체 부품",
                "34010190": "기타 집적회로반도체",
            },
        },
    }
    url = "https://apis.data.go.kr/1220000/nnewtempertrade/getNnewtempertradeList"
    start_year = 2016
    end_year = date.today().year
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0"})
    result: dict[str, list[dict]] = {}
    for indicator_key, spec in specs.items():
        by_period: dict[str, dict[str, float]] = {}
        code_names: dict[str, str] = spec["codes"]
        for code, code_name in code_names.items():
            for year in range(start_year, end_year + 1):
                params = {
                    "serviceKey": service_key,
                    "strtYymm": f"{year}01",
                    "endYymm": f"{year}12",
                    "imexTpcd": "1",
                    "imexTmprUnfcClsfCd": code,
                    "cntyCd": "VN",
                }
                try:
                    response = session.get(url, params=params, timeout=30)
                    response.raise_for_status()
                    result_code, result_msg, items = _parse_customs_xml_items(response.text)
                    if result_code and result_code != "00":
                        print(f"[VietnamNnew] skip {indicator_key}/{code}/{year}: {result_code} {result_msg}")
                        continue
                except Exception as exc:
                    print(f"[VietnamNnew] skip {indicator_key}/{code}/{year}: {exc}")
                    continue
                for item in items:
                    ym_raw = str(item.get("year") or "")
                    ym_match = re.search(r"(\d{4})[. -]?(\d{2})", ym_raw)
                    if not ym_match:
                        continue
                    period = f"{ym_match.group(1)}-{ym_match.group(2)}"
                    value_usd = safe_float(item.get("dlr")) or 0.0
                    weight_kg = safe_float(item.get("wgt")) or 0.0
                    slot = by_period.setdefault(period, {"value_usd": 0.0, "weight_kg": 0.0})
                    slot["value_usd"] += value_usd
                    slot["weight_kg"] += weight_kg
        rows: list[dict] = []
        code_note = ",".join(code_names.keys())
        for period, metrics in sorted(by_period.items()):
            value_usd = metrics["value_usd"]
            weight_kg = metrics["weight_kg"]
            common = {
                "period": period,
                "source_name": "Korea_Customs_nnewtempertrade",
                "source_detail": f"cntyCd=VN imexTpcd=1 unified product codes={code_note}",
                "quality": "official_country_product_export_proxy_not_vietnam_global_export",
            }
            rows.append({
                **common,
                "series_name": f"{spec['label']}_수출액",
                "value": round(value_usd / 1000.0, 3),
                "unit": "천달러",
            })
            if weight_kg > 0:
                rows.append({
                    **common,
                    "series_name": f"{spec['label']}_수출중량",
                    "value": round(weight_kg, 3),
                    "unit": "kg",
                })
                rows.append({
                    **common,
                    "series_name": f"{spec['label']}_수출단가",
                    "value": round(value_usd / weight_kg, 4),
                    "unit": "USD/kg",
                })
        result[indicator_key] = rows
        print(f"[VietnamNnew] {indicator_key} {spec['label']}: {len(rows)}행")
    return result


def hyundai_sales_file_map(year: str) -> dict[str, str]:
    response = requests.post(
        "https://www.hyundai.com/wsvc/ww/salesPerformance.list.do",
        data={"lang": "ko", "year": year},
        headers={
            "User-Agent": "Mozilla/5.0",
            "Referer": "https://www.hyundai.com/worldwide/ko/company/ir/ir-resources/sales-results",
        },
        timeout=20,
    )
    response.raise_for_status()
    payload = response.json()
    rows = ((payload or {}).get("data") or {}).get("list") or []
    if not rows:
        return {}
    item = rows[0]
    return {
        "sales_model": item.get("attrSalesModelValue") or "",
        "us_retail": item.get("attrUSRetailValue") or "",
    }


def fetch_xlsx_bytes(url_path: str) -> bytes:
    if not url_path:
        return b""
    url = f"https://www.hyundai.com{url_path}" if url_path.startswith("/") else url_path
    response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
    response.raise_for_status()
    return response.content


def parse_hyundai_domestic_model_sales(xlsx_bytes: bytes, year: int, max_month: int | None = None) -> list[dict]:
    if not xlsx_bytes:
        return []
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    month_labels = {
        "Jan.": "01", "Feb.": "02", "Mar.": "03", "Apr.": "04", "May.": "05", "Jun.": "06",
        "Jul.": "07", "Aug.": "08", "Sep.": "09", "Oct.": "10", "Nov.": "11", "Dec.": "12",
        "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04", "May": "05", "Jun": "06",
        "Jul": "07", "Aug": "08", "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
    }
    header_row = next(ws.iter_rows(min_row=1, max_row=6, values_only=True))
    month_cols: list[tuple[int, str]] = []
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        if row and any(str(cell).strip() in month_labels for cell in row if cell):
            header_row = row
            break
    for idx, cell in enumerate(header_row):
        label = str(cell).strip() if cell is not None else ""
        if label in month_labels:
            month_cols.append((idx, month_labels[label]))
    rows: list[dict] = []
    section = None
    group_name = ""
    for excel_row in ws.iter_rows(min_row=6, values_only=True):
        label = excel_row[1]
        model = excel_row[2]
        if label == "Domestic":
            section = "domestic"
            group_name = ""
            continue
        if label == "Export":
            break
        if section != "domestic":
            continue
        if label in {"PC", "RV", "CV"} and model:
            group_name = str(label)
        if not model or str(model).strip() in {"Sub-total", "Total"}:
            continue
        model_name = str(model).strip()
        for col_idx, mm in month_cols:
            if max_month is not None and int(mm) > max_month:
                continue
            value = excel_row[col_idx]
            if value in (None, ""):
                continue
            rows.append(
                {
                    "period": f"{year}-{mm}",
                    "series_name": model_name,
                    "value": float(value),
                    "unit": "대",
                    "source_name": "HYUNDAI_IR_SALES_BY_MODEL",
                    "source_detail": f"현대차 국내 모델별 판매 ({group_name})",
                    "quality": "official",
                }
            )
    return rows


def parse_hyundai_us_retail_sales(xlsx_bytes: bytes, year: int, max_month: int | None = None) -> list[dict]:
    if not xlsx_bytes:
        return []
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    month_labels = {
        "Jan.": "01", "Feb.": "02", "Mar.": "03", "Apr.": "04", "May.": "05", "Jun.": "06",
        "Jul.": "07", "Aug.": "08", "Sep.": "09", "Oct.": "10", "Nov.": "11", "Dec.": "12",
        "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04", "May": "05", "Jun": "06",
        "Jul": "07", "Aug": "08", "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
    }
    header_row = next(ws.iter_rows(min_row=1, max_row=6, values_only=True))
    month_cols: list[tuple[int, str]] = []
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        if row and any(str(cell).strip() in month_labels for cell in row if cell):
            header_row = row
            break
    for idx, cell in enumerate(header_row):
        label = str(cell).strip() if cell is not None else ""
        if label in month_labels:
            month_cols.append((idx, month_labels[label]))
    rows: list[dict] = []
    group_name = ""
    for excel_row in ws.iter_rows(min_row=6, values_only=True):
        label = excel_row[1]
        model = excel_row[2]
        if label in {"PC", "RV", "LCV"} and model:
            group_name = str(label)
        if not model or str(model).strip() in {"Sub-total", "Total"}:
            continue
        model_name = str(model).strip()
        for col_idx, mm in month_cols:
            if max_month is not None and int(mm) > max_month:
                continue
            value = excel_row[col_idx]
            if value in (None, ""):
                continue
            rows.append(
                {
                    "period": f"{year}-{mm}",
                    "series_name": model_name,
                    "value": float(value),
                    "unit": "대",
                    "source_name": "HYUNDAI_IR_US_RETAIL",
                    "source_detail": f"현대차 미국 소매판매 ({group_name})",
                    "quality": "official",
                }
            )
    return rows


def infer_report_month_from_path(url_path: str) -> int | None:
    lower = (url_path or "").lower()
    month_tokens = {
        "jan": 1, "january": 1,
        "feb": 2, "february": 2,
        "mar": 3, "march": 3,
        "apr": 4, "april": 4,
        "may": 5,
        "jun": 6, "june": 6,
        "jul": 7, "july": 7,
        "aug": 8, "august": 8,
        "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10,
        "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }
    for token, month in month_tokens.items():
        if f"-{token}-" in lower or f"_{token}_" in lower or f"/{token}-" in lower:
            return month
    return None


def normalize_int(value) -> int | None:
    if value is None:
        return None
    s = str(value).strip().replace(",", "")
    if not s or s in {"-", "—"}:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def parse_korean_unit_number(text: str) -> int | None:
    if not text:
        return None
    import re

    s = str(text).strip().replace(",", "").replace(" ", "")
    total = 0
    m = re.search(r"(\d+)만", s)
    if m:
        total += int(m.group(1)) * 10000
        s = s[m.end():]
    m = re.search(r"(\d+)", s)
    if m:
        total += int(m.group(1))
    return total or None


def normalize_company_label(text: str | None) -> str | None:
    if not text:
        return None
    s = re.sub(r"\s+", "", str(text))
    if not s or "검산" in s:
        return None
    if "현대" in s:
        return "현대차"
    if "기아" in s:
        return "기아"
    if "한국지엠" in s or "한국GM" in s or ("지엠" in s and "한국" in s):
        return "한국GM"
    if "KG모빌리티" in s or "쌍용" in s:
        return "KG모빌리티"
    if "르노코리아" in s or "르노삼성" in s:
        return "르노코리아"
    if "대우버스" in s:
        return "대우버스"
    if "타타대우" in s:
        return "타타대우"
    return None


def normalize_header_cell(text) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", "", str(text))


def infer_excel_engine(org_file_name: str) -> str | None:
    lower = (org_file_name or "").lower()
    if lower.endswith(".xls"):
        return "xlrd"
    if lower.endswith(".xlsx"):
        return "openpyxl"
    return None


def find_sheet_name(sheet_names: list[str], prefix: str) -> str | None:
    target = prefix.replace(" ", "")
    for name in sheet_names:
        if normalize_header_cell(name).startswith(target):
            return name
    return None


def find_kama_model_sheet_name(sheet_names: list[str]) -> str | None:
    candidates = []
    for name in sheet_names:
        normalized = normalize_header_cell(name)
        if "업체별.모델별" in normalized and "생산.내수.수출" in normalized:
            candidates.append(name)
    if not candidates:
        return None
    candidates.sort(key=lambda x: normalize_header_cell(x))
    return candidates[0]


def parse_kama_period(text: str) -> tuple[int, int] | None:
    if not text:
        return None
    s = str(text).strip()
    m = re.search(r"(20\d{2})[.\-_년 ]+\s*(\d{1,2})", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None


def fetch_kama_monthly_file_index(start_year: int = 2016, max_pages: int = 40) -> list[dict]:
    items: list[dict] = []
    seen_periods: set[tuple[int, int]] = set()
    for page in range(1, max_pages + 1):
        r = requests.get(
            "https://www.kama.or.kr/NewsController",
            params={"boardmaster_id": "Produce", "cmd": "L", "menunum": "0003", "pagenum": page},
            timeout=30,
            headers={"User-Agent": "Mozilla/5.0"},
        )
        r.raise_for_status()
        r.encoding = "cp949"
        text = r.text
        matches = re.findall(
            r'<a href="/NewsController\?cmd=V&amp;boardmaster_id=Produce&amp;board_id=(\d+)[^"]*">([^<]+)</a>.*?'
            r"fileDown\('([^']+)','([^']+)','([^']+)','Produce'\)",
            text,
            flags=re.S,
        )
        if not matches:
            continue
        for board_id, raw_title, org_file, server_file, path in matches:
            title = re.sub(r"&[#A-Za-z0-9]+;", " ", raw_title).strip()
            if "자동차통계월보" not in title:
                continue
            if not org_file.lower().startswith("monthly"):
                continue
            period = parse_kama_period(title) or parse_kama_period(org_file)
            if not period:
                continue
            year, month = period
            if year < start_year:
                continue
            if (year, month) in seen_periods:
                continue
            seen_periods.add((year, month))
            items.append(
                {
                    "board_id": int(board_id),
                    "title": title,
                    "period": f"{year:04d}-{month:02d}",
                    "year": year,
                    "month": month,
                    "org_file": org_file,
                    "server_file": server_file,
                    "path": path,
                }
            )
    items.sort(key=lambda x: (x["year"], x["month"], x["board_id"]))
    return items


def download_kama_attachment(item: dict) -> bytes:
    r = requests.post(
        "https://www.kama.or.kr/jsp/common/FileDown.jsp",
        data={
            "org_fileName": item["org_file"],
            "server_fileName": item["server_file"],
            "boardmaster_id": "Produce",
            "path": item["path"],
        },
        timeout=(10, 20),
        headers={"User-Agent": "Mozilla/5.0", "Referer": "https://www.kama.or.kr/"},
    )
    r.raise_for_status()
    return r.content


def load_kama_company_sheet(file_bytes: bytes, org_file_name: str) -> pd.DataFrame:
    engine = infer_excel_engine(org_file_name)
    return pd.read_excel(BytesIO(file_bytes), sheet_name="1-3업체별총괄", header=None, engine=engine)


def parse_kama_company_month_rows(df: pd.DataFrame) -> list[dict]:
    company_row = None
    for i in range(min(10, len(df))):
        row_values = [normalize_company_label(v) for v in df.iloc[i].tolist()]
        if "현대차" in row_values and "기아" in row_values:
            company_row = i
            break
    if company_row is None:
        return []

    company_positions: list[tuple[int, str]] = []
    for col, val in enumerate(df.iloc[company_row].tolist()):
        name = normalize_company_label(val)
        if name:
            company_positions.append((col, name))
    if not company_positions:
        return []

    metric_row = company_row + 2
    metric_sub_row = company_row + 3
    company_metrics: dict[str, dict[str, int]] = {}
    for idx, (start_col, company) in enumerate(company_positions):
        end_col = company_positions[idx + 1][0] if idx + 1 < len(company_positions) else len(df.columns)
        metrics: dict[str, int] = {}
        for col in range(start_col, end_col):
            top = normalize_header_cell(df.iat[metric_row, col] if metric_row < len(df) else None)
            sub = normalize_header_cell(df.iat[metric_sub_row, col] if metric_sub_row < len(df) else None)
            if top == "계" and "total" not in metrics:
                metrics["total"] = col
            elif top == "내수" and "domestic" not in metrics:
                metrics["domestic"] = col
            elif top == "수출" and sub in {"수량", ""} and "export" not in metrics:
                metrics["export"] = col
        company_metrics[company] = metrics

    month_rows: list[dict] = []
    current_block_year: int | None = None
    data_start = metric_sub_row + 1
    for i in range(data_start, len(df)):
        left0 = df.iat[i, 0] if 0 < len(df.columns) else None
        left1 = df.iat[i, 1] if 1 < len(df.columns) else None
        year_month = parse_kama_period(left0) or parse_kama_period(left1)
        month_num = None
        if year_month:
            current_block_year, month_num = year_month
        elif current_block_year is not None:
            month_num = normalize_int(left0)
            if month_num is None:
                month_num = normalize_int(left1)
        if current_block_year is None or month_num is None or not (1 <= int(month_num) <= 12):
            continue

        period = f"{current_block_year:04d}-{int(month_num):02d}"
        company_values: dict[str, dict[str, int | None]] = {}
        nonzero = False
        for company, metrics in company_metrics.items():
            total = normalize_int(df.iat[i, metrics["total"]]) if "total" in metrics else None
            domestic = normalize_int(df.iat[i, metrics["domestic"]]) if "domestic" in metrics else None
            export_qty = normalize_int(df.iat[i, metrics["export"]]) if "export" in metrics else None
            company_values[company] = {
                "total": total,
                "domestic": domestic,
                "export": export_qty,
            }
            if any((v or 0) > 0 for v in (total, domestic, export_qty)):
                nonzero = True
        if not nonzero:
            continue
        month_rows.append({"period": period, "companies": company_values})
    return month_rows


def parse_kama_model_family_rows(df: pd.DataFrame) -> list[dict]:
    month_label = None
    model_col = prod_col = prod_ytd_col = domestic_col = domestic_ytd_col = export_col = export_ytd_col = None
    start_row = None

    for idx in range(min(len(df), 40)):
        row = [normalize_header_cell(v) for v in df.iloc[idx].tolist()]
        if "업체" in row and "모델" in row and "생산" in row and "내수" in row and "수출" in row:
            start_row = idx + 2
            try:
                model_col = row.index("모델")
                prod_col = row.index("생산")
                domestic_col = row.index("내수")
                export_col = row.index("수출")
                prod_ytd_col = prod_col + 1
                domestic_ytd_col = domestic_col + 1
                export_ytd_col = export_col + 1
            except ValueError:
                return []
            next_row = [normalize_header_cell(v) for v in df.iloc[idx + 1].tolist()]
            month_label = next_row[domestic_col] if domestic_col is not None and domestic_col < len(next_row) else None
            break

    if start_row is None or month_label is None:
        return []

    company = None
    out: list[dict] = []
    for idx in range(start_row, len(df)):
        company_cell = normalize_company_label(df.iat[idx, 0] if 0 < len(df.columns) else "")
        if company_cell:
            company = company_cell
        if not company:
            continue

        model = str(df.iat[idx, model_col]).strip() if model_col is not None and model_col < len(df.columns) and pd.notna(df.iat[idx, model_col]) else ""
        if not model:
            continue

        normalized_model = re.sub(r"\s+", " ", model)
        if normalized_model in {"소 계", "소계", "총 계", "총계", "국산", "OEM 수입"}:
            continue
        if normalized_model.endswith("EXPORT"):
            continue
        if not normalized_model.endswith("계"):
            continue

        family = normalized_model[:-1].strip()
        family = family.replace("TORESS", "TORRES")
        if not family:
            continue

        def _to_num(col_idx):
            if col_idx is None or col_idx >= len(df.columns):
                return None
            return safe_float(df.iat[idx, col_idx])

        out.append(
            {
                "month_label": month_label,
                "company": company,
                "model_family": family,
                "production": _to_num(prod_col),
                "production_ytd": _to_num(prod_ytd_col),
                "domestic": _to_num(domestic_col),
                "domestic_ytd": _to_num(domestic_ytd_col),
                "export": _to_num(export_col),
                "export_ytd": _to_num(export_ytd_col),
            }
        )
    return out


def parse_kama_global_country_rows(df: pd.DataFrame, max_year: int | None = None, max_month: int | None = None) -> list[dict]:
    if len(df) < 8:
        return []

    header_country = [str(x).strip() if x == x else "" for x in df.iloc[4, :].tolist()]
    header_type = [str(x).strip() if x == x else "" for x in df.iloc[5, :].tolist()]
    total_cols: list[tuple[int, str]] = []
    current_country = ""
    for col in range(2, len(header_type)):
        if header_country[col]:
            current_country = re.sub(r"\s+", "", header_country[col])
        if header_type[col] == "계" and current_country:
            total_cols.append((col, current_country))

    if not total_cols:
        return []

    rows: list[dict] = []
    current_year = None
    for idx in range(7, len(df)):
        label = str(df.iat[idx, 1]).strip() if 1 < len(df.columns) and pd.notna(df.iat[idx, 1]) else ""
        if not label:
            continue

        m_full = re.match(r"^(\d{4})\.\s*(\d{1,2})$", label)
        m_ytd = re.match(r"^(\d{4})\.\s*1~(\d{1,2})$", label)
        m_month = re.match(r"^(\d{1,2})$", label)

        period = None
        if m_full:
            year = int(m_full.group(1))
            month = int(m_full.group(2))
            current_year = year
            period = (year, month)
        elif m_ytd:
            current_year = int(m_ytd.group(1))
            continue
        elif m_month and current_year is not None:
            period = (current_year, int(m_month.group(1)))
        else:
            continue

        year, month = period
        if max_year is not None and year > max_year:
            continue
        if max_year is not None and max_month is not None and year == max_year and month > max_month:
            continue

        for col, country in total_cols:
            value = safe_float(df.iat[idx, col]) if col < len(df.columns) else None
            passenger = safe_float(df.iat[idx, col + 1]) if col + 1 < len(df.columns) else None
            commercial = safe_float(df.iat[idx, col + 2]) if col + 2 < len(df.columns) else None
            if value == 0 and passenger is None and commercial is None:
                continue
            if value is None:
                continue
            rows.append(
                {
                    "period": f"{year:04d}-{month:02d}",
                    "series_name": country,
                    "value": float(value),
                    "unit": "대",
                    "source_name": "KAMA_GLOBAL_NEW_REGISTRATIONS",
                    "source_detail": f"KAMA 주요국 신차등록 계 ({year:04d}-{month:02d})",
                    "quality": "official_association",
                }
            )
    return rows


def collect_kama_auto_company_indicators(start_year: int = 2016) -> tuple[list[dict], list[dict], dict[str, list[dict]], dict[str, list[dict]], list[dict]]:
    files = fetch_kama_monthly_file_index(start_year=start_year)
    company_sales_rows: list[dict] = []
    market_share_rows: list[dict] = []
    specific_rows: dict[str, list[dict]] = {"epic:0:19": [], "epic:0:20": [], "epic:0:21": []}
    model_rows: dict[str, list[dict]] = {"epic:0:17": [], "epic:0:112": [], "epic:0:113": []}
    global_country_rows: list[dict] = []
    company_indicator_map = {
        "KG모빌리티": "epic:0:19",
        "르노코리아": "epic:0:20",
        "한국GM": "epic:0:21",
    }
    model_indicator_map = {
        ("기아", "domestic"): "epic:0:17",
        ("KG모빌리티", "domestic"): "epic:0:112",
        ("KG모빌리티", "export"): "epic:0:113",
    }

    for item in files:
        try:
            content = download_kama_attachment(item)
            df = load_kama_company_sheet(content, item["org_file"])
            month_rows = parse_kama_company_month_rows(df)
        except Exception:
            continue
        model_family_rows: list[dict] = []
        global_rows_for_file: list[dict] = []
        try:
            engine = infer_excel_engine(item["org_file"])
            xf = pd.ExcelFile(BytesIO(content), engine=engine)
            sheet_name = find_kama_model_sheet_name(xf.sheet_names)
            if sheet_name:
                model_df = xf.parse(sheet_name, header=None)
                model_family_rows = parse_kama_model_family_rows(model_df)
            global_sheet = find_sheet_name(xf.sheet_names, "10-2")
            if global_sheet:
                global_df = xf.parse(global_sheet, header=None)
                global_rows_for_file = parse_kama_global_country_rows(global_df, max_year=item["year"], max_month=item["month"])
        except Exception:
            model_family_rows = []
            global_rows_for_file = []
        for month_row in month_rows:
            period = month_row["period"]
            companies = month_row["companies"]
            domestic_sum = sum((vals.get("domestic") or 0) for vals in companies.values())
            for company, vals in companies.items():
                total = vals.get("total")
                domestic = vals.get("domestic")
                export_qty = vals.get("export")
                if total is not None:
                    company_sales_rows.append(
                        {
                            "period": period,
                            "series_name": company,
                            "value": float(total),
                            "unit": "대",
                            "source_name": "KAMA_MONTHLY_COMPANY_TOTAL",
                            "source_detail": f"KAMA 자동차통계월보 업체별 총괄 판매계 ({period})",
                            "quality": "official_association",
                        }
                    )
                if domestic_sum > 0 and domestic is not None:
                    market_share_rows.append(
                        {
                            "period": period,
                            "series_name": company,
                            "value": round(domestic * 100.0 / domestic_sum, 4),
                            "unit": "%",
                            "source_name": "KAMA_MONTHLY_MARKET_SHARE",
                            "source_detail": f"KAMA 자동차통계월보 업체별 내수 기준 시장점유율 ({period})",
                            "quality": "derived_from_official_sales",
                        }
                    )
                indicator_key = company_indicator_map.get(company)
                if indicator_key:
                    if total is not None:
                        specific_rows[indicator_key].append(
                            {
                                "period": period,
                                "series_name": "total_sales",
                                "value": float(total),
                                "unit": "대",
                                "source_name": "KAMA_MONTHLY_TOTAL",
                                "source_detail": f"{company} 판매계 ({period})",
                                "quality": "official_association",
                            }
                        )
                    if domestic is not None:
                        specific_rows[indicator_key].append(
                            {
                                "period": period,
                                "series_name": "domestic_sales",
                                "value": float(domestic),
                                "unit": "대",
                                "source_name": "KAMA_MONTHLY_DOMESTIC",
                                "source_detail": f"{company} 내수 ({period})",
                                "quality": "official_association",
                            }
                        )
                    if export_qty is not None:
                        specific_rows[indicator_key].append(
                            {
                                "period": period,
                                "series_name": "export_units",
                                "value": float(export_qty),
                                "unit": "대",
                                "source_name": "KAMA_MONTHLY_EXPORT",
                                "source_detail": f"{company} 수출 ({period})",
                                "quality": "official_association",
                            }
                        )
        for model_row in model_family_rows:
            period = f"{item['year']:04d}-{item['month']:02d}"
            company = model_row["company"]
            model_family = model_row["model_family"]
            for metric_name in ("domestic", "export"):
                indicator_key = model_indicator_map.get((company, metric_name))
                value = model_row.get(metric_name)
                if indicator_key and value is not None:
                    model_rows[indicator_key].append(
                        {
                            "period": period,
                            "series_name": model_family,
                            "value": float(value),
                            "unit": "대",
                            "source_name": f"KAMA_MODEL_{metric_name.upper()}",
                            "source_detail": f"KAMA 자동차통계월보 모델별 {company} {metric_name} ({period})",
                            "quality": "official_association",
                        }
                    )
        global_country_rows.extend(global_rows_for_file)
    return company_sales_rows, market_share_rows, specific_rows, model_rows, global_country_rows


def collect_hyundai_model_indicators() -> tuple[list[dict], list[dict]]:
    domestic_rows: list[dict] = []
    us_rows: list[dict] = []
    current_year = date.today().year
    for year in range(2016, current_year + 1):
        file_map = hyundai_sales_file_map(str(year))
        if not file_map:
            continue
        max_month = infer_report_month_from_path(file_map.get("sales_model", "")) or infer_report_month_from_path(file_map.get("us_retail", ""))
        sales_model_bytes = fetch_xlsx_bytes(file_map.get("sales_model", ""))
        us_retail_bytes = fetch_xlsx_bytes(file_map.get("us_retail", ""))
        domestic_rows.extend(parse_hyundai_domestic_model_sales(sales_model_bytes, year, max_month=max_month))
        us_rows.extend(parse_hyundai_us_retail_sales(us_retail_bytes, year, max_month=max_month))
    return domestic_rows, us_rows


def fetch_kia_newsroom_list(page_size: int = 50) -> list[dict]:
    out: list[dict] = []
    page = 0
    while True:
        r = requests.get(
            "https://worldwide.kia.com/api/newsroom",
            params={"pageSize": page_size, "p": page},
            timeout=30,
            headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"},
        )
        r.raise_for_status()
        js = r.json()
        data = js.get("data") or {}
        items = data.get("list") or []
        if not items:
            break
        out.extend(items)
        paging = data.get("paging") or {}
        total = int(paging.get("totalCount") or 0)
        if len(out) >= total:
            break
        page += 1
        if page > 20:
            break
    return out


def fetch_kia_newsroom_detail(article_id: int) -> dict | None:
    r = requests.get(
        f"https://worldwide.kia.com/api/newsroom/id/{article_id}",
        timeout=30,
        headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"},
    )
    r.raise_for_status()
    js = r.json()
    return ((js.get("data") or {}).get("present")) or None


def collect_kia_domestic_company_sales() -> list[dict]:
    import re

    rows: list[dict] = []
    seen_periods: set[str] = set()
    items = fetch_kia_newsroom_list(50)
    for item in items:
        title = str(item.get("title") or "")
        m = re.search(r"기아,\s*(\d{4})년\s*(\d{1,2})월\s*[\d만,\s]+대\s*판매", title)
        if not m:
            continue
        sales_year = int(m.group(1))
        sales_month = int(m.group(2))
        period = f"{sales_year:04d}-{sales_month:02d}"
        if period in seen_periods:
            continue
        detail = fetch_kia_newsroom_detail(int(item["id"]))
        desc = str((detail or {}).get("description") or "")
        lines = [x.strip() for x in desc.replace("\r", "\n").split("\n") if x.strip()]
        domestic = None
        for line in lines:
            if "국내 시장에서" in line and domestic is None:
                mm = re.search(r"국내 시장에서\s*([\d만,\s]+)대\s*판매", line)
                if mm:
                    domestic = parse_korean_unit_number(mm.group(1))
        if domestic is not None:
            rows.append(
                {
                    "period": period,
                    "series_name": "기아",
                    "value": domestic,
                    "unit": "대",
                    "source_name": "KIA_NEWSROOM_DOMESTIC_TOTAL",
                    "source_detail": f"기아 국내 판매 합계 ({period})",
                    "quality": "official",
                }
            )
            seen_periods.add(period)
    return rows


def collect_hyundai_domestic_company_sales(conn: sqlite3.Connection) -> list[dict]:
    sums = conn.execute(
        """
        select period, round(sum(coalesce(value, 0)), 0) as total
          from quant_major_indicator_series
         where indicator_key = 'epic:0:14'
           and source_name = 'HYUNDAI_IR_SALES_BY_MODEL'
         group by period
         order by period
        """
    ).fetchall()
    return [
        {
            "period": period,
            "series_name": "현대차",
            "value": float(total),
            "unit": "대",
            "source_name": "HYUNDAI_IR_DOMESTIC_TOTAL",
            "source_detail": f"현대차 국내 판매 합계 ({period})",
            "quality": "derived_official",
        }
        for period, total in sums
    ]


def collect_kia_us_model_sales() -> list[dict]:
    today = date.today()
    out: list[dict] = []
    for year in range(2017, today.year + 1):
        max_month = today.month if year == today.year else 12
        compare_year = year - 1 if year > 2017 else year
        for month in range(1, max_month + 1):
            url = (
                "https://www.kiamedia.com/us/en/sales/"
                f"salesbymonthexport?month={month}&year={year}&yeartocompare={compare_year}"
            )
            r = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
            r.raise_for_status()
            wb = openpyxl.load_workbook(BytesIO(r.content), data_only=True)
            ws = wb[wb.sheetnames[0]]
            period = f"{year:04d}-{month:02d}"
            for row in ws.iter_rows(min_row=3, values_only=True):
                model = row[0]
                if not model:
                    continue
                model = str(model).strip()
                if model.upper() == "TOTAL":
                    continue
                value = normalize_int(row[1])
                if value is None:
                    continue
                out.append(
                    {
                        "period": period,
                        "series_name": model,
                        "value": value,
                        "unit": "대",
                        "source_name": "KIA_US_SALES_BY_MODEL",
                        "source_detail": f"Kia America monthly sales by model ({period})",
                        "quality": "official",
                    }
                )
    return out


def collect_semiconductor_export_from_hs(conn: sqlite3.Connection) -> dict[str, list[dict]]:
    """
    hs_trade_lab DB에서 반도체 관련 HS 코드 월별 수출금액 집계
    epic:3:2 = 한국 반도체(집적회로) 수출금액 (HS 8542, 달러)
    epic:3:20 = 한국 반도체 제조장비 수출금액 (HS 8486, 달러)
    """
    hs_db = ROOT / "hs_trade_lab" / "data" / "hs_trade_lab.db"
    if not hs_db.exists():
        print("[반도체수출] hs_trade_lab.db 없음 — 스킵")
        return {}

    hconn = sqlite3.connect(str(hs_db), timeout=20)
    hconn.row_factory = sqlite3.Row

    result = {}
    queries = [
        ("epic:3:2", "8542%", "반도체(집적회로) 수출금액", "USD"),
        ("epic:3:20", "8486%", "반도체제조장비 수출금액", "USD"),
        ("epic:3:21", "8541%", "전자부품(다이오드/트랜지스터) 수출금액", "USD"),
    ]

    for indicator_key, hs_like, label, unit in queries:
        rows = hconn.execute("""
            SELECT period_ym, SUM(export_value) AS total_exp
            FROM customs_monthly_record
            WHERE hs_code LIKE ?
            GROUP BY period_ym
            HAVING total_exp > 0
            ORDER BY period_ym
        """, (hs_like,)).fetchall()

        series = []
        for r in rows:
            ym = r["period_ym"]  # "YYYY-MM"
            if not ym or len(ym) != 7:
                continue
            val = r["total_exp"] or 0
            series.append({
                "period": ym,
                "series_name": label,
                "value": round(val / 1_000_000, 2),  # 백만달러 단위
                "unit": "million_usd",
                "source_name": "customs_hs_trade_lab",
                "source_detail": f"관세청 HS {hs_like} 월별 수출금액 합산",
                "quality": "official_customs_hs_prefix_aggregate",
            })

        result[indicator_key] = series
        print(f"[반도체수출] {indicator_key} ({label}): {len(series)}행")

    hconn.close()

    # catalog에 항목 추가/업데이트
    for ik, series in result.items():
        if not series:
            continue
        existing = conn.execute(
            "SELECT 1 FROM quant_major_indicator_catalog WHERE indicator_key=?", (ik,)
        ).fetchone()
        if not existing:
            conn.execute("""
                INSERT OR IGNORE INTO quant_major_indicator_catalog
                (indicator_key, epic_category_code, epic_sub_code, epic_indicator_name,
                 base_unit, source_system, status, priority, notes)
                VALUES (?, '3', ?, ?, ?, '관세청(hs_trade_lab)', 'ready_existing', 'P1', '월별 수출금액')
            """, (ik, ik.split(":")[2] if ":" in ik else "0",
                  series[0]["series_name"], series[0]["unit"]))
        else:
            conn.execute(
                "UPDATE quant_major_indicator_catalog SET status='ready_existing' WHERE indicator_key=?",
                (ik,)
            )
    conn.commit()

    return result


def collect_dram_price_monthly() -> list[dict]:
    """
    DRAM Exchange (dramexchange.com) or DRAMeXchange 공개 데이터 대체로
    World Bank Commodity Price (CMO) API의 메모리 관련 항목 또는
    FRED API의 반도체 관련 PPI 데이터 수집.

    대안: FRED - PCU33441133441103 (Memory Chip Producer Price Index, US)
    또는 OECD.Stat - TEiTARGET / 한국 통계청 반도체 수출단가지수

    epic:3:1 = DRAM 현물 가격 (DDR4 4Gb)
    """
    rows = []

    # 시도 1: FRED 반도체 PPI (메모리칩 생산자 물가지수)
    try:
        fred_key = os.environ.get("FRED_API_KEY", "")
        # FRED에 DRAM 전용 직접 지표 없음 → 한국 통계청 반도체 수출단가지수 사용
        # KOSIS: 반도체 수출단가지수 (1-5 수출입물가지수, "반도체" 품목)
        url = ("https://ecos.bok.or.kr/api/StatisticSearch/sample/json/kr/"
               "1/100/901Y067/M/202001/202612/S23A")
        r = requests.get(url, timeout=15)
        if r.ok:
            data = r.json()
            items = data.get("StatisticSearch", {}).get("row", [])
            for item in items:
                try:
                    ym = item.get("TIME", "")  # e.g. "202301"
                    val = float(item.get("DATA_VALUE", 0) or 0)
                    if len(ym) == 6 and val > 0:
                        rows.append({
                            "period": f"{ym[:4]}-{ym[4:]}-01",
                            "value": val,
                            "unit": "index",
                            "label": "투자자예탁금",
                            "source": "ECOS_BOK",
                        })
                except Exception:
                    pass
        if rows:
            return rows
    except Exception:
        pass

    # 시도 2: KOSIS 반도체 수출단가지수 (공개 API)
    try:
        kosis_url = (
            "https://kosis.kr/openapi/Param/statisticsParamData.do"
            "?method=getList&apiKey=&itmId=T1&objL1=A08&objL2=&objL3=&objL4=&objL5=&objL6=&objL7=&objL8="
            "&format=json&jsonVD=Y&prdSe=M&startPrdDe=202001&endPrdDe=202612"
            "&orgId=101&tblId=DT_1KE10030&vwCd=MT_ZTITLE&listId="
        )
        # KOSIS API key 없으면 skip
        kosis_key = os.environ.get("KOSIS_API_KEY", "")
        if not kosis_key:
            raise ValueError("No KOSIS API key")
    except Exception:
        pass

    # 시도 3: World Bank commodity DRAM proxy via 반도체 수출단가 ECOS
    try:
        # 한국 수출입물가지수 - 반도체 (ECOS 공개 샘플)
        url2 = "https://ecos.bok.or.kr/api/StatisticSearch/sample/json/kr/1/200/402Y014/M/202001/202612/1000000"
        r2 = requests.get(url2, timeout=15)
        if r2.ok:
            data2 = r2.json()
            items2 = data2.get("StatisticSearch", {}).get("row", [])
            for item in items2:
                try:
                    ym = item.get("TIME", "")
                    val = float(item.get("DATA_VALUE", 0) or 0)
                    if len(ym) == 6 and val > 0:
                        rows.append({
                            "period": f"{ym[:4]}-{ym[4:]}-01",
                            "value": val,
                            "unit": "index",
                            "label": "반도체수출단가지수",
                            "source": "ECOS_EXPORT_PRICE",
                        })
                except Exception:
                    pass
    except Exception:
        pass

    return rows


def collect_semiconductor_export_value() -> list[dict]:
    """
    한국 반도체 수출금액 (월별) — 관세청/산업통상자원부 공개 데이터
    epic:3:2 = 한국 반도체 수출금액 (백만달러)

    출처: 관세청 수출입무역통계 (unipass.customs.go.kr) 또는
          산업통상자원부 월간 수출동향 보도자료
    """
    rows = []
    try:
        # 관세청 수출입통계 공공 API (open.customs.go.kr)
        customs_key = os.environ.get("CUSTOMS_API_KEY", "")
        if not customs_key:
            # API 키 없으면 KITA (무역협회) 공개 통계 시도
            # KITA 자유무역통계 - 반도체 품목 (HS 8542)
            kita_url = "https://stat.kita.net/stat/istat/cts/CtsWholeList.screen"
            # 직접 파싱 복잡 → 한국무역통계진흥원 TRASS API 시도
            pass

        # 대안: ECOS 무역수지통계에서 IT품목 추출
        ecos_url = ("https://ecos.bok.or.kr/api/StatisticSearch/sample/json/kr/"
                    "1/200/902Y013/M/202001/202612/S16")
        r = requests.get(ecos_url, timeout=15)
        if r.ok:
            data = r.json()
            items = data.get("StatisticSearch", {}).get("row", [])
            for item in items:
                try:
                    ym = item.get("TIME", "")
                    val = float(item.get("DATA_VALUE", 0) or 0)
                    if len(ym) == 6 and val > 0:
                        rows.append({
                            "period": f"{ym[:4]}-{ym[4:]}-01",
                            "value": val,
                            "unit": "100mn_usd",
                            "label": "수출(IT)",
                            "source": "ECOS_TRADE",
                        })
                except Exception:
                    pass
    except Exception as e:
        print(f"[반도체수출] 오류: {e}")

    return rows


def collect_korea_semiconductor_export_molit() -> list[dict]:
    """
    산업통상자원부 수출통계 (월간 수출입 동향)에서 반도체 수출 추출
    epic:3:2 대용 — KITA 품목별 수출통계 HTML 스크래핑
    """
    rows = []
    try:
        # KITA 무역통계 - HS 8542 (집적회로) 수출
        # 공개 URL: https://stat.kita.net/stat/istat/cts/ItemImpExpList.screen
        # POST 요청 필요
        import re as _re
        # 산업부 보도자료 대신 KOTRA 무역통계 API 시도 (공개키)
        # 현재 환경에서는 API 키 없이 접근 제한 → KOSIS 수출 통계 사용
        kosis_url = (
            "https://kosis.kr/openapi/statisticsData.do"
            "?method=getList&apiKey=SAMPLE&format=json"
            "&orgId=360&tblId=DT_1R11B01&vwCd=MT_ZTITLE"
            "&prdSe=M&startPrdDe=202001&endPrdDe=202612"
            "&itmId=T1&objL1=A08"
        )
        # API 키 없으면 빈 반환
    except Exception:
        pass
    return rows


def collect_ecos_agricultural_ppi() -> dict:
    """ECOS 생산자물가지수(기본분류) 404Y014에서 농산물 관련 월별 지수 수집.

    수집 대상:
    - epic:11:90  한국 농산물 가격지수 (생산자물가 농림수산품 1AA / 농산물 1011AA)
    - epic:11:91  한국 채소 가격지수  (채소및과실 10112AA / 채소 101121AA)
    - epic:11:92  한국 과일 가격지수  (채소및과실 중 과실 부분)

    Returns dict: indicator_key → list[{period, value, unit, series_name}]
    """
    key = get_ecos_key()
    if not key:
        print("[ECOS-AgriPPI] ECOS API 키 없음")
        return {}

    start_ym = "201001"
    end_ym = date.today().strftime("%Y%m")

    # 관심 item_code → (indicator_key, series_name)
    TARGET_MAP = {
        "1AA":       ("epic:11:90", "농림수산품_총지수"),
        "1011AA":    ("epic:11:90", "농산물_지수"),
        "10112AA":   ("epic:11:91", "채소및과실_지수"),
        "101121AA":  ("epic:11:91", "채소_지수"),
        "1012AA":    ("epic:11:90", "축산물_지수"),
    }

    result: dict = {}
    # 각 item_code별 직접 조회 (전체 조회 시 10000건 한도 초과 방지)
    for item_code, (indicator_key, series_name) in TARGET_MAP.items():
        try:
            url = f"https://ecos.bok.or.kr/api/StatisticSearch/{key}/json/kr/1/500/404Y014/M/{start_ym}/{end_ym}/{item_code}"
            resp = requests.get(url, timeout=15)
            data = resp.json()
            if "StatisticSearch" not in data:
                continue
            rows = data["StatisticSearch"].get("row", [])
            for r in rows:
                period = r.get("TIME", "")
                val_str = r.get("DATA_VALUE", "")
                if not period or not val_str:
                    continue
                try:
                    val = float(val_str)
                except ValueError:
                    continue
                result.setdefault(indicator_key, []).append({
                    "period": period,
                    "series_name": series_name,
                    "value": val,
                    "unit": "지수(2020=100)",
                    "source_name": "ECOS_404Y014",
                    "source_detail": f"한국은행 ECOS 404Y014/{item_code}",
                    "quality": "official_ecos_ppi",
                })
        except Exception as e:
            print(f"[ECOS-AgriPPI] {item_code} 수집 실패: {e}")

    if not result:
        print("[ECOS-AgriPPI] 데이터 없음")
        return {}

    # 채소및과실(10112AA) - 채소(101121AA) = 과실 프록시
    vegetable_fruit = {r["period"]: r["value"] for r in result.get("epic:11:91", []) if r["series_name"] == "채소및과실_지수"}
    vegetable = {r["period"]: r["value"] for r in result.get("epic:11:91", []) if r["series_name"] == "채소_지수"}
    if vegetable_fruit and vegetable:
        fruit_rows = []
        for period, vf_val in vegetable_fruit.items():
            v_val = vegetable.get(period)
            if v_val is not None:
                fruit_rows.append({
                    "period": period,
                    "series_name": "과실_지수_proxy",
                    "value": round(vf_val - v_val + 100, 2),  # 과실 프록시 (근사값)
                    "unit": "지수(2020=100, proxy)",
                    "source_name": "ECOS_404Y014_derived",
                    "source_detail": "ECOS 404Y014 채소및과실 지수와 채소 지수 차이를 이용한 과실 프록시",
                    "quality": "official_ecos_derived_proxy",
                })
        if fruit_rows:
            result["epic:11:92"] = fruit_rows

    for k, v in result.items():
        print(f"[ECOS-AgriPPI] {k}: {len(v)}행")
    return result


def collect_hira_medical_subject_annual_proxy() -> dict[str, list[dict]]:
    """Collect annual HIRA care-cost rows for medical-subject card proxies.

    EPIC labels these as monthly card-spending estimates. We do not have a
    public card-company exact source, so this collector stores a clearly marked
    annual official proxy from HIRA/data.go.kr. Do not relabel these rows as
    exact card spending.
    """
    url = (
        "https://www.data.go.kr/cmm/cmm/fileDownload.do"
        "?atchFileId=FILE_000000003547909&fileDetailSn=1&insertDataPrcus=N"
    )
    try:
        response = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
        response.raise_for_status()
        df = pd.read_csv(BytesIO(response.content), encoding="cp949")
    except Exception as exc:
        print(f"[HIRA medical proxy] 수집 실패: {exc}")
        return {}

    required = {
        "진료년도",
        "진료과목(표시과목)",
        "환자수",
        "명세서청구건수",
        "입내원일수",
        "보험자부담금(선별포함)",
        "요양급여비용총액(선별포함)",
    }
    if not required.issubset(set(df.columns)):
        print("[HIRA medical proxy] 예상 컬럼 불일치")
        return {}

    numeric_cols = [
        "환자수",
        "명세서청구건수",
        "입내원일수",
        "보험자부담금(선별포함)",
        "요양급여비용총액(선별포함)",
    ]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    subject = df["진료과목(표시과목)"].astype(str)
    specs = {
        "epic:16:110": {
            "label": "피부과",
            "mask": subject.eq("피부과"),
            "basis": "진료과목(표시과목)=피부과 exact rows",
        },
        "epic:16:111": {
            "label": "성형외과",
            "mask": subject.eq("성형외과"),
            "basis": "진료과목(표시과목)=성형외과 exact rows",
        },
        "epic:16:112": {
            "label": "치과",
            "mask": subject.str.contains("치과", na=False),
            "basis": "진료과목(표시과목)에 치과 포함 rows",
        },
    }
    metric_specs = {
        "patients": ("환자수", "명"),
        "claims": ("명세서청구건수", "건"),
        "visit_days": ("입내원일수", "일"),
        "insurer_payment": ("보험자부담금(선별포함)", "원"),
        "total_benefit_cost": ("요양급여비용총액(선별포함)", "원"),
    }

    out: dict[str, list[dict]] = {key: [] for key in specs}
    for indicator_key, spec in specs.items():
        sub = df.loc[spec["mask"]].copy()
        if sub.empty:
            continue
        for year, group in sub.groupby("진료년도"):
            period = str(int(year))
            row_count = len(group)
            for series_name, (column, unit) in metric_specs.items():
                value = safe_float(group[column].sum(skipna=True))
                if value is None:
                    continue
                out[indicator_key].append(
                    {
                        "period": period,
                        "series_name": series_name,
                        "value": value,
                        "unit": unit,
                        "source_name": "data.go.kr_HIRA_medical_subject_annual",
                        "source_detail": (
                            "공공데이터포털 건강보험심사평가원 진료과목별 진료 현황 "
                            f"- {spec['label']} / {spec['basis']} / aggregated_rows={row_count}"
                        ),
                        "quality": "official_hira_annual_medical_proxy",
                    }
                )
    total_rows = sum(len(rows) for rows in out.values())
    if total_rows:
        print(f"[HIRA medical proxy] {total_rows}행 수집")
    return out


def collect_gkl_visitors_from_publicdata() -> list[dict]:
    """Collect GKL monthly visitor counts from the official public data CSV.

    The public file is daily branch/gender/nationality visitor counts. We
    aggregate only completed months so a mid-month file refresh never appears
    as a confirmed monthly collapse in the dashboard.
    """
    url = (
        "https://www.data.go.kr/cmm/cmm/fileDownload.do"
        "?atchFileId=FILE_000000003622303&fileDetailSn=1&insertDataPrcus=N"
    )
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://www.data.go.kr/data/15131132/fileData.do?recommendDataYn=Y",
    }
    response = requests.get(url, headers=headers, timeout=40)
    response.raise_for_status()

    df = pd.read_csv(BytesIO(response.content), encoding="cp949")
    if df.empty or "영업일" not in df.columns or "영업장명" not in df.columns:
        return []

    df["영업일"] = pd.to_datetime(df["영업일"], errors="coerce")
    df = df.dropna(subset=["영업일"])
    visitor_cols = [c for c in df.columns if str(c).endswith("고객입장수")]
    if not visitor_cols:
        return []

    for col in visitor_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    df["방문객수"] = df[visitor_cols].sum(axis=1)
    df["period"] = df["영업일"].dt.to_period("M").astype(str)
    month_max = df.groupby("period")["영업일"].max().reset_index()
    month_max["month_end"] = month_max["영업일"].dt.to_period("M").dt.to_timestamp("M")
    complete_periods = set(month_max.loc[month_max["영업일"] == month_max["month_end"], "period"])
    df = df[df["period"].isin(complete_periods)].copy()
    if df.empty:
        return []

    rows: list[dict] = []
    source_detail = "공공데이터포털 그랜드코리아레저(주)_국적별 입장객 수 CSV, 완성 월만 집계"

    def append_row(period: str, series_name: str, value: float) -> None:
        rows.append(
            {
                "period": period,
                "series_name": series_name,
                "value": float(value),
                "unit": "명",
                "source_name": "data.go.kr_GKL_visitor_file",
                "source_detail": source_detail,
                "quality": "official_complete_month",
            }
        )

    for period, g in df.groupby("period", sort=True):
        append_row(period, "전체 입장객", g["방문객수"].sum())
        for branch, bg in g.groupby("영업장명", sort=True):
            append_row(period, f"{branch} 입장객", bg["방문객수"].sum())
        for col in visitor_cols:
            nationality = (
                str(col)
                .replace("고고객입장수", "")
                .replace("고객입장수", "")
                .strip()
            )
            append_row(period, f"{nationality} 입장객", g[col].sum())

    print(f"[GKL Visitors] {len(rows)}행, {min(complete_periods)}~{max(complete_periods)} 완성 월")
    return rows


def collect_paradise_segment_drop_from_ir_excel() -> list[dict]:
    """Collect Paradise monthly drop by segment from official IR Pack Excel.

    EPIC's target is monthly drop by nationality/segment. DART fair-disclosure
    text only has total drop, but Paradise's official Monthly IR Pack Segment
    sheet provides CN VIP, JP VIP, Other VIP, Mass, and Total in KRW mn.
    """
    # Keep newest known file first. The latest official workbook contains the
    # historical monthly table, while older files are retained as a fallback.
    download_ids = ["27479", "23046"]
    headers = {"User-Agent": "Mozilla/5.0"}
    last_error = None
    for file_id in download_ids:
        url = f"https://www.paradise.co.kr/download/{file_id}"
        try:
            response = requests.get(url, headers=headers, timeout=60, verify=False)
            response.raise_for_status()
            if not response.content.startswith(b"PK"):
                continue
            wb = openpyxl.load_workbook(BytesIO(response.content), data_only=True, read_only=True)
            if "Segment" not in wb.sheetnames:
                continue
            ws = wb["Segment"]
            rows: list[dict] = []
            metric_cols = {
                3: "CN VIP 드롭액",
                4: "JP VIP 드롭액",
                5: "Other VIP 드롭액",
                6: "Mass 드롭액",
                7: "Total 드롭액",
            }
            for row_idx in range(1, ws.max_row + 1):
                period_raw = ws.cell(row_idx, 2).value
                if not isinstance(period_raw, datetime):
                    continue
                period = period_raw.strftime("%Y-%m")
                for col_idx, series_name in metric_cols.items():
                    value = safe_float(ws.cell(row_idx, col_idx).value)
                    if value is None:
                        continue
                    rows.append(
                        {
                            "period": period,
                            "series_name": series_name,
                            "value": value,
                            "unit": "백만원",
                            "source_name": "Paradise_Monthly_IR_Pack_XLSX",
                            "source_detail": f"Paradise official Monthly IR Pack Segment sheet; download_id={file_id}; {url}",
                            "quality": "company_ir_excel_official_segment_drop",
                        }
                    )
            if rows:
                print(f"[Paradise segment drop] {len(rows)}행 수집 (download_id={file_id})")
                return rows
        except Exception as exc:
            last_error = exc
            continue
    if last_error:
        print(f"[Paradise segment drop] 수집 실패: {last_error}")
    return []


def collect_dreamtower_visitors_from_ir_excel() -> list[dict]:
    """Collect Jeju Dream Tower casino visitors from Lotte Tour IR Pack XLSX.

    The latest monthly IR pack contains the full historical table. We parse the
    newest Excel first and use older files only for missing historical periods.
    """
    base = "http://ir.lottetour.com"
    page_url = f"{base}/kor/IrPresentationA"
    headers = {"User-Agent": "Mozilla/5.0", "Referer": page_url}
    try:
        page = requests.get(page_url, headers=headers, timeout=25)
        page.raise_for_status()
    except Exception:
        return []
    soup = BeautifulSoup(page.text, "html.parser")
    links: list[str] = []
    for anchor in soup.find_all("a", href=True):
        href = str(anchor.get("href") or "")
        if "드림타워카지노 IR Pack" in href and href.lower().endswith(".xlsx"):
            links.append(href)
    if not links:
        return []

    rows_by_period: dict[str, dict] = {}
    from urllib.parse import quote

    for href in links:
        url = base + quote(href, safe="/:")
        try:
            response = requests.get(url, headers=headers, timeout=35)
            response.raise_for_status()
            wb = openpyxl.load_workbook(BytesIO(response.content), data_only=True, read_only=True)
            if "DreamtowerCasino" not in wb.sheetnames:
                continue
            ws = wb["DreamtowerCasino"]
        except Exception as exc:
            print(f"[DreamtowerVisitors] skip {href}: {exc}")
            continue
        for row in ws.iter_rows(min_row=5, values_only=True):
            period_value = row[1] if len(row) > 1 else None
            visitor_value = row[10] if len(row) > 10 else None
            if not isinstance(period_value, (datetime, pd.Timestamp)):
                continue
            visitors = safe_float(visitor_value)
            if visitors is None:
                continue
            period = period_value.strftime("%Y-%m")
            rows_by_period.setdefault(
                period,
                {
                    "period": period,
                    "series_name": "드림타워 카지노 방문객",
                    "value": visitors,
                    "unit": "명",
                    "source_name": "LotteTour_IR_Pack_XLSX",
                    "source_detail": f"롯데관광개발 IR Pack DreamtowerCasino sheet: {href}",
                    "quality": "company_ir_excel_official",
                },
            )
    rows = [rows_by_period[p] for p in sorted(rows_by_period)]
    if rows:
        print(f"[DreamtowerVisitors] {len(rows)}행, {rows[0]['period']}~{rows[-1]['period']}")
    return rows


def main() -> None:
    conn = sqlite3.connect(DB_PATH, timeout=60)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA busy_timeout=60000")
        init_tables(conn)
        seeded = seed_catalog(conn)

        base_rate_rows = collect_bok_base_rate()
        liquidity_rows = collect_bok_market_liquidity()
        short_rows = collect_monthly_short_balance(conn)
        smp_rows = collect_kpx_monthly_smp()
        online_shopping_rows = collect_kosis_online_shopping_total()
        online_shopping_breakdown_rows = collect_kosis_online_shopping_breakdown()
        video_subscription_proxy_rows = derive_video_subscription_proxy_from_online_breakdown(online_shopping_breakdown_rows)
        music_subscription_proxy_rows = derive_music_subscription_proxy_from_online_breakdown(online_shopping_breakdown_rows)
        online_consumption_proxy_rows = derive_online_consumption_proxies_from_online_breakdown(online_shopping_breakdown_rows)
        retail_store_proxy_rows = collect_kosis_retail_store_sales_proxy()
        pharmacy_goods_proxy_rows = collect_kosis_pharmacy_goods_sales_proxy()
        service_industry_proxy_rows = collect_kosis_service_industry_index_proxy()
        child_education_proxy_rows = derive_child_education_proxy_from_service_index(service_industry_proxy_rows)
        macao_ggr_rows = collect_macao_ggr()
        kto_foreign_visitors_rows = collect_kto_foreign_visitors_purpose_file()
        kto_regional_visitors_rows = collect_kto_regional_visitors_monthly()
        itstat_iptv_rows = collect_itstat_iptv_subscribers_annual()
        eia_us_rig_count_rows = collect_eia_us_rig_count_monthly()
        bdry_shipping_proxy_rows = collect_bdry_dry_bulk_shipping_proxy_monthly()
        kline_dry_bulk_rows = collect_kline_dry_bulk_indices_weekly()
        worldbank_iron_ore_rows = collect_worldbank_iron_ore_monthly()
        steelbenchmarker_china_rows = collect_steelbenchmarker_china_latest()
        sunsirs_china_steel_rows = collect_sunsirs_china_steel_daily()
        dart_casino_rows = collect_dart_casino_monthly(conn)
        seoul_subway_rows = collect_seoul_subway_monthly()
        kric_rail_line_rows = collect_kric_rail_line_passenger_monthly()
        mtrace_pork_price_rows = collect_mtrace_pork_auction_price_monthly()
        skipjack_import_price_rows = collect_skipjack_import_unit_price_from_hs()
        kto_korean_outbound_rows = collect_kto_korean_outbound_transport_monthly()
        modetour_package_rows = collect_modetour_press_release_package_outbound()
        hyundai_domestic_rows, hyundai_us_rows = collect_hyundai_model_indicators()
        kia_us_rows = collect_kia_us_model_sales()
        kama_company_rows, kama_market_share_rows, kama_specific_rows, kama_model_rows, kama_global_rows = collect_kama_auto_company_indicators(start_year=2016)
        semiconductor_export_rows = collect_semiconductor_export_from_hs(conn)
        agri_ppi_rows = collect_ecos_agricultural_ppi()
        hira_medical_proxy_rows = collect_hira_medical_subject_annual_proxy()
        market_breadth_rows = derive_market_breadth_from_price_history(conn)
        ecos_macro_extension_rows = collect_ecos_macro_quant_extensions()
        customs_sector_rows = collect_customs_sector_quant_extensions()
        local_market_structure_rows = collect_local_market_structure_indicators(conn)
        vietnam_country_product_rows = collect_vietnam_country_product_exports()
        gkl_visitor_rows = collect_gkl_visitors_from_publicdata()
        paradise_segment_drop_rows = collect_paradise_segment_drop_from_ir_excel()
        dreamtower_visitor_rows = collect_dreamtower_visitors_from_ir_excel()

        conn.execute("DELETE FROM quant_major_indicator_series WHERE indicator_key IN ('epic:0:1', 'epic:0:14', 'epic:0:55', 'epic:0:57', 'epic:0:2', 'epic:0:4', 'epic:0:17', 'epic:0:19', 'epic:0:20', 'epic:0:21', 'epic:0:112', 'epic:0:113', 'epic:6:18', 'epic:2:98', 'epic:2:22', 'epic:2:23', 'epic:2:93', 'epic:2:94', 'epic:2:95', 'epic:2:96', 'epic:2:97', 'epic:16:113', 'epic:8:14', 'epic:8:15', 'epic:12:5', 'epic:12:6', 'epic:12:10', 'epic:15:11', 'epic:11:155', 'epic:11:156', 'epic:13:20', 'epic:13:21', 'epic:13:22', 'epic:9:13', 'epic:3:70', 'epic:3:71', 'epic:3:97', 'epic:3:98', 'epic:3:36', 'epic:10:11', 'epic:19:50', 'epic:1:37', 'epic:1:25', 'epic:1:26', 'epic:1:27', 'epic:1:28', 'epic:1:29', 'epic:1:30', 'epic:7:14', 'epic:7:15', 'epic:7:16', 'epic:7:17', 'epic:22:9', 'epic:22:10', 'epic:7:36', 'epic:11:105', 'epic:11:69', 'epic:3:2', 'epic:3:20', 'epic:3:21', 'epic:9:24', 'epic:9:37', 'epic:11:90', 'epic:11:91', 'epic:11:92', 'public:21:1', 'public:21:2', 'public:21:3', 'public:21:4', 'public:20:101', 'public:20:102', 'public:20:103', 'public:20:104', 'public:20:105')")
        conn.execute("DELETE FROM quant_major_indicator_series WHERE indicator_key IN ('public:21:5', 'public:21:6', 'public:20:106', 'public:20:107', 'public:20:108')")
        conn.execute("DELETE FROM quant_major_indicator_series WHERE indicator_key LIKE 'public:23:%'")
        conn.commit()

        counts = {
            "catalog_seeded": len(seeded),
            "base_rate_rows": upsert_series(conn, "epic:20:1", base_rate_rows) if base_rate_rows else 0,
            "liquidity_rows": upsert_series(conn, "epic:20:99", liquidity_rows) if liquidity_rows else 0,
            "short_balance_rows": upsert_series(conn, "epic:20:22", short_rows) if short_rows else 0,
            "kpx_smp_rows": upsert_series(conn, "epic:6:18", smp_rows) if smp_rows else 0,
            "kosis_online_shopping_rows": upsert_series(conn, "epic:2:98", online_shopping_rows) if online_shopping_rows else 0,
            "kosis_online_internet_group_rows": upsert_series(conn, "epic:2:22", online_shopping_breakdown_rows.get("epic:2:22", [])) if online_shopping_breakdown_rows.get("epic:2:22") else 0,
            "kosis_online_mobile_group_rows": upsert_series(conn, "epic:2:23", online_shopping_breakdown_rows.get("epic:2:23", [])) if online_shopping_breakdown_rows.get("epic:2:23") else 0,
            "kosis_video_subscription_proxy_rows": upsert_series(conn, "epic:3:97", video_subscription_proxy_rows.get("epic:3:97", [])) if video_subscription_proxy_rows.get("epic:3:97") else 0,
            "kosis_music_subscription_proxy_rows": upsert_series(conn, "epic:3:98", music_subscription_proxy_rows.get("epic:3:98", [])) if music_subscription_proxy_rows.get("epic:3:98") else 0,
            "kosis_online_food_service_proxy_rows": upsert_series(conn, "epic:11:155", online_consumption_proxy_rows.get("epic:11:155", [])) if online_consumption_proxy_rows.get("epic:11:155") else 0,
            "kosis_online_education_goods_proxy_rows": upsert_series(conn, "epic:13:22", online_consumption_proxy_rows.get("epic:13:22", [])) if online_consumption_proxy_rows.get("epic:13:22") else 0,
            "kosis_internet_cosmetics_rows": upsert_series(conn, "epic:8:14", online_shopping_breakdown_rows.get("epic:8:14", [])) if online_shopping_breakdown_rows.get("epic:8:14") else 0,
            "kosis_mobile_cosmetics_rows": upsert_series(conn, "epic:8:15", online_shopping_breakdown_rows.get("epic:8:15", [])) if online_shopping_breakdown_rows.get("epic:8:15") else 0,
            "kosis_internet_apparel_fashion_rows": upsert_series(conn, "epic:12:5", online_shopping_breakdown_rows.get("epic:12:5", [])) if online_shopping_breakdown_rows.get("epic:12:5") else 0,
            "kosis_mobile_apparel_fashion_rows": upsert_series(conn, "epic:12:6", online_shopping_breakdown_rows.get("epic:12:6", [])) if online_shopping_breakdown_rows.get("epic:12:6") else 0,
            "kosis_department_store_proxy_rows": upsert_series(conn, "epic:2:93", retail_store_proxy_rows.get("epic:2:93", [])) if retail_store_proxy_rows.get("epic:2:93") else 0,
            "kosis_discount_super_proxy_rows": upsert_series(conn, "epic:2:94", retail_store_proxy_rows.get("epic:2:94", [])) if retail_store_proxy_rows.get("epic:2:94") else 0,
            "kosis_convenience_proxy_rows": upsert_series(conn, "epic:2:95", retail_store_proxy_rows.get("epic:2:95", [])) if retail_store_proxy_rows.get("epic:2:95") else 0,
            "kosis_duty_free_proxy_rows": upsert_series(conn, "epic:2:96", retail_store_proxy_rows.get("epic:2:96", [])) if retail_store_proxy_rows.get("epic:2:96") else 0,
            "kosis_nonstore_proxy_rows": upsert_series(conn, "epic:2:97", retail_store_proxy_rows.get("epic:2:97", [])) if retail_store_proxy_rows.get("epic:2:97") else 0,
            "kosis_pharmacy_goods_proxy_rows": upsert_series(conn, "epic:16:113", pharmacy_goods_proxy_rows.get("epic:16:113", [])) if pharmacy_goods_proxy_rows.get("epic:16:113") else 0,
            "kosis_food_service_index_proxy_rows": upsert_series(conn, "epic:11:156", service_industry_proxy_rows.get("epic:11:156", [])) if service_industry_proxy_rows.get("epic:11:156") else 0,
            "kosis_education_service_index_proxy_rows": upsert_series(conn, "epic:13:20", service_industry_proxy_rows.get("epic:13:20", [])) if service_industry_proxy_rows.get("epic:13:20") else 0,
            "kosis_child_education_proxy_rows": upsert_series(conn, "epic:13:21", child_education_proxy_rows.get("epic:13:21", [])) if child_education_proxy_rows.get("epic:13:21") else 0,
            "macao_ggr_rows": upsert_series(conn, "epic:9:13", macao_ggr_rows) if macao_ggr_rows else 0,
            "kto_foreign_visitors_rows": upsert_series(conn, "epic:3:70", kto_foreign_visitors_rows) if kto_foreign_visitors_rows else 0,
            "kto_regional_visitors_rows": upsert_series(conn, "epic:3:71", kto_regional_visitors_rows) if kto_regional_visitors_rows else 0,
            "itstat_iptv_subscribers_rows": upsert_series(conn, "epic:10:11", itstat_iptv_rows) if itstat_iptv_rows else 0,
            "eia_us_rig_count_rows": upsert_series(conn, "epic:19:50", eia_us_rig_count_rows) if eia_us_rig_count_rows else 0,
            "kline_bdi_rows": upsert_series(conn, "epic:7:14", kline_dry_bulk_rows.get("epic:7:14", [])) if kline_dry_bulk_rows.get("epic:7:14") else (upsert_series(conn, "epic:7:14", bdry_shipping_proxy_rows) if bdry_shipping_proxy_rows else 0),
            "kline_bci_rows": upsert_series(conn, "epic:7:15", kline_dry_bulk_rows.get("epic:7:15", [])) if kline_dry_bulk_rows.get("epic:7:15") else 0,
            "kline_bpi_rows": upsert_series(conn, "epic:7:16", kline_dry_bulk_rows.get("epic:7:16", [])) if kline_dry_bulk_rows.get("epic:7:16") else 0,
            "kline_bsi_rows": upsert_series(conn, "epic:7:17", kline_dry_bulk_rows.get("epic:7:17", [])) if kline_dry_bulk_rows.get("epic:7:17") else 0,
            "worldbank_iron_ore_rows": upsert_series(conn, "epic:1:37", worldbank_iron_ore_rows) if worldbank_iron_ore_rows else 0,
            "steelbenchmarker_china_hrb_rows": upsert_series(conn, "epic:1:25", steelbenchmarker_china_rows.get("epic:1:25", [])) if steelbenchmarker_china_rows.get("epic:1:25") else 0,
            "steelbenchmarker_china_crc_rows": upsert_series(conn, "epic:1:26", steelbenchmarker_china_rows.get("epic:1:26", [])) if steelbenchmarker_china_rows.get("epic:1:26") else 0,
            "steelbenchmarker_china_plate_rows": upsert_series(conn, "epic:1:27", steelbenchmarker_china_rows.get("epic:1:27", [])) if steelbenchmarker_china_rows.get("epic:1:27") else 0,
            "sunsirs_china_gi_rows": upsert_series(conn, "epic:1:28", sunsirs_china_steel_rows.get("epic:1:28", [])) if sunsirs_china_steel_rows.get("epic:1:28") else 0,
            "steelbenchmarker_china_rebar_rows": upsert_series(conn, "epic:1:29", steelbenchmarker_china_rows.get("epic:1:29", [])) if steelbenchmarker_china_rows.get("epic:1:29") else 0,
            "sunsirs_china_wire_rod_rows": upsert_series(conn, "epic:1:30", sunsirs_china_steel_rows.get("epic:1:30", [])) if sunsirs_china_steel_rows.get("epic:1:30") else 0,
            "paradise_casino_sales_rows": upsert_series(conn, "epic:9:18", dart_casino_rows.get("epic:9:18", [])) if dart_casino_rows.get("epic:9:18") else 0,
            "gkl_visitor_rows": upsert_series(conn, "epic:9:19", gkl_visitor_rows) if gkl_visitor_rows else 0,
            "gkl_casino_sales_rows": upsert_series(conn, "epic:9:20", dart_casino_rows.get("epic:9:20", [])) if dart_casino_rows.get("epic:9:20") else 0,
            "gkl_drop_rows": upsert_series(conn, "epic:9:21", dart_casino_rows.get("epic:9:21", [])) if dart_casino_rows.get("epic:9:21") else 0,
            "gkl_hold_rows": upsert_series(conn, "epic:9:22", dart_casino_rows.get("epic:9:22", [])) if dart_casino_rows.get("epic:9:22") else 0,
            "paradise_drop_rows": upsert_series(conn, "epic:9:23", dart_casino_rows.get("epic:9:23", [])) if dart_casino_rows.get("epic:9:23") else 0,
            "paradise_segment_drop_rows": upsert_series(conn, "epic:9:24", paradise_segment_drop_rows) if paradise_segment_drop_rows else 0,
            "paradise_hold_rows": upsert_series(conn, "epic:9:25", dart_casino_rows.get("epic:9:25", [])) if dart_casino_rows.get("epic:9:25") else 0,
            "dreamtower_casino_sales_rows": upsert_series(conn, "epic:9:35", dart_casino_rows.get("epic:9:35", [])) if dart_casino_rows.get("epic:9:35") else 0,
            "dreamtower_drop_rows": upsert_series(conn, "epic:9:36", dart_casino_rows.get("epic:9:36", [])) if dart_casino_rows.get("epic:9:36") else 0,
            "dreamtower_visitor_rows": upsert_series(conn, "epic:9:37", dreamtower_visitor_rows) if dreamtower_visitor_rows else 0,
            "dreamtower_hold_rows": upsert_series(conn, "epic:9:38", dart_casino_rows.get("epic:9:38", [])) if dart_casino_rows.get("epic:9:38") else 0,
            "seoul_subway_total_rows": upsert_series(conn, "epic:22:9", seoul_subway_rows.get("epic:22:9", [])) if seoul_subway_rows.get("epic:22:9") else 0,
            "seoul_subway_line_rows": upsert_series(conn, "epic:22:10", seoul_subway_rows.get("epic:22:10", [])) if seoul_subway_rows.get("epic:22:10") else 0,
            "kric_rail_line_passenger_rows": upsert_series(conn, "epic:7:36", kric_rail_line_rows) if kric_rail_line_rows else 0,
            "mtrace_pork_price_rows": upsert_series(conn, "epic:11:105", mtrace_pork_price_rows) if mtrace_pork_price_rows else 0,
            "skipjack_import_unit_price_rows": upsert_series(conn, "epic:11:69", skipjack_import_price_rows) if skipjack_import_price_rows else 0,
            "kto_korean_outbound_rows": upsert_series(conn, "epic:3:34", kto_korean_outbound_rows) if kto_korean_outbound_rows else 0,
            "modetour_package_press_rows": upsert_series(conn, "epic:3:36", modetour_package_rows) if modetour_package_rows else 0,
            "global_auto_country_rows": upsert_series(conn, "epic:0:1", kama_global_rows) if kama_global_rows else 0,
            "hyundai_domestic_model_rows": upsert_series(conn, "epic:0:14", hyundai_domestic_rows) if hyundai_domestic_rows else 0,
            "hyundai_us_retail_rows": upsert_series(conn, "epic:0:55", hyundai_us_rows) if hyundai_us_rows else 0,
            "kia_us_model_rows": upsert_series(conn, "epic:0:57", kia_us_rows) if kia_us_rows else 0,
            "kama_company_sales_rows": upsert_series(conn, "epic:0:2", kama_company_rows) if kama_company_rows else 0,
            "kama_market_share_rows": upsert_series(conn, "epic:0:4", kama_market_share_rows) if kama_market_share_rows else 0,
            "kia_domestic_model_rows": upsert_series(conn, "epic:0:17", kama_model_rows["epic:0:17"]) if kama_model_rows["epic:0:17"] else 0,
            "kgm_sales_rows": upsert_series(conn, "epic:0:19", kama_specific_rows["epic:0:19"]) if kama_specific_rows["epic:0:19"] else 0,
            "renault_sales_rows": upsert_series(conn, "epic:0:20", kama_specific_rows["epic:0:20"]) if kama_specific_rows["epic:0:20"] else 0,
            "gm_sales_rows": upsert_series(conn, "epic:0:21", kama_specific_rows["epic:0:21"]) if kama_specific_rows["epic:0:21"] else 0,
            "kgm_domestic_model_rows": upsert_series(conn, "epic:0:112", kama_model_rows["epic:0:112"]) if kama_model_rows["epic:0:112"] else 0,
            "kgm_export_model_rows": upsert_series(conn, "epic:0:113", kama_model_rows["epic:0:113"]) if kama_model_rows["epic:0:113"] else 0,
            "semiconductor_ic_export_rows": upsert_series(conn, "epic:3:2", semiconductor_export_rows.get("epic:3:2", [])),
            "semiconductor_equipment_export_rows": upsert_series(conn, "epic:3:20", semiconductor_export_rows.get("epic:3:20", [])),
            "electronic_parts_export_rows": upsert_series(conn, "epic:3:21", semiconductor_export_rows.get("epic:3:21", [])),
            "ecos_agri_ppi_rows": upsert_series(conn, "epic:11:90", agri_ppi_rows.get("epic:11:90", [])) if agri_ppi_rows.get("epic:11:90") else 0,
            "ecos_vegetable_ppi_rows": upsert_series(conn, "epic:11:91", agri_ppi_rows.get("epic:11:91", [])) if agri_ppi_rows.get("epic:11:91") else 0,
            "ecos_fruit_ppi_proxy_rows": upsert_series(conn, "epic:11:92", agri_ppi_rows.get("epic:11:92", [])) if agri_ppi_rows.get("epic:11:92") else 0,
            "hira_dermatology_proxy_rows": upsert_series(conn, "epic:16:110", hira_medical_proxy_rows.get("epic:16:110", [])) if hira_medical_proxy_rows.get("epic:16:110") else 0,
            "hira_plastic_surgery_proxy_rows": upsert_series(conn, "epic:16:111", hira_medical_proxy_rows.get("epic:16:111", [])) if hira_medical_proxy_rows.get("epic:16:111") else 0,
            "hira_dental_proxy_rows": upsert_series(conn, "epic:16:112", hira_medical_proxy_rows.get("epic:16:112", [])) if hira_medical_proxy_rows.get("epic:16:112") else 0,
            "market_breadth_kospi_rows": upsert_series(conn, "public:21:1", market_breadth_rows.get("public:21:1", [])) if market_breadth_rows.get("public:21:1") else 0,
            "market_breadth_kosdaq_rows": upsert_series(conn, "public:21:2", market_breadth_rows.get("public:21:2", [])) if market_breadth_rows.get("public:21:2") else 0,
            "market_volume_kospi_rows": upsert_series(conn, "public:21:3", market_breadth_rows.get("public:21:3", [])) if market_breadth_rows.get("public:21:3") else 0,
            "market_volume_kosdaq_rows": upsert_series(conn, "public:21:4", market_breadth_rows.get("public:21:4", [])) if market_breadth_rows.get("public:21:4") else 0,
            "ecos_consumer_sentiment_rows": upsert_series(conn, "public:20:101", ecos_macro_extension_rows.get("public:20:101", [])) if ecos_macro_extension_rows.get("public:20:101") else 0,
            "ecos_economic_sentiment_rows": upsert_series(conn, "public:20:102", ecos_macro_extension_rows.get("public:20:102", [])) if ecos_macro_extension_rows.get("public:20:102") else 0,
            "ecos_manufacturing_bsi_rows": upsert_series(conn, "public:20:103", ecos_macro_extension_rows.get("public:20:103", [])) if ecos_macro_extension_rows.get("public:20:103") else 0,
            "ecos_manufacturing_inventory_ratio_rows": upsert_series(conn, "public:20:104", ecos_macro_extension_rows.get("public:20:104", [])) if ecos_macro_extension_rows.get("public:20:104") else 0,
            "ecos_all_industry_production_rows": upsert_series(conn, "public:20:105", ecos_macro_extension_rows.get("public:20:105", [])) if ecos_macro_extension_rows.get("public:20:105") else 0,
            "local_program_market_rows": upsert_series(conn, "public:21:5", local_market_structure_rows.get("public:21:5", [])) if local_market_structure_rows.get("public:21:5") else 0,
            "local_program_stock_concentration_rows": upsert_series(conn, "public:21:6", local_market_structure_rows.get("public:21:6", [])) if local_market_structure_rows.get("public:21:6") else 0,
            "local_credit_foreign_position_rows": upsert_series(conn, "public:20:106", local_market_structure_rows.get("public:20:106", [])) if local_market_structure_rows.get("public:20:106") else 0,
            "local_investor_flow_total_rows": upsert_series(conn, "public:20:107", local_market_structure_rows.get("public:20:107", [])) if local_market_structure_rows.get("public:20:107") else 0,
            "local_short_lending_total_rows": upsert_series(conn, "public:20:108", local_market_structure_rows.get("public:20:108", [])) if local_market_structure_rows.get("public:20:108") else 0,
            "customs_vietnam_apparel_shoes_export_rows": upsert_series(conn, "epic:12:10", vietnam_country_product_rows.get("epic:12:10", [])) if vietnam_country_product_rows.get("epic:12:10") else 0,
            "customs_vietnam_it_export_rows": upsert_series(conn, "epic:15:11", vietnam_country_product_rows.get("epic:15:11", [])) if vietnam_country_product_rows.get("epic:15:11") else 0,
            "customs_auto_finished_rows": upsert_series(conn, "public:23:1", customs_sector_rows.get("public:23:1", [])) if customs_sector_rows.get("public:23:1") else 0,
            "customs_auto_parts_rows": upsert_series(conn, "public:23:2", customs_sector_rows.get("public:23:2", [])) if customs_sector_rows.get("public:23:2") else 0,
            "customs_battery_rows": upsert_series(conn, "public:23:3", customs_sector_rows.get("public:23:3", [])) if customs_sector_rows.get("public:23:3") else 0,
            "customs_memory_semiconductor_rows": upsert_series(conn, "public:23:4", customs_sector_rows.get("public:23:4", [])) if customs_sector_rows.get("public:23:4") else 0,
            "customs_system_semiconductor_rows": upsert_series(conn, "public:23:5", customs_sector_rows.get("public:23:5", [])) if customs_sector_rows.get("public:23:5") else 0,
            "customs_semiconductor_equipment_rows": upsert_series(conn, "public:23:6", customs_sector_rows.get("public:23:6", [])) if customs_sector_rows.get("public:23:6") else 0,
            "customs_shipbuilding_rows": upsert_series(conn, "public:23:7", customs_sector_rows.get("public:23:7", [])) if customs_sector_rows.get("public:23:7") else 0,
            "customs_steel_rows": upsert_series(conn, "public:23:8", customs_sector_rows.get("public:23:8", [])) if customs_sector_rows.get("public:23:8") else 0,
            "customs_cosmetics_rows": upsert_series(conn, "public:23:9", customs_sector_rows.get("public:23:9", [])) if customs_sector_rows.get("public:23:9") else 0,
            "customs_pharma_rows": upsert_series(conn, "public:23:10", customs_sector_rows.get("public:23:10", [])) if customs_sector_rows.get("public:23:10") else 0,
        }
        for indicator_key, rows in sorted(customs_sector_rows.items()):
            if indicator_key in {f"public:23:{idx}" for idx in range(1, 11)}:
                continue
            counts[f"customs_sector_{indicator_key.replace(':', '_')}_rows"] = upsert_series(conn, indicator_key, rows) if rows else 0

        custom_market_specs = [
            (
                "public:21:1",
                1,
                "KOSPI 시장폭: 상승/하락/보합 종목수",
                "KOSPI 보통주 가격이력에서 상승종목수·하락종목수·상승종목비율·중앙수익률을 계산. 커버 500종목 미만 불완전 수집일은 제외.",
            ),
            (
                "public:21:2",
                2,
                "KOSDAQ 시장폭: 상승/하락/보합 종목수",
                "KOSDAQ 보통주 가격이력에서 상승종목수·하락종목수·상승종목비율·중앙수익률을 계산. 커버 500종목 미만 불완전 수집일은 제외.",
            ),
            (
                "public:21:3",
                3,
                "KOSPI 거래량 확산: 신고가/신저가/3배 거래량",
                "KOSPI 보통주 가격이력에서 20일 신고가수·20일 신저가수·거래량 3배 종목수·총거래대금을 계산.",
            ),
            (
                "public:21:4",
                4,
                "KOSDAQ 거래량 확산: 신고가/신저가/3배 거래량",
                "KOSDAQ 보통주 가격이력에서 20일 신고가수·20일 신저가수·거래량 3배 종목수·총거래대금을 계산.",
            ),
        ]
        for indicator_key, sub_code, name, notes in custom_market_specs:
            upsert_custom_catalog(
                conn,
                indicator_key=indicator_key,
                epic_category_code=21,
                epic_sub_code=sub_code,
                epic_indicator_name=name,
                frequency="Daily",
                base_unit="종목/%/억원",
                status="ready_existing",
                replacement_family="market_breadth_volume",
                source_system="local price_history derived",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="derived_from_local_price_history",
                priority="p1",
                notes=notes,
            )

        custom_macro_specs = [
            ("public:20:101", 101, "소비자심리지수 CSI", "ECOS 소비자동향조사 소비자심리지수. 소비/내수 업종과 시장 위험선호 보조 지표."),
            ("public:20:102", 102, "경제심리지수 ESI 순환변동치", "ECOS 경제심리지수 순환변동치. 경기 방향성 및 시장 레짐 보조 지표."),
            ("public:20:103", 103, "제조업 BSI: 업황/신규수주/전망", "ECOS 기업경기조사 제조업 업황실적·신규수주실적·업황전망 BSI."),
            ("public:20:104", 104, "제조업 재고율", "ECOS 제조업 재고율. 재고 부담과 제조업 사이클 판단 보조 지표."),
            ("public:20:105", 105, "전산업생산지수 SA", "ECOS 전산업생산지수(농림어업 제외) 계절조정. 국내 경기 흐름 보조 지표."),
        ]
        for indicator_key, sub_code, name, notes in custom_macro_specs:
            upsert_custom_catalog(
                conn,
                indicator_key=indicator_key,
                epic_category_code=20,
                epic_sub_code=sub_code,
                epic_indicator_name=name,
                frequency="Monthly",
                base_unit="지수/%",
                status="ready_existing",
                replacement_family="ecos_macro_sentiment_cycle",
                source_system="한국은행 ECOS",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_ecos_exact",
                priority="p1",
                notes=notes,
            )

        for spec in LOCAL_MARKET_STRUCTURE_INDICATORS:
            indicator_key = spec["indicator_key"]
            upsert_custom_catalog(
                conn,
                indicator_key=indicator_key,
                epic_category_code=spec["epic_category_code"],
                epic_sub_code=spec["epic_sub_code"],
                epic_indicator_name=spec["epic_indicator_name"],
                frequency=spec["frequency"],
                base_unit=spec["base_unit"],
                status="ready_existing" if local_market_structure_rows.get(indicator_key) else "ready_existing_partial",
                replacement_family=spec["replacement_family"],
                source_system=spec["source_system"],
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness=spec["exactness"],
                priority=spec["priority"],
                notes=spec["notes"],
            )

        for indicator_key, sub_code, name, _label, _prefixes, notes in CUSTOMS_SECTOR_QUANT_SPECS:
            upsert_custom_catalog(
                conn,
                indicator_key=indicator_key,
                epic_category_code=23,
                epic_sub_code=sub_code,
                epic_indicator_name=name,
                frequency="Monthly",
                base_unit="백만달러/USD/kg",
                status="ready_existing",
                replacement_family="customs_sector_trade",
                source_system="관세청 수출입 HS Trade DB",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_customs_hs_prefix_aggregate",
                priority="p1",
                notes=notes + " 기업별 배분값이 아니라 HS 접두어 기준 섹터 총량이다.",
            )

        if hyundai_domestic_rows:
            update_catalog_status(
                conn,
                "epic:0:14",
                status="ready_existing",
                source_system="현대차 IR",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_exact",
                notes="현대차 IR 월간 Sales by Model 엑셀에서 국내 모델별 판매량 직접 수집.",
            )
        if kama_global_rows:
            update_catalog_status(
                conn,
                "epic:0:1",
                status="ready_existing",
                source_system="KAMA 자동차통계월보",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_association_exact",
                notes="KAMA 10-2 주요국신차등록 시트의 국가별 계(총합) 컬럼에서 월간 신차등록을 직접 수집.",
            )
        if hyundai_us_rows:
            update_catalog_status(
                conn,
                "epic:0:55",
                status="ready_existing",
                source_system="현대차 IR",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_exact",
                notes="현대차 IR 월간 US Retail Sales 엑셀에서 미국 모델별 소매판매 직접 수집.",
            )
        if kia_us_rows:
            update_catalog_status(
                conn,
                "epic:0:57",
                status="ready_existing",
                source_system="Kia America Newsroom",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_exact",
                notes="Kia America 월간 Sales By Model export 엑셀에서 미국 모델별 판매량 직접 수집.",
            )
        if kama_company_rows:
            update_catalog_status(
                conn,
                "epic:0:2",
                status="ready_existing",
                source_system="KAMA 자동차통계월보",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_association_exact",
                notes="KAMA 자동차통계월보 1-3 업체별 총괄 시트에서 7개 국내 완성차사의 월별 판매계를 직접 수집.",
            )
        if kama_market_share_rows:
            update_catalog_status(
                conn,
                "epic:0:4",
                status="ready_existing",
                source_system="KAMA 자동차통계월보",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="derived_from_official_sales",
                notes="KAMA 자동차통계월보 업체별 내수 판매를 합산해 회사별 월간 시장점유율로 계산.",
            )
        if kama_model_rows["epic:0:17"]:
            update_catalog_status(
                conn,
                "epic:0:17",
                status="ready_existing",
                source_system="KAMA 자동차통계월보",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_association_exact",
                notes="KAMA 1-4 업체별·모델별 생산·내수·수출 시트의 모델 패밀리 합계행에서 기아 국내 모델별 판매를 월 단위로 수집.",
            )
        if kama_specific_rows["epic:0:19"]:
            update_catalog_status(
                conn,
                "epic:0:19",
                status="ready_existing",
                source_system="KAMA 자동차통계월보",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_association_exact",
                notes="KAMA 업체별 총괄에서 KG모빌리티의 판매계/내수/수출을 월별 수집.",
            )
        if kama_specific_rows["epic:0:20"]:
            update_catalog_status(
                conn,
                "epic:0:20",
                status="ready_existing",
                source_system="KAMA 자동차통계월보",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_association_exact",
                notes="KAMA 업체별 총괄에서 르노코리아(구 르노삼성)의 판매계/내수/수출을 월별 수집.",
            )
        if kama_specific_rows["epic:0:21"]:
            update_catalog_status(
                conn,
                "epic:0:21",
                status="ready_existing",
                source_system="KAMA 자동차통계월보",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_association_exact",
                notes="KAMA 업체별 총괄에서 한국GM(구 한국지엠)의 판매계/내수/수출을 월별 수집.",
            )
        if kama_model_rows["epic:0:112"]:
            update_catalog_status(
                conn,
                "epic:0:112",
                status="ready_existing",
                source_system="KAMA 자동차통계월보",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_association_exact",
                notes="KAMA 1-4 모델 패밀리 합계행에서 KG모빌리티 국내 모델별 판매를 월 단위로 수집.",
            )
        if kama_model_rows["epic:0:113"]:
            update_catalog_status(
                conn,
                "epic:0:113",
                status="ready_existing",
                source_system="KAMA 자동차통계월보",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_association_exact",
                notes="KAMA 1-4 모델 패밀리 합계행에서 KG모빌리티 수출 모델별 판매를 월 단위로 수집.",
            )
        if smp_rows:
            update_catalog_status(
                conn,
                "epic:6:18",
                status="ready_existing",
                source_system="전력거래소(KPX)",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_exact",
                notes="전력거래소 월별 계통한계가격(SMP) 표에서 육지/제주/통합 SMP를 월별로 직접 수집.",
            )
        if online_shopping_rows:
            update_catalog_status(
                conn,
                "epic:2:98",
                status="ready_existing",
                source_system="통계청/KOSIS",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_exact",
                notes="KOSIS 경제상황판 내부 API(selectDetailDataList)에서 온라인쇼핑 거래액 전국 월별 시계열을 직접 수집.",
            )
        if online_shopping_breakdown_rows.get("epic:2:22"):
            update_catalog_status(
                conn,
                "epic:2:22",
                status="ready_existing",
                source_system="통계청/KOSIS",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_exact",
                notes="KOSIS DT_1KE10071 공식 statHtml 표에서 판매매체=인터넷쇼핑, 상품군별 거래액을 월별 직접 수집.",
            )
        if online_shopping_breakdown_rows.get("epic:2:23"):
            update_catalog_status(
                conn,
                "epic:2:23",
                status="ready_existing",
                source_system="통계청/KOSIS",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_exact",
                notes="KOSIS DT_1KE10071 공식 statHtml 표에서 판매매체=모바일쇼핑, 상품군별 거래액을 월별 직접 수집.",
            )
        if online_shopping_breakdown_rows.get("epic:8:14"):
            update_catalog_status(
                conn,
                "epic:8:14",
                status="ready_existing",
                source_system="통계청/KOSIS",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_exact",
                notes="KOSIS DT_1KE10071 공식 statHtml 표에서 인터넷쇼핑/화장품 거래액을 월별 직접 수집.",
            )
        if online_shopping_breakdown_rows.get("epic:8:15"):
            update_catalog_status(
                conn,
                "epic:8:15",
                status="ready_existing",
                source_system="통계청/KOSIS",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_exact",
                notes="KOSIS DT_1KE10071 공식 statHtml 표에서 모바일쇼핑/화장품 거래액을 월별 직접 수집.",
            )
        if online_shopping_breakdown_rows.get("epic:12:5"):
            update_catalog_status(
                conn,
                "epic:12:5",
                status="ready_existing",
                source_system="통계청/KOSIS",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_derived_sum",
                notes="KOSIS 공식 상품군 의복+신발+가방+패션용품 및 액세서리를 인터넷쇼핑 채널별 월 단위 합산.",
            )
        if online_shopping_breakdown_rows.get("epic:12:6"):
            update_catalog_status(
                conn,
                "epic:12:6",
                status="ready_existing",
                source_system="통계청/KOSIS",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_derived_sum",
                notes="KOSIS 공식 상품군 의복+신발+가방+패션용품 및 액세서리를 모바일쇼핑 채널별 월 단위 합산.",
            )
        retail_proxy_specs = {
            "epic:2:93": "백화점 판매액",
            "epic:2:94": "대형마트+슈퍼마켓 및 잡화점 판매액 합산",
            "epic:2:95": "편의점 판매액",
            "epic:2:96": "면세점 판매액",
            "epic:2:97": "무점포 소매 판매액",
        }
        for indicator_key, label in retail_proxy_specs.items():
            if retail_store_proxy_rows.get(indicator_key):
                update_catalog_status(
                    conn,
                    indicator_key,
                    status="ready_existing_partial",
                    source_system="통계청/KOSIS",
                    collector_path="scripts/ops/sync_quant_major_indicators.py",
                    exactness="official_retail_sales_proxy",
                    notes=f"EPIC 카드결제액 추정치 exact 원천은 미확정. 대체지표로 KOSIS DT_1K41003 소매업태별 판매액의 {label}을 월별 공식 프록시로 수집.",
                )
        if pharmacy_goods_proxy_rows.get("epic:16:113"):
            update_catalog_status(
                conn,
                "epic:16:113",
                status="ready_existing_partial",
                source_system="통계청/KOSIS",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_medicine_retail_sales_proxy_not_pharmacy_card_exact",
                notes="EPIC 원 지표는 약국 카드 결제액이나 exact 카드 업종 원천은 미확정. 대체지표로 KOSIS DT_1K41002 재별 및 상품군별 판매액의 의약품 월별 판매액을 공식 프록시로 수집. 약국 카드 결제액 exact로 해석 금지.",
            )

        service_index_specs = {
            "epic:11:156": "숙박 및 음식점업",
            "epic:13:20": "교육 서비스업",
        }
        for indicator_key, label in service_index_specs.items():
            if service_industry_proxy_rows.get(indicator_key):
                update_catalog_status(
                    conn,
                    indicator_key,
                    status="ready_existing_partial",
                    source_system="통계청/KOSIS",
                    collector_path="scripts/ops/sync_quant_major_indicators.py",
                    exactness="official_quarterly_service_index_proxy",
                    notes=f"EPIC 카드결제액 exact 원천은 미확정. 대체지표로 KOSIS DT_1KC2023 시도별 서비스업생산지수의 {label} 경상/불변지수를 분기 공식 프록시로 수집.",
                )
        if child_education_proxy_rows.get("epic:13:21"):
            update_catalog_status(
                conn,
                "epic:13:21",
                status="ready_existing_partial",
                source_system="통계청/KOSIS",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_quarterly_education_service_proxy_not_child_card_exact",
                notes="EPIC 원 지표는 유아교육 카드 결제액이나 exact 카드 업종 원천은 미확정. 대체지표로 KOSIS DT_1KC2023 교육 서비스업 생산지수를 분기 공식 프록시로 수집. 유아교육 또는 카드 결제 exact로 해석 금지.",
            )

        if video_subscription_proxy_rows.get("epic:3:97"):
            update_catalog_status(
                conn,
                "epic:3:97",
                status="ready_existing_partial",
                source_system="통계청/KOSIS",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_online_content_spending_proxy",
                notes="EPIC 원 지표는 영상 구독 서비스 지출건수/금액이나 exact 카드 결제 원천은 미확정. 대체지표로 KOSIS DT_1KE10071 온라인쇼핑 문화 및 레저서비스 거래액(인터넷+모바일)을 월별 공식 프록시로 수집. 음원 구독과 동일시 금지.",
            )
        if music_subscription_proxy_rows.get("epic:3:98"):
            update_catalog_status(
                conn,
                "epic:3:98",
                status="ready_existing_partial",
                source_system="통계청/KOSIS",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_online_content_spending_proxy_not_music_exact",
                notes="EPIC 원 지표는 음원 구독 서비스 지출건수/금액이나 exact 카드/음원 플랫폼 원천은 미확정. 대체지표로 KOSIS DT_1KE10071 온라인쇼핑 문화 및 레저서비스 거래액(인터넷+모바일)을 월별 공식 프록시로 수집. 음원 구독 exact로 해석 금지.",
            )
        online_consumption_proxy_specs = {
            "epic:11:155": (
                "official_online_food_service_proxy",
                "EPIC 원 지표는 제과/커피/패스트푸드 카드 결제액이나 exact 카드 업종 원천은 미확정. 대체지표로 KOSIS DT_1KE10071 온라인쇼핑 음식서비스 거래액(인터넷+모바일)을 월별 공식 프록시로 수집. 외식/배달 포함 가능성이 있어 카드 업종 exact로 해석 금지.",
            ),
            "epic:13:22": (
                "official_online_education_goods_proxy",
                "EPIC 원 지표는 교육용품 카드 결제액이나 exact 카드 업종 원천은 미확정. 대체지표로 KOSIS DT_1KE10071 온라인쇼핑 서적+사무·문구 거래액(인터넷+모바일)을 월별 공식 프록시로 수집. 교육용품 exact로 해석 금지.",
            ),
        }
        for indicator_key, (exactness, notes) in online_consumption_proxy_specs.items():
            if online_consumption_proxy_rows.get(indicator_key):
                update_catalog_status(
                    conn,
                    indicator_key,
                    status="ready_existing_partial",
                    source_system="통계청/KOSIS",
                    collector_path="scripts/ops/sync_quant_major_indicators.py",
                    exactness=exactness,
                    notes=notes,
                )
        if macao_ggr_rows:
            update_catalog_status(
                conn,
                "epic:9:13",
                status="ready_existing",
                source_system="DICJ Macau",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_exact",
                notes="마카오 Gaming Inspection and Coordination Bureau(DICJ) 공식 XML에서 월별 Gross Revenue from Games of Fortune을 수집.",
            )
        if kto_foreign_visitors_rows:
            update_catalog_status(
                conn,
                "epic:3:70",
                status="ready_existing_partial",
                source_system="한국관광공사/data.go.kr",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_partial",
                notes="공공데이터포털 한국관광공사_방한 외래관광객 목적별 월별 집계 CSV에서 2023-08~2024-07 구간을 부분 수집. 장기/최신 구간은 한국관광데이터랩 또는 ODCloud 정상 인증키 필요.",
            )
        if kto_korean_outbound_rows:
            update_catalog_status(
                conn,
                "epic:3:34",
                status="ready_existing_partial",
                source_system="한국관광공사/data.go.kr",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_outbound_travel_demand_proxy_not_modetour_exact",
                notes="EPIC 원 지표는 모두투어 월별 송출객 현황이나 회사별 exact 원천은 미확정. 공식 대체지표로 한국관광공사_국민 해외관광객 교통수단별 월별 집계 2023-08~2024-07의 전체/공항/항구/출국장별 국민 해외관광객 수를 수집. 모두투어 자체 송출객 또는 패키지 송출객으로 해석 금지.",
            )
        if modetour_package_rows:
            update_catalog_status(
                conn,
                "epic:3:36",
                status="ready_existing_partial",
                source_system="모두투어 보도자료/뉴스와이어",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="company_press_release_partial_exact",
                notes="모두투어 보도자료에서 월별 해외 패키지 송출객 숫자가 명확히 기재된 달만 수집. 누락 월은 보간하지 않으며, 회사 보도자료 텍스트 파싱 기반이라 IR 원표 exact 전체 히스토리로 해석 금지.",
            )
        if kto_regional_visitors_rows:
            update_catalog_status(
                conn,
                "epic:3:71",
                status="ready_existing_partial",
                source_system="한국관광공사 DataLab",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_datalab_trend",
                notes="한국관광공사 DataLab 지역별 관광현황 LN_01_01_016에서 월별 17개 시도 방문자수 추세를 수집. DataLab 안내상 총량보다 추세 분석 지표로 활용 권장.",
            )
        if itstat_iptv_rows:
            update_catalog_status(
                conn,
                "epic:10:11",
                status="ready_existing_partial",
                source_system="ICT통계포털/ITSTAT",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_annual_partial",
                notes="EPIC 원 지표는 월별 IPTV 가입자 수이나, 무료 공개 공식 원천으로 ICT통계포털 DT_164_27 유료방송 가입자(단자기준)의 IPTV 소계 연간 시계열을 우선 수집. 월간 exact로 해석 금지.",
            )
        if eia_us_rig_count_rows:
            update_catalog_status(
                conn,
                "epic:19:50",
                status="ready_existing_partial",
                source_system="EIA",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_monthly_proxy",
                notes="Baker Hughes 주간 rig count exact 원천은 접속 제한/타임아웃으로 미연결. EIA NG_ENR_DRILL_S1_M 월간 U.S. rotary rigs in operation을 공식 월간 프록시로 수집.",
            )
        if kline_dry_bulk_rows.get("epic:7:14"):
            dry_bulk_status_specs = {
                "epic:7:14": "BDI(Baltic Dry Index)",
                "epic:7:15": "BCI(Baltic Capesize Index)",
                "epic:7:16": "BPI(Baltic Panamax Index)",
                "epic:7:17": "BSI(Baltic Supramax Index)",
            }
            for indicator_key, label in dry_bulk_status_specs.items():
                if kline_dry_bulk_rows.get(indicator_key):
                    update_catalog_status(
                        conn,
                        indicator_key,
                        status="ready_existing_partial",
                        source_system="K-Line Shipping Market Information",
                        collector_path="scripts/ops/sync_quant_major_indicators.py",
                        exactness="third_party_weekly_index_republication",
                        notes=f"EPIC 원 지표는 Baltic Exchange {label} 일별 지수이나 공식 licensed feed는 미연결. K-Line IR 공개 차트에 포함된 주간 Friday label 시계열을 수집. 제3자 재게시 자료이므로 exact 원천으로 과신 금지.",
                    )
        elif bdry_shipping_proxy_rows:
            update_catalog_status(
                conn,
                "epic:7:14",
                status="ready_existing_partial",
                source_system="Yahoo Finance BDRY",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="market_proxy_monthly",
                notes="EPIC 원 지표는 일별 BDI이나 무료 공개 exact CSV가 현재 차단/불안정. BDRY ETF 월별 조정종가와 거래량을 건화물 운임 방향성 프록시로 수집. BCI/BPI/BSI 개별 지수로 해석 금지.",
            )
        if worldbank_iron_ore_rows:
            update_catalog_status(
                conn,
                "epic:1:37",
                status="ready_existing_partial",
                source_system="World Bank Pink Sheet",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_proxy_monthly",
                notes="World Bank Commodity Price Data의 Iron ore, cfr spot 월간 가격을 수집. EPIC 원 지표의 주간 China import price와 exact는 아니므로 공식 월간 프록시로 사용.",
            )

        for indicator_key, label in [
            ("epic:1:25", "China HRB/HRC"),
            ("epic:1:26", "China CRC"),
            ("epic:1:27", "China Standard Plate"),
            ("epic:1:29", "China Rebar"),
        ]:
            if steelbenchmarker_china_rows.get(indicator_key):
                update_catalog_status(
                    conn,
                    indicator_key,
                    status="ready_existing_partial",
                    source_system="SteelBenchmarker",
                    collector_path="scripts/ops/sync_quant_major_indicators.py",
                    exactness="public_report_latest_only",
                        notes=f"SteelBenchmarker 공개 PDF에서 최신 Mainland China {label} 가격 1포인트를 수집. 장기 시계열은 PDF history 표 파서 검증 후 확장 필요.",
                    )

        sunsirs_status_specs = {
            "epic:1:28": "China Galvanized sheet(G.I), HDG/DX51D+Z/1.0*1250*C",
            "epic:1:30": "China Wire Rod, HPB235/Φ8",
        }
        for indicator_key, label in sunsirs_status_specs.items():
            if sunsirs_china_steel_rows.get(indicator_key):
                update_catalog_status(
                    conn,
                    indicator_key,
                    status="ready_existing_partial",
                    source_system="SunSirs",
                    collector_path="scripts/ops/sync_quant_major_indicators.py",
                    exactness="third_party_recent_public_daily_price",
                    notes=(
                        f"SunSirs 공개 China Commodity Data Group 페이지에서 {label} 최근 일별 spot price를 수집. "
                        "로그인 없이 공개되는 최근 7일 내외만 적재 가능하므로 장기 exact 시계열은 아님. "
                        "수집 실패 시 기존 히스토리 보존 필요."
                    ),
                )

        hira_medical_proxy_specs = {
            "epic:16:110": "피부과",
            "epic:16:111": "성형외과",
            "epic:16:112": "치과",
        }
        for indicator_key, label in hira_medical_proxy_specs.items():
            if hira_medical_proxy_rows.get(indicator_key):
                update_catalog_status(
                    conn,
                    indicator_key,
                    status="ready_existing_partial",
                    source_system="공공데이터포털/HIRA",
                    collector_path="scripts/ops/sync_quant_major_indicators.py",
                    exactness="official_annual_medical_proxy_not_card_exact",
                    notes=(
                        f"EPIC 원 지표는 {label} 월별 카드 결제액 추정치이나 exact 카드사 원천은 미확정. "
                        "공식 대체지표로 건강보험심사평가원 진료과목별 진료 현황 2024년 연간 "
                        "환자수/청구건수/입내원일수/보험자부담금/요양급여비용총액을 수집. "
                        "카드 소비액 또는 월별 지표로 해석 금지."
                    ),
                )

        casino_status_specs = {
            "epic:9:18": ("파라다이스 월별 카지노 매출액", "official_disclosure_exact"),
            "epic:9:20": ("GKL 월별 카지노 매출액", "official_disclosure_exact"),
            "epic:9:21": ("GKL 월별 테이블 드롭액", "official_disclosure_exact"),
            "epic:9:22": ("GKL 월별 홀드율=카지노매출액/테이블드롭액", "derived_from_official_disclosure"),
            "epic:9:23": ("파라다이스 월별 테이블 드롭액", "official_disclosure_exact"),
            "epic:9:25": ("파라다이스 월별 홀드율=카지노매출액/테이블드롭액", "derived_from_official_disclosure"),
            "epic:9:35": ("드림타워 카지노 월별 카지노 매출액", "official_disclosure_exact"),
            "epic:9:36": ("드림타워 카지노 월별 테이블 드롭액", "official_disclosure_exact"),
            "epic:9:38": ("드림타워 카지노 월별 홀드율=카지노매출액/테이블드롭액", "derived_from_official_disclosure"),
        }
        for indicator_key, (label, exactness) in casino_status_specs.items():
            if dart_casino_rows.get(indicator_key):
                update_catalog_status(
                    conn,
                    indicator_key,
                    status="ready_existing",
                    source_system="DART 영업잠정실적 공정공시",
                    collector_path="scripts/ops/sync_quant_major_indicators.py",
                    exactness=exactness,
                    notes=f"{label}을 DART 영업잠정실적 공시 원문에서 월별 수집. 입장객은 해당 공시 본문에 없어 별도 IR 소스 확보 전까지 미수집.",
                )
        if gkl_visitor_rows:
            update_catalog_status(
                conn,
                "epic:9:19",
                status="ready_existing",
                source_system="공공데이터포털/GKL",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_publicdata_daily_aggregate",
                notes="공공데이터포털 그랜드코리아레저(주)_국적별 입장객 수 CSV를 내려받아 영업일 기준 완성 월만 월별 합산. 전체/영업장별/국적별 입장객을 저장하며, 월중 부분 데이터는 제외.",
            )
        if paradise_segment_drop_rows:
            update_catalog_status(
                conn,
                "epic:9:24",
                status="ready_existing",
                source_system="파라다이스 Monthly IR Pack",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="company_ir_excel_exact_segment_drop",
                notes=(
                    "파라다이스 공식 Monthly IR Pack 엑셀 Segment 시트에서 CN VIP/JP VIP/Other VIP/Mass/Total "
                    "월별 드롭액(KRW mn=백만원)을 직접 수집. DART 공시의 총 드롭액과 교차 확인 가능하며, "
                    "IR 엑셀 실패 시 기존 히스토리 보존 필요."
                ),
            )
        if dreamtower_visitor_rows:
            update_catalog_status(
                conn,
                "epic:9:37",
                status="ready_existing",
                source_system="롯데관광개발 IR Pack",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="company_ir_excel_exact",
                notes="롯데관광개발 IR 자료실의 최신 드림타워카지노 IR Pack 엑셀 DreamtowerCasino 시트에서 카지노 방문객 월별 값을 직접 수집. 최신 파일이 과거 전체 테이블을 포함하므로 최신 파일 우선, 과거 파일은 결측 보강용.",
            )

        if seoul_subway_rows.get("epic:22:10"):
            update_catalog_status(
                conn,
                "epic:22:10",
                status="ready_existing",
                source_system="서울 열린데이터광장",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_exact",
                notes="서울 열린데이터광장 OA-12914 지하철호선별 역별 승하차 인원 CSV를 월별/노선별로 합산 수집. 2023-01~최신 공개월 커버.",
            )
        if seoul_subway_rows.get("epic:22:9"):
            update_catalog_status(
                conn,
                "epic:22:9",
                status="ready_existing_partial",
                source_system="서울 열린데이터광장",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_partial_subway_only",
                notes="대중교통 전체 지표 중 지하철 승하차 합계만 서울 열린데이터광장 OA-12914로 우선 수집. 버스/기타 교통수단은 별도 공식 소스 연결 필요.",
            )
        if kric_rail_line_rows:
            update_catalog_status(
                conn,
                "epic:7:36",
                status="ready_existing_partial",
                source_system="철도산업정보센터/KRIC",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_korail_general_rail_partial",
                notes="철도산업정보센터 노선별 여객수송(월) HTML 표에서 한국철도공사 일반철도 노선별/열차종별 수송인원을 월별 수집. 도시철도·민자 전체 철도까지 포함한 완전 커버리지는 아니므로 부분연결로 표시.",
            )
        if mtrace_pork_price_rows:
            update_catalog_status(
                conn,
                "epic:11:105",
                status="ready_existing_partial",
                source_system="축산물품질평가원/MTRACE",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_monthly_proxy",
                notes="EPIC 원 지표는 일별 돼지고기 도매가격이나, 공개 공식 원천으로 MTRACE DT_APGS_016 돼지도체 도매시장별 등급별 경락가격의 전체 등급/전체 도매시장 월별 경락가격(원/kg)을 부분 프록시로 수집.",
            )
        if skipjack_import_price_rows:
            update_catalog_status(
                conn,
                "epic:11:69",
                status="ready_existing_partial",
                source_system="관세청 수출입/HS Trade Lab",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_customs_unit_price_proxy",
                notes="EPIC 원 지표는 가다랑어 어가추이나 무료 공식 어가 원천은 미확정. 대체지표로 HS 0303430000(냉동 가다랑어) 및 0302330000(신선/냉장 가다랑어)의 월별 수입금액/수입중량 단가(USD/kg)를 공식 관세청 수출입 프록시로 수집. 신선/냉장은 거래가 드물어 기본 판단은 냉동 시리즈 우선.",
            )
        vietnam_status_specs = {
            "epic:12:10": (
                "베트남 의류·신발 수출",
                "의류/의류부속품/신발류 성질통합분류 6개 코드",
            ),
            "epic:15:11": (
                "베트남 IT제품 수출",
                "컴퓨터/통신기기/반도체 성질통합분류 13개 코드",
            ),
        }
        for indicator_key, (label, code_desc) in vietnam_status_specs.items():
            if vietnam_country_product_rows.get(indicator_key):
                update_catalog_status(
                    conn,
                    indicator_key,
                    status="ready_existing_partial",
                    source_system="관세청 nnewtempertrade",
                    collector_path="scripts/ops/sync_quant_major_indicators.py",
                    exactness="official_country_product_export_proxy_not_vietnam_global_export",
                    notes=(
                        f"EPIC 원 지표 '{label}'의 글로벌 베트남 수출 exact 원천은 미확정. "
                        f"공식 대체지표로 관세청 nnewtempertrade에서 cntyCd=VN, imexTpcd=수출, {code_desc}를 월별 합산. "
                        "이는 한국 기준 베트남향 수출액/중량/단가이며 베트남 전체 글로벌 수출로 해석 금지."
                    ),
                )
        if agri_ppi_rows.get("epic:11:90"):
            update_catalog_status(
                conn, "epic:11:90",
                status="ready_existing",
                source_system="한국은행ECOS/404Y014",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_ppi_proxy",
                notes="한국은행 ECOS 생산자물가지수(기본분류) 404Y014에서 농림수산품(1AA)/농산물(1011AA)/축산물(1012AA) 월별 지수(2020=100) 수집. KAMIS 직접 가격이 아닌 PPI 프록시.",
            )
        if agri_ppi_rows.get("epic:11:91"):
            update_catalog_status(
                conn, "epic:11:91",
                status="ready_existing",
                source_system="한국은행ECOS/404Y014",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_ppi_proxy",
                notes="한국은행 ECOS 생산자물가지수(기본분류) 404Y014에서 채소및과실(10112AA)/채소(101121AA) 월별 지수(2020=100) 수집.",
            )
        if agri_ppi_rows.get("epic:11:92"):
            update_catalog_status(
                conn, "epic:11:92",
                status="ready_existing_partial",
                source_system="한국은행ECOS/404Y014_derived",
                collector_path="scripts/ops/sync_quant_major_indicators.py",
                exactness="official_ppi_derived_proxy",
                notes="채소및과실 지수 - 채소 지수 + 100 으로 근사한 과실 지수 프록시. 직접 과실 단독 지수 수집 불가로 파생.",
            )

        latest = {
            "base_rate_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:20:1' order by period desc limit 1"
            ).fetchone(),
            "liquidity_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:20:99' order by period desc, series_name asc limit 2"
            ).fetchall(),
            "short_balance_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:20:22' order by period desc, series_name asc limit 2"
            ).fetchall(),
            "smp_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:6:18' order by period desc, series_name asc limit 6"
            ).fetchall(),
            "online_shopping_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:2:98' order by period desc limit 6"
            ).fetchall(),
            "online_shopping_breakdown_latest": conn.execute(
                "select indicator_key, period, series_name, value from quant_major_indicator_series where indicator_key in ('epic:2:22','epic:2:23','epic:8:14','epic:8:15','epic:12:5','epic:12:6') order by indicator_key asc, period desc, series_name asc limit 18"
            ).fetchall(),
            "retail_store_proxy_latest": conn.execute(
                "select indicator_key, period, series_name, value from quant_major_indicator_series where indicator_key in ('epic:2:93','epic:2:94','epic:2:95','epic:2:96','epic:2:97') order by indicator_key asc, period desc, series_name asc limit 15"
            ).fetchall(),
            "service_industry_proxy_latest": conn.execute(
                "select indicator_key, period, series_name, value from quant_major_indicator_series where indicator_key in ('epic:11:156','epic:13:20') order by indicator_key asc, period desc, series_name asc limit 12"
            ).fetchall(),
            "macao_ggr_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:9:13' order by period desc limit 6"
            ).fetchall(),
            "kto_foreign_visitors_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:3:70' order by period desc, series_name asc limit 12"
            ).fetchall(),
            "kto_regional_visitors_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:3:71' order by period desc, series_name asc limit 17"
            ).fetchall(),
            "eia_us_rig_count_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:19:50' order by period desc limit 6"
            ).fetchall(),
            "worldbank_iron_ore_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:1:37' order by period desc limit 6"
            ).fetchall(),
            "steelbenchmarker_china_latest": conn.execute(
                "select indicator_key, period, series_name, value from quant_major_indicator_series where indicator_key in ('epic:1:25','epic:1:26','epic:1:27','epic:1:29') order by indicator_key asc, period desc"
            ).fetchall(),
            "dart_casino_latest": conn.execute(
                "select indicator_key, period, series_name, value from quant_major_indicator_series where indicator_key in ('epic:9:18','epic:9:20','epic:9:21','epic:9:22','epic:9:23','epic:9:24','epic:9:25','epic:9:35','epic:9:36','epic:9:38') order by indicator_key asc, period desc limit 32"
            ).fetchall(),
            "seoul_subway_latest": conn.execute(
                "select indicator_key, period, series_name, value from quant_major_indicator_series where indicator_key in ('epic:22:9','epic:22:10') order by indicator_key asc, period desc, series_name asc limit 18"
            ).fetchall(),
            "kric_rail_line_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:7:36' order by period desc, series_name asc limit 18"
            ).fetchall(),
            "mtrace_pork_price_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:11:105' order by period desc limit 6"
            ).fetchall(),
            "skipjack_import_unit_price_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:11:69' order by period desc, series_name asc limit 9"
            ).fetchall(),
            "hyundai_domestic_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:0:14' order by period desc, series_name asc limit 5"
            ).fetchall(),
            "global_auto_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:0:1' order by period desc, series_name asc limit 10"
            ).fetchall(),
            "hyundai_us_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:0:55' order by period desc, series_name asc limit 5"
            ).fetchall(),
            "kia_us_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:0:57' order by period desc, series_name asc limit 5"
            ).fetchall(),
            "company_sales_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:0:2' order by period desc, series_name asc limit 6"
            ).fetchall(),
            "market_share_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:0:4' order by period desc, series_name asc limit 6"
            ).fetchall(),
            "kia_domestic_model_latest": conn.execute(
                "select period, series_name, value from quant_major_indicator_series where indicator_key='epic:0:17' order by period desc, value desc limit 8"
            ).fetchall(),
            "kgm_model_latest": conn.execute(
                "select indicator_key, period, series_name, value from quant_major_indicator_series where indicator_key in ('epic:0:112','epic:0:113') order by indicator_key asc, period desc, value desc limit 12"
            ).fetchall(),
        }
        print(json.dumps({"ok": True, "counts": counts, "latest": latest}, ensure_ascii=False, indent=2, default=list))
    finally:
        conn.close()


if __name__ == "__main__":
    main()
